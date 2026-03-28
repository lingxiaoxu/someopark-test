import numpy as np
import statsmodels.api as sm
import statsmodels.tsa.stattools as ts
import pandas as pd
import datetime
from datetime import datetime, time
import yfinance as yf

import logging

import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.ticker import FuncFormatter

import openpyxl
from openpyxl.utils import get_column_letter
import os

from pykalman import KalmanFilter
import pandas_market_calendars as mcal
from collections import defaultdict

import quantstats as qs
import math

# Set up logging
logging.basicConfig(level=logging.INFO)
log = logging.getLogger(__name__)

class CurrentData:
    def __init__(self, data_slice):
        self.data_slice = data_slice

    def history(self, symbols, field, bar_count, frequency):
        if field != "price":
            raise ValueError("Only 'price' field is supported in this implementation")
        if frequency != "1d":
            raise ValueError("Only '1d' frequency is supported in this implementation")
        return self.data_slice.loc[:, symbols].tail(bar_count)


class date_rules:
    @staticmethod
    def every_day():
        return lambda date: True


class time_rules:
    @staticmethod
    def market_close(minutes):
        return lambda date: True  # For simplicity, always return True


class Trade:
    def __init__(self, date, symbol, amount, price, order_type, pair=None):
        self.date = date
        self.symbol = symbol
        self.amount = amount
        self.price = price
        self.order_type = order_type  # 'open' or 'close'
        self.direction = 'long' if amount > 0 else 'short'
        self.pair = pair  # pair key e.g. "CL/WST", set at order time


class Portfolio:
    def __init__(self, initial_cash, initial_loan, interest_rate, strategy_pairs):

        self.initial_cash = initial_cash
        self.initial_loan = initial_loan

        self.processed_dates = []

        self.strategy_pairs = strategy_pairs
        self.positions = {}
        self.trades = []
        self.volatilities_in_window = {}
        self.normalized_volatility = None
        self.current_date = None
        self.signal_history = {}
        self.interest_rate = interest_rate

        # Balance sheet categories
        self.asset_cash = self.initial_cash + self.initial_loan
        self.asset_securities = {}
        self.liability_securities = {}
        self.liability_loan = self.initial_loan

        # New history lists
        self.asset_cash_history = []
        self.asset_securities_history = {}
        self.liability_securities_history = {}
        self.liability_loan_history = []

        # History
        self.statistical_test_history = {}
        
        # Track warnings that have been shown
        self.warnings_shown = set()
        self.hedge_history = {}
        self.asset_history = []
        self.liability_history = []
        self.equity_history = []
        self.value_history = []
        self.price_history = {}
        self.share_history = {}
        self.percentage_history = {}
        self.pair_trade_history = {f"{pair[0]}/{pair[1]}": [] for pair in strategy_pairs}

        self.finished_trades_pnl = {}
        self.cost_basis_history = {}
        self.total_cost_history = {}

        # Per-pair parallel tables: {pair: {symbol: ...}}
        self.share_history_by_pair = {}
        self.cost_basis_history_by_pair = {}
        self.finished_trades_pnl_by_pair = {}
        self.total_cost_history_by_pair = {}

        # Per-pair current net position tracker (updated on each trade, used to build share_history_by_pair)
        self._pair_positions = {}  # {pair: {symbol: current_net_shares}}

        # Per-pair percentage history: {pair: {symbol: [(date, pct)]}}
        # pct = signed market value of symbol within pair / sum of abs market values of both legs
        self.percentage_history_by_pair = {}

        self.acc_security_pnl_history_by_pair = {}  # {date: {pair: {symbol: {pnl_dollar, pnl_percent}}}}

        self.interest_expense_history = []
        self.daily_pnl_history = []

        self.acc_security_pnl_history = {}
        self.acc_pair_trade_pnl_history = {}
        self.dod_security_pnl_history = {}
        self.dod_pair_trade_pnl_history = {}

        self.acc_interest_history = []
        self.acc_daily_pnl_history = []

        self.max_drawdown_history = []
        self.peak_value = initial_cash  # Initial equity (total asset - total loan)
        self.leverage = initial_loan / initial_cash  # Initial leverage

    def update_price_history(self, date, prices):
        self.current_date = date
        for symbol, price in prices.items():
            if symbol not in self.price_history:
                self.price_history[symbol] = []
            self.price_history[symbol].append((date, price))

    def update_values(self, prices):
        self.update_values_and_histories(prices)

    def update_values_and_histories(self, prices):
        # Calculate daily interest expense
        daily_interest = self.liability_loan * (self.interest_rate / 365)

        # Deduct interest expense from asset_cash
        self.asset_cash -= daily_interest

        # current_date = self.current_date
        # if current_date >= np.datetime64('2023-07-03T00:00:00.000000000'):
        #     print("it is time")
        # else:
        #     print("not yet")

        # Calculate total asset and liability values
        total_asset = self.asset_cash
        total_liability = self.liability_loan

        for symbol, current_shares in self.positions.items():
            if symbol in prices:
                current_price = prices[symbol]

                # Update price history if needed
                if (symbol not in self.price_history or
                        not self.price_history[symbol] or
                        self.price_history[symbol][-1][0] != self.current_date):
                    self.update_price_history(self.current_date, {symbol: current_price})

                # Update share history
                if symbol not in self.share_history:
                    self.share_history[symbol] = []
                self.share_history[symbol].append((self.current_date, current_shares))
                # Sync share_history_by_pair from _pair_positions (per-pair net position, not broadcast)
                for pair in self.pair_trade_history:
                    if symbol in pair.split('/'):
                        pair_shares = self._pair_positions.get(pair, {}).get(symbol, 0)
                        if pair not in self.share_history_by_pair:
                            self.share_history_by_pair[pair] = {}
                        if symbol not in self.share_history_by_pair[pair]:
                            self.share_history_by_pair[pair][symbol] = []
                        self.share_history_by_pair[pair][symbol].append((self.current_date, pair_shares))

                # Add to total asset or liability based on existing position
                if current_shares > 0:  # Long position
                    total_asset += current_shares * current_price
                    self.asset_securities[symbol] = current_shares * current_price
                    if symbol not in self.asset_securities_history:
                        self.asset_securities_history[symbol] = []
                    self.asset_securities_history[symbol].append((self.current_date, current_shares * current_price))
                elif current_shares < 0:  # Short position
                    total_liability += abs(current_shares) * current_price
                    self.liability_securities[symbol] = abs(current_shares) * current_price
                    if symbol not in self.liability_securities_history:
                        self.liability_securities_history[symbol] = []
                    self.liability_securities_history[symbol].append((self.current_date, abs(current_shares) * current_price))
            else:
                log.warning(f"No price data available for {symbol}")

        # Calculate net equity
        net_equity = total_asset - total_liability

        # Calculate and update daily P&L
        if len(self.equity_history) >= 1:
            prev_equity = self.equity_history[-1][1]
            daily_pnl = net_equity - prev_equity
        else:
            daily_pnl = -daily_interest
        self.daily_pnl_history.append((self.current_date, daily_pnl))

        # Update histories
        self.asset_cash_history.append((self.current_date, self.asset_cash))
        self.liability_loan_history.append((self.current_date, self.liability_loan))
        self.asset_history.append((self.current_date, total_asset))
        self.liability_history.append((self.current_date, total_liability))
        self.equity_history.append((self.current_date, net_equity))
        self.value_history.append((self.current_date, net_equity))

        # Update interest expense history
        self.interest_expense_history.append((self.current_date, daily_interest))
        if not self.acc_interest_history:
            self.acc_interest_history.append((self.current_date, daily_interest))
        else:
            prev_acc_interest = self.acc_interest_history[-1][1]
            self.acc_interest_history.append((self.current_date, prev_acc_interest + daily_interest))

        # Update accumulated daily P&L
        if not self.acc_daily_pnl_history:
            self.acc_daily_pnl_history.append((self.current_date, daily_pnl))
        else:
            prev_acc_pnl = self.acc_daily_pnl_history[-1][1]
            self.acc_daily_pnl_history.append((self.current_date, prev_acc_pnl + daily_pnl))

        # Update percentage history
        for symbol in self.positions:
            if symbol not in self.percentage_history:
                self.percentage_history[symbol] = []
            if net_equity > 0:
                percentage = (self.positions[symbol] * prices.get(symbol, 0)) / net_equity
            else:
                percentage = 0
            self.percentage_history[symbol].append((self.current_date, percentage))

        # Update per-pair percentage history
        # pct[sym] = signed_value(sym) / (|val_s1| + |val_s2|) within each pair
        for pair in self.pair_trade_history:
            stock1, stock2 = pair.split('/')
            pos1 = self._pair_positions.get(pair, {}).get(stock1, 0)
            pos2 = self._pair_positions.get(pair, {}).get(stock2, 0)
            price1 = prices.get(stock1, 0)
            price2 = prices.get(stock2, 0)
            val1 = pos1 * price1
            val2 = pos2 * price2
            total_abs = abs(val1) + abs(val2)
            if pair not in self.percentage_history_by_pair:
                self.percentage_history_by_pair[pair] = {}
            for sym, val in ((stock1, val1), (stock2, val2)):
                if sym not in self.percentage_history_by_pair[pair]:
                    self.percentage_history_by_pair[pair][sym] = []
                pct = val / total_abs if total_abs != 0 else 0
                self.percentage_history_by_pair[pair][sym].append((self.current_date, pct))

    def update_pnl_history(self, portfolio_analysis, data, symbols):
        current_prices = data.history(symbols, "Adj Close", 1, "1d").iloc[-1]

        # Calculate accumulated security PnL
        acc_security_pnl = portfolio_analysis.calculate_acc_security_pnl(current_prices)
        self.acc_security_pnl_history[self.current_date] = acc_security_pnl

        # Calculate accumulated pair trade PnL
        acc_pair_trade_pnl = portfolio_analysis.calculate_acc_pair_trade_pnl(current_prices)
        self.acc_pair_trade_pnl_history[self.current_date] = acc_pair_trade_pnl

        total_security_pnl, total_pair_pnl = portfolio_analysis.reconcile_pnls(acc_security_pnl, acc_pair_trade_pnl)

        current_date = self.current_date
        #if current_date >= np.datetime64('2023-07-05T00:00:00.000000000'):
        #    print("it is time")
        #else:
        #    print("not yet")

        # Calculate DoD PnL
        self.dod_security_pnl_history = portfolio_analysis.calculate_dod_security_pnl()
        self.dod_pair_trade_pnl_history = portfolio_analysis.calculate_dod_pair_trade_pnl()

        return self.daily_pnl_history[-1][1]  # Return the latest daily P&L

    def update_max_drawdown(self):
        # Calculate the current portfolio value (equity)
        total_asset = self.asset_cash + sum(self.asset_securities.values())
        total_liability = self.liability_loan + sum(self.liability_securities.values())
        current_value = total_asset - total_liability

        # Calculate leverage, handling cases where current_value is zero or negative
        if current_value != 0:
            self.leverage = abs(self.liability_loan / current_value)
        else:
            self.leverage = float('inf')  # Handle undefined leverage case

        # Find the peak and trough in equity history
        peak_value = 0
        max_drawdown_percent = 0

        for date, equity in self.equity_history:
            if equity > peak_value:
                peak_value = equity
            drawdown_dollar = peak_value - equity
            drawdown_percent = drawdown_dollar / peak_value if peak_value != 0 else 0

            # Adjust drawdown for leverage
            leveraged_drawdown_percent = self.leverage * drawdown_percent

            if leveraged_drawdown_percent > max_drawdown_percent:
                max_drawdown_percent = leveraged_drawdown_percent

        # Update max drawdown history
        if not self.max_drawdown_history or max_drawdown_percent > self.max_drawdown_history[-1][2]:
            self.max_drawdown_history.append((self.current_date, drawdown_dollar, max_drawdown_percent))
        else:
            self.max_drawdown_history.append(
                (self.current_date, self.max_drawdown_history[-1][1], self.max_drawdown_history[-1][2])
            )

    def record_trade(self, trade, pair_key=None):
        self.trades.append(trade)
        if trade.symbol not in self.signal_history:
            self.signal_history[trade.symbol] = []
        self.signal_history[trade.symbol].append((self.current_date, trade.direction))

        # Update pair trade history — use pair_key hint when available (fixes shared-symbol attribution)
        pair = self.get_pair_for_symbol(trade.symbol, hint_pair=pair_key)
        if pair:
            self.pair_trade_history[pair].append(trade)
            # Update per-pair net position tracker
            if pair not in self._pair_positions:
                self._pair_positions[pair] = {}
            prev = self._pair_positions[pair].get(trade.symbol, 0)
            self._pair_positions[pair][trade.symbol] = prev + trade.amount

    def get_pair_for_symbol(self, symbol, hint_pair=None):
        """Return the pair key for the given symbol.

        If hint_pair is supplied (e.g. "CL/GD"), use it when the symbol
        belongs to that pair — this correctly handles symbols that appear in
        multiple pairs (e.g. CL in CL/WST, CL/GD, CL/SRE).
        Falls back to the first matching pair when no valid hint is given.
        """
        if hint_pair:
            parts = hint_pair.split('/')
            if symbol in parts:
                return hint_pair
        for pair in self.strategy_pairs:
            if symbol in pair[:2]:  # Check only the first two elements (stock symbols)
                return f"{pair[0]}/{pair[1]}"
        return None


# ~~~~~~~~~~~~~~~~~~~~~~ TESTS FOR FINDING PAIR TO TRADE ON ~~~~~~~~~~~~~~~~~~~~~~
class ADF(object):
    """
    Augmented Dickey–Fuller (ADF) unit root test
    Source: http://www.pythonforfinance.net/2016/05/09/python-backtesting-mean-reversion-part-2/
    """

    def __init__(self):
        self.p_value = None
        self.five_perc_stat = None
        self.perc_stat = None
        self.p_min = .0
        self.p_max = .05
        self.look_back = 63

    def apply_adf(self, time_series):
        model = ts.adfuller(time_series, 1)
        self.p_value = model[1]
        self.five_perc_stat = model[4]['5%']
        self.perc_stat = model[0]

    def use_P(self):
        return (self.p_value > self.p_min) and (self.p_value < self.p_max)

    def use_critical(self):
        return abs(self.perc_stat) > abs(self.five_perc_stat)


class KPSS(object):
    #Kwiatkowski-Phillips-Schmidt-Shin (KPSS) stationarity tests
    def __init__(self):
        Exception("Not implemented yet")
        self.p_value = None
        self.ten_perc_stat = None
        self.perc_stat = None
        self.p_min = 0.0
        self.p_max = 0.2
        self.look_back = 50

    def apply_kpss(self, time_series):
        self.p_value = ts.adfuller(time_series, 1)[1]
        self.five_perc_stat = ts.adfuller(time_series, 1)[4]['5%'] # possibly make this 10%
        self.ten_perc_stat = ts.adfuller(time_series, 1)[4]['10%']
        self.perc_stat = ts.adfuller(time_series, 1)[0]

    def use(self):
        return (self.p_value > self.p_min) and (self.p_value < self.p_max) and (self.perc_stat > self.five_perc_stat)


class Data:
    def __init__(self, historical_data):
        self.historical_data = historical_data

    def history(self, assets, fields, bar_count, frequency):
        if frequency != '1d':
            raise ValueError("Only daily frequency is supported")

        end_date = self.historical_data.index[-1]
        start_date = end_date - pd.Timedelta(days=bar_count - 1)

        if isinstance(fields, str):
            fields = [fields]

        if 'price' in fields:
            # If 'price' is requested, use 'Close' price
            fields = [f if f != 'price' else 'Adj Close' for f in fields]

        result = self.historical_data.loc[start_date:end_date, (fields, assets)]

        if len(fields) == 1:
            result = result[fields[0]]

        return result


    def load_historical_data(start_date, end_date, symbols):
        data = yf.download(symbols, start=start_date, end=end_date)
        return data


class Half_Life(object):
    """
    Half Life test from the Ornstein-Uhlenbeck process
    Source: http://www.pythonforfinance.net/2016/05/09/python-backtesting-mean-reversion-part-2/
    """

    def __init__(self):
        self.hl_min = 1.0
        self.hl_max = 42.0
        self.look_back = 43
        self.half_life = None

    def apply_half_life(self, time_series):
        lag = np.roll(time_series, 1)
        lag[0] = 0
        ret = time_series - lag
        ret[0] = 0

        # adds intercept terms to X variable for regression
        lag2 = sm.add_constant(lag)

        model = sm.OLS(ret, lag2)
        res = model.fit()

        self.half_life = -np.log(2) / res.params[1]

    def use(self):
        return (self.half_life < self.hl_max) and (self.half_life > self.hl_min)


class Hurst():
    """
    If Hurst Exponent is under the 0.5 value of a random walk, then the series is mean reverting
    Source: https://www.quantstart.com/articles/Basics-of-Statistical-Mean-Reversion-Testing
    """

    def __init__(self):
        self.h_min = 0.0
        self.h_max = 0.4
        self.look_back = 126 #126
        self.lag_max = 100
        self.h_value = None

    def apply_hurst(self, time_series):
        """Returns the Hurst Exponent of the time series vector ts"""
        # Create the range of lag values
        lags = range(2, self.lag_max)

        # Calculate the array of the variances of the lagged differences
        tau = [np.sqrt(np.std(np.subtract(time_series[lag:], time_series[:-lag]))) for lag in lags]

        # Use a linear fit to estimate the Hurst Exponent
        poly = np.polyfit(np.log10(lags), np.log10(tau), 1)

        # Return the Hurst exponent from the polyfit output
        self.h_value = poly[0] * 2.0

    def use(self):
        return (self.h_value < self.h_max) and (self.h_value > self.h_min)


class PortfolioAnalysis:
    def __init__(self, portfolio):
        self.portfolio = portfolio

    def calculate_daily_pnl(self):
        daily_pnl = {}
        for i in range(1, len(self.portfolio.equity_history)):
            prev_date, prev_equity = self.portfolio.equity_history[i - 1]
            curr_date, curr_equity = self.portfolio.equity_history[i]
            daily_pnl[curr_date] = curr_equity - prev_equity
        return daily_pnl

    def calculate_cost_basis(self, symbol, date):
        if date is None:
            return 0

        if symbol not in self.portfolio.cost_basis_history:
            return 0

        symbol_history = self.portfolio.cost_basis_history[symbol]

        # Get the latest cost basis up to the given date
        for entry in reversed(symbol_history):
            if entry[0] <= date:  # Compare the date (first element of each entry)
                return entry[1]  # Return the cost basis (second element of each entry)

        return 0  # Return 0 if no valid entry is found

    def find_remaining_position_date(self, symbol, prev_vs_current):
        current_position = prev_vs_current['current']['current_position']
        current_date = prev_vs_current['current']['current_date']
        prev_position = prev_vs_current['previous']['prev_position']

        # Case 1: Current position is 0
        if current_position == 0:
            return None  # This will lead to new_cost_basis = 0 in the calling function

        # Case 2: Current position and previous position have different signs
        if (current_position > 0 and prev_position < 0) or (current_position < 0 and prev_position > 0):
            return current_date

        # Case 3: Current position magnitude is greater than previous position magnitude
        if abs(current_position) > abs(prev_position):
            return current_date

        # Case 4: Current position magnitude is less than or equal to previous position magnitude
        # We need to find the date when the position was equal to the current position
        share_history = self.portfolio.share_history.get(symbol, [])
        for date, shares in sorted(share_history, reverse=True):
            if shares == current_position:
                return date
            elif (current_position > 0 and shares > current_position) or \
                    (current_position < 0 and shares < current_position):
                # We've found the first date where the position was greater than or equal to the current position
                return date

        # If we couldn't find a matching date, return the earliest date in our history
        return share_history[-1][0] if share_history else current_date

    def get_previous_vs_current_security(self, symbol, current_price):
        current_date = self.portfolio.current_date

        # Find the previous processed date
        prev_date_index = self.portfolio.processed_dates.index(current_date) - 1
        prev_date = self.portfolio.processed_dates[prev_date_index] if prev_date_index >= 0 else None

        current_position = next(
            (shares for date, shares in self.portfolio.share_history.get(symbol, []) if date == current_date), 0)
        prev_position = next(
            (shares for date, shares in reversed(self.portfolio.share_history.get(symbol, [])) if date <= prev_date),
            0) if prev_date else 0

        prev_acc_pnl = self.portfolio.acc_security_pnl_history.get(prev_date, {}).get(symbol, {'pnl_dollar': 0,
                                                                                               'pnl_percent': 0}) if prev_date else {
            'pnl_dollar': 0, 'pnl_percent': 0}

        current_cost_basis = self.calculate_cost_basis(symbol, current_date)
        prev_cost_basis = self.calculate_cost_basis(symbol, prev_date) if prev_date else 0

        trades = [trade for trade in reversed(self.portfolio.trades) if trade.symbol == symbol]
        last_trade = trades[0] if trades else None
        previous_trade = trades[1] if len(trades) > 1 else None

        # Get the previous date price
        previous_price = next(
            (price for date, price in reversed(self.portfolio.price_history.get(symbol, [])) if date <= prev_date),
            None) if prev_date else None

        return {
            'previous': {
                'prev_date': prev_date,
                'prev_position': prev_position,
                'prev_cost_basis': prev_cost_basis,
                'prev_acc_pnl': prev_acc_pnl,
                'previous_trade': previous_trade,
                'previous_price': previous_price
            },
            'current': {
                'current_date': current_date,
                'current_position': current_position,
                'new_cost_basis': current_cost_basis,
                'current_acc_pnl': {'pnl_dollar': None, 'pnl_percent': None},
                'last_trade': last_trade,
                'current_price': current_price
            }
        }

    def get_previous_vs_current_pair(self, pair, current_prices):
        stock1, stock2 = pair.split('/')
        current_date = self.portfolio.current_date

        # Find the previous processed date
        prev_date_index = self.portfolio.processed_dates.index(current_date) - 1
        prev_date = self.portfolio.processed_dates[prev_date_index] if prev_date_index >= 0 else None

        # Get current positions
        current_position_1 = next(
            (shares for date, shares in self.portfolio.share_history.get(stock1, []) if date == current_date), 0)
        current_position_2 = next(
            (shares for date, shares in self.portfolio.share_history.get(stock2, []) if date == current_date), 0)

        # Get previous positions
        prev_position_1 = next(
            (shares for date, shares in reversed(self.portfolio.share_history.get(stock1, [])) if date <= prev_date),
            0) if prev_date else 0
        prev_position_2 = next(
            (shares for date, shares in reversed(self.portfolio.share_history.get(stock2, [])) if date <= prev_date),
            0) if prev_date else 0

        # Get previous accumulated pair PnL
        prev_acc_pair_pnl = self.portfolio.acc_pair_trade_pnl_history.get(prev_date, {}).get(pair, {'pnl_dollar': 0,
                                                                                                    'pnl_percent': 0}) if prev_date else {
            'pnl_dollar': 0, 'pnl_percent': 0}

        # Get cost bases
        current_cost_basis_1 = self.calculate_cost_basis(stock1, current_date)
        current_cost_basis_2 = self.calculate_cost_basis(stock2, current_date)
        prev_cost_basis_1 = self.calculate_cost_basis(stock1, prev_date) if prev_date else 0
        prev_cost_basis_2 = self.calculate_cost_basis(stock2, prev_date) if prev_date else 0

        # Get trades — use pair-specific trade history to avoid cross-pair contamination
        # (e.g. CL traded in CL/WST today should not affect trade_today for CL/SRE)
        pair_trades = list(reversed(self.portfolio.pair_trade_history.get(pair, [])))
        trades_1 = [t for t in pair_trades if t.symbol == stock1]
        trades_2 = [t for t in pair_trades if t.symbol == stock2]
        last_trade_1 = trades_1[0] if trades_1 else None
        last_trade_2 = trades_2[0] if trades_2 else None
        previous_trade_1 = trades_1[1] if len(trades_1) > 1 else None
        previous_trade_2 = trades_2[1] if len(trades_2) > 1 else None

        # Get previous prices
        previous_price_1 = next(
            (price for date, price in reversed(self.portfolio.price_history.get(stock1, [])) if date <= prev_date),
            None) if prev_date else None
        previous_price_2 = next(
            (price for date, price in reversed(self.portfolio.price_history.get(stock2, [])) if date <= prev_date),
            None) if prev_date else None

        return {
            'previous': {
                'prev_date': prev_date,
                'prev_position': [prev_position_1, prev_position_2],
                'prev_cost_basis': [prev_cost_basis_1, prev_cost_basis_2],
                'prev_acc_pair_pnl': prev_acc_pair_pnl,
                'previous_trade': [previous_trade_1, previous_trade_2],
                'previous_price': [previous_price_1, previous_price_2]
            },
            'current': {
                'current_date': current_date,
                'current_position': [current_position_1, current_position_2],
                'new_cost_basis': [current_cost_basis_1, current_cost_basis_2],
                'current_acc_pair_pnl': {'pnl_dollar': None, 'pnl_percent': None},
                'last_trade': [last_trade_1, last_trade_2],
                'current_price': [current_prices[stock1], current_prices[stock2]]
            }
        }

    def calculate_weighted_pnl_percentage(self, stock1, stock2, pnl_percent1, pnl_percent2, use_current=True, pair=None):
        index = -1 if use_current else -2

        if pair and pair in self.portfolio.percentage_history_by_pair:
            # Use per-pair percentage (avoids shared-symbol net-zero weight problem)
            pair_pct = self.portfolio.percentage_history_by_pair[pair]
            s1_hist = pair_pct.get(stock1, [])
            s2_hist = pair_pct.get(stock2, [])
            stock1_percent = s1_hist[index][1] if len(s1_hist) >= abs(index) else 0
            stock2_percent = s2_hist[index][1] if len(s2_hist) >= abs(index) else 0
        else:
            stock1_hist = self.portfolio.percentage_history.get(stock1, [])
            stock2_hist = self.portfolio.percentage_history.get(stock2, [])
            stock1_percent = stock1_hist[index][1] if len(stock1_hist) >= abs(index) else 0
            stock2_percent = stock2_hist[index][1] if len(stock2_hist) >= abs(index) else 0

        total_abs_percent = abs(stock1_percent) + abs(stock2_percent)
        if total_abs_percent != 0:
            stock1_weight = abs(stock1_percent) / total_abs_percent
            stock2_weight = abs(stock2_percent) / total_abs_percent
            weighted_pnl_percent = (
                    pnl_percent1 * stock1_weight +
                    pnl_percent2 * stock2_weight
            )
        else:
            weighted_pnl_percent = 0

        return weighted_pnl_percent

    def update_cost_basis_history(self, symbol, date, new_cost_basis):
        if symbol not in self.portfolio.cost_basis_history:
            self.portfolio.cost_basis_history[symbol] = [(date, new_cost_basis)]
        else:
            existing_entry = next((entry for entry in self.portfolio.cost_basis_history[symbol] if entry[0] == date),
                                  None)
            if existing_entry:
                if existing_entry[1] != new_cost_basis:
                    index = self.portfolio.cost_basis_history[symbol].index(existing_entry)
                    self.portfolio.cost_basis_history[symbol][index] = (date, new_cost_basis)
            else:
                self.portfolio.cost_basis_history[symbol].append((date, new_cost_basis))
        self._update_cost_basis_history_by_pair(symbol, date, new_cost_basis)

    # ── Per-pair parallel table updaters ────────────────────────────────────

    def _get_pairs_for_symbol(self, symbol):
        """Return all pair keys that contain this symbol."""
        return [pair for pair in self.portfolio.pair_trade_history if symbol in pair.split('/')]

    def _update_share_history_by_pair(self, symbol, date, shares):
        """Sync share_history_by_pair whenever share_history[symbol] is updated."""
        for pair in self._get_pairs_for_symbol(symbol):
            pbp = self.portfolio.share_history_by_pair
            if pair not in pbp:
                pbp[pair] = {}
            if symbol not in pbp[pair]:
                pbp[pair][symbol] = []
            pbp[pair][symbol].append((date, shares))
            # Reconcile: latest value in _by_pair must match share_history
            ref = self.portfolio.share_history.get(symbol, [])
            if ref and pbp[pair][symbol][-1] != ref[-1]:
                log.warning(f"share_history_by_pair reconcile fail: {pair}/{symbol} "
                            f"by_pair={pbp[pair][symbol][-1]} vs symbol={ref[-1]}")

    def _update_cost_basis_history_by_pair(self, symbol, date, new_cost_basis):
        """Recompute cost_basis_history_by_pair for all pairs containing symbol.

        Each pair's cost basis is calculated independently from its own trade history
        (weighted average of open trades up to date), so shared symbols (e.g. CL in
        CL/WST and CL/SRE) reflect their own entry prices rather than the portfolio-wide
        weighted average.

        Reconcile invariant: the weighted average across all pairs (weighted by shares)
        should equal the symbol-level cost basis.
        """
        pbp = self.portfolio.cost_basis_history_by_pair
        all_pairs = self._get_pairs_for_symbol(symbol)

        total_shares_check = 0
        total_cost_check = 0.0

        for pair in all_pairs:
            if pair not in pbp:
                pbp[pair] = {}
            if symbol not in pbp[pair]:
                pbp[pair][symbol] = []

            # Compute cost basis for this pair from its own trade history up to date
            pair_trades = [t for t in self.portfolio.pair_trade_history.get(pair, [])
                           if t.symbol == symbol and t.date <= date]

            net_shares = 0
            total_cost = 0.0
            for t in pair_trades:
                if t.order_type == 'open':
                    total_cost += abs(t.amount) * t.price
                    net_shares += t.amount
                elif t.order_type == 'close':
                    # Proportionally reduce cost basis on close
                    if net_shares != 0:
                        close_shares = min(abs(t.amount), abs(net_shares))
                        total_cost -= close_shares * (total_cost / abs(net_shares))
                    net_shares += t.amount  # t.amount is negative for close

            pair_cost_basis = (total_cost / abs(net_shares)) if net_shares != 0 else 0

            existing = next((e for e in pbp[pair][symbol] if e[0] == date), None)
            if existing:
                if existing[1] != pair_cost_basis:
                    idx = pbp[pair][symbol].index(existing)
                    pbp[pair][symbol][idx] = (date, pair_cost_basis)
            else:
                pbp[pair][symbol].append((date, pair_cost_basis))

            total_shares_check += abs(net_shares)
            total_cost_check += total_cost

        # No cross-pair reconcile for cost_basis: each pair uses its own entry price,
        # which intentionally differs from the portfolio-level weighted average.

    def _update_finished_trades_pnl_by_pair(self, pair, symbol, value):
        """Sync finished_trades_pnl_by_pair whenever finished_trades_pnl[symbol] is updated."""
        pbp = self.portfolio.finished_trades_pnl_by_pair
        if pair not in pbp:
            pbp[pair] = {}
        pbp[pair][symbol] = value
        # Reconcile: sum across all pairs for this symbol must equal symbol-level total
        all_pairs = self._get_pairs_for_symbol(symbol)
        pair_sum = sum(pbp.get(p, {}).get(symbol, 0) for p in all_pairs)
        ref = self.portfolio.finished_trades_pnl.get(symbol, 0)
        if not math.isclose(pair_sum, ref, abs_tol=1e-6):
            log.warning(f"finished_trades_pnl_by_pair reconcile fail: {symbol} "
                        f"pair_sum={pair_sum} vs symbol_total={ref}")

    def _update_total_cost_history_by_pair(self, pair, symbol, value):
        """Sync total_cost_history_by_pair whenever total_cost_history[symbol] is updated."""
        pbp = self.portfolio.total_cost_history_by_pair
        if pair not in pbp:
            pbp[pair] = {}
        pbp[pair][symbol] = value
        # Reconcile: sum across all pairs for this symbol should equal symbol-level total
        all_pairs = self._get_pairs_for_symbol(symbol)
        pair_sum = sum(pbp.get(p, {}).get(symbol, 0) for p in all_pairs)
        ref = self.portfolio.total_cost_history.get(symbol, 0)
        if ref != 0 and not math.isclose(pair_sum, ref, rel_tol=1e-6):
            log.warning(f"total_cost_history_by_pair reconcile fail: {symbol} "
                        f"pair_sum={pair_sum} vs symbol_total={ref}")

    def calculate_dod_security_pnl(self):
        acc_history = self.portfolio.acc_security_pnl_history
        dates = sorted(acc_history.keys())

        # If dod_security_pnl doesn't exist, create it
        if not hasattr(self.portfolio, 'dod_security_pnl'):
            self.portfolio.dod_security_pnl = {}

        # Find the last date that was calculated
        last_calculated_date = max(self.portfolio.dod_security_pnl.keys()) if self.portfolio.dod_security_pnl else None

        # Find the index of the last calculated date, or start from the beginning if it doesn't exist
        start_index = dates.index(last_calculated_date) + 1 if last_calculated_date in dates else 0

        # Calculate DoD PnL only for new dates
        for i in range(start_index, len(dates)):
            date = dates[i]
            self.portfolio.dod_security_pnl[date] = {}

            if i == 0:
                # For the first date, DoD PnL is the same as accumulated PnL
                self.portfolio.dod_security_pnl[date] = acc_history[date]
            else:
                prev_date = dates[i - 1]
                for symbol in acc_history[date]:
                    if symbol in acc_history[prev_date]:
                        current_price = self.portfolio.price_history[symbol][i][1]
                        prev_vs_current = self.get_previous_vs_current_security(symbol, current_price)

                        # Force modify current_acc_pnl with values from acc_history
                        prev_vs_current['current']['current_acc_pnl']['pnl_dollar'] = acc_history[date][symbol][
                            'pnl_dollar']
                        prev_vs_current['current']['current_acc_pnl']['pnl_percent'] = acc_history[date][symbol][
                            'pnl_percent']

                        prev = prev_vs_current['previous']
                        current = prev_vs_current['current']

                        curr_pnl = current['current_acc_pnl']['pnl_dollar']
                        prev_pnl = prev['prev_acc_pnl']['pnl_dollar']
                        dod_pnl_dollar = curr_pnl - prev_pnl

                        # Calculate notional previous day
                        total_cost_prev = abs(prev['prev_position']) * prev['prev_cost_basis']

                        dod_pnl_percent = dod_pnl_dollar / total_cost_prev if total_cost_prev != 0 else 0

                        self.portfolio.dod_security_pnl[date][symbol] = {
                            'pnl_dollar': dod_pnl_dollar,
                            'pnl_percent': dod_pnl_percent
                        }
                    else:
                        self.portfolio.dod_security_pnl[date][symbol] = acc_history[date][symbol]

        return self.portfolio.dod_security_pnl

    def calculate_dod_pair_trade_pnl(self):
        # Use per-pair acc_security_pnl_by_pair (not symbol-level dod_security_pnl) to avoid
        # shared-symbol double-counting (e.g. CL appears in CL/WST, CL/SRE, CL/GD simultaneously).
        acc_by_pair_history = self.portfolio.acc_security_pnl_history_by_pair
        dates = sorted(acc_by_pair_history.keys())

        if not hasattr(self.portfolio, 'dod_pair_trade_pnl_history'):
            self.portfolio.dod_pair_trade_pnl_history = {}

        last_calculated_date = max(
            self.portfolio.dod_pair_trade_pnl_history.keys()) if self.portfolio.dod_pair_trade_pnl_history else None
        start_index = dates.index(last_calculated_date) + 1 if last_calculated_date in dates else 0

        for i in range(start_index, len(dates)):
            date = dates[i]
            self.portfolio.dod_pair_trade_pnl_history[date] = {}

            prev_date_index = self.portfolio.processed_dates.index(date) - 1
            prev_date = self.portfolio.processed_dates[prev_date_index] if prev_date_index >= 0 else None

            for pair in self.portfolio.pair_trade_history.keys():
                stock1, stock2 = pair.split('/')

                curr_pair_pnl = acc_by_pair_history.get(date, {}).get(pair, {})
                prev_pair_pnl = acc_by_pair_history.get(prev_date, {}).get(pair, {}) if prev_date else {}

                if stock1 not in curr_pair_pnl or stock2 not in curr_pair_pnl:
                    continue

                # DoD dollar = current acc per-pair minus previous acc per-pair (per symbol, then sum)
                curr_s1 = curr_pair_pnl[stock1]['pnl_dollar']
                curr_s2 = curr_pair_pnl[stock2]['pnl_dollar']
                prev_s1 = prev_pair_pnl.get(stock1, {}).get('pnl_dollar', 0) if prev_pair_pnl else 0
                prev_s2 = prev_pair_pnl.get(stock2, {}).get('pnl_dollar', 0) if prev_pair_pnl else 0

                pair_pnl_dollar = (curr_s1 - prev_s1) + (curr_s2 - prev_s2)

                # DoD percent: use per-pair acc pnl_percent diff weighted by position cost
                pair_cost_s1 = self.portfolio.total_cost_history_by_pair.get(pair, {}).get(stock1, 0)
                pair_cost_s2 = self.portfolio.total_cost_history_by_pair.get(pair, {}).get(stock2, 0)
                total_pair_cost = pair_cost_s1 + pair_cost_s2
                if total_pair_cost != 0:
                    pair_pnl_percent = pair_pnl_dollar / total_pair_cost
                else:
                    pair_pnl_percent = 0

                self.portfolio.dod_pair_trade_pnl_history[date][pair] = {
                    'pnl_dollar': pair_pnl_dollar,
                    'pnl_percent': pair_pnl_percent
                }

        return self.portfolio.dod_pair_trade_pnl_history

    def calculate_acc_security_pnl(self, current_prices):
        acc_security_pnl = {}
        current_date = self.portfolio.current_date

        for symbol, current_price in current_prices.items():
            if symbol not in self.portfolio.total_cost_history:
                self.portfolio.total_cost_history[symbol] = 0

            # All trades for this symbol today, in order, each carrying their pair
            trades_today = [t for t in self.portfolio.trades
                            if t.symbol == symbol and t.date == current_date]

            if trades_today:
                # Process each trade individually so every pair's open/close is handled
                for trade in trades_today:
                    trade_pair = trade.pair
                    if trade.order_type == 'open':
                        total_cost = abs(trade.amount) * trade.price
                        self.portfolio.total_cost_history[symbol] += total_cost
                        if trade_pair:
                            prev_pair_cost = self.portfolio.total_cost_history_by_pair.get(trade_pair, {}).get(symbol, 0)
                            self._update_total_cost_history_by_pair(trade_pair, symbol, prev_pair_cost + total_cost)

                    elif trade.order_type == 'close':
                        if trade_pair:
                            cb_history = self.portfolio.cost_basis_history_by_pair.get(trade_pair, {}).get(symbol, [])
                            pair_cost_basis = next((v for d, v in reversed(cb_history) if d < current_date), 0)
                            sh_history = self.portfolio.share_history_by_pair.get(trade_pair, {}).get(symbol, [])
                            pair_prev_position = next(
                                (s for d, s in reversed(sh_history) if d < current_date), 0)
                        else:
                            pair_cost_basis = 0
                            pair_prev_position = 0

                        if pair_cost_basis != 0 and pair_prev_position != 0:
                            closing_amount = abs(trade.amount)
                            if pair_prev_position > 0:
                                pnl = (trade.price - pair_cost_basis) * closing_amount
                            else:
                                pnl = (pair_cost_basis - trade.price) * closing_amount

                            self.portfolio.finished_trades_pnl[symbol] = \
                                self.portfolio.finished_trades_pnl.get(symbol, 0) + pnl
                            if trade_pair:
                                prev_pair_finished = self.portfolio.finished_trades_pnl_by_pair.get(trade_pair, {}).get(symbol, 0)
                                self._update_finished_trades_pnl_by_pair(trade_pair, symbol, prev_pair_finished + pnl)

                # After processing all today's trades, update cost_basis_history (symbol level)
                prev_vs_current = self.get_previous_vs_current_security(symbol, current_price)
                remaining_position_date = self.find_remaining_position_date(symbol, prev_vs_current)
                new_cost_basis = self.calculate_cost_basis(symbol, remaining_position_date)
                self.update_cost_basis_history(symbol, current_date, new_cost_basis)

            else:
                # No new trade today
                prev_vs_current = self.get_previous_vs_current_security(symbol, current_price)
                current_info = prev_vs_current['current']
                new_cost_basis = current_info['new_cost_basis']
                self.update_cost_basis_history(symbol, current_date, new_cost_basis)

            # Symbol-level acc_pnl placeholder — will be overwritten from by_pair sum below
            acc_security_pnl[symbol] = {'pnl_dollar': 0, 'pnl_percent': 0}

        # Build acc_security_pnl_by_pair using per-pair finished_trades_pnl and total_cost_history
        current_date = self.portfolio.current_date
        acc_security_pnl_by_pair = {}
        for pair in self.portfolio.pair_trade_history:
            acc_security_pnl_by_pair[pair] = {}
            stock1, stock2 = pair.split('/')
            for symbol in (stock1, stock2):
                if symbol not in current_prices:
                    continue
                current_price = current_prices[symbol]
                pair_finished = self.portfolio.finished_trades_pnl_by_pair.get(pair, {}).get(symbol, 0)
                pair_cost = self.portfolio.total_cost_history_by_pair.get(pair, {}).get(symbol, 0)

                # Determine if this pair has an active position for this symbol
                # by counting open vs close trades up to current_date in pair_trade_history
                pair_sym_trades = [t for t in self.portfolio.pair_trade_history[pair]
                                   if t.symbol == symbol and t.date <= current_date]
                opens_count = sum(1 for t in pair_sym_trades if t.order_type == 'open')
                closes_count = sum(1 for t in pair_sym_trades if t.order_type == 'close')
                pair_has_position = opens_count > closes_count

                # Only compute MTM if this pair currently holds a position
                if pair_has_position:
                    cb_history = self.portfolio.cost_basis_history_by_pair.get(pair, {}).get(symbol, [])
                    cost_basis = next((v for d, v in reversed(cb_history) if d <= current_date), 0)
                    sh_history = self.portfolio.share_history_by_pair.get(pair, {}).get(symbol, [])
                    position = next((s for d, s in reversed(sh_history) if d == current_date), 0)
                    mtm_pnl = (current_price - cost_basis) * position if position != 0 and cost_basis != 0 else 0
                    # open 当天 MTM = 0
                    pair_trades_today = [t for t in pair_sym_trades if t.date == current_date]
                    if pair_trades_today and pair_trades_today[-1].order_type == 'open':
                        mtm_pnl = 0
                else:
                    mtm_pnl = 0

                acc_security_pnl_by_pair[pair][symbol] = {
                    'pnl_dollar': pair_finished + mtm_pnl,
                    'pnl_percent': (pair_finished + mtm_pnl) / pair_cost if pair_cost != 0 else 0,
                }
        self.portfolio.acc_security_pnl_history_by_pair[current_date] = acc_security_pnl_by_pair

        # Overwrite symbol-level acc_security_pnl with sum of per-pair values.
        # This enforces Invariant C: sum(acc_pnl_by_pair[pair][sym]) == acc_security_pnl[sym]
        # and ensures correctness when a symbol appears in multiple pairs (e.g. CL in CL/WST, CL/SRE, CL/GD).
        for symbol in list(acc_security_pnl.keys()):
            pairs_for = [p for p in acc_security_pnl_by_pair if symbol in p.split('/')]
            if pairs_for:
                sym_sum = sum(acc_security_pnl_by_pair[p].get(symbol, {}).get('pnl_dollar', 0) for p in pairs_for)
                total_cost = self.portfolio.total_cost_history.get(symbol, 0)
                acc_security_pnl[symbol] = {
                    'pnl_dollar': sym_sum,
                    'pnl_percent': sym_sum / total_cost if total_cost != 0 else 0,
                }

        return acc_security_pnl

    def calculate_acc_pair_trade_pnl(self, current_prices):
        acc_pair_trade_pnl = {}
        # Use per-pair security PnL (already built in calculate_acc_security_pnl)
        latest_acc_security_pnl_by_pair = self.portfolio.acc_security_pnl_history_by_pair.get(
            self.portfolio.current_date, {})

        current_date = self.portfolio.current_date

        for pair in self.portfolio.pair_trade_history.keys():
            stock1, stock2 = pair.split('/')
            # Get previous and current data
            prev_vs_current = self.get_previous_vs_current_pair(pair, current_prices)
            prev_date = prev_vs_current['previous']['prev_date']
            prev_acc_pair_pnl = prev_vs_current['previous']['prev_acc_pair_pnl']

            # Per-pair security PnL for this pair
            pair_sec_pnl = latest_acc_security_pnl_by_pair.get(pair, {})

            # Bug B fix: if either stock missing from today's per-pair security PnL, fall back to prev value
            if stock1 not in pair_sec_pnl or stock2 not in pair_sec_pnl:
                if prev_acc_pair_pnl:
                    acc_pair_trade_pnl[pair] = {
                        'pnl_dollar': prev_acc_pair_pnl['pnl_dollar'],
                        'pnl_percent': prev_acc_pair_pnl['pnl_percent'],
                    }
                continue

            last_trade_1, last_trade_2 = prev_vs_current['current']['last_trade']

            position1, position2 = prev_vs_current['current']['current_position']
            prev_position1, prev_position2 = prev_vs_current['previous']['prev_position']

            # Calculate PnL dollar from per-pair security PnL (no shared-symbol duplication)
            pnl_dollar = pair_sec_pnl[stock1]['pnl_dollar'] + pair_sec_pnl[stock2]['pnl_dollar']

            # Determine if a trade occurred on the current date
            trade_today = (last_trade_1 and last_trade_1.date == current_date) or \
                          (last_trade_2 and last_trade_2.date == current_date)

            # Determine if positions were closed on the current date
            positions_closed_today = (position1 == 0 and position2 == 0) and (
                        prev_position1 != 0 or prev_position2 != 0)

            # Compute pnl_percent using total_cost_history_by_pair as stable denominator.
            # This avoids the portfolio-level percentage_history pitfalls (shared symbols,
            # closed positions, etc.) and is consistent across all branches.
            pair_cost_s1 = self.portfolio.total_cost_history_by_pair.get(pair, {}).get(stock1, 0)
            pair_cost_s2 = self.portfolio.total_cost_history_by_pair.get(pair, {}).get(stock2, 0)
            total_pair_cost = pair_cost_s1 + pair_cost_s2

            if trade_today and not positions_closed_today:
                is_open_trade = (
                    (last_trade_1 and last_trade_1.date == current_date and last_trade_1.order_type == 'open') or
                    (last_trade_2 and last_trade_2.date == current_date and last_trade_2.order_type == 'open')
                )
                is_close_trade = (
                    (last_trade_1 and last_trade_1.date == current_date and last_trade_1.order_type == 'close') or
                    (last_trade_2 and last_trade_2.date == current_date and last_trade_2.order_type == 'close')
                )
                if is_open_trade or is_close_trade:
                    pnl_percent = pnl_dollar / total_pair_cost if total_pair_cost != 0 else 0
                else:
                    pnl_dollar = prev_acc_pair_pnl['pnl_dollar']
                    pnl_percent = prev_acc_pair_pnl['pnl_percent']
                    if abs(pnl_percent) < 1e-10 and abs(pnl_dollar) > 1e-4 and total_pair_cost != 0:
                        pnl_percent = pnl_dollar / total_pair_cost
            elif positions_closed_today or position1 != 0 or position2 != 0:
                pnl_percent = pnl_dollar / total_pair_cost if total_pair_cost != 0 else 0
            else:
                # Both positions are closed and were closed before the current date
                pnl_dollar = prev_acc_pair_pnl['pnl_dollar']
                pnl_percent = prev_acc_pair_pnl['pnl_percent']
                # Repair: if inherited percent is 0 but dollar is non-zero, recompute from total_cost
                if abs(pnl_percent) < 1e-10 and abs(pnl_dollar) > 1e-4 and total_pair_cost != 0:
                    pnl_percent = pnl_dollar / total_pair_cost

            # Calculate accumulated PnL
            acc_pair_trade_pnl[pair] = {
                'pnl_dollar': pnl_dollar,
                'pnl_percent': pnl_percent
            }

        return acc_pair_trade_pnl

    def reconcile_pnls(self, acc_security_pnl, acc_pair_trade_pnl):
        # Use per-pair security PnL to avoid double-counting shared symbols (Bug C fix)
        current_date = self.portfolio.current_date
        acc_security_pnl_by_pair = self.portfolio.acc_security_pnl_history_by_pair.get(current_date, {})

        # total_security_pnl: sum each pair's two symbols independently (no shared-symbol duplication)
        total_security_pnl_by_pair = sum(
            sec_pnl['pnl_dollar']
            for pair_sec in acc_security_pnl_by_pair.values()
            for sec_pnl in pair_sec.values()
        )
        total_pair_pnl = sum(pnl['pnl_dollar'] for pnl in acc_pair_trade_pnl.values())

        if not math.isclose(total_security_pnl_by_pair, total_pair_pnl, rel_tol=1e-9):
            log.warning(f"Total PnL mismatch on {current_date}: "
                        f"Total Security PnL (by_pair) = {total_security_pnl_by_pair}, Total Pair PnL = {total_pair_pnl}")

        # Check individual pairs using per-pair security PnL
        for pair, pair_pnl in acc_pair_trade_pnl.items():
            pair_sec = acc_security_pnl_by_pair.get(pair, {})
            if not pair_sec:
                continue
            security_pnl_sum = sum(s['pnl_dollar'] for s in pair_sec.values())
            if not math.isclose(pair_pnl['pnl_dollar'], security_pnl_sum, rel_tol=1e-9):
                log.warning(f"PnL mismatch for pair {pair} on {current_date}: "
                            f"Pair PnL = {pair_pnl['pnl_dollar']}, Security PnL sum (by_pair) = {security_pnl_sum}")

        # Also return symbol-level total for backward compatibility
        total_security_pnl = sum(pnl['pnl_dollar'] for pnl in acc_security_pnl.values())
        return total_security_pnl, total_pair_pnl

    def calculate_trading_days_percentage(self, portfolio):
        # Get the NYSE calendar
        nyse = mcal.get_calendar('NYSE')

        # Check if there are any trades
        if not any(portfolio.pair_trade_history.values()):
            log.warning("No trades found in pair_trade_history.")
            return 0

        # Find the first trade date
        first_trade_date = min(
            trade.date for pair_trades in portfolio.pair_trade_history.values() for trade in pair_trades)

        log.info(f"First trade date: {first_trade_date} --> Last test date: {portfolio.current_date}")

        # Get all NYSE trading days between first trade and last day of simulation
        trading_days = nyse.valid_days(start_date=first_trade_date, end_date=portfolio.current_date)

        # Total NYSE trading days in the period
        total_trading_days = len(trading_days)
        log.info(f"Total trading days: {total_trading_days}")

        # Log information about share_history structure
        # log.info(f"Share history keys: {list(portfolio.share_history.keys())}")
        # for key, value in portfolio.share_history.items():
        #    log.info(f"Length of share history for {key}: {len(value)}")

        # Create a dictionary to hold the share data
        share_data = defaultdict(dict)
        for stock, history in portfolio.share_history.items():
            for date, shares in history:
                share_data[date][stock] = shares

        # Create DataFrame from the dictionary
        share_df = pd.DataFrame.from_dict(share_data, orient='index')
        share_df.index = pd.to_datetime(share_df.index)
        share_df = share_df.sort_index()

        # Remove rows where all stock columns are 0, NaN, or null
        valid_share_days = share_df[(share_df != 0) & (~share_df.isna())].dropna(how='all').index

        # Convert trading days and valid share days to integers for comparison
        trading_days_int = set(int(date.strftime('%Y%m%d')) for date in trading_days)
        valid_share_days_int = set(int(date.strftime('%Y%m%d')) for date in valid_share_days)

        # Count days with open positions
        days_with_positions = sum(1 for date in trading_days_int if date in valid_share_days_int)
        log.info(f"Days with positions: {days_with_positions}")

        # Check if positions were ever opened
        if days_with_positions == 0:
            log.warning("No days with open positions found.")

        # Log a sample of the share history
        # log.info(f"Sample of share history:\n{share_df.head()}")

        return days_with_positions / total_trading_days if total_trading_days > 0 else 0


class ExportExcel:
    def __init__(self, filename):
        self.filename = filename

    def format_date(self, date):
        if isinstance(date, datetime):
            return date.strftime('%Y-%m-%d')
        return str(date)

    def format_percentage(self, value):
        return value  # Store raw float for full precision; Excel will format for display

    def should_keep_time(self, dates):
        return any(isinstance(d, datetime) and d.time() != time(0, 0, 0) for d in dates)

    def export_list_of_tuples(self, sheet, data):
        sheet.delete_rows(1, sheet.max_row)
        sheet.append(['Date', 'Value'])

        formatted_data = [(self.format_date(date), value) for date, value in data]
        for row in formatted_data:
            sheet.append(row)

        # Check if we should keep the time
        if not self.should_keep_time([date for date, _ in data]):
            for cell in sheet['A']:
                if cell.row > 1:  # Skip header
                    cell.value = cell.value.split()[0] if cell.value else cell.value

    def export_dict_of_symbol_tuples(self, sheet, data):
        sheet.delete_rows(1, sheet.max_row)
        all_dates = set()
        for symbol_data in data.values():
            if isinstance(symbol_data, list):
                all_dates.update(date for date, *_ in symbol_data)
            elif isinstance(symbol_data, dict):
                all_dates.update(symbol_data.keys())
        dates = sorted(all_dates)
        symbols = list(data.keys())

        header = ['Date'] + symbols
        sheet.append(header)

        for date in dates:
            row = [self.format_date(date)]
            for symbol in symbols:
                symbol_data = data[symbol]
                if isinstance(symbol_data, list):
                    value = next((v for d, v, *_ in symbol_data if d == date), '')
                elif isinstance(symbol_data, dict):
                    value = symbol_data.get(date, '')
                else:
                    value = ''
                row.append(value)
            sheet.append(row)

        # Check if we should keep the time
        if not self.should_keep_time(dates):
            for cell in sheet['A']:
                if cell.row > 1:  # Skip header
                    cell.value = cell.value.split()[0] if cell.value else cell.value

    def export_statistical_test_history(self, sheet, data):
        sheet.delete_rows(1, sheet.max_row)  # Clear existing data

        # Create headers
        headers = ['Date', 'Pair']
        all_attributes = set()
        for pair_data in data.values():
            for date_data in pair_data.values():
                all_attributes.update(date_data.keys())
        headers.extend(sorted(all_attributes))
        sheet.append(headers)

        # Add data
        for pair, pair_data in data.items():
            for date, test_results in pair_data.items():
                row = [self.format_date(date), pair]
                for attr in headers[2:]:
                    row.append(test_results.get(attr, ''))
                sheet.append(row)

        # Format date column
        for cell in sheet['A']:
            if cell.row > 1:  # Skip header
                cell.value = cell.value.split()[0] if cell.value else cell.value

    def export_daily_pnl_history(self, sheet, data):
        sheet.delete_rows(1, sheet.max_row)  # Clear existing data
        sheet.append(['Date', 'Daily PnL'])
        for date, pnl in data:
            sheet.append([self.format_date(date), pnl])

    def export_dod_security_pnl_history(self, sheet, data):
        sheet.delete_rows(1, sheet.max_row)  # Clear existing data
        headers = ['Date', 'Symbol', 'PnL Dollar', 'PnL Percent']
        sheet.append(headers)
        for date, securities in data.items():
            for symbol, pnl in securities.items():
                row = [self.format_date(date), symbol, pnl['pnl_dollar'], self.format_percentage(pnl['pnl_percent'])]
                sheet.append(row)

    def export_dod_pair_trade_pnl_history(self, sheet, data):
        sheet.delete_rows(1, sheet.max_row)  # Clear existing data
        headers = ['Date', 'Pair', 'PnL Dollar', 'PnL Percent']
        sheet.append(headers)
        for date, pairs in data.items():
            for pair, pnl in pairs.items():
                row = [self.format_date(date), pair, pnl['pnl_dollar'], self.format_percentage(pnl['pnl_percent'])]
                sheet.append(row)

    def export_acc_security_pnl_history(self, sheet, data):
        sheet.delete_rows(1, sheet.max_row)  # Clear existing data
        headers = ['Date', 'Symbol', 'PnL Dollar', 'PnL Percent']
        sheet.append(headers)
        for date, securities in data.items():
            for symbol, pnl in securities.items():
                row = [self.format_date(date), symbol, pnl['pnl_dollar'], self.format_percentage(pnl['pnl_percent'])]
                sheet.append(row)

    def export_acc_security_pnl_history_by_pair(self, sheet, data):
        """Export acc_security_pnl_history_by_pair: {date: {pair: {symbol: {pnl_dollar, pnl_percent}}}}"""
        sheet.delete_rows(1, sheet.max_row)
        headers = ['Date', 'Pair', 'Symbol', 'PnL Dollar', 'PnL Percent']
        sheet.append(headers)
        for date, pairs in data.items():
            for pair, securities in pairs.items():
                for symbol, pnl in securities.items():
                    row = [self.format_date(date), pair, symbol,
                           pnl['pnl_dollar'], self.format_percentage(pnl['pnl_percent'])]
                    sheet.append(row)

    def export_acc_pair_trade_pnl_history(self, sheet, data):
        sheet.delete_rows(1, sheet.max_row)  # Clear existing data
        headers = ['Date', 'Pair', 'PnL Dollar', 'PnL Percent']
        sheet.append(headers)
        for date, pairs in data.items():
            for pair, pnl in pairs.items():
                row = [self.format_date(date), pair, pnl['pnl_dollar'], self.format_percentage(pnl['pnl_percent'])]
                sheet.append(row)

    def export_recorded_vars(self, sheet, data):
        sheet.delete_rows(1, sheet.max_row)  # Clear existing data

        # Create headers
        headers = ['Date', 'Pair']
        all_attributes = set()
        for pair_data in data.values():
            for date_data in pair_data.values():
                all_attributes.update(date_data.keys())
        headers.extend(sorted(all_attributes))
        sheet.append(headers)

        # Add data
        for pair, pair_data in data.items():
            for date, vars_data in pair_data.items():
                row = [self.format_date(date), pair]
                for attr in headers[2:]:
                    row.append(vars_data.get(attr, ''))  # Use empty string for missing attributes
                sheet.append(row)

        # Format date column
        for cell in sheet['A']:
            if cell.row > 1:  # Skip header
                cell.value = cell.value.split()[0] if cell.value else cell.value

    def export_stop_loss_history(self, sheet, data):
        sheet.delete_rows(1, sheet.max_row)  # Clear existing data

        # Create headers
        headers = ['Date', 'Pair', 'Reason', 'Price 1', 'Price 2', 'Spread', 'Z-Score',
                   'Volatility Stop Loss Level', 'Price Stop Loss Level',
                   'Max Holding Period', 'Since Last Trade', 'Triggered By']
        sheet.append(headers)

        # Add data
        for pair, stop_loss_events in data.items():
            for event in stop_loss_events:
                row = [
                    self.format_date(event['date']),
                    pair,
                    event['reason'],
                    event['price_1'],
                    event['price_2'],
                    event['spread'],
                    event['z_score'],
                    event['volatility_stop_loss_level'],
                    event['price_stop_loss_level'],
                    event['max_holding_period'],
                    event['since_last_trade'],
                    event['triggered_by']
                ]
                sheet.append(row)

        # Format date column
        for cell in sheet['A']:
            if cell.row > 1:  # Skip header
                cell.value = cell.value.split()[0] if cell.value else cell.value

    def export_pair_trade_history(self, sheet, data):
        sheet.delete_rows(1, sheet.max_row)  # Clear existing data

        # Create headers
        headers = ['Pair', 'Date', 'Symbol', 'Amount', 'Price', 'Order Type', 'Direction']
        sheet.append(headers)

        # Add data
        for pair, trades in data.items():
            for trade in trades:
                row = [
                    pair,
                    self.format_date(trade.date),
                    trade.symbol,
                    trade.amount,
                    trade.price,
                    trade.order_type,
                    trade.direction
                ]
                sheet.append(row)

        # Format date column
        for cell in sheet['B']:  # Assuming 'Date' is in column B
            if cell.row > 1:  # Skip header
                cell.value = cell.value.split()[0] if cell.value else cell.value

    def export_symbol_scalar_dict(self, sheet, data):
        """Export {symbol: scalar_value} — e.g. finished_trades_pnl, total_cost_history."""
        sheet.delete_rows(1, sheet.max_row)
        sheet.append(['Symbol', 'Value'])
        for symbol, value in sorted(data.items()):
            sheet.append([symbol, value])

    def export_pair_symbol_scalar_dict(self, sheet, data):
        """Export {pair: {symbol: scalar_value}} — e.g. finished_trades_pnl_by_pair."""
        sheet.delete_rows(1, sheet.max_row)
        sheet.append(['Pair', 'Symbol', 'Value'])
        for pair, sym_dict in sorted(data.items()):
            for symbol, value in sorted(sym_dict.items()):
                sheet.append([pair, symbol, value])

    def export_pair_symbol_tuple_history(self, sheet, data):
        """Export {pair: {symbol: [(date, value), ...]}} — e.g. share_history_by_pair."""
        sheet.delete_rows(1, sheet.max_row)
        sheet.append(['Date', 'Pair', 'Symbol', 'Value'])
        rows = []
        for pair, sym_dict in data.items():
            for symbol, history in sym_dict.items():
                for date, value in history:
                    rows.append((date, pair, symbol, value))
        rows.sort(key=lambda r: r[0])
        for date, pair, symbol, value in rows:
            sheet.append([self.format_date(date), pair, symbol, value])

    def export_pair_symbol_cb_history(self, sheet, data):
        """Export cost_basis_history_by_pair: {pair: {symbol: [(date, cost_basis), ...]}}."""
        self.export_pair_symbol_tuple_history(sheet, data)

    def export_max_drawdown_history(self, sheet, data):
        sheet.delete_rows(1, sheet.max_row)  # Clear existing data
        sheet.append(['Date', 'Max Drawdown ($)', 'Max Drawdown (%)'])
        for date, drawdown_dollar, drawdown_percent in data:
            sheet.append([self.format_date(date), drawdown_dollar, f"{drawdown_percent:.2%}"])

    def export_portfolio_data(self, portfolio, context, statistical_tests=None):
        try:
            if os.path.exists(self.filename):
                workbook = openpyxl.load_workbook(self.filename)
            else:
                workbook = openpyxl.Workbook()
                workbook.remove(workbook.active)  # Remove the default sheet
        except Exception as e:
            print(f"Error opening or creating workbook: {e}")
            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)  # Remove the default sheet

        # List of all attributes to export
        all_attributes = [
            'acc_daily_pnl_history', 'acc_interest_history',
            'asset_history', 'equity_history',
            'interest_expense_history', 'liability_history', 'value_history',
            'price_history', 'share_history', 'percentage_history', 'cost_basis_history',
            'hedge_history',
            'asset_cash_history', 'asset_securities_history',
            'liability_securities_history', 'liability_loan_history',
            'share_history_by_pair', 'cost_basis_by_pair',
            'finished_trades_pnl', 'finished_trades_pnl_by_pair',
            'total_cost_history', 'total_cost_history_by_pair',
            'acc_sec_pnl_by_pair',
            'percentage_history_by_pair',
        ]

        # Remove any sheets that are not in the all_attributes list, but keep at least one
        sheets_to_remove = [sheet for sheet in workbook.sheetnames if sheet not in all_attributes]
        for sheet_name in sheets_to_remove:
            workbook.remove(workbook[sheet_name])

        # Ensure at least one sheet exists
        if len(workbook.sheetnames) == 0:
            workbook.create_sheet("Sheet1")

        # Export list of tuples data
        list_attributes = [
            'acc_daily_pnl_history', 'acc_interest_history',
            'asset_history', 'equity_history',
            'interest_expense_history', 'liability_history', 'value_history',
            'asset_cash_history', 'liability_loan_history'
        ]
        for attr in list_attributes:
            if hasattr(portfolio, attr):
                data = getattr(portfolio, attr)
                if data:  # Only process non-empty data
                    if attr not in workbook.sheetnames:
                        sheet = workbook.create_sheet(attr)
                    else:
                        sheet = workbook[attr]
                    self.export_list_of_tuples(sheet, data)

        # Export dict of symbol tuples data
        dict_attributes = [
            'price_history', 'share_history', 'percentage_history', 'cost_basis_history',
            'hedge_history',
            'asset_securities_history', 'liability_securities_history'
        ]
        for attr in dict_attributes:
            if hasattr(portfolio, attr):
                data = getattr(portfolio, attr)
                if data:  # Only process non-empty data
                    if attr not in workbook.sheetnames:
                        sheet = workbook.create_sheet(attr)
                    else:
                        sheet = workbook[attr]
                    self.export_dict_of_symbol_tuples(sheet, data)

        # Export per-pair parallel tables (immediately after their symbol-level counterparts)
        for sheet_name, attr, method in [
            ('share_history_by_pair',        'share_history_by_pair',        self.export_pair_symbol_tuple_history),
            ('cost_basis_by_pair',           'cost_basis_history_by_pair',   self.export_pair_symbol_cb_history),
            ('finished_trades_pnl',          'finished_trades_pnl',          self.export_symbol_scalar_dict),
            ('finished_trades_pnl_by_pair',  'finished_trades_pnl_by_pair',  self.export_pair_symbol_scalar_dict),
            ('total_cost_history',           'total_cost_history',           self.export_symbol_scalar_dict),
            ('total_cost_history_by_pair',   'total_cost_history_by_pair',   self.export_pair_symbol_scalar_dict),
            ('percentage_history_by_pair',   'percentage_history_by_pair',   self.export_pair_symbol_tuple_history),
        ]:
            if hasattr(portfolio, attr):
                data = getattr(portfolio, attr)
                if data:
                    if sheet_name not in workbook.sheetnames:
                        sheet = workbook.create_sheet(sheet_name)
                    else:
                        sheet = workbook[sheet_name]
                    method(sheet, data)

        # Export new histories
        if 'daily_pnl_history' not in workbook.sheetnames:
            sheet = workbook.create_sheet('daily_pnl_history')
        else:
            sheet = workbook['daily_pnl_history']
        self.export_daily_pnl_history(sheet, portfolio.daily_pnl_history)

        if 'acc_security_pnl_history' not in workbook.sheetnames:
            sheet = workbook.create_sheet('acc_security_pnl_history')
        else:
            sheet = workbook['acc_security_pnl_history']
        self.export_acc_security_pnl_history(sheet, portfolio.acc_security_pnl_history)

        if 'acc_sec_pnl_by_pair' not in workbook.sheetnames:
            sheet = workbook.create_sheet('acc_sec_pnl_by_pair')
        else:
            sheet = workbook['acc_sec_pnl_by_pair']
        self.export_acc_security_pnl_history_by_pair(sheet, portfolio.acc_security_pnl_history_by_pair)

        if 'acc_pair_trade_pnl_history' not in workbook.sheetnames:
            sheet = workbook.create_sheet('acc_pair_trade_pnl_history')
        else:
            sheet = workbook['acc_pair_trade_pnl_history']
        self.export_acc_pair_trade_pnl_history(sheet, portfolio.acc_pair_trade_pnl_history)

        if 'dod_security_pnl_history' not in workbook.sheetnames:
            sheet = workbook.create_sheet('dod_security_pnl_history')
        else:
            sheet = workbook['dod_security_pnl_history']
        self.export_dod_security_pnl_history(sheet, portfolio.dod_security_pnl_history)

        if 'dod_pair_trade_pnl_history' not in workbook.sheetnames:
            sheet = workbook.create_sheet('dod_pair_trade_pnl_history')
        else:
            sheet = workbook['dod_pair_trade_pnl_history']
        self.export_dod_pair_trade_pnl_history(sheet, portfolio.dod_pair_trade_pnl_history)

        # Export statistical test history
        if 'statistical_test_history' not in workbook.sheetnames:
            sheet = workbook.create_sheet('statistical_test_history')
        else:
            sheet = workbook['statistical_test_history']
        self.export_statistical_test_history(sheet, portfolio.statistical_test_history)

        # Export recorded vars
        if 'recorded_vars' not in workbook.sheetnames:
            sheet = workbook.create_sheet('recorded_vars')
        else:
            sheet = workbook['recorded_vars']
        self.export_recorded_vars(sheet, context.recorded_vars)

        # Add this new section to export stop_loss_history
        if 'stop_loss_history' not in workbook.sheetnames:
            sheet = workbook.create_sheet('stop_loss_history')
        else:
            sheet = workbook['stop_loss_history']
        self.export_stop_loss_history(sheet, context.execution.stop_loss_history)

        # Add this new section to export pair_trade_history
        if 'pair_trade_history' not in workbook.sheetnames:
            sheet = workbook.create_sheet('pair_trade_history')
        else:
            sheet = workbook['pair_trade_history']
        self.export_pair_trade_history(sheet, portfolio.pair_trade_history)

        if 'max_drawdown_history' not in workbook.sheetnames:
            sheet = workbook.create_sheet('max_drawdown_history')
        else:
            sheet = workbook['max_drawdown_history']
        self.export_max_drawdown_history(sheet, portfolio.max_drawdown_history)

        # Save and close the workbook
        try:
            workbook.save(self.filename)
        except Exception as e:
            print(f"Error saving workbook: {e}")
        finally:
            workbook.close()


class Context:
    def __init__(self):
        self.portfolio = None
        self.execution = None
        self.recorded_vars = {}
        self.strategy_pairs = None
        self.num_pairs = None
        # Add any other necessary attributes


class Execution:
    def __init__(self):
        self.z_back = 36
        self.v_back = 32
        self.mean_back = 30
        self.std_back = 30
        self.entry_z = 0.5
        self.exit_z = 0.0
        self.hedge_lag = None
        self.volatility_stop_loss_multiplier = 2
        self.max_holding_period = 12  # in days
        self.cooling_off_period = 2
        self.volatility_stop_loss_level = None
        self.price_stop_loss_level = None
        self.price_level_stop_loss = {}  # Will be set for each pair
        self.stop_loss_history = {}

        self.amplifier = 2
        self.capital_utilization = 0.70       # fraction of net equity to deploy (70% default)

        self.base_entry_z = 0.75
        self.base_exit_z = 0.0
        self.entry_volatility_factor = 2.25  # the larger volatility_factor is, the harder to enter a long position and a short position
        self.exit_volatility_factor = 0.75  # the larger volatility_factor is, the easier to exit in a long position and a short position
        self.max_entry_z = 2.5
        self.max_exit_z = 0.5


class PortfolioVisualizer:
    def __init__(self, portfolio, context, excel_filename, chart_dir=None):
        self.portfolio = portfolio
        self.context = context
        self.excel_filename = excel_filename
        self.chart_dir = chart_dir  # If set, save charts to this directory instead of plt.show()

    def plot_all_histories(self):
        self.plot_portfolio_history()
        self.plot_individual_stocks()
        self.plot_pair_trades()
        self.plot_additional_histories()

    def plot_portfolio_history(self):
        fig, ax = plt.subplots(figsize=(15, 10))

        dates = [date for date, _ in self.portfolio.equity_history]
        equity = [value for _, value in self.portfolio.equity_history]
        assets = [value for _, value in self.portfolio.asset_history]
        liabilities = [value for _, value in self.portfolio.liability_history]

        ax.plot(dates, equity, label='Net Equity')
        ax.plot(dates, assets, label='Total Assets')
        ax.plot(dates, liabilities, label='Total Liabilities')

        ax.set_title('Portfolio History')
        ax.legend()
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
        ax.xaxis.set_major_locator(mdates.MonthLocator())
        ax.grid(True)
        ax.yaxis.set_major_formatter(FuncFormatter(lambda x, p: f'${x:,.0f}'))

        plt.xticks(rotation=45)
        plt.tight_layout()
        if self.chart_dir:
            plt.savefig(os.path.join(self.chart_dir, 'portfolio_history.png'), dpi=150, bbox_inches='tight')
            plt.close(fig)
        else:
            plt.show()

    def plot_individual_stocks(self):
        fig, axs = plt.subplots(len(self.context.strategy_pairs), 1, figsize=(15, 5 * len(self.context.strategy_pairs)),
                                sharex=True)

        trade_types = {'long_open': '^', 'short_open': 'v', 'long_close': 'v', 'short_close': '^'}
        trade_colors = {'long_open': 'green', 'short_open': 'red', 'long_close': 'blue', 'short_close': 'purple'}

        for i, pair in enumerate(self.context.strategy_pairs):
            stock_1, stock_2 = pair[0], pair[1]
            ax = axs[i] if len(self.context.strategy_pairs) > 1 else axs

            dates = [date for date, _ in self.portfolio.price_history[stock_1]]
            prices_1 = [price for _, price in self.portfolio.price_history[stock_1]]
            prices_2 = [price for _, price in self.portfolio.price_history[stock_2]]

            ax.plot(dates, prices_1, label=f'{stock_1} Price')
            ax.plot(dates, prices_2, label=f'{stock_2} Price')
            ax.set_title(f'Pair: {stock_1} - {stock_2}')

            trade_labels = set()
            for trade in self.portfolio.trades:
                if trade.symbol in [stock_1, stock_2]:
                    trade_type = f"{trade.direction}_{trade.order_type}"
                    label = trade_type.replace('_', ' ').title()
                    if label not in trade_labels:
                        ax.plot(trade.date, trade.price, trade_types[trade_type], markersize=7,
                                color=trade_colors[trade_type], label=label)
                        trade_labels.add(label)
                    else:
                        ax.plot(trade.date, trade.price, trade_types[trade_type], markersize=7,
                                color=trade_colors[trade_type])

            ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
            ax.xaxis.set_major_locator(mdates.MonthLocator())
            ax.grid(True)
            ax.yaxis.set_major_formatter(FuncFormatter(lambda x, p: f'${x:,.2f}'))
            ax.legend()

        plt.xticks(rotation=45)
        plt.tight_layout()
        if self.chart_dir:
            plt.savefig(os.path.join(self.chart_dir, 'individual_stocks.png'), dpi=150, bbox_inches='tight')
            plt.close(fig)
        else:
            plt.show()

    def plot_pair_trades(self):
        # Read the Excel file
        xls = pd.ExcelFile(self.excel_filename)

        price_history = pd.read_excel(xls, 'price_history', parse_dates=['Date'])
        hedge_history = pd.read_excel(xls, 'hedge_history', parse_dates=['Date'])
        pair_trade_history = pd.read_excel(xls, 'pair_trade_history', parse_dates=['Date'])

        n_pairs = len(self.context.strategy_pairs)
        fig, axs = plt.subplots(n_pairs, 1, figsize=(15, 6 * n_pairs), sharex=True)
        if n_pairs == 1:
            axs = [axs]

        # Define consistent colors
        price_colors = plt.cm.Set1(np.linspace(0, 1, 3))[:2]  # Two distinct colors for prices
        spread_color = 'gray'
        trade_colors = {'open': 'g', 'close': 'r'}

        for i, pair in enumerate(self.context.strategy_pairs):
            stock_1, stock_2 = pair[0], pair[1]
            ax = axs[i]

            # Get price data
            prices_1 = price_history[['Date', stock_1]].dropna()
            prices_2 = price_history[['Date', stock_2]].dropna()

            # Get hedge data
            hedge_data = hedge_history[['Date', f'{stock_1}/{stock_2}']].dropna()

            # Merge price and hedge data
            merged_data = prices_1.merge(prices_2, on='Date', suffixes=('_1', '_2'))
            merged_data = merged_data.merge(hedge_data, on='Date', how='left')
            merged_data[f'{stock_1}/{stock_2}'] = merged_data[f'{stock_1}/{stock_2}'].ffill()

            # Calculate adjusted prices and spread
            merged_data['adjusted_price_2'] = merged_data[stock_2] * merged_data[f'{stock_1}/{stock_2}']
            merged_data['spread'] = merged_data[stock_1] - merged_data['adjusted_price_2']

            # Plot prices and spread
            ax.plot(merged_data['Date'], merged_data[stock_1], color=price_colors[0], label=f'{stock_1} Price')
            ax.plot(merged_data['Date'], merged_data['adjusted_price_2'], color=price_colors[1],
                    label=f'{stock_2} Adjusted Price')
            ax.plot(merged_data['Date'], merged_data['spread'], color=spread_color, label='Spread', linestyle='--')

            # Plot trade events
            pair_trades = pair_trade_history[pair_trade_history['Pair'] == f'{stock_1}/{stock_2}']
            for _, trade in pair_trades.iterrows():
                color = trade_colors[trade['Order Type']]
                label = f"{trade['Order Type'].capitalize()} {trade['Direction']}"
                ax.axvline(x=trade['Date'], color=color, linestyle='--', label=label)

            ax.set_title(f'Pair Trade: {stock_1} - {stock_2}', fontsize=10)
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
            ax.xaxis.set_major_locator(mdates.MonthLocator())
            ax.grid(True)
            ax.yaxis.set_major_formatter(FuncFormatter(lambda x, p: f'${x:,.2f}'))

            # Create legend with unique labels
            handles, labels = ax.get_legend_handles_labels()
            by_label = dict(zip(labels, handles))
            ax.legend(by_label.values(), by_label.keys(), fontsize=8)

            ax.tick_params(axis='both', which='major', labelsize=8)

        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        if self.chart_dir:
            plt.savefig(os.path.join(self.chart_dir, 'pair_trades.png'), dpi=150, bbox_inches='tight')
            plt.close(fig)
        else:
            plt.show()

    def plot_additional_histories(self):
        fig, axs = plt.subplots(len(self.context.strategy_pairs), 2, figsize=(20, 8 * len(self.context.strategy_pairs)),
                                sharex=True)

        for i, pair in enumerate(self.context.strategy_pairs):
            stock_1, stock_2 = pair[0], pair[1]
            pair_key = f"{stock_1}/{stock_2}"

            ax1 = axs[i, 0] if len(self.context.strategy_pairs) > 1 else axs[0]
            ax2 = axs[i, 1] if len(self.context.strategy_pairs) > 1 else axs[1]

            # Get all dates for this pair
            all_dates = sorted(set(self.context.recorded_vars[pair_key].keys()) |
                               set(self.portfolio.dod_pair_trade_pnl_history.keys()))

            # Prepare data for ax1
            z_scores, long_entry_zs, short_exit_zs, short_entry_zs, long_exit_zs, norm_vols, plot_dates = [], [], [], [], [], [], []
            for date in all_dates:
                data = self.context.recorded_vars[pair_key].get(date, {})
                z_score = data.get('Z_tech') or data.get('Z_finance') or data.get('Z_food') or data.get('Z_industrial') or data.get('Z_energy')
                entry_z = data.get('Entry_Z')
                exit_z = data.get('Exit_Z')
                norm_vol = data.get('Normalized_Spread_Sigma')

                if all(v is not None for v in [z_score, entry_z, exit_z, norm_vol]):
                    z_scores.append(z_score)
                    long_entry_zs.append(entry_z)
                    short_exit_zs.append(exit_z)
                    short_entry_zs.append(entry_z * -1)
                    long_exit_zs.append(exit_z * -1)
                    norm_vols.append(norm_vol)
                    plot_dates.append(date)

            # Plot data for ax1
            ax1.plot(plot_dates, z_scores, label='Z-score')
            ax1.plot(plot_dates, long_entry_zs, label='Long Entry Z', linestyle='--')
            ax1.plot(plot_dates, short_exit_zs, label='Short Exit Z', linestyle='--')
            ax1.plot(plot_dates, short_entry_zs, label='Short Entry Z', linestyle='--')
            ax1.plot(plot_dates, long_exit_zs, label='Long Exit Z', linestyle='--')
            ax1.plot(plot_dates, norm_vols, label='Normalized Volatility', alpha=0.5)
            ax1.set_title(f'{pair_key} Z-scores and Volatility')
            ax1.legend()
            ax1.grid(True)

            # Prepare and plot data for ax2
            dod_pnl = self.portfolio.dod_pair_trade_pnl_history
            dod_dates, dod_values = [], []
            for date in all_dates:
                value = dod_pnl.get(date, {}).get(pair_key, {}).get('pnl_dollar', 0)
                if value != 0:
                    dod_dates.append(date)
                    dod_values.append(value)

            ax2.bar(dod_dates, dod_values)
            ax2.set_title(f'{pair_key} DoD Pair Trade PnL')
            ax2.yaxis.set_major_formatter(FuncFormatter(lambda x, p: f'${x:,.0f}'))
            ax2.grid(True)

            # Set x-axis for both subplots
            min_date = min(min(plot_dates or [datetime.max]), min(dod_dates or [datetime.max]))
            max_date = max(max(plot_dates or [datetime.min]), max(dod_dates or [datetime.min]))
            for ax in [ax1, ax2]:
                ax.set_xlim(min_date, max_date)
                ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
                ax.xaxis.set_major_locator(mdates.AutoDateLocator())

            plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45, ha='right')
            plt.setp(ax2.xaxis.get_majorticklabels(), rotation=45, ha='right')

        plt.tight_layout()
        if self.chart_dir:
            plt.savefig(os.path.join(self.chart_dir, 'z_scores_and_pnl.png'), dpi=150, bbox_inches='tight')
            plt.close(fig)
        else:
            plt.show()


class PortfolioStopLossFunction:
    def __init__(self, constant):
        self.constant = 1
        self.portfolio_order = PortfolioMakeOrder(constant=1)

    def check_volatility_stop_loss(self, context, pair, spread, z_score):
        pair_key = f"{pair[0]}/{pair[1]}"

        if len(spread) < max(context.execution.mean_back, context.execution.std_back):
            return False, None, None

        # Use the last mean_back number of spreads for mean calculation
        mean_spread = np.mean(spread[-context.execution.mean_back:])

        # Use the last std_back number of spreads for std calculation
        std_spread = np.std(spread[-context.execution.std_back:])

        if pair[2]['in_long']:
            context.execution.volatility_stop_loss_level = mean_spread - (
                        context.execution.volatility_stop_loss_multiplier * std_spread)
            if spread[-1] < context.execution.volatility_stop_loss_level:
                return True, "Volatility Stop Loss (Long)", context.execution.volatility_stop_loss_level
        elif pair[2]['in_short']:
            context.execution.volatility_stop_loss_level = mean_spread + (
                        context.execution.volatility_stop_loss_multiplier * std_spread)
            if spread[-1] > context.execution.volatility_stop_loss_level:
                return True, "Volatility Stop Loss (Short)", context.execution.volatility_stop_loss_level

        return False, None, None

    def check_time_based_stop_loss(self, context, pair):
        pair_key = f"{pair[0]}/{pair[1]}"
        if pair_key in context.portfolio.pair_trade_history:
            last_trade = context.portfolio.pair_trade_history[pair_key][-1]

            # Get the historical data for one of the pair's stocks (either will do)
            stock = pair[0]
            historical_data = context.data.history([stock], "price", context.execution.max_holding_period * 2, "1d")

            # If last_trade.date is outside the history window, position held longer than window > max_holding_period
            if last_trade.date not in historical_data.index:
                return True, "Time-based Stop Loss", None

            # Find the index of the last trade date and the current date
            last_trade_index = historical_data.index.get_loc(last_trade.date)
            current_date_index = historical_data.index.get_loc(context.portfolio.current_date)

            # Calculate the number of trading days between the two dates
            trading_days_since_last_trade = current_date_index - last_trade_index

            if trading_days_since_last_trade > context.execution.max_holding_period:
                return True, "Time-based Stop Loss", None
        return False, None, None

    def check_price_level_stop_loss(self, context, pair, spread, z_score):
        pair_key = f"{pair[0]}/{pair[1]}"
        context.execution.price_stop_loss_level = context.execution.price_level_stop_loss.get(pair_key)

        if context.execution.price_stop_loss_level is not None:
            if pair[2]['in_long'] and spread[-1] < context.execution.price_stop_loss_level:
                return True, "Price Level Stop Loss (Long)", context.execution.price_stop_loss_level
            elif pair[2]['in_short'] and spread[-1] > context.execution.price_stop_loss_level:
                return True, "Price Level Stop Loss (Short)", context.execution.price_stop_loss_level

        return False, None, None

    def handle_stop_loss(self, context, pair, reason, z_score):
        stock_1, stock_2 = pair[0], pair[1]
        pair_key = f"{stock_1}/{stock_2}"

        # Get current prices and spread
        current_price_1 = PortfolioMakeOrder.get_current_price(context, stock_1)
        current_price_2 = PortfolioMakeOrder.get_current_price(context, stock_2)
        current_spread = context.strategy_pairs[context.strategy_pairs.index(pair)][2]['spread'][-1]

        # Close the position
        self.portfolio_order.order_target(context, stock_1, 0)
        self.portfolio_order.order_target(context, stock_2, 0)

        # Record the stop loss event
        if pair_key not in context.execution.stop_loss_history:
            context.execution.stop_loss_history[pair_key] = []

        context.execution.stop_loss_history[pair_key].append({
            'date': context.portfolio.current_date.strftime('%Y-%m-%d'),
            'reason': reason,
            'price_1': current_price_1,
            'price_2': current_price_2,
            'spread': current_spread,
            'z_score': z_score,
            'volatility_stop_loss_level': context.execution.volatility_stop_loss_level,
            'price_stop_loss_level': context.execution.price_stop_loss_level,
            'max_holding_period': context.execution.max_holding_period,
            'since_last_trade': (context.portfolio.current_date - context.portfolio.pair_trade_history[pair_key][
                -1].date).days if context.portfolio.pair_trade_history[pair_key] else None,
            'triggered_by': reason
        })

        # Reset the pair's trading state
        for p in context.strategy_pairs:
            if p[0] == stock_1 and p[1] == stock_2:
                p[2]['in_short'] = False
                p[2]['in_long'] = False
                break

        log.info(f"Stop Loss triggered for pair {pair_key}")

        # Implement re-evaluation logic here
        # For now, we'll just avoid re-entering for a certain period
        context.execution.price_level_stop_loss[pair_key] = None  # Reset price-level stop loss

        return [stock_1, stock_2,
                {'in_short': False, 'in_long': False, 'spread': p[2]['spread'], 'hedge_history': p[2]['hedge_history']}]

    def re_evaluate_pair(self, context, pair):
        stock_1, stock_2 = pair[0], pair[1]
        pair_key = f"{stock_1}/{stock_2}"

        # Implement your re-evaluation logic here
        # For example, you might want to:
        # 1. Check for any significant news or events
        # 2. Re-run your cointegration tests
        # 3. Analyze recent price movements

        # For now, we'll implement a simple cooling-off period
        last_stop_loss = context.execution.stop_loss_history[pair_key][-1]['date']

        # Convert last_stop_loss to Timestamp if it's a string
        if isinstance(last_stop_loss, str):
            last_stop_loss = pd.Timestamp(last_stop_loss)

        if (context.portfolio.current_date - last_stop_loss).days < context.execution.cooling_off_period:
            return False  # Don't re-enter yet

        # Instead of clearing the history, add a new event
        context.execution.stop_loss_history[pair_key].append({
            'date': context.portfolio.current_date.strftime('%Y-%m-%d'),
            'reason': "Cooling-off period ended",
            'price_1': None,
            'price_2': None,
            'spread': None,
            'z_score': None,
            'volatility_stop_loss_level': None,
            'price_stop_loss_level': None,
            'max_holding_period': None,
            'since_last_trade': None,
            'triggered_by': "Re-evaluation"
        })

        return True  # Allow re-entry


class PortfolioMakeOrder:
    def __init__(self, constant):
        self.constant = constant
        self.open_orders = {}

    def get_current_price(context, asset):
        if asset in context.portfolio.price_history and context.portfolio.price_history[asset]:
            return context.portfolio.price_history[asset][-1][1]
        else:
            # If we don't have the price in our history, try to get it from the data
            prices = context.data.history([asset], "Adj Close", 1, "1d")
            if not prices.empty:
                price = prices[asset][-1]
                # Update the price history
                if asset not in context.portfolio.price_history:
                    context.portfolio.price_history[asset] = []
                context.portfolio.price_history[asset].append((context.portfolio.current_date, price))
                return price
            else:
                raise ValueError(f"No price data available for {asset}")

    def get_open_orders(self):
        return self.open_orders

    def update_balance_sheet(self, context, asset, shares_to_trade, current_price, current_position):
        cost = shares_to_trade * current_price

        if abs(shares_to_trade) > 0:
            if shares_to_trade > 0:  # Buy or Short Cover
                if current_position >= 0:  # Buy (open or increase long position)
                    context.asset_cash -= cost
                    context.asset_securities[asset] = context.asset_securities.get(asset, 0) + cost
                else:  # Short Cover (close or reduce short position)
                    context.asset_cash -= cost
                    context.liability_securities[asset] = max(0, context.liability_securities.get(asset, 0) - cost)
            else:  # Sell or Short
                if current_position > 0:  # Sell (close or reduce long position)
                    context.asset_cash += abs(cost)
                    context.asset_securities[asset] = max(0, context.asset_securities.get(asset, 0) - abs(cost))
                else:  # Short (open or increase short position)
                    context.asset_cash += abs(cost)
                    context.liability_securities[asset] = context.liability_securities.get(asset, 0) + abs(cost)


    def update_position(self, context, asset, shares_to_trade, current_price):
        current_position = context.positions.get(asset, 0)
        new_position = current_position + shares_to_trade

        if asset not in context.cost_basis_history:
            context.cost_basis_history[asset] = []

        if current_position == 0:
            # Opening a new position (long or short)
            context.cost_basis_history[asset].append((context.current_date, current_price))
        elif (current_position > 0 and shares_to_trade > 0) or (current_position < 0 and shares_to_trade < 0):
            # Same direction: weighted average
            prev_cb = context.cost_basis_history[asset][-1][1] if context.cost_basis_history[asset] else current_price
            total_cost = abs(current_position) * prev_cb + abs(shares_to_trade) * current_price
            new_cost_basis = total_cost / abs(new_position)
            context.cost_basis_history[asset].append((context.current_date, new_cost_basis))
        elif (current_position > 0 and shares_to_trade < 0) or (current_position < 0 and shares_to_trade > 0):
            # Opposite direction: new trade offsets existing position
            if new_position == 0:
                # Fully closed
                context.cost_basis_history[asset].append((context.current_date, 0))
            elif abs(shares_to_trade) > abs(current_position):
                # Reversed: old position fully consumed, remainder is entirely new → use new price
                context.cost_basis_history[asset].append((context.current_date, current_price))
            else:
                # Partially reduced: old position survives, keep old cost basis
                pass

        context.positions[asset] = new_position


    def computeHoldingsPct(self, stock_1_shares, stock_2_shares, stock_1_price, stock_2_price):
        stock_1_Dol = stock_1_shares * stock_1_price
        stock_2_Dol = stock_2_shares * stock_2_price
        notionalDol = abs(stock_1_Dol) + abs(stock_2_Dol)
        stock_1_perc = stock_1_Dol / notionalDol
        stock_2_perc = stock_2_Dol / notionalDol
        return (stock_1_perc, stock_2_perc)


    def _apply_order_constraints(self, context, asset, shares_to_trade, current_price, current_position, target):
        # Only constrain opening long positions; closing short (shares_to_trade > 0 but target == 0) is unrestricted
        is_opening_long = target > 0 and current_position >= 0
        cost = shares_to_trade * current_price
        if is_opening_long and cost > context.portfolio.asset_cash:
            shares_to_trade = context.portfolio.asset_cash // current_price
        return shares_to_trade

    def order_target(self, context, asset, target):
        current_pair_key = (f"{context.current_pair[0]}/{context.current_pair[1]}"
                            if hasattr(context, 'current_pair') and context.current_pair else None)

        # Use pair-level position as the baseline so that shared symbols (e.g. CL in CL/WST and CL/SRE)
        # are treated independently per pair rather than saturating on the portfolio-level net position.
        if current_pair_key:
            pair_position = context.portfolio._pair_positions.get(current_pair_key, {}).get(asset, 0)
        else:
            pair_position = context.portfolio.positions.get(asset, 0)

        shares_to_trade = target - pair_position
        current_price = PortfolioMakeOrder.get_current_price(context, asset)
        portfolio_position = context.portfolio.positions.get(asset, 0)

        shares_to_trade = self._apply_order_constraints(context, asset, shares_to_trade, current_price, portfolio_position, target)

        if abs(shares_to_trade) > 0:
            self.update_balance_sheet(context.portfolio, asset, shares_to_trade, current_price, portfolio_position)
            self.update_position(context.portfolio, asset, shares_to_trade, current_price)
            trade = Trade(context.portfolio.current_date, asset, shares_to_trade, current_price,
                          'open' if target != 0 else 'close', pair=current_pair_key)
            context.portfolio.record_trade(trade, pair_key=current_pair_key)

        return shares_to_trade


    def order_target_percent(self, context, asset, target_percent):
        # Calculate total assets
        total_assets = context.portfolio.asset_cash + sum(
            max(shares, 0) * PortfolioMakeOrder.get_current_price(context, symbol)
            for symbol, shares in context.portfolio.positions.items()
        )

        # Calculate total liabilities
        total_liabilities = sum(
            abs(min(shares, 0)) * PortfolioMakeOrder.get_current_price(context, symbol)
            for symbol, shares in context.portfolio.positions.items()
        )

        # Calculate net equity
        net_equity = total_assets - total_liabilities

        # Calculate target value based on net equity
        target_value = net_equity * target_percent

        current_price = PortfolioMakeOrder.get_current_price(context, asset)
        target_shares = int(target_value / current_price)

        return self.order_target(context, asset, target_shares)


class PortfolioConstruct:
    def __init__(self, constant):
        self.constant = constant

    def hedge_ratio(sel, Y, X):
        # Prepare the data for Kalman Filter
        delta = 1e-5
        trans_cov = delta / (1 - delta) * np.eye(2)
        obs_mat = np.vstack([X, np.ones(X.shape)]).T[:, np.newaxis]

        # Initialize and run Kalman Filter
        kf = KalmanFilter(
            n_dim_obs=1,
            n_dim_state=2,
            initial_state_mean=np.zeros(2),
            initial_state_covariance=np.ones((2, 2)),
            transition_matrices=np.eye(2),
            observation_matrices=obs_mat,
            observation_covariance=1.0,
            transition_covariance=trans_cov
        )

        state_means, _ = kf.filter(Y.values)

        # Return the last slope estimate as the hedge ratio
        return state_means[-1, 0]

    def calculate_dynamic_z_scores_entry_exit(self, context, pair_key, current_volatility_of_spreads):
        # print(f"Start of function. volatilities_in_window for {pair_key}: {context.portfolio.volatilities_in_window[pair_key]}")

        volatilities_in_window = context.portfolio.volatilities_in_window[pair_key]

        # Calculate min and max volatilities from the rolling window
        min_volatility_in_window = min(volatilities_in_window)
        max_volatility_in_window = max(volatilities_in_window)

        # Normalize the current volatility
        if max_volatility_in_window > min_volatility_in_window:
            context.portfolio.normalized_volatility = (current_volatility_of_spreads - min_volatility_in_window) / (
                        max_volatility_in_window - min_volatility_in_window)
        else:
            context.portfolio.normalized_volatility = 0.5  # Default to middle value if min and max are the same

        # Calculate dynamic entry and exit z-scores

        context.execution.entry_z = context.execution.base_entry_z + context.execution.entry_volatility_factor * context.portfolio.normalized_volatility
        context.execution.exit_z = context.execution.base_exit_z + context.execution.exit_volatility_factor * context.portfolio.normalized_volatility

        # Ensure z-scores are within reasonable bounds
        context.execution.entry_z = max(context.execution.base_entry_z,
                                        min(context.execution.max_entry_z, context.execution.entry_z))
        context.execution.exit_z = max(context.execution.base_exit_z,
                                       min(context.execution.max_exit_z, context.execution.exit_z))

        # print(f"End of function. volatilities_in_window for {pair_key}: {context.portfolio.volatilities_in_window[pair_key]}")
        return context.portfolio.normalized_volatility, context.execution.entry_z, context.execution.exit_z


# ══════════════════════════════════════════════════════════════════════════════
# MTFS (Momentum Trend Following Strategy) Classes
# These classes are additive — existing MRPT classes above are NOT modified.
# ══════════════════════════════════════════════════════════════════════════════


class MomentumSignal:
    """Computes multi-window momentum scores, VAMS, trend confirmation,
    and trend-reversal detection for a single stock."""

    # Default lookback windows (trading days)
    DEFAULT_WINDOWS = [6, 12, 30, 60, 120, 150]
    # Default weights for composite scoring (tilt toward medium-term)
    DEFAULT_WEIGHTS = [0.10, 0.15, 0.20, 0.20, 0.20, 0.15]
    # Skip most-recent 21 trading days for windows >= this threshold
    SKIP_THRESHOLD = 60
    SKIP_DAYS = 21
    # SMA periods for trend confirmation
    SMA_SHORT = 50
    SMA_LONG = 200

    def __init__(self, windows=None, weights=None, skip_days=None):
        self.windows = windows or self.DEFAULT_WINDOWS
        self.weights = weights or self.DEFAULT_WEIGHTS
        self.skip_days = skip_days if skip_days is not None else self.SKIP_DAYS
        if len(self.windows) != len(self.weights):
            raise ValueError("windows and weights must have the same length")

    # ── single-stock helpers ──────────────────────────────────────────────

    def compute_return(self, prices, window):
        """Raw return over *window* trading days with optional skip-month.
        prices: pd.Series indexed by date, sorted ascending.
        Returns scalar float or np.nan if insufficient data."""
        skip = self.skip_days if window >= self.SKIP_THRESHOLD else 0
        required = window + skip
        if len(prices) < required + 1:
            return np.nan
        end_price = prices.iloc[-(1 + skip)] if skip > 0 else prices.iloc[-1]
        start_price = prices.iloc[-(1 + window)]
        if start_price == 0:
            return np.nan
        return (end_price / start_price) - 1.0

    def compute_realized_vol(self, prices, window):
        """Annualised realized volatility over *window* days."""
        if len(prices) < window + 1:
            return np.nan
        daily_ret = prices.pct_change().dropna().iloc[-window:]
        if len(daily_ret) < 2:
            return np.nan
        return daily_ret.std() * np.sqrt(252)

    def compute_vams(self, prices, window):
        """Volatility-Adjusted Momentum Score for one window."""
        ret = self.compute_return(prices, window)
        vol = self.compute_realized_vol(prices, window)
        if np.isnan(ret) or np.isnan(vol) or vol == 0:
            return np.nan
        return ret / vol

    def compute_sma(self, prices, period):
        """Simple moving average of the last *period* prices."""
        if len(prices) < period:
            return np.nan
        return prices.iloc[-period:].mean()

    @staticmethod
    def apply_llt(prices, n=10):
        """Low-Lag Trendline (LLT) filter — second-order low-pass filter.
        Reduces MA lag while preserving trend direction.
        n: period equivalent (alpha = 2/(n+1)).
        Returns a pd.Series of the same length as prices."""
        alpha = 2.0 / (n + 1)
        llt = prices.copy().astype(float).values
        p = prices.values
        for i in range(2, len(p)):
            llt[i] = (
                (alpha - alpha ** 2 / 4.0) * p[i]
                + (alpha ** 2 / 2.0) * p[i - 1]
                - (alpha - 3.0 * alpha ** 2 / 4.0) * p[i - 2]
                + 2.0 * (1.0 - alpha) * llt[i - 1]
                - (1.0 - alpha) ** 2 * llt[i - 2]
            )
        return pd.Series(llt, index=prices.index)

    def compute_llt_return(self, prices, window, llt_n=None):
        """Return over *window* days computed on LLT-smoothed prices.
        llt_n: LLT period; defaults to window//3 (capped 5..30) if None."""
        if llt_n is None:
            llt_n = max(5, min(30, window // 3))
        skip = self.skip_days if window >= self.SKIP_THRESHOLD else 0
        required = window + skip
        if len(prices) < required + 1:
            return np.nan
        llt = self.apply_llt(prices, n=llt_n)
        end_price = llt.iloc[-(1 + skip)] if skip > 0 else llt.iloc[-1]
        start_price = llt.iloc[-(1 + window)]
        if start_price == 0:
            return np.nan
        return float(end_price / start_price) - 1.0

    def compute_llt_vams(self, prices, window, llt_n=None):
        """VAMS computed on LLT-smoothed prices."""
        ret = self.compute_llt_return(prices, window, llt_n=llt_n)
        vol = self.compute_realized_vol(prices, window)
        if np.isnan(ret) or np.isnan(vol) or vol == 0:
            return np.nan
        return ret / vol

    # ── composite scoring ─────────────────────────────────────────────────

    def composite_raw_momentum(self, prices):
        """Weighted average of raw returns across all windows."""
        scores = []
        for w, wt in zip(self.windows, self.weights):
            r = self.compute_return(prices, w)
            if np.isnan(r):
                return np.nan
            scores.append(r * wt)
        return sum(scores)

    def composite_vams(self, prices):
        """Weighted average of VAMS across all windows."""
        scores = []
        for w, wt in zip(self.windows, self.weights):
            v = self.compute_vams(prices, w)
            if np.isnan(v):
                return np.nan
            scores.append(v * wt)
        return sum(scores)

    def composite_llt_momentum(self, prices, use_vams=True):
        """Weighted composite using LLT-smoothed returns (parallel to composite_vams/raw).
        use_vams=True: LLT return / realized vol; False: raw LLT return."""
        scores = []
        for w, wt in zip(self.windows, self.weights):
            v = self.compute_llt_vams(prices, w) if use_vams else self.compute_llt_return(prices, w)
            if np.isnan(v):
                return np.nan
            scores.append(v * wt)
        return sum(scores)

    def per_window_returns(self, prices):
        """Dict of {window: raw_return} for all windows."""
        return {w: self.compute_return(prices, w) for w in self.windows}

    def per_window_vams(self, prices):
        """Dict of {window: vams} for all windows."""
        return {w: self.compute_vams(prices, w) for w in self.windows}

    # ── trend confirmation ────────────────────────────────────────────────

    def trend_confirmed_long(self, prices):
        """True if Price > SMA_short AND SMA_short > SMA_long."""
        sma_s = self.compute_sma(prices, self.SMA_SHORT)
        sma_l = self.compute_sma(prices, self.SMA_LONG)
        if np.isnan(sma_s) or np.isnan(sma_l):
            return False
        return prices.iloc[-1] > sma_s and sma_s > sma_l

    def trend_confirmed_short(self, prices):
        """True if Price < SMA_short AND SMA_short < SMA_long."""
        sma_s = self.compute_sma(prices, self.SMA_SHORT)
        sma_l = self.compute_sma(prices, self.SMA_LONG)
        if np.isnan(sma_s) or np.isnan(sma_l):
            return False
        return prices.iloc[-1] < sma_s and sma_s < sma_l

    # ── trend reversal detection ──────────────────────────────────────────

    def momentum_decay_detected(self, prices, short_window=10, long_window=90):
        """True if short-term momentum has flipped sign relative to long-term.
        This is the primary momentum crash early-warning signal."""
        r_short = self.compute_return(prices, short_window)
        r_long = self.compute_return(prices, long_window)
        if np.isnan(r_short) or np.isnan(r_long):
            return False
        # Decay: long-term is positive but short-term has turned negative, or vice versa
        return (r_long > 0 and r_short < 0) or (r_long < 0 and r_short > 0)

    def sma_crossover_reversal(self, prices, lookback=5):
        """True if SMA_short crossed below SMA_long in the last *lookback* days
        (bearish for longs) OR crossed above (bearish for shorts).
        Returns (reversal_detected: bool, direction: str or None)."""
        if len(prices) < max(self.SMA_SHORT, self.SMA_LONG) + lookback:
            return False, None
        sma_s = prices.rolling(self.SMA_SHORT).mean()
        sma_l = prices.rolling(self.SMA_LONG).mean()
        recent_diff = (sma_s - sma_l).dropna().iloc[-lookback:]
        if len(recent_diff) < 2:
            return False, None
        # Check for sign change in the recent window
        signs = np.sign(recent_diff.values)
        for i in range(1, len(signs)):
            if signs[i] != signs[i - 1] and signs[i - 1] != 0:
                direction = 'bearish' if signs[i] < 0 else 'bullish'
                return True, direction
        return False, None

    def momentum_consistency(self, prices):
        """Returns a 0-1 score measuring how consistent momentum is across windows.
        1 = all windows agree on direction, 0 = maximum disagreement."""
        returns = self.per_window_returns(prices)
        valid = [r for r in returns.values() if not np.isnan(r)]
        if not valid:
            return 0.0
        positive_count = sum(1 for r in valid if r > 0)
        negative_count = sum(1 for r in valid if r < 0)
        return max(positive_count, negative_count) / len(valid)

    def trailing_factor_volatility(self, daily_pnl_series, window=60):
        """Annualised volatility of recent strategy daily returns.
        Used for momentum crash filter: if > 90th percentile of history, scale down."""
        if len(daily_pnl_series) < window:
            return np.nan
        return np.std(daily_pnl_series[-window:]) * np.sqrt(252)


# ════════════════════════════════════════════════════════════════════════════
# MTFS Statistical Test Classes
# Mirrors the ADF / Half_Life / Hurst / KPSS pattern used in MRPT,
# but tests momentum-specific properties instead of cointegration.
# ════════════════════════════════════════════════════════════════════════════

class MomentumDecayTest:
    """Detects short-term vs long-term momentum divergence.
    Primary trend-reversal early-warning signal.
    Analogous to Half_Life — both measure how quickly a regime is changing."""

    def __init__(self):
        self.short_window = 6
        self.long_window = 60
        self.short_return = None
        self.long_return = None
        self.decay_ratio = None      # short_ret / long_ret  (< 0 = divergence)
        self.is_decaying = None
        self.look_back = 60

    def apply(self, prices, ms):
        """prices: pd.Series, ms: MomentumSignal instance."""
        self.short_return = ms.compute_return(prices, self.short_window)
        self.long_return = ms.compute_return(prices, self.long_window)
        if np.isnan(self.short_return) or np.isnan(self.long_return):
            self.decay_ratio = np.nan
            self.is_decaying = None
            return
        if self.long_return == 0:
            self.decay_ratio = 0.0
        else:
            self.decay_ratio = self.short_return / self.long_return
        # Decay: long positive but short negative, or vice versa
        self.is_decaying = (self.long_return > 0 and self.short_return < 0) or \
                           (self.long_return < 0 and self.short_return > 0)

    def use(self):
        """True if momentum is NOT decaying (safe to hold)."""
        if self.is_decaying is None:
            return False
        return not self.is_decaying


class TrendStrengthTest:
    """ADX-like directional strength measure using rolling returns.
    High value = strong trend, low value = choppy / range-bound.
    Analogous to Hurst — both characterise the nature of the time series."""

    def __init__(self):
        self.look_back = 30
        self.strength_min = 0.3    # minimum to consider trend tradeable
        self.strength_max = 1.0
        self.trend_strength = None
        self.avg_abs_daily_return = None
        self.cumulative_return = None

    def apply(self, prices):
        """Efficiency ratio: |cum_return| / sum(|daily_returns|).
        1.0 = perfectly straight trend, 0.0 = pure noise."""
        if len(prices) < self.look_back + 1:
            self.trend_strength = np.nan
            return
        recent = prices.iloc[-(self.look_back + 1):]
        daily_ret = recent.pct_change().dropna()
        self.cumulative_return = (recent.iloc[-1] / recent.iloc[0]) - 1.0
        sum_abs = daily_ret.abs().sum()
        self.avg_abs_daily_return = daily_ret.abs().mean()
        if sum_abs == 0:
            self.trend_strength = 0.0
        else:
            self.trend_strength = abs(self.cumulative_return) / sum_abs

    def use(self):
        if self.trend_strength is None or np.isnan(self.trend_strength):
            return False
        return self.trend_strength >= self.strength_min


class MomentumConsistencyTest:
    """Measures cross-window directional agreement.
    Analogous to ADF — both are pass/fail gating tests for entry."""

    def __init__(self):
        self.consistency = None
        self.positive_windows = None
        self.negative_windows = None
        self.total_windows = None
        self.consistency_min = 0.6   # at least 60% windows must agree

    def apply(self, prices, ms):
        """ms: MomentumSignal instance."""
        returns = ms.per_window_returns(prices)
        valid = {w: r for w, r in returns.items() if not np.isnan(r)}
        self.total_windows = len(valid)
        if self.total_windows == 0:
            self.consistency = 0.0
            self.positive_windows = 0
            self.negative_windows = 0
            return
        self.positive_windows = sum(1 for r in valid.values() if r > 0)
        self.negative_windows = sum(1 for r in valid.values() if r < 0)
        self.consistency = max(self.positive_windows, self.negative_windows) / self.total_windows

    def use(self):
        if self.consistency is None:
            return False
        return self.consistency >= self.consistency_min


class VolatilityRegimeTest:
    """Tests whether current realized vol is in a normal or crash regime.
    Analogous to KPSS — both characterise the current regime state."""

    def __init__(self):
        self.look_back = 60
        self.current_vol = None          # annualised realized vol
        self.vol_percentile = None       # percentile rank vs full history
        self.crash_threshold = 0.90      # vol > 90th pctl = crash regime
        self.is_crash_regime = None

    def apply(self, prices):
        if len(prices) < self.look_back + 1:
            self.current_vol = np.nan
            self.vol_percentile = np.nan
            self.is_crash_regime = None
            return
        daily_ret = prices.pct_change().dropna()
        self.current_vol = daily_ret.iloc[-self.look_back:].std() * np.sqrt(252)
        # Rolling vol history for percentile
        if len(daily_ret) >= self.look_back * 2:
            rolling_vol = daily_ret.rolling(self.look_back).std().dropna() * np.sqrt(252)
            self.vol_percentile = (rolling_vol < self.current_vol).mean()
        else:
            self.vol_percentile = 0.5  # neutral if not enough history
        self.is_crash_regime = self.vol_percentile > self.crash_threshold

    def use(self):
        """True if NOT in crash regime (safe to trade)."""
        if self.is_crash_regime is None:
            return False
        return not self.is_crash_regime


class SMACrossoverTest:
    """Tests SMA crossover status for trend confirmation / reversal.
    Records the current SMA gap and recent crossover events."""

    def __init__(self):
        self.sma_short_period = 50
        self.sma_long_period = 200
        self.crossover_lookback = 5
        self.sma_short = None
        self.sma_long = None
        self.sma_gap = None              # (sma_short - sma_long) / price
        self.crossover_detected = None
        self.crossover_direction = None  # 'bullish' or 'bearish'

    def apply(self, prices, ms):
        if len(prices) < self.sma_long_period:
            self.sma_short = np.nan
            self.sma_long = np.nan
            self.sma_gap = np.nan
            self.crossover_detected = None
            self.crossover_direction = None
            return
        self.sma_short = ms.compute_sma(prices, self.sma_short_period)
        self.sma_long = ms.compute_sma(prices, self.sma_long_period)
        current_price = prices.iloc[-1]
        if current_price != 0:
            self.sma_gap = (self.sma_short - self.sma_long) / current_price
        else:
            self.sma_gap = 0.0
        self.crossover_detected, self.crossover_direction = \
            ms.sma_crossover_reversal(prices, self.crossover_lookback)

    def use(self):
        """True if trend alignment is bullish (SMA_short > SMA_long)."""
        if self.sma_gap is None or np.isnan(self.sma_gap):
            return False
        return self.sma_gap > 0


class MTFSExecution:
    """Execution parameters specific to Momentum Trend Following Strategy.
    Analogous to the MRPT Execution class but with momentum-specific params."""

    def __init__(self):
        # ── Momentum windows & scoring ────────────────────────────────────
        self.momentum_windows = [6, 12, 30, 60, 120, 150]
        self.momentum_weights = [0.20, 0.20, 0.20, 0.15, 0.15, 0.10]
        self.skip_days = 21  # skip-month for windows >= 60
        self.use_vams = True  # volatility-adjusted momentum scoring
        self.use_llt = False  # LLT-smoothed prices for momentum scoring

        # ── Trend confirmation ────────────────────────────────────────────
        self.sma_short = 20
        self.sma_long = 50
        self.require_trend_confirmation = True  # gate opens on SMA alignment

        # ── Momentum entry/exit thresholds ────────────────────────────────
        # Composite momentum score must exceed this to open a position.
        # For the long stock: composite > entry_momentum_threshold
        # For the short stock: composite < -entry_momentum_threshold
        self.entry_momentum_threshold = 0.0   # 0 means always enter (rely on ranking)
        self.exit_momentum_decay_threshold = 0.5  # consistency < this triggers exit

        # ── Trend reversal thresholds ─────────────────────────────────────
        self.reversal_sma_lookback = 3       # days to check for SMA crossover
        self.momentum_decay_short_window = 5
        self.momentum_decay_long_window = 30
        self.exit_on_reversal = True         # close if SMA crossover reversal detected
        self.exit_on_momentum_decay = True   # close if short/long momentum diverge

        # ── Volatility scaling (portfolio-level crash filter) ─────────────
        self.target_annual_vol = 0.10        # 10% target portfolio volatility
        self.vol_scale_window = 40           # trailing days for realized vol
        self.max_vol_scale_factor = 1.5      # cap leverage from vol scaling
        self.crash_vol_percentile = 0.85     # if trailing vol > this percentile, reduce
        self.crash_scale_factor = 0.20       # scale to 20% during crash regime

        # ── Position sizing & leverage ────────────────────────────────────
        self.amplifier = 2
        self.capital_utilization = 0.70       # fraction of net equity to deploy (70% default)
        self.use_vol_weighted_sizing = False  # True = inverse-vol weight within pair

        # ── Hedge ratio (pair dollar-neutrality) ──────────────────────────
        self.hedge_method = 'dollar_neutral'  # 'dollar_neutral' | 'beta_neutral' | 'kalman'
        self.hedge_lag = 1

        # ── Stop loss (analogous to MRPT) ─────────────────────────────────
        self.volatility_stop_loss_multiplier = 1.5
        self.max_holding_period = 10         # ~2 weeks for momentum
        self.cooling_off_period = 3
        self.pair_stop_loss_pct = 0.03       # -3% per pair
        self.price_level_stop_loss = {}
        self.stop_loss_history = {}

        # ── Rebalancing ───────────────────────────────────────────────────
        self.rebalance_frequency = 10        # trading days between full rebalances

        # ── Lookback for spread/vol (reused from MRPT for stop-loss calcs) ─
        self.mean_back = 20
        self.std_back = 20
        self.v_back = 20

        # ── Dynamic threshold bounds (analogous to MRPT) ─────────────────
        self.volatility_stop_loss_level = None
        self.price_stop_loss_level = None


class MTFSPortfolioConstruct:
    """Pair construction and hedge-ratio computation for MTFS.
    Handles dollar-neutral pairing, volatility-weighted sizing,
    and dynamic momentum-based entry/exit thresholds."""

    def __init__(self, constant=1):
        self.constant = constant
        self.momentum_signal = MomentumSignal()

    def compute_pair_hedge_ratio(self, stock_1_prices, stock_2_prices, method='dollar_neutral'):
        """Compute sizing ratio for an MTFS momentum pair.

        IMPORTANT: In MTFS the returned value is a *sizing ratio* (always > 0), NOT a
        cointegration hedge ratio.  The trade direction (long/short) is already determined
        by the momentum signal before this function is called.  The ratio only controls
        the relative notional allocation between the two legs:
            stock_1_shares =  1        (direction: long winner)
            stock_2_shares = -ratio    (direction: short loser)
        A negative ratio would flip the short leg to long, creating an unhedged
        directional bet — so we always enforce ratio > 0.

        Methods:
            'dollar_neutral':  ratio = price_1 / price_2  (equal dollar notional)
            'vol_neutral':     ratio = vol_2 / vol_1  (inverse-volatility weight;
                               allocates less notional to the more volatile leg)
            'beta_neutral':    |rolling OLS beta|, falling back to dollar_neutral
                               when beta <= 0 (MTFS pairs are often negatively
                               correlated by construction, so raw beta is unreliable)
            'kalman':          |Kalman-filter beta|, same fallback as beta_neutral
        """
        p1 = stock_1_prices.iloc[-1]
        p2 = stock_2_prices.iloc[-1]
        dollar_neutral_ratio = p1 / p2 if p2 != 0 else 1.0

        if method == 'dollar_neutral':
            return dollar_neutral_ratio

        elif method == 'vol_neutral':
            # Inverse-vol sizing: give less notional to the more volatile leg.
            r1 = stock_1_prices.pct_change().dropna()
            r2 = stock_2_prices.pct_change().dropna()
            window = min(20, len(r1), len(r2))
            if window < 5:
                return dollar_neutral_ratio
            vol_1 = r1.iloc[-window:].std()
            vol_2 = r2.iloc[-window:].std()
            if vol_1 == 0 or vol_2 == 0:
                return dollar_neutral_ratio
            # ratio = (vol_2 / vol_1) * (p1 / p2) keeps dollar-notional roughly equal
            # after adjusting for volatility differences
            return (vol_2 / vol_1) * dollar_neutral_ratio

        elif method == 'beta_neutral':
            # Rolling OLS beta: cov(r1, r2) / var(r2).
            # MTFS pairs are often negatively correlated (winner vs loser), so beta
            # can be negative.  We fall back to dollar_neutral in that case.
            r1 = stock_1_prices.pct_change().dropna()
            r2 = stock_2_prices.pct_change().dropna()
            window = min(60, len(r1), len(r2))
            if window < 10:
                return dollar_neutral_ratio
            r1_w = r1.iloc[-window:]
            r2_w = r2.iloc[-window:]
            cov = np.cov(r1_w, r2_w)[0, 1]
            var = np.var(r2_w)
            if var == 0:
                return dollar_neutral_ratio
            beta = cov / var
            if beta <= 0:
                # Negative beta: pair is negatively correlated over rolling window,
                # which is expected for MTFS.  Dollar-neutral is the safe fallback.
                return dollar_neutral_ratio
            return beta

        elif method == 'kalman':
            # Kalman-filter beta (same as MRPT).  Can be negative for MTFS pairs;
            # fall back to dollar_neutral when that happens.
            kalman_beta = PortfolioConstruct(constant=1).hedge_ratio(stock_1_prices, stock_2_prices)
            if kalman_beta <= 0:
                return dollar_neutral_ratio
            return kalman_beta

        else:
            raise ValueError(f"Unknown hedge method: {method}")

    def compute_momentum_scores(self, prices_dict):
        """Compute composite momentum scores for all stocks.

        Args:
            prices_dict: dict of {symbol: pd.Series of prices}

        Returns:
            dict of {symbol: {'composite_raw': float, 'composite_vams': float,
                              'per_window': dict, 'consistency': float,
                              'trend_long': bool, 'trend_short': bool}}
        """
        scores = {}
        for symbol, prices in prices_dict.items():
            composite_raw = self.momentum_signal.composite_raw_momentum(prices)
            composite_vams = self.momentum_signal.composite_vams(prices)
            per_window = self.momentum_signal.per_window_returns(prices)
            consistency = self.momentum_signal.momentum_consistency(prices)
            trend_long = self.momentum_signal.trend_confirmed_long(prices)
            trend_short = self.momentum_signal.trend_confirmed_short(prices)
            scores[symbol] = {
                'composite_raw': composite_raw,
                'composite_vams': composite_vams,
                'per_window': per_window,
                'consistency': consistency,
                'trend_long': trend_long,
                'trend_short': trend_short,
            }
        return scores

    def compute_vol_scale_factor(self, daily_pnl_history, execution):
        """Portfolio-level volatility scaling factor.
        Implements Barroso & Santa-Clara (2015) vol-targeting.

        Returns a multiplier in [crash_scale_factor, max_vol_scale_factor]."""
        if len(daily_pnl_history) < execution.vol_scale_window:
            return 1.0  # not enough history yet

        pnl_values = [v for _, v in daily_pnl_history]
        recent_pnl = pnl_values[-execution.vol_scale_window:]
        realized_vol = np.std(recent_pnl) * np.sqrt(252)

        if realized_vol == 0:
            return 1.0

        # Check for momentum crash regime
        all_vols = []
        for i in range(execution.vol_scale_window, len(pnl_values)):
            chunk = pnl_values[i - execution.vol_scale_window:i]
            all_vols.append(np.std(chunk) * np.sqrt(252))

        if all_vols:
            percentile_rank = sum(1 for v in all_vols if v <= realized_vol) / len(all_vols)
            if percentile_rank >= execution.crash_vol_percentile:
                return execution.crash_scale_factor

        # Normal vol-targeting
        # Convert realized_vol from dollar-space to return-space (approx)
        scale = execution.target_annual_vol / realized_vol
        return max(execution.crash_scale_factor, min(scale, execution.max_vol_scale_factor))

    def calculate_dynamic_momentum_thresholds(self, context, pair_key, stock_1_prices, stock_2_prices):
        """Calculate dynamic entry/exit thresholds based on pair momentum characteristics.

        Analogous to MRPT's calculate_dynamic_z_scores_entry_exit but for momentum.
        Returns (momentum_strength, entry_threshold, exit_threshold)."""
        ms = self.momentum_signal

        # Compute momentum strength as average of both stocks' consistency
        consistency_1 = ms.momentum_consistency(stock_1_prices)
        consistency_2 = ms.momentum_consistency(stock_2_prices)
        momentum_strength = (consistency_1 + consistency_2) / 2.0

        # Compute pair-level volatility from price ratio
        if len(stock_1_prices) > 30 and len(stock_2_prices) > 30:
            ratio = stock_1_prices / stock_2_prices
            ratio = ratio.dropna()
            if len(ratio) > 30:
                ratio_vol = ratio.pct_change().dropna().iloc[-30:].std() * np.sqrt(252)
            else:
                ratio_vol = 0.5
        else:
            ratio_vol = 0.5

        # Track volatilities for normalization (like MRPT)
        if pair_key not in context.portfolio.volatilities_in_window:
            context.portfolio.volatilities_in_window[pair_key] = []
        context.portfolio.volatilities_in_window[pair_key].append(ratio_vol)
        max_v_back = context.execution.v_back // 2
        context.portfolio.volatilities_in_window[pair_key] = \
            context.portfolio.volatilities_in_window[pair_key][-max_v_back:]

        vols = context.portfolio.volatilities_in_window[pair_key]
        min_v = min(vols)
        max_v = max(vols)
        if max_v > min_v:
            normalized_vol = (ratio_vol - min_v) / (max_v - min_v)
        else:
            normalized_vol = 0.5

        context.portfolio.normalized_volatility = normalized_vol

        # Dynamic entry: higher vol → harder to enter (raise threshold)
        entry_threshold = context.execution.entry_momentum_threshold + 0.3 * normalized_vol
        # Dynamic exit: higher vol → easier to exit (lower consistency required)
        exit_threshold = max(0.2, context.execution.exit_momentum_decay_threshold - 0.2 * normalized_vol)

        return momentum_strength, entry_threshold, exit_threshold


class MTFSStopLossFunction:
    """Stop-loss functions adapted for Momentum Trend Following Strategy.
    Mirrors PortfolioStopLossFunction interface but uses momentum-based triggers."""

    def __init__(self, constant=1):
        self.constant = constant
        self.portfolio_order = PortfolioMakeOrder(constant=1)
        self.momentum_signal = MomentumSignal()

    def check_momentum_reversal_stop(self, context, pair, stock_1_prices, stock_2_prices):
        """Stop loss triggered by trend reversal detection.
        This is the MTFS-specific critical risk control."""
        pair_key = f"{pair[0]}/{pair[1]}"
        in_long = pair[2].get('in_long', False)
        in_short = pair[2].get('in_short', False)

        if not in_long and not in_short:
            return False, None, None

        ms = self.momentum_signal

        # Check momentum decay (short vs long window divergence)
        if context.execution.exit_on_momentum_decay:
            stock_1_sym = pair[0]
            stock_2_sym = pair[1]

            if in_long:
                # Long stock_1 (winner), short stock_2 (loser)
                # Danger: winner's short-term momentum turns negative
                decay_1 = ms.momentum_decay_detected(
                    stock_1_prices,
                    context.execution.momentum_decay_short_window,
                    context.execution.momentum_decay_long_window
                )
                # Danger: loser's short-term momentum turns positive (recovering)
                decay_2 = ms.momentum_decay_detected(
                    stock_2_prices,
                    context.execution.momentum_decay_short_window,
                    context.execution.momentum_decay_long_window
                )
                if decay_1 and decay_2:
                    return True, "Momentum Decay (both stocks reversing)", None

            elif in_short:
                decay_1 = ms.momentum_decay_detected(
                    stock_1_prices,
                    context.execution.momentum_decay_short_window,
                    context.execution.momentum_decay_long_window
                )
                decay_2 = ms.momentum_decay_detected(
                    stock_2_prices,
                    context.execution.momentum_decay_short_window,
                    context.execution.momentum_decay_long_window
                )
                if decay_1 and decay_2:
                    return True, "Momentum Decay (both stocks reversing)", None

        # Check SMA crossover reversal
        if context.execution.exit_on_reversal:
            if in_long:
                rev_1, dir_1 = ms.sma_crossover_reversal(
                    stock_1_prices, context.execution.reversal_sma_lookback)
                rev_2, dir_2 = ms.sma_crossover_reversal(
                    stock_2_prices, context.execution.reversal_sma_lookback)
                # Long winner reversing bearish = danger
                if rev_1 and dir_1 == 'bearish':
                    return True, f"SMA Reversal ({pair[0]} bearish crossover)", None
                # Short loser reversing bullish = danger
                if rev_2 and dir_2 == 'bullish':
                    return True, f"SMA Reversal ({pair[1]} bullish crossover)", None

            elif in_short:
                rev_1, dir_1 = ms.sma_crossover_reversal(
                    stock_1_prices, context.execution.reversal_sma_lookback)
                rev_2, dir_2 = ms.sma_crossover_reversal(
                    stock_2_prices, context.execution.reversal_sma_lookback)
                if rev_1 and dir_1 == 'bullish':
                    return True, f"SMA Reversal ({pair[0]} bullish crossover)", None
                if rev_2 and dir_2 == 'bearish':
                    return True, f"SMA Reversal ({pair[1]} bearish crossover)", None

        return False, None, None

    def check_pair_pnl_stop_loss(self, context, pair):
        """Stop loss based on per-pair P&L percentage.
        Triggers if pair has lost more than pair_stop_loss_pct."""
        pair_key = f"{pair[0]}/{pair[1]}"
        stock_1, stock_2 = pair[0], pair[1]

        # Get current positions and prices
        pos_1 = context.portfolio.positions.get(stock_1, 0)
        pos_2 = context.portfolio.positions.get(stock_2, 0)
        if pos_1 == 0 and pos_2 == 0:
            return False, None, None

        price_1 = PortfolioMakeOrder.get_current_price(context, stock_1)
        price_2 = PortfolioMakeOrder.get_current_price(context, stock_2)

        # Get cost basis
        cb_1 = context.portfolio.cost_basis_history.get(stock_1, [(None, 0)])[-1][1]
        cb_2 = context.portfolio.cost_basis_history.get(stock_2, [(None, 0)])[-1][1]

        # Calculate pair P&L
        pnl_1 = pos_1 * (price_1 - cb_1) if cb_1 != 0 else 0
        pnl_2 = pos_2 * (price_2 - cb_2) if cb_2 != 0 else 0
        pair_pnl = pnl_1 + pnl_2

        # Calculate notional
        notional = abs(pos_1) * cb_1 + abs(pos_2) * cb_2
        if notional == 0:
            return False, None, None

        pnl_pct = pair_pnl / notional
        if pnl_pct < -context.execution.pair_stop_loss_pct:
            return True, f"Pair P&L Stop Loss ({pnl_pct:.2%})", pnl_pct

        return False, None, None

    def check_volatility_stop_loss(self, context, pair, stock_1_prices, stock_2_prices):
        """Volatility-based stop loss using price ratio volatility.
        Analogous to MRPT's spread-based volatility stop."""
        pair_key = f"{pair[0]}/{pair[1]}"

        # Use price ratio as the "spread" analog
        if len(stock_1_prices) < context.execution.mean_back or len(stock_2_prices) < context.execution.mean_back:
            return False, None, None

        ratio = (stock_1_prices / stock_2_prices).dropna()
        if len(ratio) < context.execution.mean_back:
            return False, None, None

        mean_ratio = ratio.iloc[-context.execution.mean_back:].mean()
        std_ratio = ratio.iloc[-context.execution.std_back:].std()
        current_ratio = ratio.iloc[-1]

        in_long = pair[2].get('in_long', False)
        in_short = pair[2].get('in_short', False)

        if in_long:
            stop_level = mean_ratio - context.execution.volatility_stop_loss_multiplier * std_ratio
            context.execution.volatility_stop_loss_level = stop_level
            if current_ratio < stop_level:
                return True, "Volatility Stop Loss (Long)", stop_level
        elif in_short:
            stop_level = mean_ratio + context.execution.volatility_stop_loss_multiplier * std_ratio
            context.execution.volatility_stop_loss_level = stop_level
            if current_ratio > stop_level:
                return True, "Volatility Stop Loss (Short)", stop_level

        return False, None, None

    def check_time_based_stop_loss(self, context, pair):
        """Reuse the same time-based stop loss logic as MRPT."""
        pair_key = f"{pair[0]}/{pair[1]}"
        if pair_key in context.portfolio.pair_trade_history:
            trades = context.portfolio.pair_trade_history[pair_key]
            if not trades:
                return False, None, None
            last_trade = trades[-1]

            stock = pair[0]
            historical_data = context.data.history(
                [stock], "price", context.execution.max_holding_period * 2, "1d")

            if last_trade.date not in historical_data.index:
                return True, "Time-based Stop Loss", None

            last_trade_index = historical_data.index.get_loc(last_trade.date)
            current_date_index = historical_data.index.get_loc(context.portfolio.current_date)
            trading_days_since = current_date_index - last_trade_index

            if trading_days_since > context.execution.max_holding_period:
                return True, "Time-based Stop Loss", None
        return False, None, None

    def handle_stop_loss(self, context, pair, reason, momentum_score=None):
        """Close positions and record the stop loss event.
        Same interface as MRPT's handle_stop_loss."""
        stock_1, stock_2 = pair[0], pair[1]
        pair_key = f"{stock_1}/{stock_2}"

        current_price_1 = PortfolioMakeOrder.get_current_price(context, stock_1)
        current_price_2 = PortfolioMakeOrder.get_current_price(context, stock_2)

        # Close both positions
        self.portfolio_order.order_target(context, stock_1, 0)
        self.portfolio_order.order_target(context, stock_2, 0)

        # Record stop loss event
        if pair_key not in context.execution.stop_loss_history:
            context.execution.stop_loss_history[pair_key] = []

        context.execution.stop_loss_history[pair_key].append({
            'date': context.portfolio.current_date.strftime('%Y-%m-%d'),
            'reason': reason,
            'price_1': current_price_1,
            'price_2': current_price_2,
            'spread': current_price_1 / current_price_2 if current_price_2 != 0 else None,
            'z_score': momentum_score,
            'volatility_stop_loss_level': context.execution.volatility_stop_loss_level,
            'price_stop_loss_level': context.execution.price_stop_loss_level,
            'max_holding_period': context.execution.max_holding_period,
            'since_last_trade': (
                context.portfolio.current_date - context.portfolio.pair_trade_history[pair_key][-1].date
            ).days if context.portfolio.pair_trade_history.get(pair_key) else None,
            'triggered_by': reason
        })

        # Reset pair state
        for p in context.strategy_pairs:
            if p[0] == stock_1 and p[1] == stock_2:
                p[2]['in_short'] = False
                p[2]['in_long'] = False
                break

        log.info(f"MTFS Stop Loss triggered for pair {pair_key}: {reason}")
        context.execution.price_level_stop_loss[pair_key] = None

        return [stock_1, stock_2,
                {'in_short': False, 'in_long': False,
                 'momentum_history': pair[2].get('momentum_history', {})}]

    def re_evaluate_pair(self, context, pair):
        """Cooling-off period check — same logic as MRPT."""
        stock_1, stock_2 = pair[0], pair[1]
        pair_key = f"{stock_1}/{stock_2}"

        last_stop_loss = context.execution.stop_loss_history[pair_key][-1]['date']
        if isinstance(last_stop_loss, str):
            last_stop_loss = pd.Timestamp(last_stop_loss)

        if (context.portfolio.current_date - last_stop_loss).days < context.execution.cooling_off_period:
            return False

        context.execution.stop_loss_history[pair_key].append({
            'date': context.portfolio.current_date.strftime('%Y-%m-%d'),
            'reason': "Cooling-off period ended",
            'price_1': None, 'price_2': None, 'spread': None, 'z_score': None,
            'volatility_stop_loss_level': None, 'price_stop_loss_level': None,
            'max_holding_period': None, 'since_last_trade': None,
            'triggered_by': "Re-evaluation"
        })
        return True

