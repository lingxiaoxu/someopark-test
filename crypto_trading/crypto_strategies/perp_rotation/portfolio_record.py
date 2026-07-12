"""
portfolio_record.py — Perp-rotation Excel record infrastructure (Plan 05 §5)
============================================================================
COPIED from qlib-main/sector_rotation/portfolio_record.py (read-only template,
880 lines). Produces the same standardized Excel outputs:

  1. Portfolio History Excel (26 template sheets + 1 crypto sheet) — full record
  2. Monitor Excel (5 sheets) — daily rebalance snapshot
  3. WF Diagnostic Excel (5-6 sheets) — walk-forward analysis

Pure recording/export layer over daily_engine.BacktestResult — no engine logic.

ADAPTATIONS (only these; sheet names/layout verbatim otherwise):
  * Output dir → trading_signals/perp_rotation/records/ (inside the tree).
  * "+ funding-paid column" (Plan 05 reuse map): portfolio_history and
    daily_pnl gain Daily_Funding / Cum_Funding columns and a new
    ``funding_history`` sheet — sourced from BacktestResult.funding_pnl_daily.
  * Leverage interest accrual /252 → /365.
  * rebalance_trades Est_Cost: flat 5bps → taker fee via costs.load_fee_rates.
  * regime_indicators macro columns → crypto regime features.
  * stop_loss_history: spy_return_3d → benchmark_return_3d (KXBTCPERP 3d).
  * strategy_vars: sector config keys → perp config keys (carry replaces
    relative_value; rvol flags replace VIX flags).
"""

from __future__ import annotations

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict

import numpy as np
import pandas as pd

from crypto_trading.crypto_common import config as _config
from crypto_trading.crypto_common.costs import load_fee_rates

log = logging.getLogger(__name__)


def _records_dir() -> Path:
    return _config.SIGNALS_DIR / "perp_rotation" / "records"


# ═══════════════════════════════════════════════════════════════════════════
#  Helper: daily weight reconstruction (template verbatim)
# ═══════════════════════════════════════════════════════════════════════════

def _reconstruct_daily_weights(weights_history: pd.DataFrame,
                               prices: pd.DataFrame,
                               equity_curve: pd.Series) -> pd.DataFrame:
    """Rebalance-date weights → daily drifting weights (shares held constant)."""
    tickers = weights_history.columns.tolist()
    all_dates = equity_curve.index
    daily_weights = pd.DataFrame(0.0, index=all_dates, columns=tickers)
    shares_held = pd.Series(0.0, index=tickers)
    reb_dates = sorted(weights_history.index)
    reb_idx = 0
    for dt in all_dates:
        if reb_idx < len(reb_dates) and dt >= reb_dates[reb_idx]:
            reb_dt = reb_dates[reb_idx]
            w = weights_history.loc[reb_dt]
            eq = equity_curve.get(dt, equity_curve.iloc[-1])
            for t in tickers:
                p = prices[t].get(dt, np.nan) if t in prices.columns else np.nan
                if pd.notna(p) and p > 0 and eq > 0:
                    shares_held[t] = w.get(t, 0.0) * eq / p
                else:
                    shares_held[t] = 0.0
            reb_idx += 1
            while reb_idx < len(reb_dates) and reb_dates[reb_idx] <= dt:
                reb_idx += 1
        total_value = 0.0
        values = {}
        for t in tickers:
            p = prices[t].get(dt, np.nan) if t in prices.columns else np.nan
            if pd.notna(p) and shares_held[t] != 0:
                v = shares_held[t] * p
                values[t] = v
                total_value += v
        if total_value > 0:
            for t in tickers:
                daily_weights.loc[dt, t] = values.get(t, 0.0) / total_value
    return daily_weights


def _compute_daily_shares(weights_history: pd.DataFrame, equity_curve: pd.Series,
                          prices: pd.DataFrame) -> pd.DataFrame:
    """Daily contracts held per perp (template's share tracker)."""
    tickers = weights_history.columns.tolist()
    all_dates = equity_curve.index
    shares = pd.DataFrame(0.0, index=all_dates, columns=tickers)
    current_shares = pd.Series(0.0, index=tickers)
    reb_set = set(sorted(weights_history.index))
    for dt in all_dates:
        if dt in reb_set:
            w = weights_history.loc[dt]
            eq = equity_curve.get(dt, 0)
            for t in tickers:
                p = prices[t].get(dt, np.nan) if t in prices.columns else np.nan
                if pd.notna(p) and p > 0 and eq > 0:
                    current_shares[t] = w.get(t, 0.0) * eq / p
                else:
                    current_shares[t] = 0.0
        shares.loc[dt] = current_shares
    return shares


def _compute_cost_basis(weights_history: pd.DataFrame, prices: pd.DataFrame,
                        equity_curve: pd.Series) -> pd.DataFrame:
    """Entry-price tracker: set on 0→pos, weighted-avg on adds, reset on exit."""
    tickers = weights_history.columns.tolist()
    all_dates = equity_curve.index
    cost_basis = pd.DataFrame(0.0, index=all_dates, columns=tickers)
    cb = pd.Series(0.0, index=tickers)
    prev_shares = pd.Series(0.0, index=tickers)
    reb_set = set(sorted(weights_history.index))
    for dt in all_dates:
        if dt in reb_set:
            w = weights_history.loc[dt]
            eq = equity_curve.get(dt, 0)
            for t in tickers:
                p = prices[t].get(dt, np.nan) if t in prices.columns else np.nan
                new_w = w.get(t, 0.0)
                if new_w < 1e-4:
                    cb[t] = 0.0
                    prev_shares[t] = 0.0
                elif prev_shares[t] < 1e-6:
                    cb[t] = p if pd.notna(p) else 0.0
                    prev_shares[t] = new_w * eq / p if pd.notna(p) and p > 0 else 0.0
                else:
                    new_shares = new_w * eq / p if pd.notna(p) and p > 0 else 0.0
                    if new_shares > prev_shares[t] + 1e-6:
                        added = new_shares - prev_shares[t]
                        total = prev_shares[t] + added
                        if total > 0:
                            cb[t] = (cb[t] * prev_shares[t] + p * added) / total
                    prev_shares[t] = new_shares
        cost_basis.loc[dt] = cb
    return cost_basis


# ═══════════════════════════════════════════════════════════════════════════
#  Main export class (template shape)
# ═══════════════════════════════════════════════════════════════════════════

class PerpRotationRecord:
    """Extract data from daily_engine.BacktestResult → standardized Excel."""

    def __init__(self, result, prices: pd.DataFrame, macro: pd.DataFrame = None,
                 param_set: str = "", signal_version: str = "v1",
                 leverage_ratio: float = 0.0, interest_rate: float = 0.05):
        self.result = result
        self.prices = prices
        self.macro = macro if macro is not None else pd.DataFrame()
        self.param_set = param_set
        self.signal_version = signal_version
        self.leverage_ratio = leverage_ratio
        self.interest_rate = interest_rate
        self.equity_curve = result.equity_curve
        self.daily_returns = result.daily_returns
        self.weights_history = result.weights_history
        self.metrics = result.metrics
        self.tickers = [c for c in self.weights_history.columns
                        if c not in ("cash", "CASH")]

    # ──────────────────────────────────────────────────────────────
    #  Portfolio History Excel (26 template sheets + funding_history)
    # ──────────────────────────────────────────────────────────────

    def export_portfolio_excel(self, output_path: Path = None,
                               mode: str = "batch", span: str = "IS") -> Path:
        if output_path is None:
            _records_dir().mkdir(parents=True, exist_ok=True)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = _records_dir() / (
                f"pr_portfolio_{self.param_set}_{self.signal_version}_{span}_{mode}_{ts}.xlsx")

        from openpyxl import Workbook
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        ec = self.equity_curve.dropna()
        dr = self.daily_returns.dropna()
        wh = self.weights_history
        m = self.metrics

        daily_pnl = dr * ec.shift(1).reindex(dr.index).fillna(ec.iloc[0])
        cum_pnl = daily_pnl.cumsum()
        running_max = ec.cummax()
        dd_pct = (ec - running_max) / running_max
        dd_dollar = ec - running_max

        lr = self.leverage_ratio
        asset = ec * (1 + lr)
        liability = ec * lr
        # ADAPTED: 24/7 interest accrual (template /252)
        interest_daily = liability * self.interest_rate / 365 if lr > 0 else ec * 0

        # crypto funding series (Plan 05 "+ funding-paid column")
        funding_daily = (self.result.funding_pnl_daily
                         if getattr(self.result, "funding_pnl_daily", None) is not None
                         else pd.Series(0.0, index=ec.index))
        funding_daily = funding_daily.reindex(ec.index).fillna(0.0)
        funding_cum = funding_daily.cumsum()

        shares_df = daily_w = cb_df = None

        def _get_shares():
            nonlocal shares_df
            if shares_df is None:
                shares_df = _compute_daily_shares(wh, ec, self.prices)
            return shares_df

        def _get_daily_w():
            nonlocal daily_w
            if daily_w is None:
                daily_w = _reconstruct_daily_weights(wh, self.prices, ec)
            return daily_w

        def _get_cb():
            nonlocal cb_df
            if cb_df is None:
                cb_df = _compute_cost_basis(wh, self.prices, ec)
            return cb_df

        # ── 1. summary ──
        ws = wb.create_sheet("summary")
        ws.append(["Metric", "Value"])
        for k in ["sharpe", "calmar", "annual_return", "annual_vol", "max_drawdown",
                  "total_return", "monthly_win_rate", "info_ratio", "beta",
                  "funding_pnl_usd", "ew_total_return"]:
            v = m.get(k)
            ws.append([k, round(v, 4) if isinstance(v, (int, float)) and v is not None
                       and not (isinstance(v, float) and np.isnan(v)) else "N/A"])
        ws.append(["param_set", self.param_set])
        ws.append(["signal_version", self.signal_version])
        ws.append(["leverage_ratio", self.leverage_ratio])

        # ── 2. portfolio_history (+ funding columns) ──
        ws = wb.create_sheet("portfolio_history")
        ws.append(["Date", "Equity", "Asset", "Liability", "Daily_PnL", "Cum_PnL",
                   "Daily_Funding", "Cum_Funding", "Drawdown_Pct"])
        for dt in ec.index:
            ws.append([str(dt.date()), round(float(ec[dt]), 2),
                       round(float(asset[dt]), 2), round(float(liability[dt]), 2),
                       round(float(daily_pnl.get(dt, 0)), 2),
                       round(float(cum_pnl.get(dt, 0)), 2),
                       round(float(funding_daily.get(dt, 0)), 4),
                       round(float(funding_cum.get(dt, 0)), 4),
                       round(float(dd_pct.get(dt, 0)), 4)])

        # ── 3-5. asset / liability / equity history ──
        for name, series in (("asset_history", asset), ("liability_history", liability),
                             ("equity_history", ec)):
            ws = wb.create_sheet(name)
            col = {"asset_history": "Total_Asset", "liability_history": "Total_Liability",
                   "equity_history": "Net_Equity"}[name]
            ws.append(["Date", col])
            for dt in ec.index:
                ws.append([str(dt.date()), round(float(series[dt]), 2)])

        # ── 6. asset_cash_history ──
        ws = wb.create_sheet("asset_cash_history")
        ws.append(["Date", "Cash"])
        cash_w = pd.Series(0.0, index=ec.index)
        for rf in self.result.risk_flags:
            d = rf.get("date")
            if d is not None:
                try:
                    cash_w.loc[d:] = rf.get("cash_pct", 0.0)
                except (KeyError, TypeError):
                    pass
        for dt in ec.index:
            ws.append([str(dt.date()), round(float(ec[dt] * cash_w.get(dt, 0.0)), 2)])

        # ── 7. sector_prices (perp contract mids) ──
        ws = wb.create_sheet("sector_prices")
        ws.append(["Date"] + self.tickers)
        etf_p = self.prices.reindex(columns=self.tickers).reindex(ec.index)
        for dt in ec.index:
            row = [str(dt.date())]
            for t in self.tickers:
                v = etf_p[t].get(dt, np.nan)
                row.append(round(float(v), 4) if pd.notna(v) else "")
            ws.append(row)

        # ── 8. share_history (contracts) ──
        ws = wb.create_sheet("share_history")
        ws.append(["Date"] + self.tickers)
        sh = _get_shares()
        for dt in ec.index:
            ws.append([str(dt.date())] + [round(float(sh.loc[dt, t]), 2)
                                          if t in sh.columns else 0 for t in self.tickers])

        # ── 9. sector_weights (rebalance dates) ──
        ws = wb.create_sheet("sector_weights")
        ws.append(["Date"] + self.tickers + ["Cash"])
        for dt in wh.index:
            row = [str(dt.date())]
            total_w = 0.0
            for t in self.tickers:
                w = float(wh.loc[dt].get(t, 0.0))
                row.append(round(w, 4))
                total_w += w
            row.append(round(1.0 - total_w, 4))
            ws.append(row)

        # ── 10. sector_weight_pct (daily drift) ──
        ws = wb.create_sheet("sector_weight_pct")
        ws.append(["Date"] + self.tickers)
        dw = _get_daily_w()
        for dt in ec.index:
            ws.append([str(dt.date())] + [round(float(dw.loc[dt, t]), 4)
                                          if t in dw.columns else 0 for t in self.tickers])

        # ── 11. cost_basis ──
        ws = wb.create_sheet("cost_basis")
        ws.append(["Date"] + self.tickers)
        cb = _get_cb()
        for dt in ec.index:
            ws.append([str(dt.date())] + [round(float(cb.loc[dt, t]), 4)
                                          if t in cb.columns else 0 for t in self.tickers])

        # ── 12. sector_ratio_matrix ──
        ws = wb.create_sheet("sector_ratio_matrix")
        for dt in wh.index:
            ws.append([f"=== {dt.date()} ==="])
            ws.append([""] + self.tickers)
            w_row = wh.loc[dt]
            for t_i in self.tickers:
                row = [t_i]
                wi = float(w_row.get(t_i, 0.0))
                for t_j in self.tickers:
                    wj = float(w_row.get(t_j, 0.0))
                    row.append(round(wi / wj, 3) if wi > 1e-4 and wj > 1e-4 else 0)
                ws.append(row)
            ws.append([])

        # ── 13-14. sector_pnl_acc + sector_pnl_daily (proportional attribution)
        ws_acc = wb.create_sheet("sector_pnl_acc")
        ws_dod = wb.create_sheet("sector_pnl_daily")
        ws_acc.append(["Date"] + [f"{t}_pnl" for t in self.tickers])
        ws_dod.append(["Date"] + [f"{t}_pnl" for t in self.tickers])
        etf_ret = etf_p.pct_change().reindex(ec.index).fillna(0)
        dw = _get_daily_w()
        weighted_ret = dw * etf_ret
        total_weighted = weighted_ret.sum(axis=1).replace(0, np.nan)
        sector_daily_pnl = pd.DataFrame(0.0, index=ec.index, columns=self.tickers)
        for t in self.tickers:
            sector_daily_pnl[t] = (weighted_ret[t] / total_weighted).fillna(0) * daily_pnl
        sector_cum_pnl = sector_daily_pnl.cumsum()
        for dt in ec.index:
            ws_acc.append([str(dt.date())] + [round(float(sector_cum_pnl.loc[dt, t]), 2)
                                              for t in self.tickers])
            ws_dod.append([str(dt.date())] + [round(float(sector_daily_pnl.loc[dt, t]), 2)
                                              for t in self.tickers])

        # ── 15. sector_contribution ──
        ws = wb.create_sheet("sector_contribution")
        ws.append(["Date"] + [f"{t}_contrib_pct" for t in self.tickers])
        for dt in ec.index:
            dp = float(daily_pnl.get(dt, 0))
            row = [str(dt.date())]
            for t in self.tickers:
                sp = float(sector_daily_pnl.loc[dt, t])
                row.append(round(sp / dp * 100, 2) if abs(dp) > 0.01 else 0)
            ws.append(row)

        # ── 16. daily_pnl (+ funding columns) ──
        ws = wb.create_sheet("daily_pnl")
        ws.append(["Date", "Daily_PnL", "Cum_PnL", "Daily_Funding", "Cum_Funding"])
        for dt in ec.index:
            ws.append([str(dt.date()), round(float(daily_pnl.get(dt, 0)), 2),
                       round(float(cum_pnl.get(dt, 0)), 2),
                       round(float(funding_daily.get(dt, 0)), 4),
                       round(float(funding_cum.get(dt, 0)), 4)])

        # ── 17-18. interest ──
        ws = wb.create_sheet("interest_expense")
        ws.append(["Date", "Daily_Interest"])
        for dt in ec.index:
            ws.append([str(dt.date()), round(float(interest_daily.get(dt, 0)), 2)])
        ws = wb.create_sheet("acc_interest")
        ws.append(["Date", "Cum_Interest"])
        acc_int = interest_daily.cumsum()
        for dt in ec.index:
            ws.append([str(dt.date()), round(float(acc_int.get(dt, 0)), 2)])

        # ── 19. realized_pnl ──
        ws = wb.create_sheet("realized_pnl")
        ws.append(["Perp", "Realized_PnL"])
        for t, v in self._compute_realized_pnl(sector_cum_pnl).items():
            ws.append([t, round(v, 2)])

        # ── 20. total_notional ──
        ws = wb.create_sheet("total_notional")
        ws.append(["Perp", "Total_Notional"])
        for t, v in self._compute_total_notional().items():
            ws.append([t, round(v, 2)])

        # ── 21. drawdown_history ──
        ws = wb.create_sheet("drawdown_history")
        ws.append(["Date", "Drawdown_Dollar", "Drawdown_Pct"])
        for dt in ec.index:
            ws.append([str(dt.date()), round(float(dd_dollar.get(dt, 0)), 2),
                       round(float(dd_pct.get(dt, 0)), 4)])

        # ── 22. rebalance_trades ──
        ws = wb.create_sheet("rebalance_trades")
        ws.append(["Date", "Perp", "Direction", "Old_Weight", "New_Weight",
                   "Contracts", "Price", "Est_Cost"])
        self._write_rebalance_trades(ws)

        # ── 23. regime_indicators (crypto features) ──
        ws = wb.create_sheet("regime_indicators")
        macro_cols = [c for c in ["btc_rvol", "funding", "basis_dispersion",
                                  "btc_dominance", "offshore_funding", "stress", "flow"]
                      if c in self.macro.columns]
        ws.append(["Date", "Regime"] + macro_cols)
        regime = self.result.regime_history
        for dt in ec.index:
            r = regime.get(dt, "") if dt in regime.index else ""
            row = [str(dt.date()), str(r)]
            for c in macro_cols:
                v = self.macro[c].get(dt, np.nan) if dt in self.macro.index else np.nan
                row.append(round(float(v), 4) if pd.notna(v) else "")
            ws.append(row)

        # ── 24. strategy_vars ──
        ws = wb.create_sheet("strategy_vars")
        self._write_strategy_vars(ws)

        # ── 25. stop_loss_history ──
        ws = wb.create_sheet("stop_loss_history")
        ws.append(["Date", "Perp", "Type", "Reason", "Entry_Price", "Current_Price",
                   "Threshold", "PnL_Pct", "Days_Held", "BTC_3d_Return"])
        if self.result.stop_loss_events:
            for ev in self.result.stop_loss_events:
                ws.append([str(ev.date.date()), ev.ticker, ev.stop_type, ev.reason,
                           round(ev.entry_price, 4), round(ev.current_price, 4),
                           round(ev.threshold, 4), round(ev.pnl_pct, 4),
                           ev.days_held, round(ev.benchmark_return_3d, 4)])

        # ── 26. config ──
        ws = wb.create_sheet("config")
        ws.append(["Key", "Value"])
        self._write_config_flat(ws, self.result.config)

        # ── 27. funding_history (crypto addition) ──
        ws = wb.create_sheet("funding_history")
        ws.append(["Date", "Daily_Funding_PnL", "Cum_Funding_PnL"])
        for dt in ec.index:
            ws.append([str(dt.date()), round(float(funding_daily.get(dt, 0)), 4),
                       round(float(funding_cum.get(dt, 0)), 4)])

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        log.info(f"[PORTFOLIO RECORD] Excel ({len(wb.sheetnames)} sheets) → {output_path}")
        return output_path

    # ──────────────────────────────────────────────────────────────
    #  Helper methods (template verbatim, fee model adapted)
    # ──────────────────────────────────────────────────────────────

    def _compute_realized_pnl(self, sector_cum_pnl: pd.DataFrame) -> Dict[str, float]:
        realized = {t: 0.0 for t in self.tickers}
        wh = self.weights_history
        reb_dates = sorted(wh.index)
        for i in range(1, len(reb_dates)):
            dt = reb_dates[i]
            prev_dt = reb_dates[i - 1]
            for t in self.tickers:
                prev_w = float(wh.loc[prev_dt].get(t, 0.0))
                curr_w = float(wh.loc[dt].get(t, 0.0))
                if prev_w > 1e-4 and curr_w < 1e-4 and dt in sector_cum_pnl.index:
                    realized[t] += float(sector_cum_pnl.loc[dt, t])
        return realized

    def _compute_total_notional(self) -> Dict[str, float]:
        notional = {t: 0.0 for t in self.tickers}
        wh = self.weights_history
        ec = self.equity_curve
        prev_w = pd.Series(0.0, index=self.tickers)
        for dt in sorted(wh.index):
            eq = ec.get(dt, 0)
            for t in self.tickers:
                curr_w = float(wh.loc[dt].get(t, 0.0))
                if curr_w > 1e-4 and prev_w[t] < 1e-4:
                    notional[t] += curr_w * eq
                prev_w[t] = curr_w
        return notional

    def _write_rebalance_trades(self, ws):
        wh = self.weights_history
        ec = self.equity_curve
        prev_w = pd.Series(0.0, index=self.tickers)
        for dt in sorted(wh.index):
            eq = ec.get(dt, 0)
            for t in self.tickers:
                new_w = float(wh.loc[dt].get(t, 0.0))
                old_w = float(prev_w.get(t, 0.0))
                delta = new_w - old_w
                if abs(delta) > 1e-4:
                    p = self.prices[t].get(dt, np.nan) if t in self.prices.columns else np.nan
                    contracts = abs(delta) * eq / p if pd.notna(p) and p > 0 else 0
                    # ADAPTED: taker fee from the live schedule (template: flat 5bps)
                    _, taker = load_fee_rates(t)
                    cost = abs(delta) * eq * taker
                    ws.append([str(dt.date()), t, "BUY" if delta > 0 else "SELL",
                               round(old_w, 4), round(new_w, 4), round(contracts, 2),
                               round(float(p), 4) if pd.notna(p) else "",
                               round(cost, 4)])
            prev_w = wh.loc[dt]

    def _write_strategy_vars(self, ws):
        cfg = self.result.config
        sig_cfg = cfg.get("signals", {})
        port_cfg = cfg.get("portfolio", {})
        reb_cfg = cfg.get("rebalance", {})
        risk_cfg = cfg.get("risk", {})
        sl_cfg = cfg.get("stop_loss", {})

        score_cols = [f"score_{t}" for t in self.tickers]
        header = (["Date", "param_set", "signal_version"] +
                  ["cs_mom_weight", "ts_mom_weight", "carry_weight", "regime_weight"] +
                  ["optimizer", "top_n", "max_weight"] +
                  ["frequency", "target_vol", "max_daily_loss"] +
                  ["sl_enabled", "sl_circuit_btc3d", "sl_collapse_dd", "sl_trailing_peak"] +
                  score_cols +
                  ["regime_label", "btc_rvol", "vol_triggered", "rvol_emergency",
                   "dd_circuit", "beta_adj"])
        ws.append(header)

        sig_hist = self.result.signals_history
        rf_by_date = {rf["date"]: rf for rf in self.result.risk_flags if "date" in rf}
        weights_cfg = sig_cfg.get("weights", {})
        for dt in sig_hist.index:
            scores = sig_hist.loc[dt]
            rf = rf_by_date.get(dt, rf_by_date.get(str(dt), {}))
            regime = (self.result.regime_history.get(dt, "")
                      if dt in self.result.regime_history.index else "")
            row = [str(dt.date()), self.param_set, self.signal_version,
                   weights_cfg.get("cross_sectional_momentum", ""),
                   weights_cfg.get("ts_momentum", ""),
                   weights_cfg.get("carry", ""),
                   weights_cfg.get("regime_adjustment", ""),
                   port_cfg.get("optimizer", ""),
                   port_cfg.get("top_n", ""),
                   port_cfg.get("constraints", {}).get("max_weight", ""),
                   reb_cfg.get("frequency", ""),
                   risk_cfg.get("target_vol_annual", ""),
                   risk_cfg.get("max_daily_loss_pct", ""),
                   sl_cfg.get("enabled", ""),
                   sl_cfg.get("portfolio_circuit_breaker", {}).get("btc_3d_limit", ""),
                   sl_cfg.get("sector_collapse", {}).get("max_dd_from_entry", ""),
                   sl_cfg.get("trailing_stop", {}).get("max_dd_from_peak", "")]
            for t in self.tickers:
                row.append(round(float(scores.get(t, 0)), 4))
            row += [str(regime), rf.get("current_rvol", ""),
                    rf.get("vol_scaling_triggered", ""),
                    rf.get("rvol_emergency_triggered", rf.get("rvol_emergency", "")),
                    rf.get("dd_circuit_triggered", ""), rf.get("beta_adjusted", "")]
            ws.append(row)

    def _write_config_flat(self, ws, cfg: dict, prefix: str = ""):
        for k, v in cfg.items():
            key = f"{prefix}.{k}" if prefix else k
            if isinstance(v, dict):
                self._write_config_flat(ws, v, key)
            elif isinstance(v, list):
                ws.append([key, json.dumps(v)])
            else:
                ws.append([key, str(v)])

    # ──────────────────────────────────────────────────────────────
    #  Monitor Excel (5 sheets — template verbatim)
    # ──────────────────────────────────────────────────────────────

    def export_monitor_excel(self, signal_date, report_data: dict,
                             output_path: Path = None) -> Path:
        if output_path is None:
            _records_dir().mkdir(parents=True, exist_ok=True)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = _records_dir() / (
                f"monitor_pr_{self.param_set}_{self.signal_version}_daily_{ts}.xlsx")

        from openpyxl import Workbook
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        ws = wb.create_sheet("snapshot")
        regime = report_data.get("regime", {})
        ws.append(["Field", "Value"])
        ws.append(["date", str(signal_date)])
        ws.append(["param_set", self.param_set])
        ws.append(["signal_version", self.signal_version])
        ws.append(["regime", regime.get("label", "")])
        ws.append(["btc_rvol", regime.get("btc_rvol", "")])
        ws.append(["cash_weight", report_data.get("cash_weight", "")])
        ws.append(["n_positions",
                   report_data.get("holdings_summary", {}).get("n_positions", "")])
        ws.append(["rebalance",
                   report_data.get("rebalance_decision", {}).get("rebalance", "")])
        ws.append(["reason",
                   report_data.get("rebalance_decision", {}).get("reason", "")])

        ws = wb.create_sheet("holdings")
        ws.append(["Perp", "Weight", "Price", "Composite_Score", "Action"])
        for sig in report_data.get("signals", []):
            ws.append([sig.get("ticker", ""), round(sig.get("target_weight", 0), 4),
                       round(sig.get("price", 0), 4),
                       round(sig.get("composite_score", 0), 4), sig.get("action", "")])

        ws = wb.create_sheet("signals")
        ws.append(["Perp", "CS_Mom", "TS_Mult", "Carry", "Composite"])
        for sig in report_data.get("signals", []):
            ws.append([sig.get("ticker", ""), round(sig.get("cs_mom", 0), 4),
                       round(sig.get("ts_mult", 0), 4), round(sig.get("carry", 0), 4),
                       round(sig.get("composite_score", 0), 4)])

        ws = wb.create_sheet("smart_select")
        ws.append(["Field", "Value"])
        for k, v in report_data.get("smart_select", {}).items():
            ws.append([k, str(v) if isinstance(v, (dict, list)) else v])

        ws = wb.create_sheet("risk_flags")
        ws.append(["Field", "Value"])
        ws.append(["risk_flags", str(report_data.get("risk_flags", ""))])

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        log.info(f"[MONITOR] Excel (5 sheets) → {output_path}")
        return output_path


# ═══════════════════════════════════════════════════════════════════════════
#  WF Diagnostic Excel (template verbatim over crypto WFResult)
# ═══════════════════════════════════════════════════════════════════════════

def export_wf_diagnostic_excel(wf_result, output_path: Path = None,
                               mode: str = "wf", signal_version: str = "v1") -> Path:
    if output_path is None:
        _records_dir().mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        wf_mode = wf_result.mode if hasattr(wf_result, "mode") else "anchored"
        output_path = _records_dir() / (
            f"wf_diagnostic_pr_{signal_version}_IS-OOS_{wf_mode}_{mode}_{ts}.xlsx")

    from openpyxl import Workbook
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    folds = wf_result.folds

    ws = wb.create_sheet("fold_summary")
    ws.append(["Fold", "IS_Start", "IS_End", "OOS_Start", "OOS_End",
               "Selected", "Method", "IS_Sharpe", "OOS_Sharpe", "WFE", "OOS_Regime"])
    for fr in folds:
        ws.append([fr.fold.fold_id, str(fr.fold.is_start.date()),
                   str(fr.fold.is_end.date()), str(fr.fold.oos_start.date()),
                   str(fr.fold.oos_end.date()), fr.is_best_name, fr.selection_method,
                   round(fr.is_best_sharpe, 4) if not np.isnan(fr.is_best_sharpe) else "",
                   round(fr.oos_metrics.get("sharpe", float("nan")), 4)
                   if not np.isnan(fr.oos_metrics.get("sharpe", float("nan"))) else "",
                   round(fr.wfe, 4) if not np.isnan(fr.wfe) else "", fr.oos_regime])

    ws = wb.create_sheet("param_oos_matrix")
    all_params = sorted({p for fr in folds for p in fr.all_oos_sharpes})
    ws.append(["Param"] + [f"Fold_{fr.fold.fold_id}" for fr in folds])
    for param in all_params:
        row = [param]
        for fr in folds:
            v = fr.all_oos_sharpes.get(param, float("nan"))
            row.append(round(v, 4) if not np.isnan(v) else "")
        ws.append(row)

    ws = wb.create_sheet("param_by_regime")
    ws.append(["Param", "Regime", "Mean_OOS_Sharpe", "N_Folds"])
    for param, regimes in sorted(wf_result.param_oos_by_regime().items()):
        for regime, stats in sorted(regimes.items()):
            ws.append([param, regime, stats["mean_oos_sharpe"], stats["n_folds"]])

    ws = wb.create_sheet("synthetic_equity")
    se = wf_result.synthetic_equity
    oracle_eq = getattr(wf_result, "oracle_equity", None)
    static_eq = getattr(wf_result, "static_best_equity", None)
    _orc = oracle_eq.reindex(se.index) if oracle_eq is not None and not oracle_eq.empty else None
    _sta = static_eq.reindex(se.index) if static_eq is not None and not static_eq.empty else None
    ws.append(["Date", "Synthetic_Value", "Oracle_Value", "Static_Best_Value"])
    for dt in se.index:
        ov = "" if _orc is None or pd.isna(_orc.get(dt)) else round(float(_orc[dt]), 4)
        sv = "" if _sta is None or pd.isna(_sta.get(dt)) else round(float(_sta[dt]), 4)
        ws.append([str(dt.date()), round(float(se[dt]), 4), ov, sv])

    ws = wb.create_sheet("selection_log")
    if wf_result.selection_log:
        keys = list(wf_result.selection_log[0].keys())
        ws.append(keys)
        for entry in wf_result.selection_log:
            ws.append([str(entry.get(k, "")) for k in keys])

    cmp = getattr(wf_result, "comparison", None)
    if cmp:
        ws = wb.create_sheet("oracle_vs_realized")
        ws.append(["Layer", "Sharpe", "CAGR", "MaxDD", "Calmar", "Note"])
        _note = {
            "static_best": f"single full-period best param "
                           f"[{getattr(wf_result, 'static_best_name', '')}] — IS/full (optimistic)",
            "synthetic": "realizable dynamic selection (OOS)",
            "oracle": "theoretical ceiling: best-OOS param per fold (hindsight)",
        }

        def _f(x):
            return "" if x is None or (isinstance(x, float) and np.isnan(x)) else round(float(x), 4)

        for layer in ("static_best", "synthetic", "oracle"):
            d = cmp.get(layer, {})
            ws.append([layer, _f(d.get("sharpe")), _f(d.get("cagr")),
                       _f(d.get("maxdd")), _f(d.get("calmar")), _note.get(layer, "")])
        ws.append([])
        ws.append(["capture_ratio_sharpe", cmp.get("capture_ratio_sharpe")])
        ws.append(["capture_ratio_cagr", cmp.get("capture_ratio_cagr")])
        ws.append(["mean_regret_sharpe", cmp.get("mean_regret_sharpe")])
        ws.append(["n_folds_optimal", cmp.get("n_folds_optimal"), f"of {len(folds)} folds"])
        ws.append([])
        olog = getattr(wf_result, "oracle_selection_log", []) or []
        if olog:
            ws.append(["Fold", "OOS_Start", "OOS_End", "Oracle_Param", "Oracle_OOS_SR",
                       "Selected_Param", "Selected_OOS_SR", "Regret", "Optimal"])
            for e in olog:
                ws.append([e.get("fold"), e.get("oos_start"), e.get("oos_end"),
                           e.get("oracle_param"), e.get("oracle_oos_sharpe"),
                           e.get("selected_param"), e.get("selected_oos_sharpe"),
                           e.get("regret"), "Y" if e.get("optimal") else ""])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    log.info(f"[WF DIAGNOSTIC] Excel → {output_path}")
    return output_path
