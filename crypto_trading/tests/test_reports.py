"""Report generation smoke tests (Plan 07 §2/§3) — tmp dirs, no network."""
import json

import numpy as np
import pandas as pd
import pytest

from crypto_trading.crypto_common import config as _config
from crypto_trading.crypto_common import risk_kill as _rk
from crypto_trading.crypto_common.reporting import ledger as ledger_mod
from crypto_trading.crypto_common.reporting import pnl_report, risk_report


@pytest.fixture
def sig_env(tmp_path, monkeypatch):
    """Redirect SIGNALS_DIR (+ risk_kill STATE_DIR) into tmp and seed artifacts."""
    sig = tmp_path / "signals"
    monkeypatch.setattr(_config, "SIGNALS_DIR", sig)
    monkeypatch.setattr(_rk, "STATE_DIR", sig / "state")

    # inventory (schema per ledger docstring)
    inv_dir = sig / "inventory"
    inv_dir.mkdir(parents=True)
    (inv_dir / "inventory_basis_meanrev.json").write_text(json.dumps({
        "strategy": "basis_meanrev", "updated_ts": 1783400000.0,
        "equity": 1004.47, "equity_sod": 1000.0, "equity_peak": 1006.0,
        "positions": [{"ticker": "KXBTCPERP", "side": "long", "contracts": 5,
                       "entry_price": 6.39, "entry_ts": 1783390000.0,
                       "funding_accrued": -0.0032, "liq_price": 3.20}]}))

    # a backtest artifact
    bt_dir = sig / "basis_meanrev" / "backtests"
    bt_dir.mkdir(parents=True)
    (bt_dir / "backtest_20260707_010000.json").write_text(json.dumps({
        "final_equity": 1004.47, "net_pnl": 4.47, "fee_scenario": "zero",
        "n_fills": 52, "note": "v1 candle tape"}))

    # dry-run orders
    dr_dir = sig / "orders_dryrun" / "basis_meanrev"
    dr_dir.mkdir(parents=True)
    (dr_dir / "2026-07-07.jsonl").write_text(json.dumps({
        "ts": 1783400100.0, "mode": "dry_run", "strategy": "basis_meanrev",
        "order": {"ticker": "KXBTCPERP", "side": "buy", "count": "5",
                  "price": "6.3900", "client_order_id": "t1"}}) + "\n")
    return sig


def test_build_ledger_from_artifacts(sig_env):
    led = ledger_mod.build_ledger(marks={"KXBTCPERP": 6.45})
    assert "basis_meanrev" in led["strategies"]
    entry = led["strategies"]["basis_meanrev"]
    assert entry["fill_source"] == "dryrun_hypothetical"
    assert entry["inventory"]["equity"] == pytest.approx(1004.47)
    row = entry["rows"][0]
    assert row["ticker"] == "KXBTCPERP" and row["n_fills"] == 1
    assert led["totals"]["net_projected"] == pytest.approx(
        led["totals"]["net_zero"] - led["totals"]["fees_projected"])


def test_pnl_report_generates_pdf(sig_env, tmp_path):
    rng = np.random.default_rng(1)
    rets = pd.Series(rng.normal(1e-3, 5e-3, 30),
                     index=pd.date_range("2026-06-05", periods=30, freq="D", tz="UTC"))
    report = pnl_report.build_report_data(
        "2026-06-05", "2026-07-07",
        led=ledger_mod.build_ledger(marks={"KXBTCPERP": 6.45}),
        bench={"KXBTCPERP-HODL": rets * 1.5}, returns=rets)
    assert report["perf"] is not None and np.isfinite(report["perf"]["sharpe"])
    out = tmp_path / "pnl.pdf"
    pnl_report.build_pdf(report, str(out))
    assert out.exists() and out.stat().st_size > 5000


def test_risk_report_four_artifacts(sig_env):
    res = risk_report.generate(trip=False, marks={"KXBTCPERP": 6.45})
    paths = res["paths"]
    for kind in ("xlsx", "pdf", "json", "txt"):
        p = paths[kind]
        assert p and json is not None
        import os
        assert os.path.exists(p) and os.path.getsize(p) > 0
    payload = json.loads(open(paths["json"]).read())
    assert "summary" in payload and "report" in payload
    assert payload["summary"]["overall_status"] in ("green", "amber", "red")
    assert "limits" in payload["report"]

    from openpyxl import load_workbook
    wb = load_workbook(paths["xlsx"])
    assert len(wb.sheetnames) == 19          # plan §3: every sheet always present

    txt = open(paths["txt"]).read()
    assert "PORTFOLIO RISK SUMMARY" in txt and "LIMITS" in txt


def test_risk_report_no_trip_by_default(sig_env):
    # equity below sod by >10% would be red daily-loss; ensure report-only mode
    inv = sig_env / "inventory" / "inventory_basis_meanrev.json"
    data = json.loads(inv.read_text())
    data["equity"] = 850.0                    # −15% vs sod 1000 → red
    inv.write_text(json.dumps(data))
    res = risk_report.generate(trip=False, marks={"KXBTCPERP": 6.45})
    assert "tripped" not in res["report"]
    from crypto_trading.crypto_common.risk_kill import RiskKill
    assert not RiskKill("basis_meanrev").halted()
