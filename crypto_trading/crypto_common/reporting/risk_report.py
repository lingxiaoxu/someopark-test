"""Crypto risk report (Plan 07 §3) — output contract mirrors /RiskManager.py
(read-only template): read-only w.r.t. strategy state, four artifacts per run:

    trading_signals/reports/risk_workbook_<ts>.xlsx   (19 sheets, plan §3 table)
    trading_signals/reports/risk_report_<ts>.pdf      (PnLReport-styled)
    trading_signals/reports/risk_report_<ts>.json     (machine-readable)
    trading_signals/reports/risk_report_<ts>.txt      (quick human summary)

Consumes the Plan 06 Layer-2 aggregator (risk/aggregator.py). By default this
REPORTING entry point runs the aggregator with ``trip_on_red=False`` — reports
observe, the scheduled aggregator governs (pass --trip to let a reporting run
trip the kill-switch too).

Workbook styling (navy/gold, breach fills) ported from
RiskManager.RiskWorkbookExporter; PDF style reuses reporting.pnl_report's
ported helpers. Theory footnotes per Plan 06 §2.
"""
from __future__ import annotations

import argparse
import json
import logging
from datetime import datetime, timezone

import pandas as pd

from crypto_trading.crypto_common import config as _config
from crypto_trading.crypto_common.reporting import ledger as _ledger
from crypto_trading.crypto_common.risk.aggregator import PortfolioAggregator, StrategyState

logger = logging.getLogger(__name__)

THEORY_FOOTNOTES = (
    "Artzner et al. (coherent risk) · Rockafellar–Uryasev (CVaR/ES) · "
    "Zangari (Cornish–Fisher VaR) · Chekhlov–Uryasev–Zabarankin (CDaR) · "
    "Litterman (risk contribution) · Almgren–Chriss (liquidity) · "
    "Bailey–López de Prado (DSR/PBO) · Khandani–Lo (deleveraging contagion) · "
    "Kelly (sizing)."
)

SHEETS = [  # plan §3 — every sheet always present; thin ones carry a stub row
    "Limit Dashboard", "Exposure & Leverage", "Net Delta", "Liquidation",
    "Concentration", "Factor Beta", "VaR CVaR", "Stress", "Correlation",
    "Risk Contribution", "Drawdown CDaR", "Funding Exposure", "Basis Exposure",
    "Liquidity", "Per-Strategy Risk", "Strategy Health", "Financial Statements",
    "Reconciliation", "Venue Ops",
]


# ── inputs: StrategyState from artifacts (marks injectable for tests) ───────

def _latest_marks(tickers: set[str]) -> dict[str, float]:
    from crypto_trading.crypto_common.loader import load_perp_candles
    marks = {}
    for t in tickers:
        try:
            c = load_perp_candles(t, "1h")
            marks[t] = float(c.price_close.dropna().iloc[-1])
        except Exception:
            continue
    return marks


def states_from_artifacts(*, marks: dict[str, float] | None = None) -> list[StrategyState]:
    """Build aggregator inputs from inventories + backtests (read-only)."""
    states = []
    for s in _ledger.strategies_present():
        inv = _ledger.load_inventory(s)
        bts = _ledger.load_backtests(s)
        equity = None
        positions: dict[str, float] = {}
        liq_prices: dict[str, float] = {}
        if inv:
            equity = inv.get("equity")
            for p in inv.get("positions") or []:
                sign = 1 if p.get("side") == "long" else -1
                positions[p["ticker"]] = positions.get(p["ticker"], 0.0) \
                    + sign * float(p.get("contracts", 0))
                if p.get("liq_price"):
                    liq_prices[p["ticker"]] = float(p["liq_price"])
        if equity is None and bts:
            equity = bts[-1].get("final_equity")
        if equity is None:
            continue
        tickers = set(positions)
        mk = marks if marks is not None else _latest_marks(tickers)
        states.append(StrategyState(
            name=s, equity=float(equity), positions=positions,
            marks={t: mk[t] for t in tickers if t in mk},
            liq_prices=liq_prices,
            equity_sod=(inv or {}).get("equity_sod"),
            equity_peak=(inv or {}).get("equity_peak")))
    return states


# ── summary / writers (template _build_summary/_write_json/_write_txt) ─────

def build_summary(report: dict) -> dict:
    lim = report.get("limits") or []
    worst = "green"
    for c in lim:
        if c["status"] == "red":
            worst = "red"
            break
        if c["status"] == "amber":
            worst = "amber"
    return {"generated": datetime.now(timezone.utc).isoformat(),
            "overall_status": worst,
            "equity": report.get("equity"),
            "gross_leverage": (report.get("exposure") or {}).get("gross_leverage"),
            "net_btc_delta": report.get("net_btc_delta"),
            "var_95": (report.get("var") or {}).get("var_95_hist"),
            "worst_liq_distance_pct": report.get("worst_liq_distance_pct"),
            "n_limits": len(lim),
            "n_amber": sum(1 for c in lim if c["status"] == "amber"),
            "n_red": sum(1 for c in lim if c["status"] == "red")}


def write_json(path, report: dict, summary: dict) -> None:
    payload = {"summary": summary, "report": report,
               "theory": THEORY_FOOTNOTES}
    with open(path, "w") as fh:
        json.dump(payload, fh, indent=1, default=str)


def write_txt(path, report: dict, summary: dict) -> None:
    lines = [
        "KALSHI CRYPTO PERPS — PORTFOLIO RISK SUMMARY",
        f"generated: {summary['generated']}   overall: {summary['overall_status'].upper()}",
        "-" * 64,
        f"equity            : {summary['equity']:.2f}" if summary.get("equity") is not None else "equity            : N/A",
    ]
    exp = report.get("exposure") or {}
    if exp.get("gross_leverage") is not None:
        lines.append(f"gross leverage    : {exp['gross_leverage']:.2f}x "
                     f"(net {exp.get('net_leverage', float('nan')):.2f}x)")
    if report.get("net_btc_delta") is not None:
        lines.append(f"net BTC delta     : ${report['net_btc_delta']:.2f}")
    if report.get("worst_liq_distance_pct") is not None:
        lines.append(f"worst liq distance: {report['worst_liq_distance_pct'] * 100:.1f}%")
    lines.append("-" * 64)
    lines.append("LIMITS (value | amber | red | status | red action)")
    for c in report.get("limits") or []:
        lines.append(f"  {c['name']:<24} {c['value']:.3g} | {c['amber']} | {c['red']} "
                     f"| {c['status'].upper():<5} | {c.get('red_action') or '-'}")
    if not (report.get("limits") or []):
        lines.append("  (no evaluable limits yet — flat book / no return series)")
    lines.append("-" * 64)
    lines.append(THEORY_FOOTNOTES)
    with open(path, "w") as fh:
        fh.write("\n".join(lines) + "\n")


# ── workbook (RiskWorkbookExporter styling, crypto sheets) ──────────────────

class CryptoRiskWorkbookExporter:
    NAVY = 'FF1A1A2E'
    ROW_ALT = 'FFF7F9FC'
    GOLD = 'FFD4A843'
    POS = 'FF1A7A4A'
    NEG = 'FFC0392B'
    AMBER_FILL = 'FFFFF3CD'
    RED_FILL = 'FFF8D7DA'
    GREEN_FILL = 'FFE6F4EA'

    def __init__(self, report: dict, summary: dict, states: list[StrategyState],
                 led: dict | None = None):
        self.r = report
        self.s = summary
        self.states = states
        self.led = led or {"strategies": {}, "totals": {}}

    def export(self, path) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        self._Font, self._Fill = Font, PatternFill
        self._Align, self._Border, self._Side = Alignment, Border, Side
        self._col = get_column_letter

        wb = Workbook()
        wb.remove(wb.active)
        builders = {
            "Limit Dashboard": self._sheet_limits,
            "Exposure & Leverage": self._sheet_exposure,
            "Net Delta": self._sheet_delta,
            "Liquidation": self._sheet_liq,
            "Concentration": self._sheet_concentration,
            "VaR CVaR": self._sheet_var,
            "Stress": self._sheet_stress,
            "Correlation": self._sheet_corr,
            "Risk Contribution": self._sheet_contrib,
            "Funding Exposure": self._sheet_funding,
            "Basis Exposure": self._sheet_basis,
            "Per-Strategy Risk": self._sheet_per_strategy,
            "Reconciliation": self._sheet_recon,
            "Venue Ops": self._sheet_venue_ops,
        }
        for name in SHEETS:
            ws = wb.create_sheet(name[:31])
            fn = builders.get(name)
            if fn:
                fn(ws)
            else:
                self._stub(ws, name)
        wb.save(path)

    # styling helpers (template ports)
    def _title(self, ws, text, ncols=6):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c = ws.cell(1, 1, text)
        c.font = self._Font(bold=True, size=13, color='FF1A1A2E')

    def _hdr(self, ws, row, headers):
        for j, h in enumerate(headers, 1):
            c = ws.cell(row, j, h)
            c.font = self._Font(bold=True, size=9, color='FFFFFFFF')
            c.fill = self._Fill('solid', fgColor=self.NAVY)
            c.alignment = self._Align(horizontal='center', wrap_text=True)
            c.border = self._Border(bottom=self._Side(style='medium', color=self.GOLD))

    def _row(self, ws, row, vals, alt=False, fill=None):
        for j, v in enumerate(vals, 1):
            c = ws.cell(row, j, v)
            c.font = self._Font(size=9)
            c.alignment = self._Align(horizontal='right' if j > 1 else 'left')
            if fill:
                c.fill = self._Fill('solid', fgColor=fill)
            elif alt:
                c.fill = self._Fill('solid', fgColor=self.ROW_ALT)

    def _autofit(self, ws, widths):
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[self._col(j)].width = w

    def _stub(self, ws, name):
        self._title(ws, name)
        ws.cell(3, 1, "no data yet — populated as live data accumulates")

    # sheets
    def _sheet_limits(self, ws):
        self._title(ws, "Limit Dashboard — amber/red (Plan 06 §3)")
        self._hdr(ws, 3, ["Limit", "Value", "Amber", "Red", "Status", "Red action"])
        r = 4
        for i, c in enumerate(self.r.get("limits") or []):
            fill = {"red": self.RED_FILL, "amber": self.AMBER_FILL,
                    "green": self.GREEN_FILL}[c["status"]]
            self._row(ws, r, [c["name"], round(c["value"], 4), c["amber"], c["red"],
                              c["status"].upper(), c.get("red_action") or "-"],
                      fill=fill)
            r += 1
        if r == 4:
            ws.cell(4, 1, "no evaluable limits yet")
        self._autofit(ws, [26, 12, 10, 10, 10, 24])

    def _sheet_exposure(self, ws):
        self._title(ws, "Exposure & Leverage")
        e = self.r.get("exposure") or {}
        self._hdr(ws, 3, ["Metric", "Value"])
        rows = [("equity", self.r.get("equity")), ("gross $", e.get("gross")),
                ("gross $ (netted)", e.get("gross_netted")), ("net $", e.get("net")),
                ("gross leverage ×", e.get("gross_leverage")),
                ("net leverage ×", e.get("net_leverage"))]
        for i, (k, v) in enumerate(rows):
            self._row(ws, 4 + i, [k, v if v is not None else "N/A"], alt=i % 2 == 1)
        self._autofit(ws, [26, 18])

    def _sheet_delta(self, ws):
        self._title(ws, "Net Delta — per asset + BTC-beta rollup")
        self._hdr(ws, 3, ["Asset", "Dollar delta"])
        r = 4
        for a, d in (self.r.get("per_asset_delta") or {}).items():
            self._row(ws, r, [a, round(d, 2)], alt=r % 2 == 0)
            r += 1
        self._row(ws, r + 1, ["NET BTC-DELTA (beta-weighted)",
                              round(self.r.get("net_btc_delta", 0.0), 2)])
        self._autofit(ws, [22, 18])

    def _sheet_liq(self, ws):
        self._title(ws, "Liquidation")
        w = self.r.get("worst_liq_distance_pct")
        ws.cell(3, 1, "worst-case liquidation distance")
        ws.cell(3, 2, f"{w * 100:.1f}%" if w is not None else "N/A (flat/unknown)")

    def _sheet_concentration(self, ws):
        self._title(ws, "Concentration")
        deltas = self.r.get("per_asset_delta") or {}
        gross = (self.r.get("exposure") or {}).get("gross") or 0
        self._hdr(ws, 3, ["Asset", "|delta| $", "% of gross"])
        r = 4
        for a, d in sorted(deltas.items(), key=lambda kv: -abs(kv[1])):
            pctg = abs(d) / gross * 100 if gross else None
            self._row(ws, r, [a, round(abs(d), 2),
                              round(pctg, 2) if pctg is not None else "N/A"],
                      alt=r % 2 == 0)
            r += 1
        ws.cell(r + 1, 1, "venue concentration: 100% Kalshi (structural — monitored)")
        self._autofit(ws, [22, 16, 14])

    def _sheet_var(self, ws):
        self._title(ws, "VaR / CVaR (historical · parametric · Cornish-Fisher)")
        v = self.r.get("var") or {}
        self._hdr(ws, 3, ["Measure", "Value ($/day)"])
        rows = [("VaR 95 historical", v.get("var_95_hist")),
                ("VaR 95 parametric", v.get("var_95_param")),
                ("CVaR 97.5 historical", v.get("cvar_975_hist")),
                ("Cornish-Fisher VaR 95", (v.get("cornish_fisher") or {}).get("var_cf")
                 if isinstance(v.get("cornish_fisher"), dict) else v.get("cornish_fisher")),
                ("Stress-correlation VaR 95 (ρ→1)", v.get("stress_correlation_var_95"))]
        for i, (k, val) in enumerate(rows):
            self._row(ws, 4 + i, [k, round(val, 2) if isinstance(val, (int, float))
                                  else "N/A"], alt=i % 2 == 1)
        self._autofit(ws, [34, 18])

    def _sheet_stress(self, ws):
        self._title(ws, "Stress (crypto-sized scenarios)")
        self._hdr(ws, 3, ["Scenario", "Channel", "P&L $", "% equity"])
        r = 4
        for row in self.r.get("stress") or []:
            self._row(ws, r, [row.get("scenario"), row.get("channel"),
                              round(row.get("pnl", 0), 2),
                              round(row.get("pct_equity", 0), 2)], alt=r % 2 == 0)
            r += 1
        if r == 4:
            ws.cell(4, 1, "no stress rows (flat book)")
        self._autofit(ws, [30, 22, 14, 12])

    def _sheet_corr(self, ws):
        self._title(ws, "Cross-strategy correlation")
        corr = self.r.get("correlation")
        if not corr:
            ws.cell(3, 1, "needs ≥2 strategies with return series")
            return
        names = list(corr)
        self._hdr(ws, 3, [""] + names)
        for i, n in enumerate(names):
            self._row(ws, 4 + i, [n] + [round(corr[m].get(n, float("nan")), 3)
                                        for m in names], alt=i % 2 == 1)

    def _sheet_contrib(self, ws):
        self._title(ws, "Risk Contribution (Litterman marginal VaR)")
        rc = self.r.get("risk_contribution") or []
        self._hdr(ws, 3, ["Strategy", "vol $", "mctr", "trc", "% of portfolio risk"])
        for i, row in enumerate(rc):
            self._row(ws, 4 + i, [row.get("name"), round(row.get("vol", 0), 2),
                                  round(row.get("mctr", 0), 4),
                                  round(row.get("trc", 0), 4),
                                  round(row.get("prc_pct", 0), 2)], alt=i % 2 == 1)
        if not rc:
            ws.cell(4, 1, "needs strategy return series")
        self._autofit(ws, [22, 12, 12, 12, 18])

    def _sheet_funding(self, ws):
        self._title(ws, "Funding Exposure (next-cycle expected, 04/12/20 UTC)")
        f = self.r.get("funding_exposure") or {}
        per = f.get("per_ticker") or {}
        self._hdr(ws, 3, ["Ticker", "expected $ next cycle"])
        r = 4
        for t, v in per.items():
            self._row(ws, r, [t, round(v, 6)], alt=r % 2 == 0)
            r += 1
        self._row(ws, r + 1, ["TOTAL", round(f.get("total_next_cycle", 0.0), 6)])
        self._autofit(ws, [22, 22])

    def _sheet_basis(self, ws):
        self._title(ws, "Basis Exposure (mark-vs-index sensitivity)")
        b = self.r.get("basis_exposure") or {}
        per = b.get("per_ticker_per_bp") or {}
        self._hdr(ws, 3, ["Ticker", "$ per 1bp basis move"])
        r = 4
        for t, v in per.items():
            self._row(ws, r, [t, round(v, 6)], alt=r % 2 == 0)
            r += 1
        self._row(ws, r + 1, ["NET / GROSS per bp",
                              f'{b.get("net_per_bp", 0):.6f} / '
                              f'{b.get("gross_per_bp", 0):.6f}'])
        self._autofit(ws, [22, 26])

    def _sheet_per_strategy(self, ws):
        self._title(ws, "Per-Strategy Risk (Plan 06 §4 dominant risks)")
        self._hdr(ws, 3, ["Strategy", "Equity", "Gross $", "#Tickers", "Halted?"])
        from crypto_trading.crypto_common.risk_kill import RiskKill
        for i, s in enumerate(self.states):
            gross = sum(abs(c) * s.marks.get(t, 0.0) for t, c in s.positions.items())
            self._row(ws, 4 + i, [s.name, round(s.equity, 2), round(gross, 2),
                                  len(s.positions),
                                  "YES" if RiskKill(s.name).halted() else "no"],
                      alt=i % 2 == 1)
        if not self.states:
            ws.cell(4, 1, "no strategies with state yet")
        self._autofit(ws, [22, 14, 14, 10, 10])

    def _sheet_recon(self, ws):
        self._title(ws, "Reconciliation (intended vs actual · funding tie-out)")
        ws.cell(3, 1, "live fills not enabled yet (margin opt-in pending) — "
                      "dry-run intents logged; recon activates with live fills")

    def _sheet_venue_ops(self, ws):
        self._title(ws, "Venue / Ops — recorder & poller heartbeats")
        self._hdr(ws, 3, ["Heartbeat", "age_s", "detail"])
        import time
        r = 4
        pd_dir = _config.PRICE_DATA
        for hb in sorted(pd_dir.rglob("heartbeat.json")):
            try:
                data = json.loads(hb.read_text())
                age = time.time() - float(data.get("ts", 0))
                self._row(ws, r, [str(hb.relative_to(pd_dir)), round(age, 0),
                                  json.dumps(data.get("counts", {}))[:80]],
                          alt=r % 2 == 0)
                r += 1
            except Exception:
                continue
        if r == 4:
            ws.cell(4, 1, "no heartbeats found")
        self._autofit(ws, [46, 10, 50])


# ── PDF (reuses pnl_report style ports) ─────────────────────────────────────

def build_pdf(report: dict, summary: dict, output_path: str) -> str:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import HRFlowable, Paragraph, SimpleDocTemplate

    from crypto_trading.crypto_common.reporting.pnl_report import (C, C_GOLD, H,
                                                                   body_style,
                                                                   footer_style,
                                                                   h1_style,
                                                                   make_kv_table,
                                                                   make_table,
                                                                   sub_style,
                                                                   title_style)
    doc = SimpleDocTemplate(output_path, pagesize=A4, topMargin=1.4 * cm,
                            bottomMargin=1.4 * cm, leftMargin=1.5 * cm,
                            rightMargin=1.5 * cm)
    W = A4[0] - 3.0 * cm
    el = [Paragraph('Kalshi Crypto Perps — Portfolio Risk Report', title_style),
          Paragraph(f'生成 {summary["generated"]} · overall: '
                    f'{summary["overall_status"].upper()} · read-only', sub_style),
          HRFlowable(width='100%', thickness=1, color=C_GOLD, spaceAfter=8),
          Paragraph('1 · Limit Dashboard（amber/red）', h1_style)]
    lim = report.get("limits") or []
    if lim:
        rows = [[H('Limit'), H('Value', 'RIGHT'), H('Amber', 'RIGHT'),
                 H('Red', 'RIGHT'), H('Status'), H('Red action')]]
        for c in lim:
            rows.append([C(c["name"]), C(f'{c["value"]:.3g}', 'RIGHT'),
                         C(str(c["amber"]), 'RIGHT'), C(str(c["red"]), 'RIGHT'),
                         C(c["status"].upper()), C(c.get("red_action") or '-')])
        el.append(make_table(rows, [W * 0.26, W * 0.12, W * 0.1, W * 0.1,
                                    W * 0.12, W * 0.3]))
    else:
        el.append(Paragraph('（暂无可评估 limit — flat book / 无收益序列）', body_style))

    el.append(Paragraph('2 · Tier 摘要', h1_style))
    exp = report.get("exposure") or {}
    v = report.get("var") or {}
    el.append(make_kv_table([
        ('Equity', f'{report.get("equity", 0):.2f}'),
        ('Gross / Net leverage',
         f'{exp.get("gross_leverage", float("nan")):.2f}× / '
         f'{exp.get("net_leverage", float("nan")):.2f}×'),
        ('Net BTC-delta', f'${report.get("net_btc_delta", 0):.2f}'),
        ('VaR95 hist / CVaR97.5',
         f'{v.get("var_95_hist") or "N/A"} / {v.get("cvar_975_hist") or "N/A"}'),
        ('Worst liquidation distance',
         f'{report["worst_liq_distance_pct"] * 100:.1f}%'
         if report.get("worst_liq_distance_pct") is not None else 'N/A'),
    ]))

    el.append(Paragraph('3 · Stress（crypto-sized）', h1_style))
    st = report.get("stress") or []
    if st:
        rows = [[H('Scenario'), H('Channel'), H('P&L $', 'RIGHT'), H('% equity', 'RIGHT')]]
        for r_ in st:
            rows.append([C(r_.get("scenario")), C(r_.get("channel")),
                         C(f'{r_.get("pnl", 0):.2f}', 'RIGHT'),
                         C(f'{r_.get("pct_equity", 0):.2f}', 'RIGHT')])
        el.append(make_table(rows, [W * 0.35, W * 0.3, W * 0.18, W * 0.17]))
    else:
        el.append(Paragraph('（flat book — 无压力敞口）', body_style))

    el.append(Paragraph(THEORY_FOOTNOTES, footer_style))
    doc.build(el)
    return output_path


# ── entry point ─────────────────────────────────────────────────────────────

def generate(*, trip: bool = False, marks: dict[str, float] | None = None,
             states: list[StrategyState] | None = None) -> dict:
    states = states if states is not None else states_from_artifacts(marks=marks)
    agg = PortfolioAggregator(states)
    report = agg.compute()
    if trip:
        report["tripped"] = agg.kill_switch(report["limits"])
    summary = build_summary(report)

    out_dir = _config.SIGNALS_DIR / "reports"
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')
    paths = {"xlsx": out_dir / f'risk_workbook_{ts}.xlsx',
             "pdf": out_dir / f'risk_report_{ts}.pdf',
             "json": out_dir / f'risk_report_{ts}.json',
             "txt": out_dir / f'risk_report_{ts}.txt'}
    led = _ledger.build_ledger()
    CryptoRiskWorkbookExporter(report, summary, states, led).export(paths["xlsx"])
    build_pdf(report, summary, str(paths["pdf"]))
    write_json(paths["json"], report, summary)
    write_txt(paths["txt"], report, summary)
    logger.info("risk report → %s", {k: str(v) for k, v in paths.items()})
    return {"summary": summary, "paths": {k: str(v) for k, v in paths.items()},
            "report": report}


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument('--trip', action='store_true',
                    help='allow this reporting run to trip the kill-switch on red')
    args = ap.parse_args(argv)
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    res = generate(trip=args.trip)
    print(json.dumps(res["summary"], indent=2, default=str))
    for k, p in res["paths"].items():
        print(f"{k}: {p}")
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
