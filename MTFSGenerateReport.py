"""
MTFSGenerateReport.py — Paired Backtest vs Validation Period Report Generator (MTFS)

Usage:
    python MTFSGenerateReport.py <backtest_excel> <forward_excel> [output_excel]

Examples:
    python MTFSGenerateReport.py \\
        historical_runs/portfolio_history_MTFS_all15_default_default_20260307.xlsx \\
        historical_runs/portfolio_history_MTFS_fwd_all15_default_default_20260307.xlsx

Generates a multi-sheet Excel report comparing every pair across backtest and
validation periods: PnL, Sharpe, win rate, drawdown, trade-by-trade detail, etc.
Mirrors MRPTGenerateReport.py but adapted for Momentum Trend Following Strategy.

Requirements:
    - Same pairs must appear in both files
    - Backtest end date <= Validation start date (no overlap)
"""

import sys
import os
import warnings
import numpy as np
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint

warnings.filterwarnings('ignore')


# ─────────────────────────────────────────────────────────────────────────────
# Colour palette (same as MRPT)
# ─────────────────────────────────────────────────────────────────────────────
C_HEADER_BG   = 'FF1F3864'   # dark navy
C_HEADER_FG   = 'FFFFFFFF'   # white
C_SUBHDR_BG   = 'FF2E75B6'   # medium blue
C_SUBHDR_FG   = 'FFFFFFFF'
C_ALT_ROW     = 'FFD9E1F2'   # light blue
C_POSITIVE    = 'FF375623'   # dark green text
C_POSITIVE_BG = 'FFE2EFDA'   # light green bg
C_NEGATIVE    = 'FF9C0006'   # dark red text
C_NEGATIVE_BG = 'FFFFC7CE'   # light red bg
C_NEUTRAL_BG  = 'FFFFF2CC'   # light yellow
C_SECTION_BG  = 'FFBDD7EE'   # section header blue
C_TOTAL_BG    = 'FF1F3864'   # same as header
C_BORDER      = 'FF8EA9C1'

thin  = Side(style='thin',   color=C_BORDER)
thick = Side(style='medium', color='FF1F3864')
THIN_BORDER  = Border(left=thin, right=thin, top=thin, bottom=thin)
THICK_BORDER = Border(left=thick, right=thick, top=thick, bottom=thick)


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────
def _read(path, sheet):
    try:
        df = pd.read_excel(path, sheet_name=sheet)
        if 'Date' in df.columns:
            df['Date'] = pd.to_datetime(df['Date'])
        return df
    except Exception:
        return pd.DataFrame()


def _style_cell(cell, bold=False, italic=False, font_color='FF000000',
                bg_color=None, align='left', num_fmt=None, border=None,
                size=11, wrap=False, color_sign=False):
    if color_sign and cell.value is not None:
        try:
            v = float(cell.value)
            font_color = C_POSITIVE if v >= 0 else C_NEGATIVE
            bg_color = bg_color or (C_POSITIVE_BG if v >= 0 else C_NEGATIVE_BG)
        except (TypeError, ValueError):
            pass
    cell.font = Font(bold=bold, italic=italic, color=font_color, size=size)
    if bg_color:
        cell.fill = PatternFill('solid', fgColor=bg_color)
    cell.alignment = Alignment(horizontal=align, vertical='center',
                                wrap_text=wrap)
    if num_fmt:
        cell.number_format = num_fmt
    if border:
        cell.border = border


def _header(ws, row, col, value, span=1, sub=False):
    bg = C_SUBHDR_BG if sub else C_HEADER_BG
    cell = ws.cell(row=row, column=col, value=value)
    _style_cell(cell, bold=True, font_color=C_HEADER_FG, bg_color=bg,
                align='center', border=THIN_BORDER, size=10 if sub else 11)
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    return cell


def _value_cell(ws, row, col, value, fmt=None, bold=False,
                color_sign=False, alt=False, align='right'):
    cell = ws.cell(row=row, column=col, value=value)
    bg = C_ALT_ROW if alt else None
    fg = 'FF000000'
    if color_sign and isinstance(value, (int, float)) and not pd.isna(value):
        if value > 0:
            fg = C_POSITIVE
            bg = C_POSITIVE_BG
        elif value < 0:
            fg = C_NEGATIVE
            bg = C_NEGATIVE_BG
    _style_cell(cell, bold=bold, font_color=fg, bg_color=bg,
                align=align, num_fmt=fmt, border=THIN_BORDER)
    return cell


# ─────────────────────────────────────────────────────────────────────────────
# MTFS-specific mappings
# ─────────────────────────────────────────────────────────────────────────────
# Sector-based Momentum_Spread column mapping (mirrors record_vars in PortfolioMTFSRun.py)
from pair_universe import mtfs_spread_col_map, mtfs_pair_keys

SPREAD_COL_MAP = mtfs_spread_col_map()
PAIR_ORDER     = mtfs_pair_keys()

# PARAM_MAP is used only as a display fallback; actual params come from the Step 2/3 Excel files.
PARAM_MAP = {}


# ─────────────────────────────────────────────────────────────────────────────
# Data extraction
# ─────────────────────────────────────────────────────────────────────────────
def extract_pair_stats(bt_path, fwd_path):
    """
    Returns dict: pair -> {bt: {...}, fwd: {...}}
    Each sub-dict contains all per-pair metrics.
    """
    # ── load sheets ──────────────────────────────────────────────────────────
    def load_all(path):
        return {
            'pt':    _read(path, 'pair_trade_history'),
            'rv':    _read(path, 'recorded_vars'),
            'acc':   _read(path, 'acc_pair_trade_pnl_history'),
            'dod':   _read(path, 'dod_pair_trade_pnl_history'),
            'eq':    _read(path, 'equity_history'),
            'dpnl':  _read(path, 'daily_pnl_history'),
            'int':   _read(path, 'acc_interest_history'),
            'sl':    _read(path, 'stop_loss_history'),
            'apnl':  _read(path, 'acc_daily_pnl_history'),
            'mdd':   _read(path, 'max_drawdown_history'),
            'price': _read(path, 'price_history'),
        }

    bt  = load_all(bt_path)
    fwd = load_all(fwd_path)

    # ── date ranges ──────────────────────────────────────────────────────────
    def date_range(sheets):
        eq = sheets['eq']
        if eq.empty:
            return None, None
        return eq['Date'].iloc[0], eq['Date'].iloc[-1]

    bt_start,  bt_end  = date_range(bt)
    fwd_start, fwd_end = date_range(fwd)

    # ── portfolio-level metrics ───────────────────────────────────────────────
    def port_metrics(sheets):
        eq    = sheets['eq']
        dpnl  = sheets['dpnl']
        apnl  = sheets['apnl']
        intdf = sheets['int']
        mdd   = sheets['mdd']
        acc   = sheets['acc']

        final_equity   = eq['Value'].iloc[-1]   if not eq.empty   else 0
        initial_equity = eq['Value'].iloc[0]    if not eq.empty   else 500000
        total_pnl      = apnl['Value'].iloc[-1] if not apnl.empty else 0
        interest       = intdf['Value'].iloc[-1] if not intdf.empty else 0
        trading_days   = len(dpnl)

        # Realized PnL = sum of each pair's last acc_pair_trade_pnl value
        realized_pnl = 0.0
        if not acc.empty and 'Pair' in acc.columns:
            for pair in acc['Pair'].unique():
                realized_pnl += acc[acc['Pair'] == pair]['PnL Dollar'].iloc[-1]

        trade_pnl    = total_pnl + interest
        unrealized_pnl = trade_pnl - realized_pnl

        # Sharpe (annualised daily)
        if not dpnl.empty and trading_days > 1:
            rets = dpnl['Daily PnL'] / initial_equity
            sharpe = (rets.mean() / rets.std() * np.sqrt(252)
                      if rets.std() > 0 else np.nan)
        else:
            sharpe = np.nan

        # Max drawdown
        max_dd_dollar = max_dd_pct = 0
        if not mdd.empty:
            max_dd_dollar = mdd['Max Drawdown ($)'].max()
            pct_col = 'Max Drawdown (%)'
            if pct_col in mdd.columns:
                pcts = pd.to_numeric(
                    mdd[pct_col].astype(str).str.replace('%',''), errors='coerce')
                max_dd_pct = pcts.max() / 100 if not pcts.isna().all() else 0

        return {
            'final_equity': final_equity,
            'initial_equity': initial_equity,
            'total_pnl': total_pnl,
            'interest': interest,
            'trade_pnl': trade_pnl,
            'realized_pnl': realized_pnl,
            'unrealized_pnl': unrealized_pnl,
            'trading_days': trading_days,
            'sharpe': sharpe,
            'max_dd_dollar': max_dd_dollar,
            'max_dd_pct': max_dd_pct,
        }

    bt_port  = port_metrics(bt)
    fwd_port = port_metrics(fwd)

    # ── per-pair extraction ───────────────────────────────────────────────────
    def pair_metrics(sheets, pair):
        pt    = sheets['pt']
        rv    = sheets['rv']
        acc   = sheets['acc']
        dod   = sheets['dod']
        sl    = sheets['sl']
        price = sheets['price']

        if acc.empty:
            return None

        s1, s2 = pair.split('/')
        p      = pt[pt['Pair'] == pair].copy().sort_values('Date') if not pt.empty else pd.DataFrame()

        opens  = p[(p['Order Type'] == 'open')  & (p['Symbol'] == s1)].reset_index(drop=True)
        closes = p[(p['Order Type'] == 'close') & (p['Symbol'] == s1)].reset_index(drop=True)
        opens2  = p[(p['Order Type'] == 'open')  & (p['Symbol'] == s2)].reset_index(drop=True)
        closes2 = p[(p['Order Type'] == 'close') & (p['Symbol'] == s2)].reset_index(drop=True)

        a = acc[acc['Pair'] == pair].sort_values('Date').reset_index(drop=True)
        d = dod[dod['Pair'] == pair].sort_values('Date').reset_index(drop=True)
        r = rv[rv['Pair']  == pair].sort_values('Date').set_index('Date')

        # latest prices for open-position valuation
        def _latest_price(sym):
            if price.empty or sym not in price.columns:
                return np.nan
            col = price[sym].dropna()
            return float(col.iloc[-1]) if not col.empty else np.nan

        # ── trade-level ──
        n = min(len(opens), len(closes))
        trades = []
        for i in range(n):
            od   = opens.iloc[i]['Date']
            cd   = closes.iloc[i]['Date']
            odir = opens.iloc[i]['Direction']
            # s1 leg
            s1_shares    = opens.iloc[i]['Amount']
            s1_open_px   = opens.iloc[i]['Price']
            s1_close_px  = closes.iloc[i]['Price'] if i < len(closes) else np.nan
            # s2 leg
            s2_shares    = opens2.iloc[i]['Amount']    if i < len(opens2)  else np.nan
            s2_open_px   = opens2.iloc[i]['Price']     if i < len(opens2)  else np.nan
            s2_close_px  = closes2.iloc[i]['Price']    if i < len(closes2) else np.nan

            # MTFS-specific: get momentum scores at open
            rv_o = r.loc[od] if od in r.index else None
            # Handle potential duplicate index (multiple pairs same date)
            if rv_o is not None and isinstance(rv_o, pd.DataFrame):
                rv_o = rv_o.iloc[0]

            spread_col = SPREAD_COL_MAP.get(pair, '')
            mom_1  = float(rv_o['Momentum_1'])  if rv_o is not None and 'Momentum_1'  in r.columns and not pd.isna(rv_o.get('Momentum_1', np.nan)) else np.nan
            mom_2  = float(rv_o['Momentum_2'])  if rv_o is not None and 'Momentum_2'  in r.columns and not pd.isna(rv_o.get('Momentum_2', np.nan)) else np.nan
            spread = float(rv_o[spread_col])    if rv_o is not None and spread_col     in r.columns and not pd.isna(rv_o.get(spread_col, np.nan)) else np.nan

            # PnL for this trade
            a_after  = a[a['Date'] <= cd]
            a_before = a[a['Date'] <  od]
            pnl_after  = a_after['PnL Dollar'].iloc[-1]  if not a_after.empty  else 0
            pnl_before = a_before['PnL Dollar'].iloc[-1] if not a_before.empty else 0
            trade_pnl  = pnl_after - pnl_before
            hold_days  = (cd - od).days

            # close reason
            sl_dates = set(sl['Date'].dropna()) if not sl.empty else set()
            reason = 'Stop Loss' if cd in sl_dates else 'Signal'

            trades.append({
                'open_date': od, 'close_date': cd,
                'direction': odir,
                's1': s1, 's2': s2,
                's1_shares': s1_shares,  's1_open_px': s1_open_px,  's1_close_px': s1_close_px,
                's2_shares': s2_shares,  's2_open_px': s2_open_px,  's2_close_px': s2_close_px,
                'mom_1': mom_1, 'mom_2': mom_2, 'momentum_spread': spread,
                'hold_days': hold_days, 'pnl': trade_pnl, 'close_reason': reason,
            })

        # open positions at end
        open_positions = []
        if len(opens) > n:
            for i in range(n, len(opens)):
                od   = opens.iloc[i]['Date']
                odir = opens.iloc[i]['Direction']
                rv_o = r.loc[od] if od in r.index else None
                if rv_o is not None and isinstance(rv_o, pd.DataFrame):
                    rv_o = rv_o.iloc[0]
                spread_col = SPREAD_COL_MAP.get(pair, '')
                mom_1  = float(rv_o['Momentum_1']) if rv_o is not None and 'Momentum_1' in r.columns and not pd.isna(rv_o.get('Momentum_1', np.nan)) else np.nan
                spread = float(rv_o[spread_col])   if rv_o is not None and spread_col    in r.columns and not pd.isna(rv_o.get(spread_col, np.nan)) else np.nan
                s1_shares   = opens.iloc[i]['Amount']
                s1_open_px  = opens.iloc[i]['Price']
                s2_shares   = opens2.iloc[i]['Amount']   if i < len(opens2) else np.nan
                s2_open_px  = opens2.iloc[i]['Price']    if i < len(opens2) else np.nan
                open_positions.append({
                    'open_date': od, 'direction': odir,
                    'mom_1': mom_1, 'momentum_spread': spread,
                    's1': s1, 's2': s2,
                    's1_shares': s1_shares,  's1_open_px': s1_open_px,
                    's1_cur_px':  _latest_price(s1),
                    's2_shares': s2_shares,  's2_open_px': s2_open_px,
                    's2_cur_px':  _latest_price(s2),
                })

        # ── aggregate metrics ──
        total_pnl = a['PnL Dollar'].iloc[-1] if not a.empty else 0
        pnls = [t['pnl'] for t in trades]
        wins = sum(1 for x in pnls if x > 0)
        losses = sum(1 for x in pnls if x <= 0)
        win_rate = wins / n if n > 0 else np.nan
        avg_win  = np.mean([x for x in pnls if x > 0]) if wins > 0 else 0
        avg_loss = np.mean([x for x in pnls if x <= 0]) if losses > 0 else 0
        profit_factor = abs(sum(x for x in pnls if x > 0) / sum(x for x in pnls if x < 0)) \
            if any(x < 0 for x in pnls) else np.inf
        avg_hold = np.mean([t['hold_days'] for t in trades]) if trades else 0

        # long vs short breakdown
        long_trades  = [t for t in trades if t['direction'] == 'long']
        short_trades = [t for t in trades if t['direction'] == 'short']
        long_pnl  = sum(t['pnl'] for t in long_trades)
        short_pnl = sum(t['pnl'] for t in short_trades)
        long_wins  = sum(1 for t in long_trades  if t['pnl'] > 0)
        short_wins = sum(1 for t in short_trades if t['pnl'] > 0)

        # Momentum spread stats at open
        spread_vals = [abs(t['momentum_spread']) for t in trades if not np.isnan(t['momentum_spread'])]
        avg_spread_open = np.mean(spread_vals) if spread_vals else np.nan

        # per-pair Sharpe from daily DoD PnL
        sharpe = np.nan
        if not d.empty and len(d) > 1:
            dod_vals = d['PnL Dollar']
            if dod_vals.std() > 0:
                sharpe = dod_vals.mean() / dod_vals.std() * np.sqrt(252)

        # max drawdown of this pair's acc PnL curve
        max_dd = 0
        if not a.empty:
            curve = a['PnL Dollar'].values
            peak = curve[0]
            for v in curve:
                peak = max(peak, v)
                dd = peak - v
                max_dd = max(max_dd, dd)

        stop_losses = len(sl[sl['Pair'] == pair]) if not sl.empty and 'Pair' in sl.columns else 0

        return {
            'total_pnl': total_pnl,
            'n_trades': n,
            'wins': wins,
            'losses': losses,
            'win_rate': win_rate,
            'avg_win': avg_win,
            'avg_loss': avg_loss,
            'profit_factor': profit_factor,
            'avg_hold_days': avg_hold,
            'long_trades': len(long_trades),
            'short_trades': len(short_trades),
            'long_pnl': long_pnl,
            'short_pnl': short_pnl,
            'long_win_rate': long_wins / len(long_trades) if long_trades else np.nan,
            'short_win_rate': short_wins / len(short_trades) if short_trades else np.nan,
            'avg_spread_open': avg_spread_open,
            'sharpe': sharpe,
            'max_dd_dollar': max_dd,
            'stop_losses': stop_losses,
            'open_positions': len(open_positions),
            'trades': trades,
            'open_pos_detail': open_positions,
        }

    result = {}
    all_pairs = PAIR_ORDER.copy()
    if not bt['pt'].empty:
        for p in bt['pt']['Pair'].unique():
            if p not in all_pairs:
                all_pairs.append(p)

    for pair in all_pairs:
        bm = pair_metrics(bt,  pair)
        fm = pair_metrics(fwd, pair)
        if bm is None and fm is None:
            continue
        result[pair] = {
            'bt': bm,
            'fwd': fm,
            'param': PARAM_MAP.get(pair, 'default'),
        }

    return result, bt_port, fwd_port, bt_start, bt_end, fwd_start, fwd_end


# ─────────────────────────────────────────────────────────────────────────────
# Sheet builders
# ─────────────────────────────────────────────────────────────────────────────

def build_summary_sheet(wb, data, bt_port, fwd_port,
                        bt_start, bt_end, fwd_start, fwd_end):
    ws = wb.create_sheet('Summary')
    ws.sheet_view.showGridLines = False

    # ── Title ────────────────────────────────────────────────────────────────
    ws.merge_cells('A1:P1')
    t = ws['A1']
    t.value = 'MTFS (Momentum Trend Following) — Backtest vs Validation Report'
    _style_cell(t, bold=True, font_color=C_HEADER_FG, bg_color=C_HEADER_BG,
                align='center', size=14)
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:P2')
    sub = ws['A2']
    bt_range  = f"{bt_start.date()} → {bt_end.date()}"
    fwd_range = f"{fwd_start.date()} → {fwd_end.date()}"
    sub.value = f"Backtest: {bt_range}    |    Validation: {fwd_range}    |    Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    _style_cell(sub, italic=True, bg_color=C_SUBHDR_BG, font_color=C_HEADER_FG,
                align='center', size=10)
    ws.row_dimensions[2].height = 18

    # ── Portfolio totals ─────────────────────────────────────────────────────
    r = 4
    ws.merge_cells(f'A{r}:P{r}')
    _style_cell(ws.cell(r, 1, 'PORTFOLIO LEVEL'), bold=True,
                bg_color=C_SECTION_BG, align='center', size=11, border=THIN_BORDER)
    ws.row_dimensions[r].height = 20
    r += 1

    port_metrics_rows = [
        ('Period',            bt_range, fwd_range, 'text',         False),
        ('Trading Days',      bt_port['trading_days'], fwd_port['trading_days'], 'int', False),
        ('Initial Equity',    bt_port['initial_equity'], fwd_port['initial_equity'], 'dollar', False),
        ('Final Equity',      bt_port['final_equity'], fwd_port['final_equity'], 'dollar', False),
        ('Total PnL (incl. interest)',
                              bt_port['total_pnl'], fwd_port['total_pnl'], 'dollar_sign', False),
        ('Interest Expense',  -bt_port['interest'], -fwd_port['interest'], 'dollar_sign', False),
        ('Pair Trade PnL (MTM)',  bt_port['trade_pnl'], fwd_port['trade_pnl'], 'dollar_sign', False),
        ('  ↳ Realized Pair PnL', bt_port['realized_pnl'], fwd_port['realized_pnl'], 'dollar_sign', True),
        ('  ↳ Unrealized (MTM)',  bt_port['unrealized_pnl'], fwd_port['unrealized_pnl'], 'dollar_sign', True),
        ('Sharpe Ratio',      bt_port['sharpe'], fwd_port['sharpe'], 'sharpe',        False),
        ('Max Drawdown ($)',   bt_port['max_dd_dollar'], fwd_port['max_dd_dollar'], 'dollar', False),
        ('Max Drawdown (%)',   bt_port['max_dd_pct'], fwd_port['max_dd_pct'], 'pct',  False),
    ]

    for label, bv, fv, typ, indent in port_metrics_rows:
        ws.cell(r, 1, label)
        _style_cell(ws.cell(r, 1), bold=not indent, italic=indent, align='left',
                    border=THIN_BORDER,
                    bg_color=C_ALT_ROW if r % 2 == 0 else None)
        for col, val in [(2, bv), (3, fv)]:
            cell = ws.cell(r, col, val)
            if typ == 'text':
                _style_cell(cell, align='center', border=THIN_BORDER)
            elif typ == 'dollar':
                _style_cell(cell, num_fmt='#,##0', align='right', border=THIN_BORDER)
            elif typ == 'dollar_sign':
                _style_cell(cell, num_fmt='#,##0;[Red]-#,##0', align='right',
                            color_sign=True, border=THIN_BORDER,
                            italic=indent)
            elif typ == 'pct':
                _style_cell(cell, num_fmt='0.00%', align='right', border=THIN_BORDER)
            elif typ == 'sharpe':
                _style_cell(cell, num_fmt='0.0000', align='right',
                            color_sign=True, border=THIN_BORDER)
            else:
                _style_cell(cell, align='right', border=THIN_BORDER)
        r += 1

    r += 1

    # ── Per-pair table ────────────────────────────────────────────────────────
    ws.merge_cells(f'A{r}:P{r}')
    _style_cell(ws.cell(r, 1, 'PER-PAIR SUMMARY'), bold=True,
                bg_color=C_SECTION_BG, align='center', size=11, border=THIN_BORDER)
    ws.row_dimensions[r].height = 20
    r += 1

    COLS = [
        ('Pair',            8,  'text'),
        ('Param Set',      16,  'text'),
        ('BT PnL',         10,  'dollar_sign'),
        ('BT Trades',       8,  'int'),
        ('BT Win%',         8,  'pct'),
        ('BT Sharpe',       9,  'sharpe'),
        ('BT MaxDD$',      10,  'dollar'),
        ('BT Long PnL',    10,  'dollar_sign'),
        ('BT Short PnL',   10,  'dollar_sign'),
        ('FWD PnL',        10,  'dollar_sign'),
        ('FWD Trades',      8,  'int'),
        ('FWD Win%',        8,  'pct'),
        ('FWD Sharpe',      9,  'sharpe'),
        ('FWD MaxDD$',     10,  'dollar'),
        ('FWD Long PnL',   10,  'dollar_sign'),
        ('FWD Short PnL',  10,  'dollar_sign'),
    ]

    for ci, (hdr, w, _) in enumerate(COLS, 1):
        _header(ws, r, ci, hdr, sub=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
    r += 1

    bt_total_pnl = fwd_total_pnl = 0
    for pi, pair in enumerate(data.keys()):
        m    = data[pair]
        bm   = m['bt']  or {}
        fm   = m['fwd'] or {}
        alt  = pi % 2 == 1
        bg   = C_ALT_ROW if alt else None

        def gv(d, k):
            return d.get(k, np.nan) if d else np.nan

        row_vals = [
            (pair,                       'text',        'FF000000', False),
            (m['param'],                 'text',        'FF000000', False),
            (gv(bm,'total_pnl'),         'dollar_sign', None,       True),
            (gv(bm,'n_trades'),          'int',         None,       False),
            (gv(bm,'win_rate'),          'pct',         None,       False),
            (gv(bm,'sharpe'),            'sharpe',      None,       True),
            (gv(bm,'max_dd_dollar'),     'dollar',      None,       False),
            (gv(bm,'long_pnl'),          'dollar_sign', None,       True),
            (gv(bm,'short_pnl'),         'dollar_sign', None,       True),
            (gv(fm,'total_pnl'),         'dollar_sign', None,       True),
            (gv(fm,'n_trades'),          'int',         None,       False),
            (gv(fm,'win_rate'),          'pct',         None,       False),
            (gv(fm,'sharpe'),            'sharpe',      None,       True),
            (gv(fm,'max_dd_dollar'),     'dollar',      None,       False),
            (gv(fm,'long_pnl'),          'dollar_sign', None,       True),
            (gv(fm,'short_pnl'),         'dollar_sign', None,       True),
        ]

        for ci, ((val, typ, fc, cs), (_, w, _t)) in enumerate(zip(row_vals, COLS), 1):
            cell = ws.cell(r, ci, val)
            bg_c = bg
            fg_c = fc or 'FF000000'
            if cs and isinstance(val, float) and not np.isnan(val):
                if val > 0:   fg_c = C_POSITIVE; bg_c = C_POSITIVE_BG
                elif val < 0: fg_c = C_NEGATIVE; bg_c = C_NEGATIVE_BG
            fmt = None
            if   typ == 'dollar_sign': fmt = '#,##0;[Red]-#,##0'
            elif typ == 'dollar':      fmt = '#,##0'
            elif typ == 'pct':         fmt = '0.0%'
            elif typ == 'sharpe':      fmt = '0.00'
            elif typ == 'int':         fmt = '#,##0'
            _style_cell(cell, font_color=fg_c, bg_color=bg_c,
                        align='left' if typ=='text' else 'right',
                        num_fmt=fmt, border=THIN_BORDER)

        bt_total_pnl  += gv(bm, 'total_pnl')  if not np.isnan(gv(bm,'total_pnl'))  else 0
        fwd_total_pnl += gv(fm, 'total_pnl')  if not np.isnan(gv(fm,'total_pnl'))  else 0
        r += 1

    # Totals row
    totals = ['TOTAL', '', bt_total_pnl, '', '', '', '', '', '',
              fwd_total_pnl, '', '', '', '', '', '']
    for ci, (val, (_, w, typ)) in enumerate(zip(totals, COLS), 1):
        cell = ws.cell(r, ci, val)
        fg = C_HEADER_FG
        bg = C_TOTAL_BG
        if isinstance(val, float) and val != 0:
            fmt = '#,##0;[Red]-#,##0'
        else:
            fmt = None
        _style_cell(cell, bold=True, font_color=fg, bg_color=bg,
                    align='center', num_fmt=fmt, border=THICK_BORDER)

    ws.freeze_panes = ws.cell(r - len(data) - 1, 1)
    return ws


def build_pair_sheet(wb, pair, m, bt_start, bt_end, fwd_start, fwd_end):
    safe = pair.replace('/', '_')
    ws = wb.create_sheet(safe)
    ws.sheet_view.showGridLines = False

    # Total columns in the trade table
    NCOLS = 18

    bm = m['bt']  or {}
    fm = m['fwd'] or {}

    def gv(d, k, default=np.nan):
        return d.get(k, default) if d else default

    # ── Title ────────────────────────────────────────────────────────────────
    ws.merge_cells(f'A1:{get_column_letter(NCOLS)}1')
    _style_cell(ws.cell(1, 1, f'{pair}  |  Param: {m["param"]}  |  Strategy: MTFS'),
                bold=True, font_color=C_HEADER_FG, bg_color=C_HEADER_BG,
                align='center', size=13, border=THIN_BORDER)
    ws.row_dimensions[1].height = 26

    # ── Metrics table ─────────────────────────────────────────────────────────
    r = 3
    _header(ws, r, 1, 'Metric',    span=3)
    _header(ws, r, 4, f'Backtest  ({bt_start.date()} → {bt_end.date()})', span=4, sub=True)
    _header(ws, r, 8, f'Validation  ({fwd_start.date()} → {fwd_end.date()})', span=4, sub=True)
    ws.row_dimensions[r].height = 18
    r += 1

    def stat_row(label, bval, fval, fmt='#,##0', color_sign=False, alt=False):
        bg = C_ALT_ROW if alt else None
        ws.cell(r, 1, label)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        _style_cell(ws.cell(r, 1), bold=True, align='left',
                    bg_color=bg, border=THIN_BORDER)
        for col, val in [(4, bval), (8, fval)]:
            cell = ws.cell(r, col, val)
            ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col+3)
            fg = 'FF000000'
            bg2 = bg
            if color_sign and isinstance(val, (int, float)) and not np.isnan(val):
                if val > 0:   fg = C_POSITIVE;  bg2 = C_POSITIVE_BG
                elif val < 0: fg = C_NEGATIVE;  bg2 = C_NEGATIVE_BG
            _style_cell(cell, font_color=fg, bg_color=bg2,
                        num_fmt=fmt, align='right', border=THIN_BORDER)

    metrics = [
        ('Total PnL ($)',        gv(bm,'total_pnl'), gv(fm,'total_pnl'), '#,##0;[Red]-#,##0', True),
        ('Sharpe Ratio',         gv(bm,'sharpe'),    gv(fm,'sharpe'),    '0.0000', True),
        ('Max Drawdown ($)',      gv(bm,'max_dd_dollar'), gv(fm,'max_dd_dollar'), '#,##0', False),
        ('Total Trades',         gv(bm,'n_trades'),  gv(fm,'n_trades'),  '#,##0', False),
        ('Win Rate',             gv(bm,'win_rate'),  gv(fm,'win_rate'),  '0.0%', False),
        ('Avg Win ($)',          gv(bm,'avg_win'),   gv(fm,'avg_win'),   '#,##0', True),
        ('Avg Loss ($)',         gv(bm,'avg_loss'),  gv(fm,'avg_loss'),  '#,##0;[Red]-#,##0', True),
        ('Profit Factor',        gv(bm,'profit_factor'), gv(fm,'profit_factor'), '0.00', False),
        ('Avg Hold (days)',       gv(bm,'avg_hold_days'), gv(fm,'avg_hold_days'), '0.0', False),
        ('Long Trades',          gv(bm,'long_trades'),  gv(fm,'long_trades'), '#,##0', False),
        ('Long PnL ($)',         gv(bm,'long_pnl'),  gv(fm,'long_pnl'),  '#,##0;[Red]-#,##0', True),
        ('Long Win Rate',        gv(bm,'long_win_rate'), gv(fm,'long_win_rate'), '0.0%', False),
        ('Short Trades',         gv(bm,'short_trades'), gv(fm,'short_trades'), '#,##0', False),
        ('Short PnL ($)',        gv(bm,'short_pnl'), gv(fm,'short_pnl'), '#,##0;[Red]-#,##0', True),
        ('Short Win Rate',       gv(bm,'short_win_rate'), gv(fm,'short_win_rate'), '0.0%', False),
        ('Avg |Mom Spread| at Entry', gv(bm,'avg_spread_open'), gv(fm,'avg_spread_open'), '0.000', False),
        ('Stop Losses Triggered', gv(bm,'stop_losses'), gv(fm,'stop_losses'), '#,##0', False),
        ('Open Positions at End', gv(bm,'open_positions'), gv(fm,'open_positions'), '#,##0', False),
    ]

    for i, (lbl, bval, fval, fmt, cs) in enumerate(metrics):
        stat_row(lbl, bval, fval, fmt=fmt, color_sign=cs, alt=(i%2==1))
        r += 1

    r += 1

    # ── Trade-by-trade tables ─────────────────────────────────────────────────
    # Columns: # | Dir | Open Date | Close Date | Hold(d) | Mom1 | Mom2
    #          | S1 | Shrs1 | S1 Open | S1 Close | S2 | Shrs2 | S2 Open | S2 Close
    #          | PnL($) | Close Reason | Cum PnL

    def trade_table(trades, open_pos, label):
        nonlocal r
        ws.merge_cells(f'A{r}:{get_column_letter(NCOLS)}{r}')
        _style_cell(ws.cell(r, 1, label), bold=True,
                    bg_color=C_SECTION_BG, align='center', border=THIN_BORDER)
        ws.row_dimensions[r].height = 18
        r += 1

        # Two-row header
        groups = [
            (1, 1, '#'),
            (2, 2, 'Direction'),
            (3, 3, 'Open Date'),
            (4, 4, 'Close Date'),
            (5, 5, 'Hold(d)'),
            (6, 6, 'Mom1'),
            (7, 7, 'MomSpread'),
            (8, 11, 'Long Leg'),
            (12, 15, 'Short Leg'),
            (16, 16, 'PnL ($)'),
            (17, 17, 'Close Reason'),
            (18, 18, 'Cum PnL'),
        ]
        for c1, c2, lbl in groups:
            cell = ws.cell(r, c1, lbl)
            if c2 > c1:
                ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
            _style_cell(cell, bold=True, font_color=C_HEADER_FG,
                        bg_color=C_HEADER_BG, align='center', border=THIN_BORDER, size=10)
        ws.row_dimensions[r].height = 16
        r += 1

        # Row 2: individual column headers
        sub_hdrs = [
            ('#',           4),
            ('Direction',   9),
            ('Open Date',  12),
            ('Close Date', 12),
            ('Hold(d)',     7),
            ('Mom1',        9),
            ('MomSpread',   9),
            ('Symbol',      7),
            ('Shares',      8),
            ('Open Px',    10),
            ('Close Px',   10),
            ('Symbol',      7),
            ('Shares',      8),
            ('Open Px',    10),
            ('Close Px',   10),
            ('PnL ($)',    12),
            ('Reason',     13),
            ('Cum PnL',    12),
        ]
        for ci, (h, w) in enumerate(sub_hdrs, 1):
            _header(ws, r, ci, h, sub=True)
            ws.column_dimensions[get_column_letter(ci)].width = w
        r += 1

        cum = 0
        for i, t in enumerate(trades):
            alt = i % 2 == 1
            bg = C_ALT_ROW if alt else None
            cum += t['pnl']
            # determine which is long leg and which is short leg
            if t['direction'] == 'long':
                ll_sym, ll_shr, ll_opx, ll_cpx = t['s1'], t['s1_shares'], t['s1_open_px'], t['s1_close_px']
                sl_sym, sl_shr, sl_opx, sl_cpx = t['s2'], t['s2_shares'], t['s2_open_px'], t['s2_close_px']
            else:
                ll_sym, ll_shr, ll_opx, ll_cpx = t['s2'], t['s2_shares'], t['s2_open_px'], t['s2_close_px']
                sl_sym, sl_shr, sl_opx, sl_cpx = t['s1'], t['s1_shares'], t['s1_open_px'], t['s1_close_px']

            row_data = [
                (i+1,                    '#,##0',              False),
                (t['direction'],         None,                 False),
                (t['open_date'],         'YYYY-MM-DD',         False),
                (t['close_date'],        'YYYY-MM-DD',         False),
                (t['hold_days'],         '#,##0',              False),
                (t['mom_1'],             '+0.000;-0.000',      False),
                (t['momentum_spread'],   '+0.000;-0.000',      False),
                (ll_sym,                 None,                 False),
                (ll_shr,                 '#,##0',              False),
                (ll_opx,                 '#,##0.00',           False),
                (ll_cpx,                 '#,##0.00',           False),
                (sl_sym,                 None,                 False),
                (sl_shr,                 '#,##0',              False),
                (sl_opx,                 '#,##0.00',           False),
                (sl_cpx,                 '#,##0.00',           False),
                (t['pnl'],              '#,##0;[Red]-#,##0',  True),
                (t['close_reason'],     None,                  False),
                (cum,                    '#,##0;[Red]-#,##0',  True),
            ]
            for ci, (val, fmt, cs) in enumerate(row_data, 1):
                cell = ws.cell(r, ci, val)
                bg2 = bg
                fg  = 'FF000000'
                if cs and isinstance(val, (int, float)) and not np.isnan(val):
                    if val > 0:   fg = C_POSITIVE;  bg2 = C_POSITIVE_BG
                    elif val < 0: fg = C_NEGATIVE;  bg2 = C_NEGATIVE_BG
                _style_cell(cell, font_color=fg, bg_color=bg2,
                            align='left' if fmt is None else 'right',
                            num_fmt=fmt, border=THIN_BORDER)
            r += 1

        # open positions
        for op in open_pos:
            if op['direction'] == 'long':
                ll_sym, ll_shr, ll_opx, ll_cpx = op['s1'], op['s1_shares'], op['s1_open_px'], op['s1_cur_px']
                sl_sym, sl_shr, sl_opx, sl_cpx = op['s2'], op['s2_shares'], op['s2_open_px'], op['s2_cur_px']
            else:
                ll_sym, ll_shr, ll_opx, ll_cpx = op['s2'], op['s2_shares'], op['s2_open_px'], op['s2_cur_px']
                sl_sym, sl_shr, sl_opx, sl_cpx = op['s1'], op['s1_shares'], op['s1_open_px'], op['s1_cur_px']

            open_data = [
                ('→',              None,             False),
                (op['direction'],  None,             False),
                (op['open_date'],  'YYYY-MM-DD',     False),
                ('[Still Open]',   None,             False),
                (None,             None,             False),
                (op.get('mom_1', np.nan),           '+0.000;-0.000', False),
                (op.get('momentum_spread', np.nan), '+0.000;-0.000', False),
                (ll_sym,           None,             False),
                (ll_shr,           '#,##0',          False),
                (ll_opx,           '#,##0.00',       False),
                (ll_cpx,           '#,##0.00',       False),
                (sl_sym,           None,             False),
                (sl_shr,           '#,##0',          False),
                (sl_opx,           '#,##0.00',       False),
                (sl_cpx,           '#,##0.00',       False),
                (None,             None,             False),
                (None,             None,             False),
                (None,             None,             False),
            ]
            for ci, (val, fmt, cs) in enumerate(open_data, 1):
                cell = ws.cell(r, ci, val)
                _style_cell(cell, italic=True, bg_color=C_NEUTRAL_BG,
                            align='left' if fmt is None else 'right',
                            num_fmt=fmt, border=THIN_BORDER)
            r += 1

    bt_trades  = gv(bm, 'trades', [])
    fwd_trades = gv(fm, 'trades', [])

    trade_table(bt_trades,  gv(bm, 'open_pos_detail', []),
                f'BACKTEST TRADES  ({bt_start.date()} → {bt_end.date()})')
    r += 1
    trade_table(fwd_trades, gv(fm, 'open_pos_detail', []),
                f'VALIDATION TRADES  ({fwd_start.date()} → {fwd_end.date()})')

    ws.freeze_panes = 'A4'
    return ws


def build_daily_pnl_sheet(wb, bt_path, fwd_path, bt_start, bt_end, fwd_start, fwd_end):
    ws = wb.create_sheet('Daily PnL (Validation)')
    ws.sheet_view.showGridLines = False

    fwd_dod = _read(fwd_path, 'dod_pair_trade_pnl_history')
    if fwd_dod.empty:
        ws['A1'] = 'No data'
        return ws

    fwd_dod['Date'] = pd.to_datetime(fwd_dod['Date'])
    pivot = fwd_dod.pivot_table(
        index='Date', columns='Pair', values='PnL Dollar', aggfunc='sum'
    ).fillna(0)
    # Reorder columns
    cols = [p for p in PAIR_ORDER if p in pivot.columns] + \
           [p for p in pivot.columns if p not in PAIR_ORDER]
    pivot = pivot[cols]
    pivot['TOTAL'] = pivot.sum(axis=1)
    pivot.index = pivot.index.strftime('%Y-%m-%d')

    ws.merge_cells(f'A1:{get_column_letter(len(pivot.columns)+2)}1')
    _style_cell(ws.cell(1, 1, f'MTFS Validation Period — Daily PnL by Pair  ({fwd_start.date()} → {fwd_end.date()})'),
                bold=True, font_color=C_HEADER_FG, bg_color=C_HEADER_BG,
                align='center', size=12, border=THIN_BORDER)
    ws.row_dimensions[1].height = 24

    # Header row
    ws.cell(2, 1, 'Date')
    _style_cell(ws.cell(2, 1), bold=True, bg_color=C_SUBHDR_BG,
                font_color=C_HEADER_FG, align='center', border=THIN_BORDER)
    ws.column_dimensions['A'].width = 12
    for ci, col in enumerate(pivot.columns, 2):
        _header(ws, 2, ci, col, sub=True)
        ws.column_dimensions[get_column_letter(ci)].width = 11

    # Data rows
    for ri, (date_str, row) in enumerate(pivot.iterrows(), 3):
        alt = ri % 2 == 0
        ws.cell(ri, 1, date_str)
        _style_cell(ws.cell(ri, 1), bold=True, border=THIN_BORDER,
                    bg_color=C_ALT_ROW if alt else None)
        for ci, val in enumerate(row.values, 2):
            cell = ws.cell(ri, ci, val)
            is_total = ci == len(pivot.columns) + 1
            fg = 'FF000000'
            bg = C_ALT_ROW if alt else None
            if isinstance(val, (int, float)):
                if val > 0:   fg = C_POSITIVE;  bg = C_POSITIVE_BG
                elif val < 0: fg = C_NEGATIVE;  bg = C_NEGATIVE_BG
            _style_cell(cell, bold=is_total, font_color=fg, bg_color=bg,
                        num_fmt='#,##0;[Red]-#,##0', align='right', border=THIN_BORDER)

    # Totals row
    total_row = pivot.sum()
    tr = len(pivot) + 3
    ws.cell(tr, 1, 'TOTAL')
    _style_cell(ws.cell(tr, 1), bold=True, bg_color=C_TOTAL_BG,
                font_color=C_HEADER_FG, align='center', border=THICK_BORDER)
    for ci, val in enumerate(total_row.values, 2):
        cell = ws.cell(tr, ci, val)
        _style_cell(cell, bold=True, bg_color=C_TOTAL_BG, font_color=C_HEADER_FG,
                    num_fmt='#,##0;[Red]-#,##0', align='right', border=THICK_BORDER)

    ws.freeze_panes = 'B3'
    return ws


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────
def generate_report(bt_path, fwd_path, out_path=None):
    print(f"Loading backtest:   {os.path.basename(bt_path)}")
    print(f"Loading validation: {os.path.basename(fwd_path)}")

    data, bt_port, fwd_port, bt_start, bt_end, fwd_start, fwd_end = \
        extract_pair_stats(bt_path, fwd_path)

    print(f"Found {len(data)} pairs")
    print(f"Backtest:   {bt_start.date()} → {bt_end.date()}  "
          f"({bt_port['trading_days']} days)")
    print(f"Validation: {fwd_start.date()} → {fwd_end.date()}  "
          f"({fwd_port['trading_days']} days)")

    wb = openpyxl.Workbook()
    del wb[wb.sheetnames[0]]

    print("Building Summary sheet...")
    build_summary_sheet(wb, data, bt_port, fwd_port,
                        bt_start, bt_end, fwd_start, fwd_end)

    print("Building Daily PnL sheet...")
    build_daily_pnl_sheet(wb, bt_path, fwd_path,
                          bt_start, bt_end, fwd_start, fwd_end)

    print("Building per-pair sheets...")
    for pair in data:
        build_pair_sheet(wb, pair, data[pair],
                         bt_start, bt_end, fwd_start, fwd_end)
        print(f"  {pair}")

    if out_path is None:
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        out_path = os.path.join(
            os.path.dirname(bt_path),
            f'mtfs_report_bt_vs_fwd_{ts}.xlsx'
        )

    wb.save(out_path)
    print(f"\nReport saved: {out_path}")
    return out_path


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print(__doc__)
        # Default: use latest matching MTFS files
        import glob as _glob
        bt_files  = sorted(_glob.glob('historical_runs/portfolio_history_MTFS_all15_*_default_*.xlsx'))
        fwd_files = sorted(_glob.glob('historical_runs/portfolio_history_MTFS_fwd_*.xlsx'))
        if bt_files and fwd_files:
            bt_path  = bt_files[-1]
            fwd_path = fwd_files[-1]
            print(f"\nAuto-detected files:")
            print(f"  Backtest:   {os.path.basename(bt_path)}")
            print(f"  Validation: {os.path.basename(fwd_path)}")
        else:
            print("Usage: python MTFSGenerateReport.py <backtest_excel> <forward_excel> [output_excel]")
            sys.exit(1)
    else:
        bt_path  = sys.argv[1]
        fwd_path = sys.argv[2]

    out = sys.argv[3] if len(sys.argv) > 3 else None
    generate_report(bt_path, fwd_path, out)
