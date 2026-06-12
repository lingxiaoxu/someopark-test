"""
RiskManager.py — Institutional-grade risk, position-management & financial-statements
module for the MRPT + MTFS pairs book (read-only).

Hooked non-fatally at the end of DailySignal.run_daily_signal(). Consumes the
in-memory objects DailySignal already computed (regime / capital / position_monitor /
strategy signal outputs) for zero-error reconciliation with the daily report, and
additionally reads (read-only): inventory_*.json (final state), inventory_history/*,
strategy_performance.json, master_portfolio_performance.json, PriceDataStore (prices +
volume), Ken French factor library, and DailySignal shared limit constants.

Outputs (ALL into trading_signals/risk_management/, never touching existing files):
    risk_workbook_<ts>.xlsx   — 19-sheet detailed workbook
    risk_report_<ts>.pdf      — PnLReport-styled PDF
    risk_report_<ts>.json     — machine-readable summary (frontend / log)
    risk_report_<ts>.txt      — quick human-readable summary

Scope: the MRPT + MTFS pairs book. "combined" = MRPT + MTFS (= strategy_performance
combined_equity, authoritative NAV). The 4-strategy master appears only in the D1
risk-contribution diagnostic (portfolio-construction view, from master JSON).

Theory references (report-only diagnostics, §4B of the plan):
    Markowitz (1952); Sharpe (1964); Kelly (1956); Jegadeesh-Titman (1993);
    Fama-French (1993, 2015); Carhart (1997); Engle-Granger (1987);
    Gatev-Goetzmann-Rouwenhorst (2006); Avellaneda-Lee (2010); Maillard-Roncalli-
    Teiletche (2010); Artzner et al. (1999); Rockafellar-Uryasev (2000);
    Litterman (1996); Chekhlov-Uryasev-Zabarankin (2005); Almgren-Chriss (2000);
    Bailey-Lopez de Prado (2012, 2014); Cont (2001); Zangari (1996); Khandani-Lo (2007).

All financial statements are model-based (prime-broker market-neutral convention),
NOT broker-reconciled.
"""

import os
import sys
import json
import glob
import math
import logging
from datetime import datetime, timedelta, date

import numpy as np
import pandas as pd

log = logging.getLogger(__name__)

# ── Paths ───────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RISK_DIR = os.path.join(BASE_DIR, 'trading_signals', 'risk_management')
FACTOR_DIR = os.path.join(BASE_DIR, 'price_data', 'factors')
PERF_JSON = os.path.join(BASE_DIR, 'someo-park-investment-management', 'public', 'data',
                         'strategy_performance.json')
MASTER_JSON = os.path.join(BASE_DIR, 'someo-park-investment-management', 'public', 'data',
                           'master_portfolio_performance.json')
INV_HISTORY_DIR = os.path.join(BASE_DIR, 'inventory_history')

# ── Risk model constants (top-level, tunable) ───────────────────────────────
ANNUALIZATION = 252
MARGIN_RATE = 0.05                  # annual margin borrowing cost (matches Portfolio)
VAR_LOOKBACK = 60                   # trading days for vol/beta/VaR covariance
BETA_LOOKBACK = 60                  # trading days for beta-to-SPY
PRICE_LOOKBACK_CAL_DAYS = 400       # calendar days to fetch (~252 trading days + buffer)
ADV_WINDOW = 20                     # trading days for average daily volume
ADV_PARTICIPATION = 0.20            # max participation rate (Almgren-Chriss liquidity horizon)
RISK_FREE_ANNUAL = 0.05             # for Sharpe excess return

# Stress scenarios (beta-implied)
STRESS_SCENARIOS = [
    ('SPY -5%', 'spy', -0.05),
    ('SPY -10%', 'spy', -0.10),
    ('VIX +10pt', 'vix', 10.0),
    ('Top sector -10%', 'sector', -0.10),
]

# Limit thresholds (amber, red). Shared/enforcement limits import from DailySignal.
LIMITS_SPEC = {
    'gross_leverage':    {'amber': 3.0,  'red': 4.0,  'fmt': 'x'},
    'net_leverage_abs':  {'amber': 0.5,  'red': 1.0,  'fmt': 'x'},
    'single_name_gross': {'amber': 15.0, 'red': 25.0, 'fmt': '%'},   # % of capital
    'sector_net':        {'amber': 30.0, 'red': 50.0, 'fmt': '%'},   # % of gross
    'max_pair':          {'amber': 20.0, 'red': 35.0, 'fmt': '%'},   # % of gross
    'net_market_beta':   {'amber': 0.30, 'red': 0.50, 'fmt': 'b'},
    'var_95_1d':         {'amber': 3.0,  'red': 5.0,  'fmt': '%'},   # % of capital
}


def _round(x, n=2):
    """NaN/inf-safe round → float or None."""
    try:
        if x is None:
            return None
        xf = float(x)
        if math.isnan(xf) or math.isinf(xf):
            return None
        return round(xf, n)
    except (TypeError, ValueError):
        return None


# ═════════════════════════════════════════════════════════════════════════════
# DATA LAYER
# ═════════════════════════════════════════════════════════════════════════════

class _DataLayer:
    """Loads & caches everything the risk/financial engines need (read-only)."""

    def __init__(self, signal_date, capital_map, regime, monitor, strat_out,
                 as_of_positions=False):
        """
        signal_date  : date — the as-of date (T).
        capital_map  : {'mrpt': cap, 'mtfs': cap, 'total': T} regime-allocated capital.
        regime       : full regime dict (in-memory, authoritative weights + indicators).
        monitor      : position_monitor dict {'mrpt':[...], 'mtfs':[...]} (real-price MTM).
        strat_out    : {'mrpt': mrpt_out, 'mtfs': mtfs_out} _run_single outputs.
        as_of_positions : if True, source positions from the inventory_history
                          snapshot on/before signal_date (for historical backfill)
                          instead of the current inventory_*.json. Default False
                          keeps the production (DailySignal hook) behaviour unchanged.
        """
        self.signal_date = signal_date if isinstance(signal_date, date) else \
            pd.Timestamp(signal_date).date()
        self.capital_map = capital_map or {}
        self.regime = regime or {}
        self.monitor = monitor or {'mrpt': [], 'mtfs': []}
        self.strat_out = strat_out or {}
        self.as_of_positions = bool(as_of_positions)

        self._prices = None          # MultiIndex DataFrame (Adj Close + Volume)
        self._adj_close = None       # DataFrame columns=tickers
        self._volume = None
        self._returns = None
        self._perf = None            # strategy_performance list
        self._master = None          # master_portfolio_performance list

        self.positions = self._collect_positions()
        # Collect tickers from current + historical positions (for price loading)
        all_tickers = {leg['ticker'] for p in self.positions for leg in p['legs']}
        # Also collect from T-1D/T-1W/T-1M inventory snapshots so historical
        # balance sheets can mark positions to market
        for lookback_days in (2, 7, 30):
            hist_date = (pd.Timestamp(self.signal_date) - pd.Timedelta(days=lookback_days)).date()
            try:
                hist_pos, _ = self.positions_on(hist_date)
                for p in hist_pos:
                    for leg in p['legs']:
                        all_tickers.add(leg['ticker'])
            except Exception:
                pass
        self.tickers = sorted(all_tickers)

    # ── positions (current final inventory) ─────────────────────────────────
    def _collect_positions(self):
        """Read final inventory for MRPT+MTFS → list of position dicts (active only).
        In as-of mode (historical backfill) positions come from the inventory_history
        snapshot ≤ signal_date instead of the current inventory."""
        if self.as_of_positions:
            pos, used = self.positions_on(self.signal_date)
            log.info(f"[RISK] as-of positions @ {self.signal_date}: "
                     f"{len(pos)} positions (snapshots {used})")
            return pos
        import DailySignal as DS
        from SelectPairs import guess_sector
        positions = []
        for strategy in ('mrpt', 'mtfs'):
            try:
                inv = DS.load_inventory(strategy)
            except Exception as e:
                log.warning(f"[RISK] load_inventory({strategy}) failed: {e}")
                continue
            for pair_key, p in inv.get('pairs', {}).items():
                if not isinstance(p, dict) or not p.get('direction'):
                    continue
                if '/' not in pair_key:
                    continue
                s1, s2 = pair_key.split('/', 1)
                s1_sh = p.get('s1_shares', 0) or 0
                s2_sh = p.get('s2_shares', 0) or 0
                positions.append({
                    'strategy': strategy,
                    'pair': pair_key,
                    's1': s1, 's2': s2,
                    'direction': p.get('direction'),
                    'days_held': p.get('days_held', 0),
                    'param_set': p.get('param_set', ''),
                    'open_date': p.get('open_date', ''),
                    'open_s1_price': p.get('open_s1_price'),
                    'open_s2_price': p.get('open_s2_price'),
                    'peak_unrealized_pnl': p.get('peak_unrealized_pnl'),
                    'legs': [
                        {'ticker': s1, 'shares': s1_sh, 'open_price': p.get('open_s1_price'),
                         'sector': guess_sector(s1)},
                        {'ticker': s2, 'shares': s2_sh, 'open_price': p.get('open_s2_price'),
                         'sector': guess_sector(s2)},
                    ],
                })
        return positions

    # ── prices ───────────────────────────────────────────────────────────────
    def load_prices(self):
        """Batch-load Adj Close + Volume for all held tickers + SPY (read-only)."""
        if self._prices is not None:
            return
        from PriceDataStore import PriceDataStore
        symbols = sorted(set(self.tickers) | {'SPY'})
        end = self.signal_date + timedelta(days=1)
        start = self.signal_date - timedelta(days=PRICE_LOOKBACK_CAL_DAYS)
        store = PriceDataStore(base_dir=BASE_DIR,
                               polygon_api_key=os.environ.get('POLYGON_API_KEY', ''))
        try:
            df = store.load(symbols, start.strftime('%Y-%m-%d'), end.strftime('%Y-%m-%d'))
        except Exception as e:
            log.warning(f"[RISK] PriceDataStore.load failed: {e}")
            df = pd.DataFrame()
        self._prices = df
        if df.empty:
            self._adj_close = pd.DataFrame()
            self._volume = pd.DataFrame()
            self._returns = pd.DataFrame()
            return
        # Extract Adj Close (fallback Close) and Volume
        def _field(field):
            cols = {}
            for sym in symbols:
                if (field, sym) in df.columns:
                    cols[sym] = df[(field, sym)]
                elif ('Close', sym) in df.columns and field == 'Adj Close':
                    cols[sym] = df[('Close', sym)]
            return pd.DataFrame(cols) if cols else pd.DataFrame()
        self._adj_close = _field('Adj Close')
        self._volume = _field('Volume')
        self._returns = self._adj_close.pct_change(fill_method=None).dropna(how='all')

    def current_price(self, ticker):
        self.load_prices()
        if self._adj_close is None or ticker not in self._adj_close.columns:
            return None
        s = self._adj_close[ticker].dropna()
        return float(s.iloc[-1]) if len(s) else None

    def price_on(self, ticker, as_of):
        """Adj Close on/before as_of date."""
        self.load_prices()
        if self._adj_close is None or ticker not in self._adj_close.columns:
            return None
        s = self._adj_close[ticker].dropna()
        s = s[s.index <= pd.Timestamp(as_of)]
        return float(s.iloc[-1]) if len(s) else None

    def returns(self, ticker, lookback=None):
        self.load_prices()
        if self._returns is None or ticker not in self._returns.columns:
            return pd.Series(dtype=float)
        s = self._returns[ticker].dropna()
        return s.tail(lookback) if lookback else s

    def adv(self, ticker, window=ADV_WINDOW):
        self.load_prices()
        if self._volume is None or ticker not in self._volume.columns:
            return None
        s = self._volume[ticker].dropna()
        return float(s.tail(window).mean()) if len(s) else None

    # ── strategy_performance (authoritative NAV / PnL series) ────────────────
    def perf(self):
        if self._perf is None:
            try:
                with open(PERF_JSON) as f:
                    rows = json.load(f)
            except Exception as e:
                log.warning(f"[RISK] strategy_performance load failed: {e}")
                rows = []
            self._perf = self._clip_to_asof(rows)
        return self._perf

    def _clip_to_asof(self, rows):
        """Point-in-time guard: in as-of mode drop any equity rows dated AFTER
        signal_date so no future information leaks into the historical report."""
        if not self.as_of_positions or not rows:
            return rows
        cutoff = pd.Timestamp(self.signal_date)
        out = [r for r in rows if r.get('date') and pd.Timestamp(r['date']) <= cutoff]
        log.info(f"[RISK] as-of clip: {len(out)}/{len(rows)} equity rows ≤ {self.signal_date}")
        return out

    def master(self):
        if self._master is None:
            try:
                with open(MASTER_JSON) as f:
                    rows = json.load(f)
            except Exception as e:
                log.warning(f"[RISK] master_performance load failed: {e}")
                rows = []
            self._master = self._clip_to_asof(rows)
        return self._master

    def nav(self, strategy, as_of=None):
        """Authoritative NAV from strategy_performance. strategy in mrpt/mtfs/combined.
        In as-of (backfill) mode an unspecified as_of defaults to signal_date so the
        NAV is the historical value, not the latest."""
        rows = self.perf()
        if not rows:
            return None
        key = f'{strategy}_equity'
        if as_of is None and self.as_of_positions:
            as_of = self.signal_date
        if as_of is None:
            for r in reversed(rows):
                if key in r:
                    return float(r[key])
            return None
        target = pd.Timestamp(as_of)
        best = None
        for r in rows:
            if pd.Timestamp(r['date']) <= target and key in r:
                best = float(r[key])
        return best

    def nav_series(self, strategy):
        """date-indexed NAV series for a strategy (combined/mrpt/mtfs)."""
        rows = self.perf()
        key = f'{strategy}_equity'
        idx, vals = [], []
        for r in rows:
            if key in r:
                idx.append(pd.Timestamp(r['date']))
                vals.append(float(r[key]))
        return pd.Series(vals, index=pd.DatetimeIndex(idx), name=strategy)

    # ── inventory history (for historical balance sheets) ────────────────────
    def positions_on(self, as_of):
        """Reconstruct active positions from inventory_history snapshot ≤ as_of.
        Returns (positions_list, actual_asof_date_str) per strategy merged."""
        from SelectPairs import guess_sector
        target = pd.Timestamp(as_of)
        out = []
        used_dates = {}
        for strategy in ('mrpt', 'mtfs'):
            files = sorted(glob.glob(os.path.join(
                INV_HISTORY_DIR, f'inventory_{strategy}_*.json')))
            # find snapshot with as_of <= target, latest such
            best_path, best_asof = None, None
            for fp in files:
                try:
                    with open(fp) as f:
                        inv = json.load(f)
                except Exception:
                    continue
                a = inv.get('as_of')
                if not a:
                    continue
                if pd.Timestamp(a) <= target:
                    if best_asof is None or pd.Timestamp(a) >= pd.Timestamp(best_asof):
                        best_asof, best_path = a, fp
            if not best_path:
                used_dates[strategy] = None
                continue
            used_dates[strategy] = best_asof
            with open(best_path) as f:
                inv = json.load(f)
            for pair_key, p in inv.get('pairs', {}).items():
                if not isinstance(p, dict) or not p.get('direction') or '/' not in pair_key:
                    continue
                s1, s2 = pair_key.split('/', 1)
                # Corporate actions：快照后执行的拆股 → 调整为当前价格口径
                # （价格源全历史回溯调整；快照文件本身不改，读取时换算）
                try:
                    from CorporateActions import adjust_position_view
                    p = adjust_position_view(p, s1, s2, best_asof)
                except Exception:
                    pass  # 模块/缓存不可用时按原值（与历史行为一致）
                out.append({
                    'strategy': strategy, 'pair': pair_key, 's1': s1, 's2': s2,
                    'direction': p.get('direction'),
                    'days_held': p.get('days_held', 0),
                    'param_set': p.get('param_set', ''),
                    'open_date': p.get('open_date', ''),
                    'peak_unrealized_pnl': p.get('peak_unrealized_pnl'),
                    'legs': [
                        {'ticker': s1, 'shares': p.get('s1_shares', 0) or 0,
                         'open_price': p.get('open_s1_price'), 'sector': guess_sector(s1)},
                        {'ticker': s2, 'shares': p.get('s2_shares', 0) or 0,
                         'open_price': p.get('open_s2_price'), 'sector': guess_sector(s2)},
                    ],
                })
        return out, used_dates


# ═════════════════════════════════════════════════════════════════════════════
# FINANCIAL STATEMENTS  (prime-broker market-neutral model; NAV authoritative)
# ═════════════════════════════════════════════════════════════════════════════

class FinancialStatements:
    """Balance sheet (multi-period), income, capital, cash-flow statements."""

    def __init__(self, data: _DataLayer):
        self.d = data

    # ── exposure for a position set, marked at a given date ──────────────────
    def _mtm_exposure(self, positions, as_of=None, strategy=None):
        """Return dict: long, short, gross, net (sum of signed leg market values)."""
        L = S = 0.0
        for p in positions:
            if strategy and p['strategy'] != strategy:
                continue
            for leg in p['legs']:
                px = (self.d.current_price(leg['ticker']) if as_of is None
                      else self.d.price_on(leg['ticker'], as_of))
                if px is None:
                    continue
                v = leg['shares'] * px
                if v >= 0:
                    L += v
                else:
                    S += -v
        return {'long': L, 'short': S, 'gross': L + S, 'net': L - S}

    def _balance_sheet_one(self, positions, nav, as_of=None, strategy=None):
        """Dual-view balance sheet: BS Method (总账) + Long/Short Book (绩效归因).

        ═══ View 1: Balance Sheet Method (资产负债表法 — 总账) ═══

        Assets:
          Free Cash       = residual after all positions funded
          Restricted Cash  = short sale proceeds held as collateral
          Long Securities  = long positions at market value
        Liabilities:
          Short Securities = obligation to return borrowed shares (at market value)
          Margin Loan      = borrowing when longs exceed equity + short proceeds
          Accrued Costs    = financing interest + borrow fees (estimated)
        NAV = Total Assets - Total Liabilities

        Cash logic (no margin case, typical for pairs book):
          Short sale → restricted cash ↑, short liability ↑
          Buy long   → free cash ↓, long MV ↑
          Restricted Cash = Short MV (proceeds held as collateral)
          Free Cash = NAV - Long MV + Short MV - Margin Loan
                    = NAV - Net Exposure - Margin Loan
          Total Cash = Free Cash + Restricted Cash = NAV + Short MV - Long MV + Margin Loan

        ═══ View 2: Long/Short Book (双账户法 — 绩效归因) ═══

        Long Book:
          Allocated Capital = NAV × (Long MV / Gross MV)   [pro-rata allocation]
          Long Cash = Allocated Capital - Long MV + Long Margin
          Long MV = long positions at market
          Long Margin = max(0, Long MV - Allocated Capital)
          Long Book NAV = Long Cash + Long MV - Long Margin = Allocated Capital
          Long Financing = Long Margin × annual_rate × days / 365
          Long Book P&L = Long MV change (unrealized) + realized - financing

        Short Book:
          Allocated Capital = NAV × (Short MV / Gross MV)  [pro-rata allocation]
          Short Collateral = Allocated Capital
          Short Proceeds = Short MV (restricted, held as collateral)
          Short MV = short positions at market (liability)
          Short Book NAV = Short Collateral + Short Proceeds - Short MV = Allocated Capital
          Short Borrow Fee = 0.0 (placeholder, not tracked)
          Short Rebate = 0.0 (placeholder)
          Short Book P&L = Short MV decrease (unrealized) + realized - borrow fee + rebate

        Total NAV = Long Book NAV + Short Book NAV + Unallocated Cash
        (must equal BS Method NAV)

        Parameters: ANNUAL_FINANCING_RATE = 0.05, BORROW_FEE_RATE = 0.0 (placeholder)
        """
        ANNUAL_FINANCING_RATE = 0.05
        BORROW_FEE_RATE = 0.0   # placeholder — interface preserved for future

        exp = self._mtm_exposure(positions, as_of, strategy)
        L, S = exp['long'], exp['short']
        E = nav if nav is not None else 0.0
        G = L + S   # gross exposure

        SHORT_COLLATERAL_RATIO = 1.02    # PB requires 102% collateral on shorts
        PORTFOLIO_MARGIN_PCT = 0.20      # Portfolio Margin: ~20% of gross

        # ── View 1: Balance Sheet Method ──────────────────────────────────
        # Professional long-short fund: short proceeds are RESTRICTED by PB.
        # Longs funded by: equity + margin loan (NOT short proceeds).
        restricted_cash = S * SHORT_COLLATERAL_RATIO  # 102% of short MV held by PB
        short_collateral_due = S * (SHORT_COLLATERAL_RATIO - 1.0)  # excess 2% over short proceeds
        # Equity funds the longs first, then the 2% excess short collateral; any
        # remaining shortfall is borrowed.  The margin loan therefore splits into a
        # long-financing part and a 2%-short-collateral part, and free cash is the
        # leftover equity FLOORED AT 0 (when levered the 2% is financed, not negative).
        long_margin  = max(0.0, L - E)                        # PB lends when longs > equity
        equity_left  = max(0.0, E - L)                        # equity remaining after longs
        short_equity = min(short_collateral_due, equity_left) # 2% funded by equity (if any left)
        short_margin = short_collateral_due - short_equity    # 2% shortfall → financed
        M = long_margin + short_margin                        # total margin loan = max(0, L+0.02S-E)
        free_cash = max(0.0, E - L - short_collateral_due)    # leftover equity, floored at 0
        # Assets = Free Cash + Restricted Cash + Long Securities
        # Liabilities = Short Obligation + Margin Loan (long financing + financed 2% collateral)
        # Identity: free_cash + 1.02S + L - S - M = E  (the equity-posted 2% lives in restricted_cash)
        total_assets = free_cash + restricted_cash + L
        total_liabilities = S + M
        # Accrued costs (daily estimates)
        accrued_financing = M * ANNUAL_FINANCING_RATE / 365
        accrued_borrow_fee = S * BORROW_FEE_RATE / 365
        accrued_costs = accrued_financing + accrued_borrow_fee
        balance_check = total_assets - total_liabilities - E
        # Margin requirements (Portfolio Margin estimate)
        pm_requirement = G * PORTFOLIO_MARGIN_PCT
        excess_liquidity = E - pm_requirement

        # ── View 2: Long/Short Book (mirrors the Balance-Sheet financing) ──
        # Reuses long_margin / short_equity / short_margin from View 1, so the total
        # margin loan M splits cleanly into Long-Book financing + Short-Book (2%
        # collateral) financing, and the two views reconcile exactly (dual_nav == E).
        long_alloc = L - long_margin                 # equity funding the longs (= L when unlevered)
        long_cash = 0.0                              # fully invested
        long_financing = long_margin * ANNUAL_FINANCING_RATE / 365
        long_book_nav = L - long_margin              # = long_alloc

        # Short Book — equity posted = equity-funded part of the 2% collateral;
        # any financed part is short_margin (rolled into the total margin loan M).
        short_alloc = short_equity
        short_collateral = short_alloc
        short_proceeds = S                # restricted cash from short sales
        short_borrow_fee = S * BORROW_FEE_RATE / 365
        short_rebate = 0.0                # placeholder
        short_financing = short_margin * ANNUAL_FINANCING_RATE / 365
        short_book_nav = short_alloc      # = 1.02S - S - short_margin = short_equity

        # Unallocated (genuinely free) equity = View-1 free_cash
        unalloc_cash = E - long_alloc - short_alloc

        dual_nav = long_book_nav + short_book_nav + unalloc_cash
        nav_alignment_check = abs(dual_nav - E)  # should be < 0.01

        return {
            'as_of': str(as_of) if as_of else str(self.d.signal_date),

            # ── View 1: Balance Sheet (backward-compatible keys) ──
            'free_cash': _round(free_cash),
            'restricted_cash': _round(restricted_cash),
            'cash': _round(free_cash),              # backward compat: "cash" = free cash
            'long_securities': _round(L),
            'total_assets': _round(total_assets),
            'short_securities': _round(S),
            'short_collateral_due': _round(short_collateral_due),
            'margin_loan': _round(M),
            'accrued_financing': _round(accrued_financing),
            'accrued_borrow_fee': _round(accrued_borrow_fee),
            'accrued_costs': _round(accrued_costs),
            'total_liabilities': _round(total_liabilities),
            'nav': _round(E),
            'balance_check': _round(balance_check, 2),
            'pm_requirement': _round(pm_requirement),
            'excess_liquidity': _round(excess_liquidity),

            # ── View 2: Long/Short Book ──
            'long_book': {
                'allocated_capital': _round(long_alloc),
                'cash': _round(long_cash),
                'market_value': _round(L),
                'margin_loan': _round(long_margin),
                'financing_cost_daily': _round(long_financing),
                'book_nav': _round(long_book_nav),
            },
            'short_book': {
                'allocated_capital': _round(short_alloc),
                'collateral_cash': _round(short_collateral),
                'short_proceeds': _round(short_proceeds),
                'market_value': _round(S),
                'margin_loan': _round(short_margin),
                'financing_cost_daily': _round(short_financing),
                'borrow_fee_daily': _round(short_borrow_fee),
                'rebate_daily': _round(short_rebate),
                'book_nav': _round(short_book_nav),
            },
            'unallocated_cash': _round(unalloc_cash),
            'dual_nav': _round(dual_nav),
            'nav_alignment_check': _round(nav_alignment_check, 4),

            # ── Exposure metrics (backward compat) ──
            'gross': _round(exp['gross']), 'net': _round(exp['net']),
            'gross_leverage': _round(exp['gross'] / E if E else None, 3),
            'net_leverage': _round(exp['net'] / E if E else None, 3),
        }

    def balance_sheet_multi(self):
        """T / T-1D / T-1W / T-1M balance sheets for combined + per-strategy."""
        # resolve as-of dates from NAV trading-day index
        combined_idx = self.d.nav_series('combined').index
        if len(combined_idx) == 0:
            t = pd.Timestamp(self.d.signal_date)
            periods = {'T': t}
        else:
            t = combined_idx[-1]
            def back(n):
                return combined_idx[-1 - n] if len(combined_idx) > n else combined_idx[0]
            periods = {'T': t, 'T_1D': back(1), 'T_1W': back(5), 'T_1M': back(21)}

        result = {'combined': {}, 'mrpt': {}, 'mtfs': {}, 'as_of_snapshots': {}}
        for label, dt in periods.items():
            if label == 'T':
                pos = self.d.positions
                result['as_of_snapshots'][label] = {'mrpt': str(self.d.signal_date),
                                                     'mtfs': str(self.d.signal_date)}
            else:
                pos, used = self.d.positions_on(dt.date())
                result['as_of_snapshots'][label] = used
            for scope in ('combined', 'mrpt', 'mtfs'):
                strat = None if scope == 'combined' else scope
                nav = self.d.nav(scope, None if label == 'T' else dt.date())
                result[scope][label] = self._balance_sheet_one(
                    pos, nav, None if label == 'T' else dt.date(), strat)
        return result

    # ── current unrealized PnL (exact, from inventory open prices) ───────────
    def current_unrealized(self, strategy=None):
        upnl = 0.0
        for p in self.d.positions:
            if strategy and p['strategy'] != strategy:
                continue
            for leg in p['legs']:
                px = self.d.current_price(leg['ticker'])
                op = leg['open_price']
                if px is None or op is None:
                    continue
                upnl += (px - op) * leg['shares']
        return upnl

    def income_statement(self):
        """Period income (1D/1W/1M/ITD). Net income authoritative from perf pnl;
        realized = net - Δunrealized - interest (residual, labeled estimated)."""
        combined_idx = self.d.nav_series('combined').index
        rows = self.d.perf()
        if not rows:
            return {}

        def period_pnl(scope, n_back):
            key = f'{scope}_pnl'
            if n_back is None:  # ITD
                return sum(float(r.get(key, 0) or 0) for r in rows)
            # last n_back trading days
            return sum(float(r.get(key, 0) or 0) for r in rows[-n_back:])

        # current margin loan per scope for the interest estimate (its own balance sheet)
        margin_by_scope = {
            'combined': self._balance_sheet_one(self.d.positions, self.d.nav('combined'))['margin_loan'] or 0.0,
            'mrpt': self._balance_sheet_one(self.d.positions, self.d.nav('mrpt'), None, 'mrpt')['margin_loan'] or 0.0,
            'mtfs': self._balance_sheet_one(self.d.positions, self.d.nav('mtfs'), None, 'mtfs')['margin_loan'] or 0.0,
        }

        out = {}
        spans = {'1D': 1, '1W': 5, '1M': 21, 'ITD': None}
        for scope in ('combined', 'mrpt', 'mtfs'):
            out[scope] = {}
            margin_now = margin_by_scope[scope]
            for label, n in spans.items():
                net = period_pnl(scope, n)
                days = (n if n else len(rows))
                interest = margin_now * MARGIN_RATE * days / 365.0
                out[scope][label] = {
                    'net_income': _round(net),
                    'interest_expense_est': _round(interest),
                    'note': 'net_income authoritative (strategy_performance); '
                            'interest estimated on each scope\'s own margin',
                }
        out['_current_unrealized'] = {
            'combined': _round(self.current_unrealized()),
            'mrpt': _round(self.current_unrealized('mrpt')),
            'mtfs': _round(self.current_unrealized('mtfs')),
        }
        return out

    def capital_statement(self):
        """NAV roll-forward over 1D/1W/1M/ITD."""
        out = {}
        spans = {'1D': 1, '1W': 5, '1M': 21, 'ITD': None}
        for scope in ('combined', 'mrpt', 'mtfs'):
            s = self.d.nav_series(scope)
            if len(s) == 0:
                continue
            out[scope] = {}
            for label, n in spans.items():
                if n is None or len(s) <= n:
                    beg = float(s.iloc[0])
                else:
                    beg = float(s.iloc[-1 - n])
                end = float(s.iloc[-1])
                # period max drawdown
                seg = s if (n is None or len(s) <= n) else s.iloc[-1 - n:]
                peak = seg.cummax()
                dd = ((seg - peak) / peak).min() * 100 if len(seg) else 0.0
                out[scope][label] = {
                    'beginning_nav': _round(beg),
                    'net_income': _round(end - beg),
                    'contributions': 0.0, 'withdrawals': 0.0,
                    'ending_nav': _round(end),
                    'return_pct': _round((end / beg - 1) * 100 if beg else None),
                    'max_drawdown_pct': _round(dd),
                }
        return out

    def cash_flow_statement(self):
        """Indirect-method approximation over 1W (and ITD). Labeled approximate."""
        # Operating = realized (net - Δunrealized) - interest; Investing = -Δ net securities;
        # Financing = Δ margin loan. Use T vs T-1W positions for deltas.
        combined_idx = self.d.nav_series('combined').index
        out = {}
        if len(combined_idx) <= 5:
            return out
        t_1w = combined_idx[-6].date()
        bs_now = self._balance_sheet_one(self.d.positions, self.d.nav('combined'))
        pos_1w, _ = self.d.positions_on(t_1w)
        # mark the 1-week-ago positions at 1-week-ago prices (as_of=t_1w), so this
        # ties to the Balance Sheet T-1W column and the statement foots.
        bs_1w = self._balance_sheet_one(pos_1w, self.d.nav('combined', t_1w), t_1w)

        rows = self.d.perf()
        net_1w = sum(float(r.get('combined_pnl', 0) or 0) for r in rows[-5:])
        margin_now = bs_now['margin_loan'] or 0.0
        interest_memo = margin_now * MARGIN_RATE * 5 / 365.0
        # ΔNAV over the week IS net income (no external flows) and already nets the
        # financing cost, so operating = net_1w. Subtracting interest again would
        # double-count it (that was the prior non-footing residual). Interest is a memo.
        operating = net_1w
        d_net_sec = (bs_now['net'] or 0) - (bs_1w['net'] or 0)
        d_margin = (bs_now['margin_loan'] or 0) - (bs_1w['margin_loan'] or 0)
        net_change = operating - d_net_sec + d_margin
        out['1W'] = {
            'operating_cash_flow': _round(operating),
            'investing_cash_flow': _round(-d_net_sec),
            'financing_cash_flow': _round(d_margin),
            'net_change_in_cash': _round(net_change),
            'beginning_cash': _round(bs_1w['cash']),
            'ending_cash': _round(bs_now['cash']),
            'interest_expense_memo': _round(interest_memo),
            'reconciliation_check': _round(bs_1w['cash'] + net_change - bs_now['cash'], 2),
            'note': 'indirect-method; operating=ΔNAV (interest already in net income, shown as memo)',
        }
        return out

    def nav_trend(self):
        """Date × {nav, gross, net, gross_leverage} series (combined)."""
        s = self.d.nav_series('combined')
        out = []
        # gross/net per date requires historical positions+prices → expensive;
        # report NAV daily + gross/net only for the periods we have snapshots.
        for dt, nav in s.items():
            out.append({'date': str(dt.date()), 'nav': _round(nav)})
        return out


# ═════════════════════════════════════════════════════════════════════════════
# FACTOR DATA LOADER  (Ken French FF5 + Momentum, ETF-proxy fallback)
# ═════════════════════════════════════════════════════════════════════════════

class FactorDataLoader:
    """Daily FF5 + Momentum factor returns. Cache to price_data/factors/, incremental;
    fallback to ETF proxies (SPY/IWM/IVE/IVW/MTUM) if the French library is unreachable."""

    FF5_URL = ('https://mba.tuck.dartmouth.edu/pages/faculty/ken.french/ftp/'
               'F-F_Research_Data_5_Factors_2x3_daily_CSV.zip')
    MOM_URL = ('https://mba.tuck.dartmouth.edu/pages/faculty/ken.french/ftp/'
               'F-F_Momentum_Factor_daily_CSV.zip')
    CACHE = os.path.join(FACTOR_DIR, 'ff_factors.parquet')

    def __init__(self, data: _DataLayer):
        self.d = data

    def load(self):
        """Return (df, source): df indexed by date, columns MKT,SMB,HML,RMW,CMA,UMD,RF
        in decimal daily returns. source = 'french' | 'proxy' | 'none'."""
        df = self._load_french()
        if df is not None and len(df) > 30:
            return df, 'french'
        proxy = self._load_proxy()
        if proxy is not None and len(proxy) > 30:
            return proxy, 'proxy'
        return None, 'none'

    def _load_french(self):
        # cache first
        try:
            if os.path.exists(self.CACHE):
                cached = pd.read_parquet(self.CACHE)
                last = cached.index.max()
                if (pd.Timestamp(self.d.signal_date) - last).days <= 7:
                    return cached
        except Exception:
            pass
        try:
            import io, zipfile, urllib.request
            def _fetch(url, valcols):
                req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
                raw = urllib.request.urlopen(req, timeout=30).read()
                zf = zipfile.ZipFile(io.BytesIO(raw))
                name = zf.namelist()[0]
                text = zf.read(name).decode('latin-1')
                # daily block: lines starting with 8-digit date
                recs = []
                for line in text.splitlines():
                    parts = [p.strip() for p in line.split(',')]
                    if len(parts) >= len(valcols) + 1 and parts[0].isdigit() and len(parts[0]) == 8:
                        try:
                            d = pd.Timestamp(parts[0])
                            vals = [float(parts[i + 1]) / 100.0 for i in range(len(valcols))]
                            recs.append((d, *vals))
                        except (ValueError, IndexError):
                            continue
                if not recs:
                    return None
                cols = ['date'] + valcols
                return pd.DataFrame(recs, columns=cols).set_index('date')
            ff5 = _fetch(self.FF5_URL, ['MKT', 'SMB', 'HML', 'RMW', 'CMA', 'RF'])
            mom = _fetch(self.MOM_URL, ['UMD'])
            if ff5 is None or mom is None:
                return None
            df = ff5.join(mom, how='inner')
            os.makedirs(FACTOR_DIR, exist_ok=True)
            try:
                df.to_parquet(self.CACHE)
            except Exception:
                pass
            return df
        except Exception as e:
            log.warning(f"[RISK] French factor fetch failed → proxy: {e}")
            return None

    def _load_proxy(self):
        """ETF-proxy Carhart factors (MKT,SMB,HML,UMD; RMW/CMA=0). Uses PriceDataStore."""
        try:
            from PriceDataStore import PriceDataStore
            tickers = ['SPY', 'IWM', 'IVE', 'IVW', 'MTUM']
            end = self.d.signal_date + timedelta(days=1)
            start = self.d.signal_date - timedelta(days=PRICE_LOOKBACK_CAL_DAYS)
            store = PriceDataStore(base_dir=BASE_DIR,
                                   polygon_api_key=os.environ.get('POLYGON_API_KEY', ''))
            df = store.load(tickers, start.strftime('%Y-%m-%d'), end.strftime('%Y-%m-%d'))
            if df.empty:
                return None
            ac = {}
            for t in tickers:
                if ('Adj Close', t) in df.columns:
                    ac[t] = df[('Adj Close', t)]
                elif ('Close', t) in df.columns:
                    ac[t] = df[('Close', t)]
            ac = pd.DataFrame(ac).dropna()
            r = ac.pct_change(fill_method=None).dropna()
            rf = RISK_FREE_ANNUAL / ANNUALIZATION
            out = pd.DataFrame({
                'MKT': r['SPY'] - rf,
                'SMB': r['IWM'] - r['SPY'],
                'HML': r['IVE'] - r['IVW'],
                'RMW': 0.0,
                'CMA': 0.0,
                'UMD': r['MTUM'] - r['SPY'],
                'RF': rf,
            }, index=r.index)
            return out
        except Exception as e:
            log.warning(f"[RISK] proxy factor build failed: {e}")
            return None


# ═════════════════════════════════════════════════════════════════════════════
# RISK ENGINE  (Tier 1-7) + DIAGNOSTICS (D1-D6)
# ═════════════════════════════════════════════════════════════════════════════

class RiskManager:
    def __init__(self, data: _DataLayer, fin: FinancialStatements):
        self.d = data
        self.fin = fin
        self.C = data.nav('combined') or data.capital_map.get('total') or 1.0
        self.C_mrpt = data.nav('mrpt') or data.capital_map.get('mrpt') or 1.0
        self.C_mtfs = data.nav('mtfs') or data.capital_map.get('mtfs') or 1.0

    # ── leg-level marked values ──────────────────────────────────────────────
    def _leg_values(self, strategy=None):
        """List of (pair, strategy, ticker, shares, price, value, sector)."""
        rows = []
        for p in self.d.positions:
            if strategy and p['strategy'] != strategy:
                continue
            for leg in p['legs']:
                px = self.d.current_price(leg['ticker'])
                if px is None:
                    continue
                rows.append({'pair': p['pair'], 'strategy': p['strategy'],
                             'ticker': leg['ticker'], 'shares': leg['shares'],
                             'price': px, 'value': leg['shares'] * px,
                             'sector': leg['sector']})
        return rows

    # ── Tier 1: exposure / leverage ──────────────────────────────────────────
    def exposure(self):
        out = {}
        for scope, C in (('combined', self.C), ('mrpt', self.C_mrpt), ('mtfs', self.C_mtfs)):
            strat = None if scope == 'combined' else scope
            legs = self._leg_values(strat)
            L = sum(v['value'] for v in legs if v['value'] > 0)
            S = sum(-v['value'] for v in legs if v['value'] < 0)
            G, N = L + S, L - S
            n_pairs = len({v['pair'] for v in legs})
            out[scope] = {
                'capital': _round(C), 'long': _round(L), 'short': _round(S),
                'gross': _round(G), 'net': _round(N),
                'gross_leverage': _round(G / C if C else None, 3),
                'net_leverage': _round(N / C if C else None, 3),
                'deployed_pct': _round(G / C * 100 if C else None),
                'margin_utilization_pct': _round(G / (2 * C) * 100 if C else None),
                'n_open_pairs': n_pairs,
            }
        return out

    # ── Tier 2: concentration ──────────────────────────────────────────────
    def concentration(self):
        legs = self._leg_values()
        G = sum(abs(v['value']) for v in legs) or 1.0
        # single-name aggregate across pairs
        by_name = {}
        for v in legs:
            d = by_name.setdefault(v['ticker'], {'gross': 0.0, 'net': 0.0, 'pairs': set()})
            d['gross'] += abs(v['value'])
            d['net'] += v['value']
            d['pairs'].add(v['pair'])
        single = sorted(
            [{'ticker': k, 'gross': _round(x['gross']), 'net': _round(x['net']),
              'gross_pct_capital': _round(x['gross'] / self.C * 100 if self.C else None),
              'pct_of_gross': _round(x['gross'] / G * 100), 'n_pairs': len(x['pairs'])}
             for k, x in by_name.items()],
            key=lambda r: -(r['gross'] or 0))
        # pair-level HHI + max pair
        by_pair = {}
        for v in legs:
            by_pair[v['pair']] = by_pair.get(v['pair'], 0) + abs(v['value'])
        hhi = sum((g / G) ** 2 for g in by_pair.values()) if by_pair else None
        max_pair_pct = max((g / G * 100 for g in by_pair.values()), default=None)
        # sector
        by_sec = {}
        for v in legs:
            s = by_sec.setdefault(v['sector'], {'long': 0.0, 'short': 0.0})
            if v['value'] > 0:
                s['long'] += v['value']
            else:
                s['short'] += -v['value']
        sector = {k: {'long': _round(x['long']), 'short': _round(x['short']),
                      'net': _round(x['long'] - x['short']),
                      'net_pct_of_gross': _round((x['long'] - x['short']) / G * 100)}
                  for k, x in by_sec.items()}
        return {'single_name_top': single[:10], 'hhi': _round(hhi, 4),
                'effective_n_pairs': _round(1 / hhi, 1) if hhi else None,
                'max_pair_pct': _round(max_pair_pct), 'sector': sector}

    # ── Tier 3: factor / beta ──────────────────────────────────────────────
    def _ticker_beta(self, ticker):
        rt = self.d.returns(ticker, BETA_LOOKBACK)
        rs = self.d.returns('SPY', BETA_LOOKBACK)
        j = pd.concat([rt, rs], axis=1, join='inner').dropna()
        if len(j) < 20 or j.iloc[:, 1].var() == 0:
            return None
        cov = np.cov(j.iloc[:, 0], j.iloc[:, 1])[0, 1]
        return cov / j.iloc[:, 1].var()

    def factor(self):
        legs = self._leg_values()
        net_beta_dollar = 0.0
        gross_beta_dollar = 0.0
        insufficient = []
        for v in legs:
            b = self._ticker_beta(v['ticker'])
            if b is None:
                insufficient.append(v['ticker'])
                continue
            net_beta_dollar += v['value'] * b
            gross_beta_dollar += abs(v['value'] * b)
        # momentum factor: net dollar in MTFS long-winner direction / capital
        mtfs_net = sum(v['value'] for v in legs if v['strategy'] == 'mtfs')
        return {
            'net_market_beta': _round(net_beta_dollar / self.C if self.C else None, 3),
            'gross_beta': _round(gross_beta_dollar / self.C if self.C else None, 3),
            'momentum_factor_net_pct': _round(mtfs_net / self.C * 100 if self.C else None),
            'sector_net_vector': {k: x['net_pct_of_gross']
                                  for k, x in self.concentration()['sector'].items()},
            'beta_data_insufficient': sorted(set(insufficient)),
        }

    # ── pair-level frozen daily $ PnL series (for VaR / component / diagnostics) ──
    def _pair_pnl_series(self):
        series = {}
        for p in self.d.positions:
            pnl = None
            for leg in p['legs']:
                self.d.load_prices()
                if self.d._adj_close is None or leg['ticker'] not in self.d._adj_close.columns:
                    pnl = None
                    break
                px = self.d._adj_close[leg['ticker']].dropna().tail(VAR_LOOKBACK + 1)
                leg_pnl = px.diff().dropna() * leg['shares']
                pnl = leg_pnl if pnl is None else pnl.add(leg_pnl, fill_value=0.0)
            if pnl is not None and len(pnl) >= 20:
                series[p['pair']] = pnl
        return series

    def _portfolio_pnl_series(self):
        ps = self._pair_pnl_series()
        if not ps:
            return pd.Series(dtype=float)
        df = pd.DataFrame(ps).fillna(0.0)
        return df.sum(axis=1)

    # ── Tier 4: VaR / volatility ─────────────────────────────────────────────
    def var(self):
        P = self._portfolio_pnl_series()
        if len(P) < 20:
            return {'note': 'insufficient price history for VaR'}
        mu, sigma = float(P.mean()), float(P.std(ddof=1))
        z95, z99 = 1.645, 2.326
        param = {
            'var_95_1d': _round(z95 * sigma), 'var_99_1d': _round(z99 * sigma),
            'vol_daily': _round(sigma), 'vol_annual': _round(sigma * math.sqrt(ANNUALIZATION)),
            'vol_annual_pct_of_capital': _round(sigma * math.sqrt(ANNUALIZATION) / self.C * 100
                                                if self.C else None),
        }
        var95 = -np.percentile(P, 5)
        tail = P[P <= -var95]
        cvar95 = -tail.mean() if len(tail) else var95
        hist = {'var_95_1d': _round(var95), 'cvar_95_1d': _round(cvar95)}
        # component VaR (per pair marginal contribution)
        ps = self._pair_pnl_series()
        comp = []
        if len(ps) >= 1:
            df = pd.DataFrame(ps).fillna(0.0)
            cov = df.cov()
            w = np.ones(len(df.columns))
            port_var = float(w @ cov.values @ w)
            port_sd = math.sqrt(port_var) if port_var > 0 else 0.0
            if port_sd > 0:
                mctr = (cov.values @ w) / port_sd      # marginal contribution to risk
                for i, pair in enumerate(df.columns):
                    ctr = w[i] * mctr[i]               # component contribution ($ sd)
                    comp.append({'pair': pair, 'component_var_95': _round(z95 * ctr),
                                 'pct_of_total': _round(ctr / port_sd * 100)})
                comp.sort(key=lambda r: -(r['component_var_95'] or 0))
        return {'method_param': param, 'method_hist': hist, 'component_var': comp[:15]}

    # ── Tier 5: liquidity ─────────────────────────────────────────────────────
    def liquidity(self):
        rows = []
        for v in self._leg_values():
            adv = self.d.adv(v['ticker'])
            if not adv or adv <= 0:
                continue
            dtl = abs(v['shares']) / (adv * ADV_PARTICIPATION)
            rows.append({'ticker': v['ticker'], 'pair': v['pair'],
                         'shares': abs(v['shares']), 'adv20': _round(adv, 0),
                         'days_to_liquidate': _round(dtl, 2)})
        rows.sort(key=lambda r: -(r['days_to_liquidate'] or 0))
        return {'worst_legs': rows[:15],
                'n_legs_over_1d': sum(1 for r in rows if (r['days_to_liquidate'] or 0) > 1)}

    # ── Tier 6: stress / scenario ─────────────────────────────────────────────
    def stress(self, factor_block):
        net_beta = (factor_block.get('net_market_beta') or 0.0)
        net_beta_dollar = net_beta * self.C
        conc = self.concentration()
        # top sector by |net|
        top_sec, top_sec_net = None, 0.0
        for k, x in conc['sector'].items():
            if abs(x['net'] or 0) > abs(top_sec_net):
                top_sec, top_sec_net = k, (x['net'] or 0)
        out = []
        for name, kind, shock in STRESS_SCENARIOS:
            if kind == 'spy':
                pnl = net_beta_dollar * shock
            elif kind == 'vix':
                # crude: VIX +10pt ≈ SPY -5% beta-implied (historical rough mapping)
                pnl = net_beta_dollar * (-0.05) * (shock / 10.0)
            elif kind == 'sector':
                pnl = top_sec_net * shock
            else:
                pnl = 0.0
            out.append({'scenario': name, 'est_pnl': _round(pnl),
                        'est_pnl_pct_of_capital': _round(pnl / self.C * 100 if self.C else None)})
        return out

    # ── Tier 7: limits ─────────────────────────────────────────────────────────
    def limits(self, exposure, concentration, factor_block, var_block):
        import DailySignal as DS
        checks = []

        def add(name, scope, value, spec):
            if value is None:
                return
            status = 'green'
            if spec['red'] is not None and abs(value) >= spec['red']:
                status = 'red'
            elif spec['amber'] is not None and abs(value) >= spec['amber']:
                status = 'amber'
            checks.append({'name': name, 'scope': scope, 'value': _round(value),
                           'amber': spec['amber'], 'red': spec['red'],
                           'fmt': spec['fmt'], 'status': status})

        for scope in ('mrpt', 'mtfs'):
            e = exposure[scope]
            add('gross_leverage', scope, e['gross_leverage'], LIMITS_SPEC['gross_leverage'])
            add('net_leverage_abs', scope, abs(e['net_leverage']) if e['net_leverage'] is not None else None,
                LIMITS_SPEC['net_leverage_abs'])
        # single-name (combined)
        for sn in concentration['single_name_top'][:3]:
            add('single_name_gross', sn['ticker'], sn['gross_pct_capital'],
                LIMITS_SPEC['single_name_gross'])
        # sector net (worst)
        worst_sec = max(concentration['sector'].items(),
                        key=lambda kv: abs(kv[1]['net_pct_of_gross'] or 0), default=None)
        if worst_sec:
            add('sector_net', worst_sec[0], abs(worst_sec[1]['net_pct_of_gross'] or 0),
                LIMITS_SPEC['sector_net'])
        add('max_pair', 'combined', concentration['max_pair_pct'], LIMITS_SPEC['max_pair'])
        add('net_market_beta', 'combined', abs(factor_block.get('net_market_beta') or 0),
            LIMITS_SPEC['net_market_beta'])
        var95 = (var_block.get('method_param') or {}).get('var_95_1d')
        if var95 is not None and self.C:
            add('var_95_1d', 'combined', var95 / self.C * 100, LIMITS_SPEC['var_95_1d'])
        # open pairs vs DS capacity (enforcement-shared)
        for scope in ('mrpt', 'mtfs'):
            cap = getattr(DS, f'_MAX_OPEN_PAIRS_{scope.upper()}', 8)
            n = exposure[scope]['n_open_pairs']
            status = 'amber' if n >= cap else 'green'
            checks.append({'name': 'open_pairs', 'scope': scope, 'value': n,
                           'amber': cap, 'red': None, 'fmt': 'n', 'status': status})
        return checks

    # ═════════════════════════════════════════════════════════════════════════
    # THEORY DIAGNOSTICS (D1-D6) — report-only
    # ═════════════════════════════════════════════════════════════════════════

    # D1 — Risk Contribution (Maillard-Roncalli-Teiletche 2010)
    def diag_risk_contribution(self):
        out = {'master': None, 'pairs': None,
               'ref': 'Maillard, Roncalli & Teiletche (2010), J. Portfolio Mgmt'}
        # master level: 4 strategies from master JSON
        rows = self.d.master()
        if rows and len(rows) > BETA_LOOKBACK:
            comps = ['mrpt', 'mtfs', 'sr', 'aiss']        # data keys ({c}_equity in master JSON)
            labels = {'mrpt': 'MRPT', 'mtfs': 'MTFS', 'sr': 'SSRS', 'aiss': 'AISS'}  # display names
            eq = {c: [] for c in comps}
            for r in rows:
                for c in comps:
                    eq[c].append(float(r.get(f'{c}_equity', np.nan)))
            edf = pd.DataFrame(eq).dropna()
            rdf = edf.pct_change(fill_method=None).dropna().tail(BETA_LOOKBACK)
            if len(rdf) >= 30:
                last = edf.iloc[-1]
                w = (last / last.sum()).values
                cov = rdf.cov().values
                port_var = float(w @ cov @ w)
                sd = math.sqrt(port_var) if port_var > 0 else 0.0
                if sd > 0:
                    mctr = (cov @ w) / sd
                    trc = w * mctr
                    prc = trc / sd
                    out['master'] = [
                        {'component': labels.get(c, c.upper()),
                         'capital_weight_pct': _round(w[i] * 100),
                         'risk_contribution_pct': _round(prc[i] * 100),
                         'erc_target_pct': _round(100.0 / len(comps)),
                         'divergence_pct': _round((prc[i] - w[i]) * 100)}
                        for i, c in enumerate(comps)]
        # pairs level within combined book
        ps = self._pair_pnl_series()
        if len(ps) >= 2:
            df = pd.DataFrame(ps).fillna(0.0)
            cov = df.cov().values
            w = np.ones(len(df.columns))
            port_var = float(w @ cov @ w)
            sd = math.sqrt(port_var) if port_var > 0 else 0.0
            if sd > 0:
                mctr = (cov @ w) / sd
                trc = w * mctr
                prc = trc / sd
                pairs = [{'pair': df.columns[i],
                          'risk_contribution_pct': _round(prc[i] * 100)}
                         for i in range(len(df.columns))]
                pairs.sort(key=lambda r: -(r['risk_contribution_pct'] or 0))
                out['pairs'] = pairs[:15]
        return out

    # D2 — Factor Attribution FF5 + UMD (Fama-French 1993/2015; Carhart 1997)
    def diag_factor_attribution(self, factor_df, factor_source):
        out = {'source': factor_source,
               'ref': 'Fama-French (1993,2015); Carhart (1997)', 'models': {}}
        if factor_df is None or len(factor_df) < 30:
            out['note'] = 'factor data unavailable'
            return out
        try:
            import statsmodels.api as sm
        except Exception:
            out['note'] = 'statsmodels unavailable'
            return out
        use_cols = [c for c in ('MKT', 'SMB', 'HML', 'RMW', 'CMA', 'UMD')
                    if c in factor_df.columns and factor_df[c].abs().sum() > 0]
        for scope in ('combined', 'mrpt', 'mtfs'):
            nav = self.d.nav_series(scope)
            if len(nav) < 30:
                continue
            r = nav.pct_change(fill_method=None).dropna()
            j = pd.concat([r.rename('y'), factor_df[use_cols + (['RF'] if 'RF' in factor_df else [])]],
                          axis=1, join='inner').dropna()
            if len(j) < 30:
                continue
            y = j['y'] - (j['RF'] if 'RF' in j else 0.0)
            X = sm.add_constant(j[use_cols])
            try:
                m = sm.OLS(y, X).fit()
            except Exception:
                continue
            loadings = {}
            for c in use_cols:
                loadings[c] = {'beta': _round(m.params.get(c), 3),
                               't_stat': _round(m.tvalues.get(c), 2)}
            out['models'][scope] = {
                'alpha_annual_pct': _round(m.params.get('const', 0) * ANNUALIZATION * 100, 2),
                'alpha_t': _round(m.tvalues.get('const'), 2),
                'loadings': loadings,
                'r_squared': _round(m.rsquared, 3),
                'n_obs': int(m.nobs),
            }
        out['note'] = ('short-sample (live ITD); t-stats indicative'
                       if factor_source != 'none' else '')
        return out

    # D3 — Return distribution & fat tails (Cont 2001; Zangari 1996 Cornish-Fisher)
    def diag_distribution(self):
        out = {'ref': 'Cont (2001); Zangari (1996) Cornish-Fisher VaR', 'scopes': {}}
        from scipy import stats as sstats
        for scope in ('combined', 'mrpt', 'mtfs'):
            nav = self.d.nav_series(scope)
            r = nav.pct_change(fill_method=None).dropna()
            if len(r) < 20:
                continue
            g1 = float(sstats.skew(r))
            g2_excess = float(sstats.kurtosis(r))           # excess (normal=0)
            jb_stat, jb_p = sstats.jarque_bera(r)
            mu, sd = float(r.mean()), float(r.std(ddof=1))
            z = -1.645
            zcf = (z + (z**2 - 1) / 6 * g1 + (z**3 - 3 * z) / 24 * g2_excess
                   - (2 * z**3 - 5 * z) / 36 * g1**2)
            var_param_pct = -(mu + z * sd) * 100
            var_cf_raw_pct = -(mu + zcf * sd) * 100
            # Cornish-Fisher is a polynomial expansion valid only for mild non-normality;
            # outside ~(|skew|≤2, excess-kurt≤7) the zcf mapping is non-monotonic and can
            # UNDERSTATE the tail (esp. with positive skew, as MTFS: zcf→-0.36). Floor the
            # reported CF VaR at the Gaussian VaR so a fat-tailed series never shows less
            # tail risk than normal, and flag when the raw expansion was out of domain.
            cf_reliable = (abs(g1) <= 2.0 and g2_excess <= 7.0 and zcf <= z)
            var_cf_pct = max(var_cf_raw_pct, var_param_pct)
            out['scopes'][scope] = {
                'skewness': _round(g1, 3), 'excess_kurtosis': _round(g2_excess, 3),
                'jarque_bera_stat': _round(float(jb_stat), 2),
                'jarque_bera_p': _round(float(jb_p), 4),
                'normal_rejected': bool(jb_p < 0.05),
                'var_95_param_pct': _round(var_param_pct, 3),
                'var_95_cornish_fisher_pct': _round(var_cf_pct, 3),
                'var_95_cornish_fisher_raw_pct': _round(var_cf_raw_pct, 3),
                'cf_reliable': bool(cf_reliable),
                'tail_underestimation_pct': _round(var_cf_pct - var_param_pct, 3),
            }
        return out

    # D4 — PSR / DSR (Bailey-Lopez de Prado 2012, 2014)
    def diag_psr(self):
        from scipy import stats as sstats
        out = {'ref': 'Bailey & Lopez de Prado (2012,2014)', 'scopes': {}}
        for scope in ('combined', 'mrpt', 'mtfs'):
            nav = self.d.nav_series(scope)
            r = nav.pct_change(fill_method=None).dropna()
            n = len(r)
            if n < 20 or r.std(ddof=1) == 0:
                continue
            sr = float(r.mean() / r.std(ddof=1))           # daily Sharpe
            g1 = float(sstats.skew(r))
            g2 = float(sstats.kurtosis(r, fisher=False))   # non-excess kurtosis
            denom = math.sqrt(max(1e-9, 1 - g1 * sr + (g2 - 1) / 4 * sr**2))
            psr0 = float(sstats.norm.cdf((sr - 0.0) * math.sqrt(n - 1) / denom))
            z95 = 1.645
            mintrl = (1 + (1 - g1 * sr + (g2 - 1) / 4 * sr**2) * (z95 / sr) ** 2) if sr != 0 else None
            out['scopes'][scope] = {
                'sharpe_daily': _round(sr, 4),
                'sharpe_annual': _round(sr * math.sqrt(ANNUALIZATION), 3),
                'psr_vs_zero': _round(psr0, 4),
                'psr_pass_95': bool(psr0 >= 0.95),
                'min_track_record_len_days': _round(mintrl, 0),
                'n_obs': n,
            }
        return out

    # D5 — CDaR (Chekhlov-Uryasev-Zabarankin 2005)
    def diag_cdar(self, alpha=0.95):
        out = {'ref': 'Chekhlov, Uryasev & Zabarankin (2005)', 'scopes': {}}
        for scope in ('combined', 'mrpt', 'mtfs'):
            nav = self.d.nav_series(scope)
            if len(nav) < 20:
                continue
            peak = nav.cummax()
            dd = ((peak - nav) / peak)                     # positive drawdown fraction
            dar = float(np.percentile(dd, alpha * 100))
            tail = dd[dd >= dar]
            cdar = float(tail.mean()) if len(tail) else dar
            tuw = float((dd > 1e-6).mean())                # time under water
            out['scopes'][scope] = {
                'max_drawdown_pct': _round(dd.max() * 100, 2),
                'dar_95_pct': _round(dar * 100, 2),
                'cdar_95_pct': _round(cdar * 100, 2),
                'avg_drawdown_pct': _round(dd.mean() * 100, 2),
                'time_under_water_pct': _round(tuw * 100, 1),
            }
        return out

    # D6 — Kelly leverage (Kelly 1956; MacLean-Thorp-Ziemba 2011)
    def diag_kelly(self, exposure):
        out = {'ref': 'Kelly (1956); MacLean-Thorp-Ziemba (2011)', 'scopes': {}}
        for scope in ('combined', 'mrpt', 'mtfs'):
            nav = self.d.nav_series(scope)
            r = nav.pct_change(fill_method=None).dropna()
            if len(r) < 20 or r.std(ddof=1) == 0:
                continue
            mu_ann = float(r.mean()) * ANNUALIZATION
            sd_ann = float(r.std(ddof=1)) * math.sqrt(ANNUALIZATION)
            excess = mu_ann - RISK_FREE_ANNUAL
            f_star = excess / (sd_ann ** 2) if sd_ann > 0 else None
            actual = exposure[scope]['gross_leverage']
            half = f_star / 2 if f_star is not None else None
            over = bool(actual is not None and half is not None and actual > half)
            out['scopes'][scope] = {
                'kelly_full': _round(f_star, 2), 'kelly_half': _round(half, 2),
                'actual_gross_leverage': actual,
                'over_half_kelly': over,
                'note': 'i.i.d./known-moments assumption; half-Kelly prudent; short-sample',
            }
        return out

    # ═════════════════════════════════════════════════════════════════════════
    # ORCHESTRATOR
    # ═════════════════════════════════════════════════════════════════════════
    def compute(self, factor_df=None, factor_source='none'):
        exposure = self.exposure()
        concentration = self.concentration()
        factor_block = self.factor()
        var_block = self.var()
        liquidity = self.liquidity()
        stress = self.stress(factor_block)
        limits = self.limits(exposure, concentration, factor_block, var_block)
        diagnostics = {
            'risk_contribution': self.diag_risk_contribution(),
            'factor_attribution': self.diag_factor_attribution(factor_df, factor_source),
            'distribution': self.diag_distribution(),
            'psr': self.diag_psr(),
            'cdar': self.diag_cdar(),
            'kelly': self.diag_kelly(exposure),
        }
        return {
            'exposure': exposure, 'concentration': concentration, 'factor': factor_block,
            'var': var_block, 'liquidity': liquidity, 'stress': stress,
            'limits': limits, 'diagnostics': diagnostics,
        }


# ═════════════════════════════════════════════════════════════════════════════
# EXCEL WORKBOOK EXPORTER (openpyxl) — 19 sheets
# ═════════════════════════════════════════════════════════════════════════════

class RiskWorkbookExporter:
    """Detailed 19-sheet institutional risk workbook. Styling mirrors the navy/gold
    PnLReport palette: navy headers, gold underline, alternating rows, breach coloring."""

    NAVY = 'FF1A1A2E'
    SUBHDR = 'FF16213E'
    ROW_ALT = 'FFF7F9FC'
    GOLD = 'FFD4A843'
    POS = 'FF1A7A4A'
    NEG = 'FFC0392B'
    AMBER_FILL = 'FFFFF3CD'
    RED_FILL = 'FFF8D7DA'
    GREEN_FILL = 'FFE6F4EA'

    def __init__(self, data, fin, risk, report):
        self.d = data
        self.fin = fin
        self.risk = risk
        self.report = report

    def export(self, path):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        self._Font = Font
        self._Fill = PatternFill
        self._Align = Alignment
        self._Border = Border
        self._Side = Side
        self._col = get_column_letter

        wb = Workbook()
        wb.remove(wb.active)
        self._sheet_cover(wb)
        self._sheet_balance_sheet(wb)
        self._sheet_income(wb)
        self._sheet_capital(wb)
        self._sheet_cashflow(wb)
        self._sheet_nav_trend(wb)
        self._sheet_exposure(wb)
        self._sheet_concentration(wb)
        self._sheet_factor_beta(wb)
        self._sheet_var(wb)
        self._sheet_liquidity(wb)
        self._sheet_stress(wb)
        self._sheet_limits(wb)
        self._sheet_position_detail(wb)
        self._sheet_pair_attribution(wb)
        self._sheet_reconciliation(wb)
        self._sheet_risk_contribution(wb)
        self._sheet_factor_attribution(wb)
        self._sheet_distribution(wb)
        wb.save(path)

    # ── styling helpers ──────────────────────────────────────────────────────
    def _title(self, ws, text, ncols=6):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c = ws.cell(1, 1, text)
        c.font = self._Font(bold=True, size=13, color='FF1A1A2E')
        c.alignment = self._Align(horizontal='left')

    def _hdr_row(self, ws, row, headers, widths=None):
        for j, h in enumerate(headers, 1):
            c = ws.cell(row, j, h)
            c.font = self._Font(bold=True, size=9, color='FFFFFFFF')
            c.fill = self._Fill('solid', fgColor=self.NAVY)
            c.alignment = self._Align(horizontal='center', wrap_text=True)
            c.border = self._Border(bottom=self._Side(style='medium', color=self.GOLD))

    def _row(self, ws, row, vals, money_cols=(), pct_cols=(), num_cols=(),
             alt=False, bold_first=False):
        """Write a data row with professional formatting.
        money_cols: accounting format $#,##0 with red/green color
        pct_cols:   percentage 0.00% with red/green color
        num_cols:   plain number #,##0.00 (shares, ratios, etc.)
        """
        for j, v in enumerate(vals, 1):
            c = ws.cell(row, j, v)
            c.font = self._Font(size=9, bold=(bold_first and j == 1))
            c.alignment = self._Align(horizontal='right' if j > 1 else 'left',
                                       vertical='center')
            if alt:
                c.fill = self._Fill('solid', fgColor=self.ROW_ALT)
            if j in money_cols and isinstance(v, (int, float)):
                c.number_format = '#,##0.00' if abs(v) < 100 else '#,##0'
                c.font = self._Font(size=9, color=(self.POS if v >= 0 else self.NEG),
                                     bold=(bold_first and j == 1))
            elif j in pct_cols and isinstance(v, (int, float)):
                c.number_format = '0.00"%"'
                c.font = self._Font(size=9, color=(self.POS if v >= 0 else self.NEG))
            elif j in num_cols and isinstance(v, (int, float)):
                c.number_format = '#,##0.00' if isinstance(v, float) else '#,##0'

    def _autofit(self, ws, widths):
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[self._col(j)].width = w

    def _kv(self, ws, start_row, kvs, label_w=34, val_w=22):
        r = start_row
        for k, v in kvs:
            ws.cell(r, 1, k).font = self._Font(size=9, bold=True)
            ws.cell(r, 1).alignment = self._Align(vertical='center')
            cell = ws.cell(r, 2, v)
            cell.font = self._Font(size=9)
            cell.alignment = self._Align(horizontal='right', vertical='center')
            if isinstance(v, (int, float)):
                if abs(v) > 1000:
                    cell.number_format = '#,##0'
                    cell.font = self._Font(size=9, color=(self.POS if v >= 0 else self.NEG))
                elif abs(v) > 1:
                    cell.number_format = '#,##0.00'
                else:
                    cell.number_format = '0.0000'
            r += 1
        self._autofit(ws, [label_w, val_w])
        return r

    # ── A1 Cover & Summary ───────────────────────────────────────────────────
    def _sheet_cover(self, wb):
        ws = wb.create_sheet('Cover & Summary')
        self._title(ws, 'SOMEO PARK — 组合风险管理报告 / Risk Management Report', 4)
        s = self.report['summary']
        ws.cell(2, 1, f"Signal date {self.report['signal_date']}  ·  Generated "
                      f"{self.report['generated_at']}  ·  Scope: MRPT+MTFS pairs book "
                      f"(model-based, broker-reconciled)").font = self._Font(size=8, italic=True)
        kvs = [
            ('NAV (equity)', s['nav']),
            ('Gross leverage (x)', s['gross_leverage']),
            ('Net leverage (x)', s['net_leverage']),
            ('Net market beta (target≈0)', s['net_beta']),
            ('VaR 95% 1-day ($)', s['var_95_1d']),
            ('VaR 95% 1-day (% NAV)', s['var_95_1d_pct']),
            ('Annual vol (% NAV)', s['annual_vol_pct']),
            ('Sharpe (annual)', s['sharpe_annual']),
            ('PSR( SR>0 )', s['psr_combined']),
            ('CDaR 95% (%)', s['cdar_95_pct']),
            ('Kelly full / half (x)', f"{s['kelly_full']} / {s['kelly_half']}"),
            ('Actual gross leverage (x)', s['kelly_actual_leverage']),
            ('Max single-name', f"{s['max_single_name']} ({s['max_single_name_pct']}% NAV)"),
            ('Max sector net (% gross)', s['max_sector_net_pct']),
            ('Open pairs', s['n_open_pairs']),
            ('Limit breaches', s['n_breaches']),
            ('Worst breach', s['worst_breach']),
        ]
        self._kv(ws, 4, kvs, label_w=34, val_w=30)

    # ── A2 Balance Sheet (T/T-1D/T-1W/T-1M) ─────────────────────────────────
    def _sheet_balance_sheet(self, wb):
        ws = wb.create_sheet('Balance Sheet')
        self._title(ws, 'BALANCE SHEET + LONG/SHORT BOOK — 资产负债表 + 双账户法', 6)
        bs = self.report['balance_sheet']
        periods = ['T', 'T_1D', 'T_1W', 'T_1M']
        plabels = {'T': 'T (今日)', 'T_1D': 'T-1D (前日)', 'T_1W': 'T-1W (上周)', 'T_1M': 'T-1M (上月)'}

        # View 1: Balance Sheet Method
        bs_items = [
            ('Free Cash 自由现金', 'free_cash', True, False),
            ('Restricted Cash 受限现金 (PB 102%)', 'restricted_cash', True, False),
            ('Long Securities 多头证券', 'long_securities', True, False),
            ('Total Assets 总资产', 'total_assets', True, True),
            ('Short Securities 空头证券 (义务)', 'short_securities', True, False),
            ('Margin Loan 保证金借款', 'margin_loan', True, False),
            ('Total Liabilities 总负债', 'total_liabilities', True, True),
            ('NAV (Equity) 净值 = TA - TL', 'nav', True, True),
            ('Gross Leverage 总杠杆 (x)', 'gross_leverage', False, False),
            ('Net Leverage 净杠杆 (x)', 'net_leverage', False, False),
            ('PM Requirement (20%)', 'pm_requirement', True, False),
            ('Excess Liquidity 超额流动性', 'excess_liquidity', True, False),
            ('Balance Check (≈0)', 'balance_check', False, False),
        ]

        # View 2: Long/Short Book
        book_items = [
            ('Long Book — Allocated Capital 分配资本', 'long_book.allocated_capital', True, False),
            ('Long Book — Cash 现金', 'long_book.cash', True, False),
            ('Long Book — Market Value 市值', 'long_book.market_value', True, False),
            ('Long Book — Margin Loan 融资', 'long_book.margin_loan', True, False),
            ('Long Book — NAV', 'long_book.book_nav', True, True),
            ('Short Book — Allocated Capital 分配资本', 'short_book.allocated_capital', True, False),
            ('Short Book — Collateral 担保金', 'short_book.collateral_cash', True, False),
            ('Short Book — Short Proceeds 做空收入', 'short_book.short_proceeds', True, False),
            ('Short Book — Market Value 市值(负债)', 'short_book.market_value', True, False),
            ('Short Book — Margin Loan 融资(2%)', 'short_book.margin_loan', True, False),
            ('Short Book — NAV', 'short_book.book_nav', True, True),
            ('Unallocated Cash 未分配现金', 'unallocated_cash', True, False),
            ('Dual-View NAV 双账户法NAV', 'dual_nav', True, True),
            ('NAV Alignment Check (≈0)', 'nav_alignment_check', False, False),
        ]

        def _resolve(d, dotkey):
            parts = dotkey.split('.')
            v = d
            for part in parts:
                v = v.get(part) if isinstance(v, dict) else None
            return v

        row = 3
        for scope in ('combined', 'mrpt', 'mtfs'):
            # View 1 header
            ws.cell(row, 1, f'{scope.upper()} — 资产负债表法').font = self._Font(bold=True, size=10, color=self.GOLD[2:])
            row += 1
            self._hdr_row(ws, row, ['Line item'] + [plabels.get(p, p) for p in periods])
            row += 1
            for i, (label, key, is_money, is_bold) in enumerate(bs_items):
                vals = [label] + [bs[scope].get(p, {}).get(key) for p in periods]
                mcols = tuple(range(2, 6)) if is_money else ()
                self._row(ws, row, vals, money_cols=mcols, alt=(i % 2 == 1), bold_first=is_bold)
                row += 1
            row += 1

            # View 2 header
            ws.cell(row, 1, f'{scope.upper()} — 双账户法').font = self._Font(bold=True, size=10, color=self.GOLD[2:])
            row += 1
            self._hdr_row(ws, row, ['Line item'] + [plabels.get(p, p) for p in periods])
            row += 1
            for i, (label, dotkey, is_money, is_bold) in enumerate(book_items):
                vals = [label] + [_resolve(bs[scope].get(p, {}), dotkey) for p in periods]
                mcols = tuple(range(2, 6)) if is_money else ()
                self._row(ws, row, vals, money_cols=mcols, alt=(i % 2 == 1), bold_first=is_bold)
                row += 1
            row += 2

        self._autofit(ws, [42, 18, 18, 18, 18])

    # ── A3 Income Statement ──────────────────────────────────────────────────
    def _sheet_income(self, wb):
        ws = wb.create_sheet('Income Statement')
        self._title(ws, 'INCOME STATEMENT — 利润表 (net income authoritative)', 6)
        inc = self.report['income_statement']
        spans = ['1D', '1W', '1M', 'ITD']
        row = 3
        cur_unreal = inc.get('_current_unrealized', {})
        for scope in ('combined', 'mrpt', 'mtfs'):
            if scope not in inc:
                continue
            ws.cell(row, 1, scope.upper()).font = self._Font(bold=True, size=10, color=self.GOLD[2:])
            row += 1
            self._hdr_row(ws, row, ['Line item'] + spans)
            row += 1
            self._row(ws, row, ['Net income'] + [inc[scope].get(s, {}).get('net_income') for s in spans],
                      money_cols=(2, 3, 4, 5), bold_first=True); row += 1
            self._row(ws, row, ['Interest expense (est)'] +
                      [inc[scope].get(s, {}).get('interest_expense_est') for s in spans],
                      money_cols=(2, 3, 4, 5), alt=True); row += 1
            self._row(ws, row, ['Current unrealized PnL', cur_unreal.get(scope)],
                      money_cols=(2,)); row += 2
        ws.cell(row, 1, 'Note: net_income authoritative (strategy_performance); interest '
                        'estimated on current margin; unrealized exact from inventory open '
                        'prices.').font = self._Font(size=8, italic=True)
        self._autofit(ws, [30, 16, 16, 16, 16])

    # ── A4 Capital Statement ─────────────────────────────────────────────────
    def _sheet_capital(self, wb):
        ws = wb.create_sheet('Capital Statement')
        self._title(ws, 'STATEMENT OF CHANGES IN CAPITAL — 资本变动表', 6)
        cap = self.report['capital_statement']
        spans = ['1D', '1W', '1M', 'ITD']
        items = [('Beginning NAV', 'beginning_nav'), ('Net income', 'net_income'),
                 ('Contributions', 'contributions'), ('Withdrawals', 'withdrawals'),
                 ('Ending NAV', 'ending_nav'), ('Return (%)', 'return_pct'),
                 ('Max drawdown (%)', 'max_drawdown_pct')]
        row = 3
        for scope in ('combined', 'mrpt', 'mtfs'):
            if scope not in cap:
                continue
            ws.cell(row, 1, scope.upper()).font = self._Font(bold=True, size=10, color=self.GOLD[2:])
            row += 1
            self._hdr_row(ws, row, ['Line item'] + spans); row += 1
            for i, (label, key) in enumerate(items):
                vals = [label] + [cap[scope].get(s, {}).get(key) for s in spans]
                mcols = (2, 3, 4, 5) if 'pct' not in key else ()
                pcols = (2, 3, 4, 5) if 'pct' in key else ()
                self._row(ws, row, vals, money_cols=mcols, pct_cols=pcols, alt=(i % 2 == 1),
                          bold_first=label in ('Beginning NAV', 'Ending NAV'))
                row += 1
            row += 1
        self._autofit(ws, [22, 16, 16, 16, 16])

    # ── A5 Cash Flow ─────────────────────────────────────────────────────────
    def _sheet_cashflow(self, wb):
        ws = wb.create_sheet('Cash Flow')
        self._title(ws, 'CASH FLOW STATEMENT — 现金流量表 (indirect, approximate)', 4)
        cf = self.report['cash_flow'].get('1W', {})
        kvs = [('Period', '1 week'),
               ('Operating cash flow', cf.get('operating_cash_flow')),
               ('Investing cash flow', cf.get('investing_cash_flow')),
               ('Financing cash flow', cf.get('financing_cash_flow')),
               ('Net change in cash', cf.get('net_change_in_cash')),
               ('Beginning cash', cf.get('beginning_cash')),
               ('Ending cash', cf.get('ending_cash'))]
        r = self._kv(ws, 3, kvs, label_w=30, val_w=20)
        ws.cell(r + 1, 1, cf.get('note', '')).font = self._Font(size=8, italic=True)

    # ── A6 NAV Trend ─────────────────────────────────────────────────────────
    def _sheet_nav_trend(self, wb):
        ws = wb.create_sheet('NAV Trend')
        self._title(ws, 'NAV TREND — 净值趋势 (combined)', 3)
        self._hdr_row(ws, 3, ['Date', 'NAV'])
        trend = self.fin.nav_trend()
        for i, p in enumerate(trend):
            self._row(ws, 4 + i, [p['date'], p['nav']], money_cols=(2,), alt=(i % 2 == 1))
        self._autofit(ws, [16, 18])
        try:
            from openpyxl.chart import LineChart, Reference
            ch = LineChart(); ch.title = 'Combined NAV'; ch.height = 8; ch.width = 20
            data = Reference(ws, min_col=2, min_row=3, max_row=3 + len(trend))
            cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(trend))
            ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
            ws.add_chart(ch, 'D3')
        except Exception:
            pass

    # ── B7 Exposure & Leverage ───────────────────────────────────────────────
    def _sheet_exposure(self, wb):
        ws = wb.create_sheet('Exposure & Leverage')
        self._title(ws, 'EXPOSURE & LEVERAGE — 敞口与杠杆', 11)
        exp = self.risk['exposure']
        self._hdr_row(ws, 3, ['Scope', 'Capital ($)', 'Long ($)', 'Short ($)', 'Gross ($)', 'Net ($)',
                              'Gross Lev (x)', 'Net Lev (x)', 'Deploy %', 'Margin %', 'Pairs'])
        for i, sc in enumerate(('combined', 'mrpt', 'mtfs')):
            e = exp[sc]
            self._row(ws, 4 + i, [sc.upper(), e['capital'], e['long'], e['short'], e['gross'],
                                  e['net'], e['gross_leverage'], e['net_leverage'],
                                  e['deployed_pct'], e['margin_utilization_pct'], e['n_open_pairs']],
                      money_cols=(2, 3, 4, 5, 6), num_cols=(7, 8, 9, 10, 11), alt=(i % 2 == 1))
        self._autofit(ws, [12, 15, 15, 15, 15, 15, 12, 11, 10, 10, 7])

    # ── B8 Concentration ─────────────────────────────────────────────────────
    def _sheet_concentration(self, wb):
        ws = wb.create_sheet('Concentration')
        self._title(ws, 'CONCENTRATION — 集中度 (single-name / sector / HHI)', 6)
        conc = self.risk['concentration']
        ws.cell(3, 1, f"HHI={conc['hhi']}  ·  effective N pairs={conc['effective_n_pairs']}  "
                      f"·  max pair={conc['max_pair_pct']}% of gross").font = self._Font(size=9, bold=True)
        self._hdr_row(ws, 5, ['Ticker', 'Gross ($)', 'Net ($)', '% of NAV', '% of Gross', 'N Pairs'])
        for i, sn in enumerate(conc['single_name_top']):
            self._row(ws, 6 + i, [sn['ticker'], sn['gross'], sn['net'], sn['gross_pct_capital'],
                                  sn['pct_of_gross'], sn['n_pairs']],
                      money_cols=(2, 3), pct_cols=(4, 5), num_cols=(6,), alt=(i % 2 == 1))
        srow = 6 + len(conc['single_name_top']) + 2
        ws.cell(srow, 1, 'SECTOR NET EXPOSURE').font = self._Font(bold=True, color=self.GOLD[2:])
        self._hdr_row(ws, srow + 1, ['Sector', 'Long', 'Short', 'Net', 'Net % of Gross'])
        for i, (sec, x) in enumerate(sorted(conc['sector'].items(),
                                            key=lambda kv: -abs(kv[1]['net'] or 0))):
            self._row(ws, srow + 2 + i, [sec, x['long'], x['short'], x['net'], x['net_pct_of_gross']],
                      money_cols=(2, 3, 4), alt=(i % 2 == 1))
        self._autofit(ws, [16, 16, 16, 12, 12, 8])

    # ── B9 Factor & Beta ─────────────────────────────────────────────────────
    def _sheet_factor_beta(self, wb):
        ws = wb.create_sheet('Factor & Beta')
        self._title(ws, 'FACTOR & BETA — 因子与市场 Beta', 4)
        f = self.risk['factor']
        kvs = [('Net market beta (target≈0)', f['net_market_beta']),
               ('Gross beta', f['gross_beta']),
               ('Momentum factor net (% NAV)', f['momentum_factor_net_pct']),
               ('Beta data insufficient', ', '.join(f['beta_data_insufficient']) or '—')]
        r = self._kv(ws, 3, kvs, label_w=32, val_w=24)
        ws.cell(r + 1, 1, 'SECTOR NET VECTOR (% of gross)').font = self._Font(bold=True, color=self.GOLD[2:])
        self._hdr_row(ws, r + 2, ['Sector', 'Net % of Gross'])
        for i, (sec, v) in enumerate(sorted(f['sector_net_vector'].items(),
                                            key=lambda kv: -abs(kv[1] or 0))):
            self._row(ws, r + 3 + i, [sec, v], pct_cols=(2,), alt=(i % 2 == 1))
        self._autofit(ws, [32, 24])

    # ── B10 VaR & Volatility ─────────────────────────────────────────────────
    def _sheet_var(self, wb):
        ws = wb.create_sheet('VaR & Volatility')
        self._title(ws, 'VALUE AT RISK & VOLATILITY', 4)
        v = self.risk['var']
        pm = v.get('method_param', {}); hm = v.get('method_hist', {})
        kvs = [('Parametric VaR 95% 1d ($)', pm.get('var_95_1d')),
               ('Parametric VaR 99% 1d ($)', pm.get('var_99_1d')),
               ('Annual vol ($)', pm.get('vol_annual')),
               ('Annual vol (% NAV)', pm.get('vol_annual_pct_of_capital')),
               ('Historical VaR 95% 1d ($)', hm.get('var_95_1d')),
               ('Historical CVaR 95% 1d ($)', hm.get('cvar_95_1d'))]
        r = self._kv(ws, 3, kvs, label_w=32, val_w=20)
        ws.cell(r + 1, 1, 'COMPONENT VaR (per pair, marginal)').font = self._Font(bold=True, color=self.GOLD[2:])
        self._hdr_row(ws, r + 2, ['Pair', 'Component VaR 95 ($)', '% of total'])
        for i, c in enumerate(v.get('component_var', [])):
            self._row(ws, r + 3 + i, [c['pair'], c['component_var_95'], c['pct_of_total']],
                      money_cols=(2,), alt=(i % 2 == 1))
        self._autofit(ws, [32, 20, 14])

    # ── B11 Liquidity ─────────────────────────────────────────────────────────
    def _sheet_liquidity(self, wb):
        ws = wb.create_sheet('Liquidity')
        self._title(ws, 'LIQUIDITY — 流动性 (Almgren-Chriss days-to-liquidate)', 5)
        lq = self.risk['liquidity']
        ws.cell(3, 1, f"Legs > 1 trading day to liquidate: {lq['n_legs_over_1d']} "
                      f"(participation cap {int(ADV_PARTICIPATION*100)}%)").font = self._Font(size=9, bold=True)
        self._hdr_row(ws, 5, ['Ticker', 'Pair', 'Shares', 'ADV(20d)', 'Days to liquidate'])
        for i, lg in enumerate(lq['worst_legs']):
            self._row(ws, 6 + i, [lg['ticker'], lg['pair'], lg['shares'], lg['adv20'],
                                  lg['days_to_liquidate']], money_cols=(3, 4), alt=(i % 2 == 1))
        self._autofit(ws, [12, 16, 12, 16, 16])

    # ── B12 Stress ─────────────────────────────────────────────────────────────
    def _sheet_stress(self, wb):
        ws = wb.create_sheet('Stress Scenarios')
        self._title(ws, 'STRESS / SCENARIO — beta-implied PnL', 3)
        self._hdr_row(ws, 3, ['Scenario', 'Est PnL ($)', 'Est PnL (% NAV)'])
        for i, sc in enumerate(self.risk['stress']):
            self._row(ws, 4 + i, [sc['scenario'], sc['est_pnl'], sc['est_pnl_pct_of_capital']],
                      money_cols=(2,), pct_cols=(3,), alt=(i % 2 == 1))
        self._autofit(ws, [22, 18, 16])

    # ── B13 Limit Monitor ───────────────────────────────────────────────────
    def _sheet_limits(self, wb):
        ws = wb.create_sheet('Limit Monitor')
        self._title(ws, 'LIMIT MONITOR — 限额监控 (shared thresholds w/ DailySignal gates)', 6)
        self._hdr_row(ws, 3, ['Limit', 'Scope', 'Value', 'Amber', 'Red', 'Status'])
        for i, l in enumerate(self.risk['limits']):
            r = 4 + i
            self._row(ws, r, [l['name'], l['scope'], l['value'], l['amber'], l['red'], l['status']])
            fill = {'red': self.RED_FILL, 'amber': self.AMBER_FILL,
                    'green': self.GREEN_FILL}.get(l['status'])
            if fill:
                for j in range(1, 7):
                    ws.cell(r, j).fill = self._Fill('solid', fgColor=fill)
        self._autofit(ws, [22, 12, 12, 10, 10, 10])

    # ── C14 Position Detail ───────────────────────────────────────────────────
    def _sheet_position_detail(self, wb):
        ws = wb.create_sheet('Position Detail')
        self._title(ws, 'POSITION DETAIL — 逐腿明细', 10)
        self._hdr_row(ws, 3, ['Strategy', 'Pair', 'Ticker', 'Dir', 'Shares', 'Price ($)',
                              'Market Value ($)', 'Sector', 'Days Held', 'Param Set'])
        r = 4
        for p in self.d.positions:
            for leg in p['legs']:
                px = self.d.current_price(leg['ticker'])
                mv = (leg['shares'] * px) if px else None
                leg_dir = 'long' if (leg['shares'] or 0) >= 0 else 'short'
                self._row(ws, r, [p['strategy'].upper(), p['pair'], leg['ticker'], leg_dir,
                                  leg['shares'], _round(px), _round(mv), leg['sector'],
                                  p['days_held'], p['param_set']],
                          num_cols=(5,), money_cols=(6, 7), alt=(r % 2 == 0))
                r += 1
        self._autofit(ws, [10, 14, 10, 6, 10, 12, 14, 12, 9, 22])

    # ── C15 Per-Pair PnL Attribution ───────────────────────────────────────
    def _sheet_pair_attribution(self, wb):
        ws = wb.create_sheet('Pair PnL Attribution')
        self._title(ws, 'PER-PAIR PnL ATTRIBUTION — 配对盈亏归因', 7)
        self._hdr_row(ws, 3, ['Strategy', 'Pair', 'Unrealized PnL ($)', 'Gross Notional ($)',
                              'Days Held', 'Param Set', 'Direction'])
        r = 4
        for p in self.d.positions:
            upnl = 0.0; gross = 0.0
            for leg in p['legs']:
                px = self.d.current_price(leg['ticker']); op = leg['open_price']
                if px and op:
                    upnl += (px - op) * leg['shares']
                if px:
                    gross += abs(leg['shares'] * px)
            self._row(ws, r, [p['strategy'].upper(), p['pair'], _round(upnl), _round(gross),
                              p['days_held'], p['param_set'], p['direction']],
                      money_cols=(3, 4), alt=(r % 2 == 0))
            r += 1
        self._autofit(ws, [10, 14, 16, 16, 9, 22, 9])

    # ── C16 Inventory Reconciliation ────────────────────────────────────────
    def _sheet_reconciliation(self, wb):
        ws = wb.create_sheet('Inventory Reconciliation')
        self._title(ws, 'INVENTORY RECONCILIATION — 快照来源审计', 4)
        snaps = self.report['balance_sheet'].get('as_of_snapshots', {})
        self._hdr_row(ws, 3, ['Period', 'MRPT snapshot as-of', 'MTFS snapshot as-of'])
        for i, (period, used) in enumerate(snaps.items()):
            mr = used.get('mrpt') if isinstance(used, dict) else used
            mt = used.get('mtfs') if isinstance(used, dict) else used
            self._row(ws, 4 + i, [period, mr, mt], alt=(i % 2 == 1))
        self._autofit(ws, [14, 26, 26])

    # ── D17 Risk Contribution ───────────────────────────────────────────────
    def _sheet_risk_contribution(self, wb):
        ws = wb.create_sheet('Risk Contribution')
        self._title(ws, 'RISK CONTRIBUTION — 风险贡献 (Maillard et al. 2010)', 6)
        rc = self.risk['diagnostics']['risk_contribution']
        ws.cell(2, 1, rc.get('ref', '')).font = self._Font(size=8, italic=True)
        row = 4
        ws.cell(row, 1, 'MASTER (4 strategies) — 等资本 vs 等风险').font = \
            self._Font(bold=True, color=self.GOLD[2:]); row += 1
        if rc.get('master'):
            self._hdr_row(ws, row, ['Component', 'Capital Wt %', 'Risk Contrib %',
                                    'ERC Target %', 'Divergence %']); row += 1
            for i, m in enumerate(rc['master']):
                self._row(ws, row, [m['component'], m['capital_weight_pct'],
                                    m['risk_contribution_pct'], m['erc_target_pct'],
                                    m['divergence_pct']], pct_cols=(5,), alt=(i % 2 == 1))
                row += 1
        else:
            ws.cell(row, 1, 'master data insufficient').font = self._Font(size=8, italic=True); row += 1
        row += 1
        ws.cell(row, 1, 'PAIRS (within combined book)').font = self._Font(bold=True, color=self.GOLD[2:])
        row += 1
        if rc.get('pairs'):
            self._hdr_row(ws, row, ['Pair', 'Risk Contribution %']); row += 1
            for i, p in enumerate(rc['pairs']):
                self._row(ws, row, [p['pair'], p['risk_contribution_pct']], alt=(i % 2 == 1))
                row += 1
        self._autofit(ws, [16, 14, 16, 14, 14])

    # ── D18 Factor Attribution ──────────────────────────────────────────────
    def _sheet_factor_attribution(self, wb):
        ws = wb.create_sheet('Factor Attribution')
        self._title(ws, 'FACTOR ATTRIBUTION — FF5 + UMD 归因 (Fama-French; Carhart)', 11)
        fa = self.risk['diagnostics']['factor_attribution']
        ws.cell(2, 1, f"{fa.get('ref','')}  ·  source={fa.get('source')}  ·  {fa.get('note','')}"
                ).font = self._Font(size=8, italic=True)
        row = 4
        factors = ['MKT', 'SMB', 'HML', 'RMW', 'CMA', 'UMD']
        self._hdr_row(ws, row, ['Scope', 'Annual α %', 'α t', 'R²', 'n'] + factors)
        row += 1
        for i, scope in enumerate(('combined', 'mrpt', 'mtfs')):
            m = fa.get('models', {}).get(scope)
            if not m:
                continue
            ld = m['loadings']
            vals = [scope.upper(), m['alpha_annual_pct'], m['alpha_t'], m['r_squared'], m['n_obs']]
            for fac in factors:
                b = ld.get(fac, {}).get('beta')
                t = ld.get(fac, {}).get('t_stat')
                vals.append(f"{b} (t={t})" if b is not None else '—')
            self._row(ws, row, vals, pct_cols=(2,), alt=(i % 2 == 1))
            row += 1
        self._autofit(ws, [11, 11, 8, 8, 6, 14, 14, 14, 14, 14, 14])

    # ── D19 Return Distribution & Tail ──────────────────────────────────────
    def _sheet_distribution(self, wb):
        ws = wb.create_sheet('Return Distribution')
        self._title(ws, 'RETURN DISTRIBUTION & TAIL — 分布/肥尾/PSR/CDaR/Kelly', 10)
        diag = self.risk['diagnostics']
        dist = diag['distribution']['scopes']
        psr = diag['psr']['scopes']
        cdar = diag['cdar']['scopes']
        kelly = diag['kelly']['scopes']
        self._hdr_row(ws, 3, ['Scope', 'Skew', 'ExKurt', 'JB p', 'NormRej',
                              'VaR95 norm%', 'VaR95 CF%', 'TailUnder%'])
        for i, sc in enumerate(('combined', 'mrpt', 'mtfs')):
            d = dist.get(sc, {})
            self._row(ws, 4 + i, [sc.upper(), d.get('skewness'), d.get('excess_kurtosis'),
                                  d.get('jarque_bera_p'), d.get('normal_rejected'),
                                  d.get('var_95_param_pct'), d.get('var_95_cornish_fisher_pct'),
                                  d.get('tail_underestimation_pct')], alt=(i % 2 == 1))
        row = 8
        self._hdr_row(ws, row, ['Scope', 'Sharpe(ann)', 'PSR(>0)', 'minTRL(d)',
                                'CDaR95%', 'MaxDD%', 'TUW%', 'Kelly f*', 'Kelly½', 'ActualLev'])
        row += 1
        for i, sc in enumerate(('combined', 'mrpt', 'mtfs')):
            p = psr.get(sc, {}); c = cdar.get(sc, {}); k = kelly.get(sc, {})
            self._row(ws, row + i, [sc.upper(), p.get('sharpe_annual'), p.get('psr_vs_zero'),
                                    p.get('min_track_record_len_days'), c.get('cdar_95_pct'),
                                    c.get('max_drawdown_pct'), c.get('time_under_water_pct'),
                                    k.get('kelly_full'), k.get('kelly_half'),
                                    k.get('actual_gross_leverage')], alt=(i % 2 == 1))
        self._autofit(ws, [11, 11, 9, 10, 9, 9, 8, 9, 9, 10])


# ═════════════════════════════════════════════════════════════════════════════
# PDF REPORTER (reportlab) — mirrors PnLReport visual design
# ═════════════════════════════════════════════════════════════════════════════

def _register_cjk_font():
    """Register a CJK TTF (mirrors PnLReport._register_cjk). Returns font name."""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    for path in ['/System/Library/Fonts/PingFang.ttc',
                 '/System/Library/Fonts/STHeiti Light.ttc',
                 '/Library/Fonts/Arial Unicode MS.ttf',
                 '/System/Library/Fonts/Supplemental/Arial Unicode.ttf']:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont('CJK', path))
                return 'CJK'
            except Exception:
                continue
    return 'Helvetica'


class RiskPDFReporter:
    """Multi-section PDF risk report; same palette/fonts/table style as PnLReport."""

    C_HEADER = '#1a1a2e'
    C_GOLD = '#d4a843'
    C_POS = '#1a7a4a'
    C_NEG = '#c0392b'
    C_ROW_ALT = '#f7f9fc'
    C_BORDER = '#cccccc'
    C_GRAY = '#888888'
    C_AMBER = '#fff3cd'
    C_RED = '#f8d7da'
    C_GREEN = '#e6f4ea'

    def __init__(self, data, fin, risk, report):
        self.d = data
        self.fin = fin
        self.risk = risk
        self.report = report
        self.FONT = _register_cjk_font()

    # ── value formatters (HTML font-color, mirrors PnLReport) ────────────────
    def _money(self, v):
        if v is None:
            return 'N/A'
        s = f'+{abs(v):,.0f}' if v >= 0 else f'-{abs(v):,.0f}'
        c = self.C_POS if v >= 0 else self.C_NEG
        return f'<font color="{c}">{s}</font>'

    def _num(self, v, dec=2, suffix=''):
        if v is None:
            return 'N/A'
        return f'{v:,.{dec}f}{suffix}'

    def export(self, path):
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_RIGHT, TA_LEFT, TA_CENTER
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph,
                                        Spacer, HRFlowable, KeepTogether)
        from reportlab.lib.styles import ParagraphStyle

        self._colors = colors
        self._cm = cm
        self._Table = Table
        self._TableStyle = TableStyle
        self._Para = Paragraph
        self._TA_RIGHT = TA_RIGHT

        W = A4[0] - 4 * cm
        doc = SimpleDocTemplate(path, pagesize=A4, leftMargin=2 * cm, rightMargin=2 * cm,
                                topMargin=1.2 * cm, bottomMargin=1.5 * cm)
        F = self.FONT
        self.title_style = ParagraphStyle('t', fontName=F, fontSize=15, leading=19,
                                           alignment=TA_CENTER)
        self.sub_style = ParagraphStyle('s', fontName=F, fontSize=8.5, leading=12,
                                        alignment=TA_CENTER, textColor=colors.HexColor(self.C_GRAY))
        self.h1 = ParagraphStyle('h1', fontName=F, fontSize=10, leading=14,
                                 spaceBefore=12, spaceAfter=5, textColor=colors.HexColor(self.C_HEADER))
        self.body = ParagraphStyle('b', fontName=F, fontSize=8, leading=11)
        self.footer = ParagraphStyle('f', fontName=F, fontSize=7, leading=9.5,
                                     textColor=colors.HexColor(self.C_GRAY))

        story = []
        s = self.report['summary']
        story.append(Paragraph('Someo Park — 组合风险管理报告', self.title_style))
        story.append(Paragraph(
            f"信号日期 {self.report['signal_date']} · 生成时间 {self.report['generated_at']}<br/>"
            f"范围：MRPT + MTFS 配对组合 · 基于模型，券商对账完成", self.sub_style))
        story.append(HRFlowable(width='100%', thickness=1, color=colors.HexColor(self.C_GOLD),
                                spaceBefore=6, spaceAfter=10))

        # Risk summary KV
        story.append(Paragraph('风险摘要', self.h1))
        story.append(self._kv_table(W, [
            ('NAV（净值）', self._money(s['nav'])),
            ('总 / 净杠杆', f"{s['gross_leverage']}x / {s['net_leverage']}x"),
            ('净市场 Beta（目标≈0）', self._num(s['net_beta'], 2)),
            ('VaR 95% 单日', f"{self._money(s['var_95_1d'])} ({s['var_95_1d_pct']}% NAV)"),
            ('年化波动率（% NAV）', self._num(s['annual_vol_pct'], 2, '%')),
            ('Sharpe（年化）/ PSR(>0)', f"{s['sharpe_annual']} / {s['psr_combined']}"),
            ('CDaR 95%', self._num(s['cdar_95_pct'], 2, '%')),
            ('Kelly 全 / 半 / 实际', f"{s['kelly_full']}x / {s['kelly_half']}x / {s['kelly_actual_leverage']}x"),
            ('最大单一标的', f"{s['max_single_name']} ({s['max_single_name_pct']}% NAV)"),
            ('持仓对数 / 超限数', f"{s['n_open_pairs']} / {s['n_breaches']}（最严重 {s['worst_breach']}）"),
        ]))

        # 一、Balance sheet
        story.append(Paragraph('一、资产负债表 + 双账户法（T 今日 / T-1D 前一交易日 / T-1W 上周 / T-1M 上月）', self.h1))
        bs_elements = self._balance_sheet_table(W)
        if isinstance(bs_elements, list):
            story.extend(bs_elements)
        else:
            story.append(bs_elements)

        # 二、Income statement
        story.append(Paragraph('二、利润表', self.h1))
        story.append(self._income_table(W))

        # 三、Capital + cash flow (header + table kept together on the same page)
        story.append(KeepTogether(
            [Paragraph('三、资本变动表 / 现金流量表', self.h1)] + self._capital_table(W)))

        # 四、Exposure
        story.append(Paragraph('四、敞口与杠杆', self.h1))
        story.append(self._exposure_table(W))

        # 五、Concentration
        story.append(Paragraph('五、集中度（单一标的前列，占 NAV %）', self.h1))
        story.append(self._concentration_table(W))

        # 六、Factor & beta
        story.append(Paragraph('六、因子与 Beta', self.h1))
        story.append(self._factor_table(W))

        # 七、VaR
        story.append(Paragraph('七、风险价值 VaR', self.h1))
        story.append(self._var_table(W))

        # 八、Liquidity + stress
        story.append(Paragraph('八、流动性与压力测试', self.h1))
        story.append(self._stress_table(W))

        # 九、Limits
        story.append(Paragraph('九、限额监控', self.h1))
        story.append(self._limits_table(W))

        # 十、Diagnostics
        story.append(Paragraph('十、组合诊断', self.h1))
        for blk in self._diagnostics_tables(W):
            story.append(blk)

        story.append(Spacer(1, 10))
        story.append(HRFlowable(width='100%', thickness=0.4, color=colors.HexColor(self.C_BORDER)))
        story.append(Paragraph(
            f"由 Someo Park RiskManager 生成 · {datetime.now().strftime('%Y-%m-%d %H:%M')} · "
            f"基于模型，券商对账完成", self.footer))
        doc.build(story)

    # ── table builders ───────────────────────────────────────────────────────
    def _P(self, text, align='LEFT', header=False, size=7.5):
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_RIGHT
        st = ParagraphStyle('_c', fontName=self.FONT, fontSize=size, leading=size + 3,
                            alignment=(TA_RIGHT if align == 'RIGHT' else 0),
                            textColor=(self._colors.white if header else self._colors.black))
        content = f'<b>{text}</b>' if header else str(text)
        return self._Para(content, st)

    def _styled_table(self, data, col_widths, header_rows=1, status_col=None, status_rows=None):
        t = self._Table(data, colWidths=col_widths, repeatRows=header_rows)
        cmds = [
            ('FONTNAME', (0, 0), (-1, -1), self.FONT), ('FONTSIZE', (0, 0), (-1, -1), 7.5),
            ('LEADING', (0, 0), (-1, -1), 10.5),
            ('BACKGROUND', (0, 0), (-1, header_rows - 1), self._colors.HexColor(self.C_HEADER)),
            ('TEXTCOLOR', (0, 0), (-1, header_rows - 1), self._colors.white),
            ('LINEBELOW', (0, header_rows - 1), (-1, header_rows - 1), 0.9, self._colors.HexColor(self.C_GOLD)),
            ('GRID', (0, 0), (-1, -1), 0.35, self._colors.HexColor(self.C_BORDER)),
            ('TOPPADDING', (0, 0), (-1, -1), 3.5), ('BOTTOMPADDING', (0, 0), (-1, -1), 3.5),
            ('LEFTPADDING', (0, 0), (-1, -1), 5), ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ]
        n = len(data)
        for i in range(header_rows, n):
            if (i - header_rows) % 2 == 1:
                cmds.append(('BACKGROUND', (0, i), (-1, i), self._colors.HexColor(self.C_ROW_ALT)))
        # status row coloring (limit monitor)
        if status_rows:
            for i, st in status_rows.items():
                col = {'red': self.C_RED, 'amber': self.C_AMBER, 'green': self.C_GREEN}.get(st)
                if col:
                    cmds.append(('BACKGROUND', (0, i), (-1, i), self._colors.HexColor(col)))
        t.setStyle(self._TableStyle(cmds))
        return t

    def _kv_table(self, W, kvs):
        data = [[self._P(k), self._P(v, 'RIGHT')] for k, v in kvs]
        t = self._Table(data, colWidths=[W * 0.55, W * 0.45])
        t.setStyle(self._TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), self.FONT), ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('LINEABOVE', (0, 0), (-1, 0), 0.6, self._colors.HexColor(self.C_GOLD)),
            ('LINEBELOW', (0, -1), (-1, -1), 0.6, self._colors.HexColor(self.C_GOLD)),
            ('TOPPADDING', (0, 0), (-1, -1), 2.5), ('BOTTOMPADDING', (0, 0), (-1, -1), 2.5),
        ]))
        for i in range(len(data)):
            if i % 2 == 1:
                t.setStyle(self._TableStyle([('BACKGROUND', (0, i), (-1, i),
                                              self._colors.HexColor(self.C_ROW_ALT))]))
        return t

    def _balance_sheet_table(self, W):
        bs = self.report['balance_sheet']['combined']
        periods = ['T', 'T_1D', 'T_1W', 'T_1M']
        _period_labels = {
            'T': 'T (今日)', 'T_1D': 'T-1D (前一交易日)',
            'T_1W': 'T-1W (上周)', 'T_1M': 'T-1M (上月)',
        }
        # ── View 1: Balance Sheet Method (资产负债表法) ──
        bs_items = [
            ('自由现金', 'free_cash', True),
            ('受限现金（PB 做空担保 102%）', 'restricted_cash', True),
            ('多头证券', 'long_securities', True),
            ('总资产', 'total_assets', True),
            ('空头证券（回补义务）', 'short_securities', True),
            ('保证金借款（多头融资 + 2% 空头担保）', 'margin_loan', True),
            ('总负债', 'total_liabilities', True),
            ('NAV（净值）= TA - TL', 'nav', True),
            ('总 / 净杠杆 (x)', 'gross_leverage', False),
            ('Portfolio Margin 要求（20%）', 'pm_requirement', True),
            ('超额流动性', 'excess_liquidity', True),
        ]
        head = [self._P('资产负债表法（合并）', header=True)] + \
               [self._P(_period_labels.get(p, p), 'RIGHT', header=True) for p in periods]
        data = [head]
        for label, key, money in bs_items:
            row = [self._P(label)]
            for p in periods:
                v = bs.get(p, {}).get(key)
                if key == 'gross_leverage':
                    nv = bs.get(p, {}).get('net_leverage')
                    row.append(self._P(f"{v}/{nv}", 'RIGHT'))
                else:
                    row.append(self._P(self._money(v), 'RIGHT'))
            data.append(row)
        t1 = self._styled_table(data, [W * 0.32] + [W * 0.17] * 4)

        # ── View 2: Long/Short Book (双账户法) ──
        book_items = [
            ('Long Book 分配资本', 'long_book.allocated_capital', True),
            ('Long Book 现金', 'long_book.cash', True),
            ('Long Book 市值', 'long_book.market_value', True),
            ('Long Book 融资', 'long_book.margin_loan', True),
            ('Long Book NAV', 'long_book.book_nav', True),
            ('Short Book 分配资本', 'short_book.allocated_capital', True),
            ('Short Book 担保金', 'short_book.collateral_cash', True),
            ('Short Book 做空收入', 'short_book.short_proceeds', True),
            ('Short Book 市值（负债）', 'short_book.market_value', True),
            ('Short Book 融资（2% 担保）', 'short_book.margin_loan', True),
            ('Short Book NAV', 'short_book.book_nav', True),
            ('未分配现金', 'unallocated_cash', True),
            ('双账户法 NAV', 'dual_nav', True),
            ('NAV 对齐误差', 'nav_alignment_check', False),
        ]
        head2 = [self._P('双账户法（合并）', header=True)] + \
                [self._P(_period_labels.get(p, p), 'RIGHT', header=True) for p in periods]
        data2 = [head2]
        for label, dotkey, money in book_items:
            row = [self._P(label)]
            for p in periods:
                d = bs.get(p, {})
                # Support dotted keys like 'long_book.cash'
                parts = dotkey.split('.')
                v = d
                for part in parts:
                    v = v.get(part) if isinstance(v, dict) else None
                if dotkey == 'nav_alignment_check':
                    row.append(self._P(self._num(v, 4) if v is not None else '—', 'RIGHT'))
                else:
                    row.append(self._P(self._money(v) if money else self._num(v, 2), 'RIGHT'))
            data2.append(row)
        t2 = self._styled_table(data2, [W * 0.32] + [W * 0.17] * 4)

        from reportlab.platypus import Spacer
        return [t1, Spacer(1, 12), t2]

    def _income_table(self, W):
        inc = self.report['income_statement']
        spans = ['1D', '1W', '1M', 'ITD']
        _span_labels = {'1D': '1D (日)', '1W': '1W (周)', '1M': '1M (月)', 'ITD': 'ITD (成立以来)'}
        head = [self._P('合并', header=True)] + [self._P(_span_labels.get(s, s), 'RIGHT', header=True) for s in spans]
        data = [head]
        c = inc.get('combined', {})
        data.append([self._P('净利润')] +
                    [self._P(self._money(c.get(s, {}).get('net_income')), 'RIGHT') for s in spans])
        data.append([self._P('利息支出（估）')] +
                    [self._P(self._money(c.get(s, {}).get('interest_expense_est')), 'RIGHT') for s in spans])
        cu = inc.get('_current_unrealized', {})
        data.append([self._P('当前未实现')] +
                    [self._P(self._money(cu.get('combined')) if s == 'ITD' else '', 'RIGHT') for s in spans])
        return self._styled_table(data, [W * 0.32] + [W * 0.17] * 4)

    def _capital_table(self, W):
        cap = self.report['capital_statement'].get('combined', {})
        cf = self.report['cash_flow'].get('1W', {})
        spans = ['1D', '1W', '1M', 'ITD']
        _span_labels = {'1D': '1D (日)', '1W': '1W (周)', '1M': '1M (月)', 'ITD': 'ITD (成立以来)'}
        head = [self._P('资本（合并）', header=True)] + [self._P(_span_labels.get(s, s), 'RIGHT', header=True) for s in spans]
        data = [head]
        for label, key, money in [('期初 NAV', 'beginning_nav', True),
                                   ('净利润', 'net_income', True),
                                   ('期末 NAV', 'ending_nav', True),
                                   ('收益率 %', 'return_pct', False),
                                   ('最大回撤 %', 'max_drawdown_pct', False)]:
            row = [self._P(label)]
            for s in spans:
                v = cap.get(s, {}).get(key)
                row.append(self._P(self._money(v) if money else self._num(v, 2, '%'), 'RIGHT'))
            data.append(row)
        t1 = self._styled_table(data, [W * 0.32] + [W * 0.17] * 4)
        cf_kv = self._kv_table(W, [
            ('现金流（1W）— 经营', self._money(cf.get('operating_cash_flow'))),
            ('投资', self._money(cf.get('investing_cash_flow'))),
            ('融资', self._money(cf.get('financing_cash_flow'))),
            ('现金净变动', self._money(cf.get('net_change_in_cash'))),
        ])
        return [t1, self._Para('<br/>', self.body), cf_kv]

    def _exposure_table(self, W):
        exp = self.risk['exposure']
        head = [self._P(h, 'RIGHT' if i else 'LEFT', header=True)
                for i, h in enumerate(['范围', '多头', '空头', '总额', '净额', '总杠杆', '净杠杆', '对数'])]
        data = [head]
        for sc in ('combined', 'mrpt', 'mtfs'):
            e = exp[sc]
            data.append([self._P(sc.upper()),
                         self._P(self._money(e['long']), 'RIGHT'),
                         self._P(self._money(e['short']), 'RIGHT'),
                         self._P(self._money(e['gross']), 'RIGHT'),
                         self._P(self._money(e['net']), 'RIGHT'),
                         self._P(f"{e['gross_leverage']}x", 'RIGHT'),
                         self._P(f"{e['net_leverage']}x", 'RIGHT'),
                         self._P(e['n_open_pairs'], 'RIGHT')])
        return self._styled_table(data, [W * 0.16, W * 0.15, W * 0.15, W * 0.15, W * 0.15,
                                         W * 0.09, W * 0.08, W * 0.07])

    def _concentration_table(self, W):
        conc = self.risk['concentration']
        head = [self._P(h, 'RIGHT' if i else 'LEFT', header=True)
                for i, h in enumerate(['标的', '总额', '净额', '% NAV', '% 总额', '对数'])]
        data = [head]
        for sn in conc['single_name_top'][:10]:
            data.append([self._P(sn['ticker']),
                         self._P(self._money(sn['gross']), 'RIGHT'),
                         self._P(self._money(sn['net']), 'RIGHT'),
                         self._P(self._num(sn['gross_pct_capital'], 1, '%'), 'RIGHT'),
                         self._P(self._num(sn['pct_of_gross'], 1, '%'), 'RIGHT'),
                         self._P(sn['n_pairs'], 'RIGHT')])
        return self._styled_table(data, [W * 0.18, W * 0.2, W * 0.2, W * 0.14, W * 0.14, W * 0.14])

    def _factor_table(self, W):
        f = self.risk['factor']
        return self._kv_table(W, [
            ('净市场 Beta（目标≈0）', self._num(f['net_market_beta'], 3)),
            ('总 Beta', self._num(f['gross_beta'], 3)),
            ('动量因子净额（% NAV）', self._num(f['momentum_factor_net_pct'], 2, '%')),
            ('Beta 数据不足', ', '.join(f['beta_data_insufficient']) or '—'),
        ])

    def _var_table(self, W):
        v = self.risk['var']
        pm = v.get('method_param', {}); hm = v.get('method_hist', {})
        return self._kv_table(W, [
            ('参数法 VaR 95% / 99% 单日', f"{self._money(pm.get('var_95_1d'))} / {self._money(pm.get('var_99_1d'))}"),
            ('年化波动率（% NAV）', self._num(pm.get('vol_annual_pct_of_capital'), 2, '%')),
            ('历史法 VaR 95% / CVaR 95%', f"{self._money(hm.get('var_95_1d'))} / {self._money(hm.get('cvar_95_1d'))}"),
            ('最大成分 VaR 配对', (v.get('component_var') or [{}])[0].get('pair', '—')),
        ])

    def _stress_table(self, W):
        head = [self._P('情景', header=True), self._P('预估盈亏', 'RIGHT', header=True),
                self._P('% NAV', 'RIGHT', header=True)]
        data = [head]
        for sc in self.risk['stress']:
            data.append([self._P(sc['scenario']),
                         self._P(self._money(sc['est_pnl']), 'RIGHT'),
                         self._P(self._num(sc['est_pnl_pct_of_capital'], 2, '%'), 'RIGHT')])
        lq = self.risk['liquidity']
        data.append([self._P(f"流动性：{lq['n_legs_over_1d']} 条腿清算需 >1 日"),
                     self._P(''), self._P('')])
        return self._styled_table(data, [W * 0.5, W * 0.25, W * 0.25])

    def _limits_table(self, W):
        head = [self._P(h, 'RIGHT' if i > 1 else 'LEFT', header=True)
                for i, h in enumerate(['限额', '范围', '数值', '黄线', '红线', '状态'])]
        data = [head]
        status_rows = {}
        _st_label = {'red': '红', 'amber': '黄', 'green': '绿'}
        for i, l in enumerate(self.risk['limits'], 1):
            status_rows[i] = l['status']
            data.append([self._P(l['name']), self._P(l['scope']),
                         self._P(l['value'], 'RIGHT'), self._P(l['amber'], 'RIGHT'),
                         self._P(l['red'], 'RIGHT'),
                         self._P(_st_label.get(l['status'], l['status']), 'RIGHT')])
        return self._styled_table(data, [W * 0.28, W * 0.16, W * 0.14, W * 0.14, W * 0.14, W * 0.14],
                                  status_rows=status_rows)

    def _diagnostics_tables(self, W):
        diag = self.risk['diagnostics']
        blocks = []
        # D1 risk contribution (master)
        rc = diag['risk_contribution']
        if rc.get('master'):
            head = [self._P(h, 'RIGHT' if i else 'LEFT', header=True)
                    for i, h in enumerate(['十.1 风险贡献（master）', 'Cap Wt%', 'Risk%', 'ERC%', 'Δ%'])]
            data = [head]
            for m in rc['master']:
                data.append([self._P(m['component']),
                             self._P(self._num(m['capital_weight_pct'], 1), 'RIGHT'),
                             self._P(self._num(m['risk_contribution_pct'], 1), 'RIGHT'),
                             self._P(self._num(m['erc_target_pct'], 1), 'RIGHT'),
                             self._P(self._num(m['divergence_pct'], 1), 'RIGHT')])
            blocks.append(self._styled_table(data, [W * 0.4, W * 0.15, W * 0.15, W * 0.15, W * 0.15]))
        # D2 factor attribution
        fa = diag['factor_attribution']
        if fa.get('models'):
            head = [self._P('十.2 因子归因', header=True), self._P('α%/yr', 'RIGHT', header=True),
                    self._P('R²', 'RIGHT', header=True), self._P('MKT', 'RIGHT', header=True),
                    self._P('UMD', 'RIGHT', header=True)]
            data = [head]
            for sc in ('combined', 'mrpt', 'mtfs'):
                m = fa['models'].get(sc)
                if not m:
                    continue
                ld = m['loadings']
                data.append([self._P(sc.upper()),
                             self._P(self._num(m['alpha_annual_pct'], 1), 'RIGHT'),
                             self._P(self._num(m['r_squared'], 2), 'RIGHT'),
                             self._P(self._num(ld.get('MKT', {}).get('beta'), 2), 'RIGHT'),
                             self._P(self._num(ld.get('UMD', {}).get('beta'), 2), 'RIGHT')])
            blocks.append(self._styled_table(data, [W * 0.3, W * 0.175, W * 0.175, W * 0.175, W * 0.175]))
        # D3/D4/D5/D6 combined scope summary
        dist = diag['distribution']['scopes'].get('combined', {})
        psr = diag['psr']['scopes'].get('combined', {})
        cdar = diag['cdar']['scopes'].get('combined', {})
        kelly = diag['kelly']['scopes'].get('combined', {})
        blocks.append(self._kv_table(W, [
            ('十.3 偏度 / 超额峰度 / 正态性被拒',
             f"{dist.get('skewness')} / {dist.get('excess_kurtosis')} / {dist.get('normal_rejected')}"),
            ('十.3 VaR95 正态 vs Cornish-Fisher (%)',
             f"{dist.get('var_95_param_pct')}% vs {dist.get('var_95_cornish_fisher_pct')}%"),
            ('十.4 Sharpe（年化）/ PSR(>0) / 最短记录期',
             f"{psr.get('sharpe_annual')} / {psr.get('psr_vs_zero')} / {psr.get('min_track_record_len_days')}天"),
            ('十.5 CDaR95 / 最大回撤 / 水下时间 (%)',
             f"{cdar.get('cdar_95_pct')}% / {cdar.get('max_drawdown_pct')}% / {cdar.get('time_under_water_pct')}%"),
            ('十.6 Kelly 全 / 半 / 实际 (x)',
             f"{kelly.get('kelly_full')} / {kelly.get('kelly_half')} / {kelly.get('actual_gross_leverage')}"),
        ]))
        return blocks


# ═════════════════════════════════════════════════════════════════════════════
# SUMMARY + JSON/TXT EXPORT
# ═════════════════════════════════════════════════════════════════════════════

def _build_summary(fin: FinancialStatements, risk: dict, data: _DataLayer):
    exp = risk['exposure']['combined']
    var_p = (risk['var'].get('method_param') or {})
    conc = risk['concentration']
    limits = risk['limits']
    breaches = [l for l in limits if l['status'] in ('amber', 'red')]
    reds = [l for l in breaches if l['status'] == 'red']
    worst = (reds or breaches or [None])[0]
    diag = risk['diagnostics']
    psr_c = (diag['psr'].get('scopes', {}).get('combined') or {})
    cdar_c = (diag['cdar'].get('scopes', {}).get('combined') or {})
    kelly_c = (diag['kelly'].get('scopes', {}).get('combined') or {})
    max_sn = (conc['single_name_top'][0] if conc['single_name_top'] else {})
    max_sec = max((abs(v['net_pct_of_gross'] or 0) for v in conc['sector'].values()),
                  default=None)
    var95 = var_p.get('var_95_1d')
    return {
        'nav': _round(data.nav('combined')),
        'gross_leverage': exp['gross_leverage'],
        'net_leverage': exp['net_leverage'],
        'net_beta': risk['factor'].get('net_market_beta'),
        'var_95_1d': var95,
        'var_95_1d_pct': _round(var95 / data.nav('combined') * 100
                                if (var95 and data.nav('combined')) else None),
        'annual_vol_pct': var_p.get('vol_annual_pct_of_capital'),
        'max_single_name': max_sn.get('ticker'),
        'max_single_name_pct': max_sn.get('gross_pct_capital'),
        'max_sector_net_pct': _round(max_sec),
        'n_open_pairs': exp['n_open_pairs'],
        'n_breaches': len(breaches),
        'worst_breach': (f"{worst['name']}:{worst['scope']} {worst['status']}"
                         if worst else None),
        'psr_combined': psr_c.get('psr_vs_zero'),
        'sharpe_annual': psr_c.get('sharpe_annual'),
        'cdar_95_pct': cdar_c.get('cdar_95_pct'),
        'kelly_full': kelly_c.get('kelly_full'),
        'kelly_half': kelly_c.get('kelly_half'),
        'kelly_actual_leverage': kelly_c.get('actual_gross_leverage'),
    }


def _write_json(path, report):
    with open(path, 'w') as f:
        json.dump(report, f, indent=2, default=str)


def _write_txt(path, report):
    s = report['summary']
    L = []
    P = L.append
    P('=' * 78)
    P('  SOMEO PARK — 组合风险管理报告 (RISK MANAGEMENT REPORT)')
    P(f"  Signal date: {report['signal_date']}   Generated: {report['generated_at']}")
    P('  Scope: MRPT + MTFS pairs book   |   model-based, broker-reconciled')
    P('=' * 78)
    P('')
    P('── RISK SUMMARY ──────────────────────────────────────────────────────────')
    P(f"  NAV:                ${s['nav']:,.0f}" if s['nav'] else "  NAV: N/A")
    P(f"  Gross leverage:     {s['gross_leverage']}x      Net leverage: {s['net_leverage']}x")
    P(f"  Net market beta:    {s['net_beta']}        (target ≈ 0, market-neutral)")
    P(f"  VaR 95% 1d:         ${s['var_95_1d']:,.0f} ({s['var_95_1d_pct']}% of NAV)"
      if s['var_95_1d'] else "  VaR 95% 1d: N/A")
    P(f"  Annual vol:         {s['annual_vol_pct']}% of NAV")
    P(f"  Sharpe (annual):    {s['sharpe_annual']}    PSR(>0): {s['psr_combined']}")
    P(f"  CDaR 95%:           {s['cdar_95_pct']}%")
    P(f"  Kelly: full={s['kelly_full']}x  half={s['kelly_half']}x  actual={s['kelly_actual_leverage']}x")
    P(f"  Max single-name:    {s['max_single_name']} ({s['max_single_name_pct']}% of NAV)")
    P(f"  Open pairs:         {s['n_open_pairs']}      Limit breaches: {s['n_breaches']}"
      f"  (worst: {s['worst_breach']})")
    P('')
    # Balance sheet (combined, T) — View 1
    bs = report['balance_sheet']['combined'].get('T', {})
    P('── BALANCE SHEET METHOD (资产负债表法, combined, T) ──────────────────────')
    P(f"  Free cash:          ${bs.get('free_cash', 0):,.0f}")
    P(f"  Restricted cash:    ${bs.get('restricted_cash', 0):,.0f}  (PB short collateral 102%)")
    P(f"  Long securities:    ${bs.get('long_securities', 0):,.0f}")
    P(f"  Total assets:       ${bs.get('total_assets', 0):,.0f}")
    P(f"  Short securities:   ${bs.get('short_securities', 0):,.0f}  (repurchase obligation)")
    P(f"  Short collat. 2%:   ${bs.get('short_collateral_due', 0):,.0f}  (equity-posted, locked in restricted cash — not a liability)")
    P(f"  Margin loan:        ${bs.get('margin_loan', 0):,.0f}  (PB financing for longs)")
    P(f"  Total liabilities:  ${bs.get('total_liabilities', 0):,.0f}")
    P(f"  NAV (equity):       ${bs.get('nav', 0):,.0f}    (balance check: {bs.get('balance_check')})")
    P(f"  PM requirement:     ${bs.get('pm_requirement', 0):,.0f}  (20% of gross)")
    P(f"  Excess liquidity:   ${bs.get('excess_liquidity', 0):,.0f}  {'⚠ MARGIN CALL' if (bs.get('excess_liquidity') or 0) < 0 else 'OK'}")
    P('')
    # Long/Short Book (combined, T) — View 2
    lb = bs.get('long_book', {})
    sb = bs.get('short_book', {})
    P('── LONG/SHORT BOOK (双账户法, combined, T) ──────────────────────────────')
    P(f"  Long Book:  capital=${lb.get('allocated_capital', 0):,.0f}  MV=${lb.get('market_value', 0):,.0f}  "
      f"margin=${lb.get('margin_loan', 0):,.0f}  NAV=${lb.get('book_nav', 0):,.0f}")
    P(f"  Short Book: capital=${sb.get('allocated_capital', 0):,.0f}  MV=${sb.get('market_value', 0):,.0f}  "
      f"collateral=${sb.get('collateral_cash', 0):,.0f}  NAV=${sb.get('book_nav', 0):,.0f}")
    P(f"  Unallocated cash:   ${bs.get('unallocated_cash', 0):,.0f}")
    P(f"  Dual-view NAV:      ${bs.get('dual_nav', 0):,.0f}  (alignment check: {bs.get('nav_alignment_check', '?')})")
    P('')
    # Limit monitor
    P('── LIMIT MONITOR ─────────────────────────────────────────────────────────')
    for l in report['limits']:
        flag = {'red': '🔴', 'amber': '⚠ ', 'green': '  '}.get(l['status'], '  ')
        P(f"  {flag} {l['name']:<20} {l['scope']:<10} = {l['value']} "
          f"(amber {l['amber']}, red {l['red']}) [{l['status']}]")
    P('')
    P('=' * 78)
    P(f"  Workbook: {report.get('xlsx_path', '—')}")
    P(f"  PDF:      {report.get('pdf_path', '—')}")
    P('=' * 78)
    with open(path, 'w') as f:
        f.write('\n'.join(L))


# ═════════════════════════════════════════════════════════════════════════════
# MAIN ENTRY  (called by DailySignal hook with in-memory data)
# ═════════════════════════════════════════════════════════════════════════════

def generate_risk_report(signal_date, total_capital, mrpt_capital, mtfs_capital,
                         regime, monitor, mrpt_out, mtfs_out, ts_str=None,
                         as_of_positions=False):
    """Build the full risk pack. Returns dict with summary + file paths.
    All inputs are in-memory objects from DailySignal (deep transfer, no file re-read).
    as_of_positions=True sources positions from inventory_history ≤ signal_date
    (historical backfill); default False = current inventory (production hook)."""
    os.makedirs(RISK_DIR, exist_ok=True)
    if ts_str is None:
        ts_str = datetime.now().strftime('%Y%m%d_%H%M%S')

    capital_map = {'mrpt': mrpt_capital, 'mtfs': mtfs_capital, 'total': total_capital}
    strat_out = {'mrpt': mrpt_out or {}, 'mtfs': mtfs_out or {}}

    data = _DataLayer(signal_date, capital_map, regime, monitor, strat_out,
                      as_of_positions=as_of_positions)
    fin = FinancialStatements(data)
    engine = RiskManager(data, fin)

    # factor data (French → proxy fallback)
    factor_df, factor_source = FactorDataLoader(data).load()

    risk = engine.compute(factor_df, factor_source)
    summary = _build_summary(fin, risk, data)

    report = {
        'report_type': 'risk',
        'signal_date': str(signal_date),
        'generated_at': datetime.now().isoformat(timespec='seconds'),
        'total_capital': _round(total_capital),
        'scope': 'MRPT + MTFS pairs book (model-based, broker-reconciled)',
        'summary': summary,
        'balance_sheet': fin.balance_sheet_multi(),
        'income_statement': fin.income_statement(),
        'capital_statement': fin.capital_statement(),
        'cash_flow': fin.cash_flow_statement(),
        'by_strategy': {'mrpt': {'exposure': risk['exposure']['mrpt']},
                        'mtfs': {'exposure': risk['exposure']['mtfs']}},
        'combined': {
            'exposure': risk['exposure']['combined'],
            'concentration': risk['concentration'],
            'factor': risk['factor'],
            'var': risk['var'],
            'liquidity': risk['liquidity'],
            'stress': risk['stress'],
        },
        'limits': risk['limits'],
        'diagnostics': risk['diagnostics'],
    }

    json_path = os.path.join(RISK_DIR, f'risk_report_{ts_str}.json')
    txt_path = os.path.join(RISK_DIR, f'risk_report_{ts_str}.txt')
    xlsx_path = os.path.join(RISK_DIR, f'risk_workbook_{ts_str}.xlsx')
    pdf_path = os.path.join(RISK_DIR, f'risk_report_{ts_str}.pdf')
    report['xlsx_path'] = os.path.relpath(xlsx_path, BASE_DIR)
    report['pdf_path'] = os.path.relpath(pdf_path, BASE_DIR)

    # Excel + PDF (added in subsequent build phases; guarded so JSON/TXT always write)
    try:
        from RiskManager import RiskWorkbookExporter
        RiskWorkbookExporter(data, fin, risk, report).export(xlsx_path)
    except Exception as e:
        log.warning(f"[RISK] workbook export skipped: {e}")
        report['xlsx_path'] = None
    try:
        from RiskManager import RiskPDFReporter
        RiskPDFReporter(data, fin, risk, report).export(pdf_path)
    except Exception as e:
        log.warning(f"[RISK] pdf export skipped: {e}")
        report['pdf_path'] = None

    _write_json(json_path, report)
    _write_txt(txt_path, report)
    report['json_path'] = os.path.relpath(json_path, BASE_DIR)
    report['txt_path'] = os.path.relpath(txt_path, BASE_DIR)
    return report


def _inputs_from_combined_signals(path=None):
    """Reconstruct generate_risk_report kwargs from a combined_signals JSON
    (standalone / testing — equivalent to what the DailySignal hook passes)."""
    if path is None:
        files = sorted(glob.glob(os.path.join(BASE_DIR, 'trading_signals',
                                              'combined_signals_*.json')))
        if not files:
            raise SystemExit('[RISK] no combined_signals_*.json found')
        path = files[-1]
    with open(path) as f:
        c = json.load(f)
    mrpt_out = c.get('mrpt', {})
    mtfs_out = c.get('mtfs', {})
    return {
        'signal_date': c.get('signal_date'),
        'total_capital': c.get('total_capital'),
        'mrpt_capital': mrpt_out.get('capital'),
        'mtfs_capital': mtfs_out.get('capital'),
        'regime': c.get('regime', {}),
        'monitor': c.get('position_monitor', {'mrpt': [], 'mtfs': []}),
        'mrpt_out': mrpt_out,
        'mtfs_out': mtfs_out,
    }, path


def main():
    import argparse
    logging.basicConfig(level=logging.INFO, format='%(levelname)s:%(name)s:%(message)s')
    ap = argparse.ArgumentParser(description='Generate institutional risk report (read-only)')
    ap.add_argument('--ts', default=None, help='timestamp suffix (default now; use TESTONLY for tests)')
    ap.add_argument('--combined', default=None, help='specific combined_signals JSON path')
    ap.add_argument('--as-of', dest='as_of', action='store_true',
                    help='historical backfill: source positions from inventory_history '
                         '≤ signal_date instead of current inventory')
    args = ap.parse_args()

    kwargs, src = _inputs_from_combined_signals(args.combined)
    # for backfills, default the timestamp to <signal_date>_000000 so files sort by
    # the date they represent and never collide with real same-day runs
    ts_str = args.ts
    if ts_str is None and args.as_of and kwargs.get('signal_date'):
        ts_str = pd.Timestamp(kwargs['signal_date']).strftime('%Y%m%d') + '_000000'
    print(f'[RISK] inputs from: {os.path.basename(src)}  signal_date={kwargs["signal_date"]}'
          f'{"  [AS-OF backfill]" if args.as_of else ""}')
    rr = generate_risk_report(ts_str=ts_str, as_of_positions=args.as_of, **kwargs)
    s = rr['summary']
    print(f"[RISK] NAV=${s['nav']:,.0f}  gross={s['gross_leverage']}x  net_beta={s['net_beta']}  "
          f"VaR95=${s['var_95_1d']}  breaches={s['n_breaches']} (worst {s['worst_breach']})")
    print(f"  json: {rr.get('json_path')}")
    print(f"  txt:  {rr.get('txt_path')}")
    print(f"  xlsx: {rr.get('xlsx_path')}")
    print(f"  pdf:  {rr.get('pdf_path')}")


if __name__ == '__main__':
    main()


