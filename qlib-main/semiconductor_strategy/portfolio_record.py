"""
portfolio_record.py — SR Portfolio Excel record infrastructure (Phase 2)
=========================================================================

Produces standardized Excel files matching MRPT/MTFS PortfolioClasses output:

  1. Portfolio History Excel (26 sheets) — full backtest record
  2. Monitor Excel (5 sheets) — daily rebalance snapshot
  3. WF Diagnostic Excel (5 sheets) — walk-forward analysis

All data extracted from BacktestResult + prices + macro. No changes to engine.py
logic — this is a pure recording/export layer.
"""

from __future__ import annotations

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import numpy as np
import pandas as pd

log = logging.getLogger(__name__)

_THIS_DIR = Path(__file__).parent.resolve()
# Mirror SSRS: Excel outputs go to repo-root historical_runs/semiconductor_strategy/
# (parallel to sector_rotation's historical_runs/sector_rotation/).
_HIST_DIR = _THIS_DIR.parent.parent / "historical_runs" / "semiconductor_strategy"


# ═══════════════════════════════════════════════════════════════════════════
#  Helper: daily weight reconstruction (forward-fill + drift)
# ═══════════════════════════════════════════════════════════════════════════

def _reconstruct_daily_weights(
    weights_history: pd.DataFrame,
    prices: pd.DataFrame,
    equity_curve: pd.Series,
) -> pd.DataFrame:
    """
    Reconstruct daily sector weights from rebalance-date weights.

    Between rebalance dates, shares are held constant → weights drift with prices.

    Returns DataFrame (index=daily dates, columns=tickers, values=weights 0-1).
    """
    tickers = weights_history.columns.tolist()
    all_dates = equity_curve.index
    daily_weights = pd.DataFrame(0.0, index=all_dates, columns=tickers)

    # Convert rebalance weights → shares at each rebalance date
    shares_held = pd.Series(0.0, index=tickers)

    reb_dates = sorted(weights_history.index)
    reb_idx = 0

    for dt in all_dates:
        # Check if this is a rebalance date
        if reb_idx < len(reb_dates) and dt >= reb_dates[reb_idx]:
            reb_dt = reb_dates[reb_idx]
            w = weights_history.loc[reb_dt]
            eq = equity_curve.get(dt, equity_curve.iloc[-1])
            for t in tickers:
                p = prices[t].get(dt, np.nan) if t in prices.columns else np.nan
                if pd.notna(p) and p > 0 and eq > 0:
                    shares_held[t] = w.get(t, 0.0) * eq / p
                else:
                    shares_held[t] = 0.0
            reb_idx += 1
            # Advance past any skipped rebalance dates
            while reb_idx < len(reb_dates) and reb_dates[reb_idx] <= dt:
                reb_idx += 1

        # Mark-to-market: value each position at today's price
        total_value = 0.0
        values = {}
        for t in tickers:
            p = prices[t].get(dt, np.nan) if t in prices.columns else np.nan
            if pd.notna(p) and shares_held[t] != 0:
                v = shares_held[t] * p
                values[t] = v
                total_value += v

        # Weights = value / total_value
        if total_value > 0:
            for t in tickers:
                daily_weights.loc[dt, t] = values.get(t, 0.0) / total_value

    return daily_weights


def _compute_daily_shares(
    weights_history: pd.DataFrame,
    equity_curve: pd.Series,
    prices: pd.DataFrame,
) -> pd.DataFrame:
    """Compute daily shares held for each sector."""
    tickers = weights_history.columns.tolist()
    all_dates = equity_curve.index
    shares = pd.DataFrame(0.0, index=all_dates, columns=tickers)

    current_shares = pd.Series(0.0, index=tickers)
    reb_dates = sorted(weights_history.index)
    reb_set = set(reb_dates)

    for dt in all_dates:
        if dt in reb_set:
            w = weights_history.loc[dt]
            eq = equity_curve.get(dt, 0)
            for t in tickers:
                p = prices[t].get(dt, np.nan) if t in prices.columns else np.nan
                if pd.notna(p) and p > 0 and eq > 0:
                    current_shares[t] = w.get(t, 0.0) * eq / p
                else:
                    current_shares[t] = 0.0
        shares.loc[dt] = current_shares

    return shares


# ═══════════════════════════════════════════════════════════════════════════
#  Cost basis tracker
# ═══════════════════════════════════════════════════════════════════════════

def _compute_cost_basis(
    weights_history: pd.DataFrame,
    prices: pd.DataFrame,
    equity_curve: pd.Series,
) -> pd.DataFrame:
    """
    Track cost basis per sector: entry price when weight 0→non-zero,
    weighted average on additions, reset to 0 when weight→0.
    """
    tickers = weights_history.columns.tolist()
    all_dates = equity_curve.index
    cost_basis = pd.DataFrame(0.0, index=all_dates, columns=tickers)

    cb = pd.Series(0.0, index=tickers)
    prev_shares = pd.Series(0.0, index=tickers)
    reb_dates = sorted(weights_history.index)
    reb_set = set(reb_dates)

    for dt in all_dates:
        if dt in reb_set:
            w = weights_history.loc[dt]
            eq = equity_curve.get(dt, 0)
            for t in tickers:
                p = prices[t].get(dt, np.nan) if t in prices.columns else np.nan
                new_w = w.get(t, 0.0)

                if new_w < 1e-4:
                    # Exited position
                    cb[t] = 0.0
                    prev_shares[t] = 0.0
                elif prev_shares[t] < 1e-6:
                    # New entry
                    cb[t] = p if pd.notna(p) else 0.0
                    prev_shares[t] = new_w * eq / p if pd.notna(p) and p > 0 else 0.0
                else:
                    # Existing position — weighted average on addition
                    new_shares = new_w * eq / p if pd.notna(p) and p > 0 else 0.0
                    if new_shares > prev_shares[t] + 1e-6:
                        # Adding to position
                        added = new_shares - prev_shares[t]
                        total = prev_shares[t] + added
                        if total > 0:
                            cb[t] = (cb[t] * prev_shares[t] + p * added) / total
                    prev_shares[t] = new_shares
        cost_basis.loc[dt] = cb

    return cost_basis


# ═══════════════════════════════════════════════════════════════════════════
#  Main export class
# ═══════════════════════════════════════════════════════════════════════════

class SectorRotationRecord:
    """
    Extract data from BacktestResult and produce standardized Excel output.
    """

    def __init__(
        self,
        result,  # BacktestResult
        prices: pd.DataFrame,
        macro: pd.DataFrame = None,
        param_set: str = "",
        signal_version: str = "v1",
        leverage_ratio: float = 0.0,
        interest_rate: float = 0.05,
    ):
        self.result = result
        self.prices = prices
        self.macro = macro if macro is not None else pd.DataFrame()
        self.param_set = param_set
        self.signal_version = signal_version
        self.leverage_ratio = leverage_ratio
        self.interest_rate = interest_rate

        self.equity_curve = result.equity_curve
        self.daily_returns = result.daily_returns
        self.weights_history = result.weights_history
        self.metrics = result.metrics

        # ETF tickers from weights columns
        self.tickers = [c for c in self.weights_history.columns
                        if c not in ("cash", "CASH")]

    # ──────────────────────────────────────────────────────────────
    #  Portfolio History Excel (26 sheets)
    # ──────────────────────────────────────────────────────────────

    def export_portfolio_excel(self, output_path: Path = None,
                              mode: str = "batch",
                              span: str = "IS") -> Path:
        """
        Export 26-sheet portfolio history Excel.

        Parameters
        ----------
        mode : entry point identifier — "batch" | "select" | "tearsheet" | "backtest"
        span : data scope — "IS" (full-period in-sample) | "IS-OOS" (walk-forward validated)
        """
        if output_path is None:
            _HIST_DIR.mkdir(parents=True, exist_ok=True)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            ver = self.signal_version
            output_path = _HIST_DIR / f"aiss_portfolio_{self.param_set}_{ver}_{span}_{mode}_{ts}.xlsx"

        from openpyxl import Workbook
        wb = Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        ec = self.equity_curve.dropna()
        dr = self.daily_returns.dropna()
        wh = self.weights_history
        m = self.metrics

        # Pre-compute derived data
        daily_pnl = dr * ec.shift(1).reindex(dr.index).fillna(ec.iloc[0])
        cum_pnl = daily_pnl.cumsum()

        # Drawdown
        running_max = ec.cummax()
        dd_pct = (ec - running_max) / running_max
        dd_dollar = ec - running_max

        # Leverage accounting
        lr = self.leverage_ratio
        asset = ec * (1 + lr)
        liability = ec * lr
        interest_daily = liability * self.interest_rate / 252 if lr > 0 else ec * 0

        # Daily shares + weights + cost basis (lazy, only if needed)
        shares_df = None
        daily_w = None
        cb_df = None

        def _get_shares():
            nonlocal shares_df
            if shares_df is None:
                shares_df = _compute_daily_shares(wh, ec, self.prices)
            return shares_df

        def _get_daily_w():
            nonlocal daily_w
            if daily_w is None:
                daily_w = _reconstruct_daily_weights(wh, self.prices, ec)
            return daily_w

        def _get_cb():
            nonlocal cb_df
            if cb_df is None:
                cb_df = _compute_cost_basis(wh, self.prices, ec)
            return cb_df

        # ── 1. summary ──
        ws = wb.create_sheet("summary")
        ws.append(["Metric", "Value"])
        for k in ["sharpe", "calmar", "annual_return", "annual_vol", "max_drawdown",
                   "total_return", "monthly_win_rate", "info_ratio", "beta"]:
            v = m.get(k)
            ws.append([k, round(v, 4) if v is not None and not np.isnan(v) else "N/A"])
        ws.append(["param_set", self.param_set])
        ws.append(["signal_version", self.signal_version])
        ws.append(["leverage_ratio", self.leverage_ratio])

        # ── 2. portfolio_history ──
        ws = wb.create_sheet("portfolio_history")
        ws.append(["Date", "Equity", "Asset", "Liability", "Daily_PnL", "Cum_PnL", "Drawdown_Pct"])
        for dt in ec.index:
            ws.append([
                str(dt.date()),
                round(float(ec[dt]), 2),
                round(float(asset[dt]), 2),
                round(float(liability[dt]), 2),
                round(float(daily_pnl.get(dt, 0)), 2),
                round(float(cum_pnl.get(dt, 0)), 2),
                round(float(dd_pct.get(dt, 0)), 4),
            ])

        # ── 3. asset_history ──
        ws = wb.create_sheet("asset_history")
        ws.append(["Date", "Total_Asset"])
        for dt in ec.index:
            ws.append([str(dt.date()), round(float(asset[dt]), 2)])

        # ── 4. liability_history ──
        ws = wb.create_sheet("liability_history")
        ws.append(["Date", "Total_Liability"])
        for dt in ec.index:
            ws.append([str(dt.date()), round(float(liability[dt]), 2)])

        # ── 5. equity_history ──
        ws = wb.create_sheet("equity_history")
        ws.append(["Date", "Net_Equity"])
        for dt in ec.index:
            ws.append([str(dt.date()), round(float(ec[dt]), 2)])

        # ── 6. asset_cash_history ──
        ws = wb.create_sheet("asset_cash_history")
        ws.append(["Date", "Cash"])
        # Forward-fill cash_weight from risk_flags
        cash_w = pd.Series(0.0, index=ec.index)
        for rf in self.result.risk_flags:
            d = rf.get("date")
            if d is not None:
                cash_w.loc[d:] = rf.get("cash_pct", 0.0)
        for dt in ec.index:
            ws.append([str(dt.date()), round(float(ec[dt] * cash_w.get(dt, 0.0)), 2)])

        # ── 7. sector_prices ──
        ws = wb.create_sheet("sector_prices")
        ws.append(["Date"] + self.tickers)
        etf_p = self.prices[self.tickers].reindex(ec.index)
        for dt in ec.index:
            row = [str(dt.date())]
            for t in self.tickers:
                v = etf_p[t].get(dt, np.nan)
                row.append(round(float(v), 4) if pd.notna(v) else "")
            ws.append(row)

        # ── 8. share_history ──
        ws = wb.create_sheet("share_history")
        ws.append(["Date"] + self.tickers)
        sh = _get_shares()
        for dt in ec.index:
            row = [str(dt.date())]
            for t in self.tickers:
                row.append(round(float(sh.loc[dt, t]), 2) if t in sh.columns else 0)
            ws.append(row)

        # ── 9. sector_weights (rebalance dates only) ──
        ws = wb.create_sheet("sector_weights")
        ws.append(["Date"] + self.tickers + ["Cash"])
        for dt in wh.index:
            row = [str(dt.date())]
            total_w = 0
            for t in self.tickers:
                w = float(wh.loc[dt].get(t, 0.0))
                row.append(round(w, 4))
                total_w += w
            row.append(round(1.0 - total_w, 4))
            ws.append(row)

        # ── 10. sector_weight_pct (daily drift) ──
        ws = wb.create_sheet("sector_weight_pct")
        ws.append(["Date"] + self.tickers)
        dw = _get_daily_w()
        for dt in ec.index:
            row = [str(dt.date())]
            for t in self.tickers:
                row.append(round(float(dw.loc[dt, t]), 4) if t in dw.columns else 0)
            ws.append(row)

        # ── 11. cost_basis ──
        ws = wb.create_sheet("cost_basis")
        ws.append(["Date"] + self.tickers)
        cb = _get_cb()
        for dt in ec.index:
            row = [str(dt.date())]
            for t in self.tickers:
                row.append(round(float(cb.loc[dt, t]), 4) if t in cb.columns else 0)
            ws.append(row)

        # ── 12. sector_ratio_matrix (at rebalance dates) ──
        ws = wb.create_sheet("sector_ratio_matrix")
        for dt in wh.index:
            ws.append([f"=== {dt.date()} ==="])
            ws.append([""] + self.tickers)
            w_row = wh.loc[dt]
            for t_i in self.tickers:
                row = [t_i]
                wi = float(w_row.get(t_i, 0.0))
                for t_j in self.tickers:
                    wj = float(w_row.get(t_j, 0.0))
                    if wi > 1e-4 and wj > 1e-4:
                        row.append(round(wi / wj, 3))
                    else:
                        row.append(0)
                ws.append(row)
            ws.append([])  # blank line

        # ── 13-14. sector_pnl_acc + sector_pnl_daily ──
        ws_acc = wb.create_sheet("sector_pnl_acc")
        ws_dod = wb.create_sheet("sector_pnl_daily")
        ws_acc.append(["Date"] + [f"{t}_pnl" for t in self.tickers])
        ws_dod.append(["Date"] + [f"{t}_pnl" for t in self.tickers])

        # Sector PnL: proportional attribution from actual portfolio daily returns.
        # total_daily_pnl = portfolio_return × equity_prev (accurate, from engine)
        # sector_contribution = (sector_weight × sector_return) / sum(weight_i × return_i) × total_pnl
        # This ensures sector PnL sums exactly to total portfolio PnL.
        etf_ret = self.prices[self.tickers].pct_change().reindex(ec.index).fillna(0)
        dw = _get_daily_w()  # relative weights among sectors
        weighted_ret = dw * etf_ret  # sector_weight × sector_return
        total_weighted = weighted_ret.sum(axis=1).replace(0, np.nan)
        # Each sector's share of total daily PnL
        sector_daily_pnl = pd.DataFrame(0.0, index=ec.index, columns=self.tickers)
        for t in self.tickers:
            sector_daily_pnl[t] = (weighted_ret[t] / total_weighted).fillna(0) * daily_pnl
        sector_cum_pnl = sector_daily_pnl.cumsum()

        for dt in ec.index:
            ws_acc.append([str(dt.date())] + [round(float(sector_cum_pnl.loc[dt, t]), 2) for t in self.tickers])
            ws_dod.append([str(dt.date())] + [round(float(sector_daily_pnl.loc[dt, t]), 2) for t in self.tickers])

        # ── 15. sector_contribution ──
        ws = wb.create_sheet("sector_contribution")
        ws.append(["Date"] + [f"{t}_contrib_pct" for t in self.tickers])
        for dt in ec.index:
            dp = float(daily_pnl.get(dt, 0))
            row = [str(dt.date())]
            for t in self.tickers:
                sp = float(sector_daily_pnl.loc[dt, t])
                row.append(round(sp / dp * 100, 2) if abs(dp) > 0.01 else 0)
            ws.append(row)

        # ── 16. daily_pnl ──
        ws = wb.create_sheet("daily_pnl")
        ws.append(["Date", "Daily_PnL", "Cum_PnL"])
        for dt in ec.index:
            ws.append([str(dt.date()),
                       round(float(daily_pnl.get(dt, 0)), 2),
                       round(float(cum_pnl.get(dt, 0)), 2)])

        # ── 17. interest_expense ──
        ws = wb.create_sheet("interest_expense")
        ws.append(["Date", "Daily_Interest"])
        for dt in ec.index:
            ws.append([str(dt.date()), round(float(interest_daily.get(dt, 0)), 2)])

        # ── 18. acc_interest ──
        ws = wb.create_sheet("acc_interest")
        ws.append(["Date", "Cum_Interest"])
        acc_int = interest_daily.cumsum()
        for dt in ec.index:
            ws.append([str(dt.date()), round(float(acc_int.get(dt, 0)), 2)])

        # ── 19. realized_pnl ──
        ws = wb.create_sheet("realized_pnl")
        ws.append(["Sector", "Realized_PnL"])
        realized = self._compute_realized_pnl(sector_cum_pnl)
        for t, v in realized.items():
            ws.append([t, round(v, 2)])

        # ── 20. total_notional ──
        ws = wb.create_sheet("total_notional")
        ws.append(["Sector", "Total_Notional"])
        notional = self._compute_total_notional()
        for t, v in notional.items():
            ws.append([t, round(v, 2)])

        # ── 21. drawdown_history ──
        ws = wb.create_sheet("drawdown_history")
        ws.append(["Date", "Drawdown_Dollar", "Drawdown_Pct"])
        for dt in ec.index:
            ws.append([str(dt.date()),
                       round(float(dd_dollar.get(dt, 0)), 2),
                       round(float(dd_pct.get(dt, 0)), 4)])

        # ── 22. rebalance_trades ──
        ws = wb.create_sheet("rebalance_trades")
        ws.append(["Date", "Sector", "Direction", "Old_Weight", "New_Weight",
                    "Shares", "Price", "Est_Cost"])
        self._write_rebalance_trades(ws)

        # ── 23. regime_indicators ──
        ws = wb.create_sheet("regime_indicators")
        macro_cols = [c for c in ["vix", "yield_curve", "hy_spread", "fin_stress",
                                  "nfci", "consumer_sent", "effr", "breakeven_10y"]
                      if c in self.macro.columns]
        ws.append(["Date", "Regime"] + macro_cols)
        regime = self.result.regime_history
        for dt in ec.index:
            r = regime.get(dt, "") if dt in regime.index else ""
            row = [str(dt.date()), str(r)]
            for c in macro_cols:
                v = self.macro[c].get(dt, np.nan) if dt in self.macro.index else np.nan
                row.append(round(float(v), 4) if pd.notna(v) else "")
            ws.append(row)

        # ── 24. strategy_vars ──
        ws = wb.create_sheet("strategy_vars")
        self._write_strategy_vars(ws)

        # ── 25. stop_loss_history ──
        ws = wb.create_sheet("stop_loss_history")
        ws.append(["Date", "Sector", "Type", "Reason", "Entry_Price", "Current_Price",
                    "Threshold", "PnL_Pct", "Days_Held", "SPY_3d_Return"])
        if self.result.stop_loss_events:
            for ev in self.result.stop_loss_events:
                ws.append([
                    str(ev.date.date()), ev.ticker, ev.stop_type, ev.reason,
                    round(ev.entry_price, 2), round(ev.current_price, 2),
                    round(ev.threshold, 4), round(ev.pnl_pct, 4),
                    ev.days_held, round(ev.spy_return_3d, 4),
                ])

        # ── 26. config ──
        ws = wb.create_sheet("config")
        ws.append(["Key", "Value"])
        self._write_config_flat(ws, self.result.config)

        # Stock decomposition sheets (explode subsector → individual stocks)
        try:
            self._write_stock_decomp_sheets(wb)
        except Exception as _sd_e:
            log.warning(f"[STOCK DECOMP] Failed: {_sd_e}")

        # Save
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        log.info(f"[PORTFOLIO RECORD] Excel → {output_path}")
        return output_path

    # ──────────────────────────────────────────────────────────────
    #  Helper methods
    # ──────────────────────────────────────────────────────────────

    def _compute_realized_pnl(self, sector_cum_pnl: pd.DataFrame) -> Dict[str, float]:
        """Realized PnL = cumulative sector PnL at the point weight → 0."""
        realized = {t: 0.0 for t in self.tickers}
        wh = self.weights_history
        reb_dates = sorted(wh.index)
        for i in range(1, len(reb_dates)):
            dt = reb_dates[i]
            prev_dt = reb_dates[i - 1]
            for t in self.tickers:
                prev_w = float(wh.loc[prev_dt].get(t, 0.0))
                curr_w = float(wh.loc[dt].get(t, 0.0))
                if prev_w > 1e-4 and curr_w < 1e-4:
                    # Position exited — lock in PnL
                    if dt in sector_cum_pnl.index:
                        realized[t] += float(sector_cum_pnl.loc[dt, t])
        return realized

    def _compute_total_notional(self) -> Dict[str, float]:
        """Total notional = sum of equity × weight at each entry (0→non-zero)."""
        notional = {t: 0.0 for t in self.tickers}
        wh = self.weights_history
        ec = self.equity_curve
        reb_dates = sorted(wh.index)
        prev_w = pd.Series(0.0, index=self.tickers)
        for dt in reb_dates:
            eq = ec.get(dt, 0)
            for t in self.tickers:
                curr_w = float(wh.loc[dt].get(t, 0.0))
                if curr_w > 1e-4 and prev_w[t] < 1e-4:
                    notional[t] += curr_w * eq
                prev_w[t] = curr_w
        return notional

    def _write_rebalance_trades(self, ws):
        """Generate trade records from weight changes."""
        wh = self.weights_history
        ec = self.equity_curve
        reb_dates = sorted(wh.index)
        prev_w = pd.Series(0.0, index=self.tickers)
        for dt in reb_dates:
            eq = ec.get(dt, 0)
            for t in self.tickers:
                new_w = float(wh.loc[dt].get(t, 0.0))
                old_w = float(prev_w.get(t, 0.0))
                delta = new_w - old_w
                if abs(delta) > 1e-4:
                    p = self.prices[t].get(dt, np.nan) if t in self.prices.columns else np.nan
                    shares = abs(delta) * eq / p if pd.notna(p) and p > 0 else 0
                    direction = "BUY" if delta > 0 else "SELL"
                    # Estimated cost (use tier 2 = 5 bps as average)
                    cost = abs(delta) * eq * 0.0005
                    ws.append([
                        str(dt.date()), t, direction,
                        round(old_w, 4), round(new_w, 4),
                        round(shares, 2),
                        round(float(p), 4) if pd.notna(p) else "",
                        round(cost, 2),
                    ])
            prev_w = wh.loc[dt]

    def _write_strategy_vars(self, ws):
        """Write strategy variables at each rebalance date."""
        cfg = self.result.config
        sig_cfg = cfg.get("signals", {})
        port_cfg = cfg.get("portfolio", {})
        reb_cfg = cfg.get("rebalance", {})
        risk_cfg = cfg.get("risk", {})
        sl_cfg = cfg.get("stop_loss", {})

        # Header: config params + composite scores + risk flags
        score_cols = [f"score_{t}" for t in self.tickers]
        header = (
            ["Date", "param_set", "signal_version"] +
            ["cs_mom_weight", "ts_mom_weight", "rv_weight", "regime_weight"] +
            ["optimizer", "top_n", "min_zscore", "max_weight"] +
            ["frequency", "zscore_threshold", "max_turnover"] +
            ["vol_scaling", "target_vol", "dd_halve"] +
            ["sl_enabled", "sl_circuit_spy3d", "sl_sector_dd", "sl_trailing_peak"] +
            score_cols +
            ["regime_label", "vix", "vol_triggered", "vix_emergency", "dd_circuit", "beta_adj"]
        )
        ws.append(header)

        sig_hist = self.result.signals_history
        risk_flags = self.result.risk_flags
        rf_by_date = {rf["date"]: rf for rf in risk_flags if "date" in rf}

        for dt in sig_hist.index:
            scores = sig_hist.loc[dt]
            rf = rf_by_date.get(dt, {})
            regime = self.result.regime_history.get(dt, "") if dt in self.result.regime_history.index else ""

            row = [
                str(dt.date()), self.param_set, self.signal_version,
                sig_cfg.get("weights", {}).get("cross_sectional_momentum", ""),
                sig_cfg.get("weights", {}).get("ts_momentum", ""),
                sig_cfg.get("weights", {}).get("relative_value", ""),
                sig_cfg.get("weights", {}).get("regime_adjustment", ""),
                port_cfg.get("optimizer", ""),
                port_cfg.get("top_n_sectors", ""),
                port_cfg.get("min_zscore", ""),
                port_cfg.get("constraints", {}).get("max_weight", ""),
                reb_cfg.get("frequency", ""),
                reb_cfg.get("zscore_change_threshold", ""),
                reb_cfg.get("max_monthly_turnover", ""),
                risk_cfg.get("vol_scaling", {}).get("enabled", ""),
                risk_cfg.get("vol_scaling", {}).get("target_vol_annual", ""),
                risk_cfg.get("drawdown", {}).get("cumulative_dd_halve", ""),
                sl_cfg.get("enabled", ""),
                sl_cfg.get("portfolio_circuit_breaker", {}).get("spy_3d_limit", ""),
                sl_cfg.get("sector_collapse", {}).get("max_dd_from_entry", ""),
                sl_cfg.get("trailing_stop", {}).get("max_dd_from_peak", ""),
            ]
            for t in self.tickers:
                row.append(round(float(scores.get(t, 0)), 4))
            row += [
                str(regime),
                rf.get("current_vix", ""),
                rf.get("vol_scaling_triggered", ""),
                rf.get("vix_emergency_triggered", ""),
                rf.get("dd_circuit_triggered", ""),
                rf.get("beta_adjusted", ""),
            ]
            ws.append(row)

    def _write_config_flat(self, ws, cfg: dict, prefix: str = ""):
        """Flatten nested config dict into key-value rows."""
        for k, v in cfg.items():
            key = f"{prefix}.{k}" if prefix else k
            if isinstance(v, dict):
                self._write_config_flat(ws, v, key)
            elif isinstance(v, list):
                ws.append([key, json.dumps(v)])
            else:
                ws.append([key, str(v)])

    # ──────────────────────────────────────────────────────────────
    #  Stock decomposition (explode subsector → individual stocks)
    # ──────────────────────────────────────────────────────────────

    def _write_stock_decomp_sheets(self, wb) -> None:
        """Add *_stock_decomp sheets for every subsector-columned sheet.

        For each of 7 sheets whose columns are subsector names, add a companion
        sheet with columns = ``{subsector}/{stock}`` (24 columns; ARM appears
        under both ai_gpu and logic_cpu) and values decomposed by the 80/15/5
        effective within-subsector weights (IPO-aware).
        """
        from semiconductor_strategy.data import universe as U
        from semiconductor_strategy.data.aiss_fetch_prices import load_prices_wide, first_available_dates

        ec = self.equity_curve
        idx = ec.index
        subs = [t for t in self.tickers if t in U.SUBSECTORS]
        if not subs:
            return

        # Load individual stock prices for the date range
        all_stk = U.all_tickers()
        stk_prices = load_prices_wide(all_stk, start=str(idx[0].date()),
                                      end=str(idx[-1].date()), auto_fetch=False)
        if stk_prices.empty:
            log.warning("[STOCK DECOMP] No individual stock prices available")
            return
        stk_prices = stk_prices.reindex(idx).ffill()

        # First-available dates for IPO-aware weights
        fa = first_available_dates(all_stk)

        # Build ordered decomp columns: [(subsector, stock), ...]
        decomp_cols = []
        for s in subs:
            for t in U.subsector_tickers(s):
                decomp_cols.append((s, t))
        col_labels = [f"{s}/{t}" for s, t in decomp_cols]

        # Precompute within-subsector weight matrix (date × decomp_col)
        # effective_weights changes at most a few times per subsector (IPO thresholds)
        within_w = pd.DataFrame(0.0, index=idx, columns=col_labels)
        for s in subs:
            tickers_s = U.subsector_tickers(s)
            # Find breakpoint dates (when effective_weights changes)
            prev_ew = None
            for dt in idx:
                ew = U.effective_weights(s, dt.date(), first_available=fa)
                if ew != prev_ew:
                    for t in tickers_s:
                        within_w.loc[dt:, f"{s}/{t}"] = ew.get(t, 0.0)
                    prev_ew = ew

        # Subsector-level data from self
        subsector_prices = self.prices[subs].reindex(idx).ffill()
        wh = self.weights_history.reindex(columns=subs, fill_value=0.0)
        daily_w = _reconstruct_daily_weights(wh, subsector_prices, ec)
        daily_sh = _compute_daily_shares(wh, ec, subsector_prices)
        daily_cb = _compute_cost_basis(wh, subsector_prices, ec)
        etf_ret = subsector_prices.pct_change().fillna(0)
        sector_daily_pnl = pd.DataFrame(0.0, index=idx, columns=subs)
        for t in subs:
            sector_daily_pnl[t] = daily_w[t].shift(1).fillna(0) * etf_ret[t] * ec
        total_daily = sector_daily_pnl.sum(axis=1).replace(0, np.nan)
        sector_contrib = sector_daily_pnl.div(total_daily, axis=0).fillna(0)

        def _expand(subsector_df, method="multiply"):
            """Expand a (date × subsector) DataFrame to (date × stock) via within_w."""
            out = pd.DataFrame(0.0, index=idx, columns=col_labels)
            for s, t in decomp_cols:
                lbl = f"{s}/{t}"
                if method == "multiply":
                    out[lbl] = subsector_df[s] * within_w[lbl]
                elif method == "price":
                    out[lbl] = stk_prices[t] if t in stk_prices else 0.0
            return out

        # 1. sector_prices → actual stock prices
        ws = wb.create_sheet("sector_prices_stock_decomp")
        sp = _expand(subsector_prices, method="price")
        ws.append(["Date"] + col_labels)
        for dt in idx:
            ws.append([str(dt.date())] + [round(float(sp.loc[dt, c]), 2) for c in col_labels])

        # 2. sector_weights → per-stock weights (W_subsector × within_w)
        ws = wb.create_sheet("sector_weights_stock_decomp")
        sw = pd.DataFrame(0.0, index=wh.index, columns=col_labels)
        for s, t in decomp_cols:
            lbl = f"{s}/{t}"
            for dt in wh.index:
                ew = U.effective_weights(s, dt.date(), first_available=fa)
                sw.loc[dt, lbl] = float(wh.loc[dt].get(s, 0.0)) * ew.get(t, 0.0)
        ws.append(["Date"] + col_labels)
        for dt in wh.index:
            ws.append([str(dt.date())] + [round(float(sw.loc[dt, c]), 4) for c in col_labels])

        # 3. sector_weight_pct → per-stock drifted weights
        ws = wb.create_sheet("sector_wt_pct_stock_decomp")
        swp = _expand(daily_w)
        ws.append(["Date"] + col_labels)
        for dt in idx:
            ws.append([str(dt.date())] + [round(float(swp.loc[dt, c]), 4) for c in col_labels])

        # 4. share_history → per-stock shares
        ws = wb.create_sheet("share_hist_stock_decomp")
        shs = pd.DataFrame(0.0, index=idx, columns=col_labels)
        for s, t in decomp_cols:
            lbl = f"{s}/{t}"
            price_t = stk_prices[t] if t in stk_prices else pd.Series(1.0, index=idx)
            shs[lbl] = (daily_sh[s] * subsector_prices[s] * within_w[lbl] / price_t.replace(0, np.nan)).fillna(0)
        ws.append(["Date"] + col_labels)
        for dt in idx:
            ws.append([str(dt.date())] + [round(float(shs.loc[dt, c]), 1) for c in col_labels])

        # 5. cost_basis → per-stock cost (stock price at subsector entry dates)
        ws = wb.create_sheet("cost_basis_stock_decomp")
        cb = _expand(daily_cb, method="price")  # show raw stock price as cost proxy
        ws.append(["Date"] + col_labels)
        for dt in idx:
            ws.append([str(dt.date())] + [round(float(cb.loc[dt, c]), 2) for c in col_labels])

        # 6. sector_pnl_daily → per-stock PnL attribution
        ws = wb.create_sheet("sector_pnl_stock_decomp")
        spnl = _expand(sector_daily_pnl)
        ws.append(["Date"] + col_labels)
        for dt in idx:
            ws.append([str(dt.date())] + [round(float(spnl.loc[dt, c]), 2) for c in col_labels])

        # 7. sector_contribution → per-stock contribution %
        ws = wb.create_sheet("sector_contrib_stock_decomp")
        sc = _expand(sector_contrib)
        ws.append(["Date"] + col_labels)
        for dt in idx:
            ws.append([str(dt.date())] + [round(float(sc.loc[dt, c]), 4) for c in col_labels])

        log.info(f"[STOCK DECOMP] Added 7 stock decomposition sheets ({len(col_labels)} columns each)")

    # ──────────────────────────────────────────────────────────────
    #  Monitor Excel (5 sheets — daily rebalance snapshot)
    # ──────────────────────────────────────────────────────────────

    def export_monitor_excel(
        self,
        signal_date,
        report_data: dict,
        output_path: Path = None,
    ) -> Path:
        """Export daily monitor snapshot (5 sheets)."""
        if output_path is None:
            _HIST_DIR.mkdir(parents=True, exist_ok=True)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            ver = self.signal_version
            output_path = _HIST_DIR / f"monitor_aiss_{self.param_set}_{ver}_daily_{ts}.xlsx"

        from openpyxl import Workbook
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # 1. snapshot
        ws = wb.create_sheet("snapshot")
        regime = report_data.get("regime", {})
        ws.append(["Field", "Value"])
        ws.append(["date", str(signal_date)])
        ws.append(["param_set", self.param_set])
        ws.append(["signal_version", self.signal_version])
        ws.append(["regime", regime.get("label", "")])
        ws.append(["vix", regime.get("vix", "")])
        ws.append(["cash_weight", report_data.get("cash_weight", "")])
        ws.append(["n_positions", report_data.get("holdings_summary", {}).get("n_positions", "")])
        ws.append(["rebalance", report_data.get("rebalance_decision", {}).get("rebalance", "")])
        ws.append(["reason", report_data.get("rebalance_decision", {}).get("reason", "")])

        # 2. holdings
        ws = wb.create_sheet("holdings")
        ws.append(["Sector", "Weight", "Price", "Composite_Score", "Action"])
        for sig in report_data.get("signals", []):
            ws.append([
                sig.get("ticker", ""),
                round(sig.get("target_weight", 0), 4),
                round(sig.get("price", 0), 4),
                round(sig.get("composite_score", 0), 4),
                sig.get("action", ""),
            ])

        # 3. signals
        ws = wb.create_sheet("signals")
        ws.append(["Sector", "CS_Mom", "TS_Mult", "Composite"])
        for sig in report_data.get("signals", []):
            ws.append([
                sig.get("ticker", ""),
                round(sig.get("cs_mom", 0), 4),
                round(sig.get("ts_mult", 0), 4),
                round(sig.get("composite_score", 0), 4),
            ])

        # 4. smart_select
        ws = wb.create_sheet("smart_select")
        ss = report_data.get("smart_select", {})
        ws.append(["Field", "Value"])
        for k, v in ss.items():
            ws.append([k, str(v) if isinstance(v, (dict, list)) else v])

        # 5. risk_flags
        ws = wb.create_sheet("risk_flags")
        ws.append(["Field", "Value"])
        rf = report_data.get("risk_flags", "")
        ws.append(["risk_flags", str(rf)])

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        log.info(f"[MONITOR] Excel (5 sheets) → {output_path}")
        return output_path


# ═══════════════════════════════════════════════════════════════════════════
#  WF Diagnostic Excel
# ═══════════════════════════════════════════════════════════════════════════

def export_wf_diagnostic_excel(
    wf_result,
    output_path: Path = None,
    mode: str = "wf",
    signal_version: str = "v1",
) -> Path:
    """
    Export WF diagnostic Excel (5 sheets).

    Parameters
    ----------
    mode : entry point — "wf" | "select" | "tearsheet" | "batch_oos"
    signal_version : "v1" | "v2"
    """
    if output_path is None:
        _HIST_DIR.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        wf_mode = wf_result.mode if hasattr(wf_result, 'mode') else "anchored"
        output_path = _HIST_DIR / f"wf_diagnostic_aiss_{signal_version}_IS-OOS_{wf_mode}_{mode}_{ts}.xlsx"

    from openpyxl import Workbook
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    folds = wf_result.folds

    # 1. fold_summary
    ws = wb.create_sheet("fold_summary")
    ws.append(["Fold", "IS_Start", "IS_End", "OOS_Start", "OOS_End",
               "Selected", "Method", "IS_Sharpe", "OOS_Sharpe", "WFE", "OOS_Regime"])
    for fr in folds:
        ws.append([
            fr.fold.fold_id,
            str(fr.fold.is_start.date()),
            str(fr.fold.is_end.date()),
            str(fr.fold.oos_start.date()),
            str(fr.fold.oos_end.date()),
            fr.is_best_name,
            fr.selection_method,
            round(fr.is_best_sharpe, 4) if not np.isnan(fr.is_best_sharpe) else "",
            round(fr.oos_metrics.get("sharpe", float("nan")), 4)
            if not np.isnan(fr.oos_metrics.get("sharpe", float("nan"))) else "",
            round(fr.wfe, 4) if not np.isnan(fr.wfe) else "",
            fr.oos_regime,
        ])

    # 2. param_oos_matrix (59 params × N folds)
    ws = wb.create_sheet("param_oos_matrix")
    all_params = sorted({p for fr in folds for p in fr.all_oos_sharpes})
    ws.append(["Param"] + [f"Fold_{fr.fold.fold_id}" for fr in folds])
    for param in all_params:
        row = [param]
        for fr in folds:
            v = fr.all_oos_sharpes.get(param, float("nan"))
            row.append(round(v, 4) if not np.isnan(v) else "")
        ws.append(row)

    # 3. param_by_regime
    ws = wb.create_sheet("param_by_regime")
    regime_data = wf_result.param_oos_by_regime()
    ws.append(["Param", "Regime", "Mean_OOS_Sharpe", "N_Folds"])
    for param, regimes in sorted(regime_data.items()):
        for regime, stats in sorted(regimes.items()):
            ws.append([param, regime, stats["mean_oos_sharpe"], stats["n_folds"]])

    # 4. synthetic_equity (+ oracle / static-best columns when available)
    ws = wb.create_sheet("synthetic_equity")
    se = wf_result.synthetic_equity
    oracle_eq = getattr(wf_result, "oracle_equity", None)
    static_eq = getattr(wf_result, "static_best_equity", None)
    _orc = oracle_eq.reindex(se.index) if oracle_eq is not None and not oracle_eq.empty else None
    _sta = static_eq.reindex(se.index) if static_eq is not None and not static_eq.empty else None
    ws.append(["Date", "Synthetic_Value", "Oracle_Value", "Static_Best_Value"])
    for dt in se.index:
        ov = "" if _orc is None or pd.isna(_orc.get(dt)) else round(float(_orc[dt]), 4)
        sv = "" if _sta is None or pd.isna(_sta.get(dt)) else round(float(_sta[dt]), 4)
        ws.append([str(dt.date()), round(float(se[dt]), 4), ov, sv])

    # 5. selection_log
    ws = wb.create_sheet("selection_log")
    if wf_result.selection_log:
        keys = list(wf_result.selection_log[0].keys())
        ws.append(keys)
        for entry in wf_result.selection_log:
            ws.append([str(entry.get(k, "")) for k in keys])

    # 6. oracle_vs_realized — 3-layer comparison + per-fold regret (when available)
    cmp = getattr(wf_result, "comparison", None)
    if cmp:
        ws = wb.create_sheet("oracle_vs_realized")
        ws.append(["Layer", "Sharpe", "CAGR", "MaxDD", "Calmar", "Note"])
        _note = {
            "static_best": f"single full-period best param [{getattr(wf_result,'static_best_name','')}] — IS/full (optimistic)",
            "synthetic":   "realizable dynamic selection (OOS)",
            "oracle":      "theoretical ceiling: best-OOS param per fold (hindsight)",
        }
        for layer in ("static_best", "synthetic", "oracle"):
            d = cmp.get(layer, {})
            def _f(x):
                return "" if x is None or (isinstance(x, float) and np.isnan(x)) else round(float(x), 4)
            ws.append([layer, _f(d.get("sharpe")), _f(d.get("cagr")),
                       _f(d.get("maxdd")), _f(d.get("calmar")), _note.get(layer, "")])
        ws.append([])
        ws.append(["capture_ratio_sharpe", cmp.get("capture_ratio_sharpe")])
        ws.append(["capture_ratio_cagr", cmp.get("capture_ratio_cagr")])
        ws.append(["mean_regret_sharpe", cmp.get("mean_regret_sharpe")])
        ws.append(["n_folds_optimal", cmp.get("n_folds_optimal"), f"of {len(folds)} folds"])
        ws.append([])
        olog = getattr(wf_result, "oracle_selection_log", []) or []
        if olog:
            ws.append(["Fold", "OOS_Start", "OOS_End", "Oracle_Param", "Oracle_OOS_SR",
                       "Selected_Param", "Selected_OOS_SR", "Regret", "Optimal"])
            for e in olog:
                ws.append([e.get("fold"), e.get("oos_start"), e.get("oos_end"),
                           e.get("oracle_param"), e.get("oracle_oos_sharpe"),
                           e.get("selected_param"), e.get("selected_oos_sharpe"),
                           e.get("regret"), "Y" if e.get("optimal") else ""])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    log.info(f"[WF DIAGNOSTIC] Excel (5 sheets) → {output_path}")
    return output_path
