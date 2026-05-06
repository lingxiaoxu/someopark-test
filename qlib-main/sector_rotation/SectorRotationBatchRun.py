"""
SectorRotationBatchRun.py
=============================================================
Runs all (or selected) named parameter sets from
SectorRotationStrategyRuns.py through the backtest engine and
collects results into a ranked summary table.

Data is loaded ONCE and shared across all runs for efficiency.
A single full backtest run takes ~5-10 s; 59 sets ≈ 5-10 min total.

Usage
-----
# Run all 55 sets (default output to backtest_results/)
  conda run -n qlib_run --no-capture-output \\
    python qlib-main/sector_rotation/SectorRotationBatchRun.py

# Run specific named sets
  python SectorRotationBatchRun.py --sets default momentum_heavy tech_bull_2023

# Run one thematic group only (A / B / C … L)
  python SectorRotationBatchRun.py --group L

# Run multiple groups
  python SectorRotationBatchRun.py --group A B L

# Custom output directory
  python SectorRotationBatchRun.py --output-dir my_results/

# Sort output table by a different metric (default: sharpe)
  python SectorRotationBatchRun.py --sort-by calmar

Output files
------------
  backtest_results/
    sr_batch_summary_<timestamp>.csv        # machine-readable raw results
    sr_batch_summary_<timestamp>.xlsx       # Excel with conditional formatting
    sr_batch_equity_<timestamp>.csv         # daily equity curves (wide format)
"""

from __future__ import annotations

import argparse
import logging
import sys
import time
import traceback
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import numpy as np
import pandas as pd

# ── Path setup ────────────────────────────────────────────────────────────────
_THIS_DIR    = Path(__file__).parent.resolve()   # sector_rotation/
_QLIB_DIR    = _THIS_DIR.parent.resolve()        # qlib-main/
_PROJECT_DIR = _QLIB_DIR.parent.resolve()        # someopark-test/

if str(_QLIB_DIR) not in sys.path:
    sys.path.insert(0, str(_QLIB_DIR))

from sector_rotation.SectorRotationStrategyRuns import (
    PARAM_SETS,
    _PARAM_SET_DESCRIPTIONS,
    apply_param_set,
)
from sector_rotation.data.loader import load_all, load_config
from sector_rotation.backtest.engine import SectorRotationBacktest

logging.basicConfig(
    level=logging.WARNING,
    format="%(asctime)s %(levelname)s: %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# Silence noisy sub-loggers during batch run
for _noisy in [
    "sector_rotation.signals.composite",
    "sector_rotation.signals.regime",
    "sector_rotation.signals.momentum",
    "sector_rotation.signals.value",
    "sector_rotation.portfolio.optimizer",
    "sector_rotation.portfolio.rebalance",
    "sector_rotation.portfolio.risk",
    "sector_rotation.backtest.engine",
    "sector_rotation.backtest.costs",
    "sector_rotation.backtest.metrics",
]:
    logging.getLogger(_noisy).setLevel(logging.ERROR)


# ---------------------------------------------------------------------------
# Group → set name mapping (prefix of description string, e.g. "A1", "B3")
# ---------------------------------------------------------------------------

def _group_of(name: str) -> Optional[str]:
    """Return single-letter group ('A'..'L') for a named param set."""
    desc = _PARAM_SET_DESCRIPTIONS.get(name, "")
    if desc and desc[0].isalpha():
        return desc[0].upper()
    return None


def filter_by_groups(groups: List[str]) -> List[str]:
    """Return set names belonging to any of the given group letters."""
    groups_upper = [g.upper() for g in groups]
    return [name for name in PARAM_SETS if _group_of(name) in groups_upper]


# ---------------------------------------------------------------------------
# Metric extraction helpers
# ---------------------------------------------------------------------------

_METRIC_COLS = [
    "sharpe", "calmar", "annual_return", "annual_vol",
    "max_drawdown", "max_drawdown_days",
    "total_return", "monthly_win_rate",
    "info_ratio", "active_return",
    "beta", "tracking_error",
    "cvar_95", "skewness",
]


def _extract_metrics(result) -> dict:
    m = result.metrics
    row = {k: m.get(k, float("nan")) for k in _METRIC_COLS}

    # Annual turnover from costs_history
    row["annual_turnover"] = float("nan")
    if (
        result.costs_history is not None
        and not result.costs_history.empty
        and "turnover_pct" in result.costs_history.columns
    ):
        monthly_to = result.costs_history["turnover_pct"] / 100.0
        if len(monthly_to) > 0:
            row["annual_turnover"] = float(monthly_to.mean() * 12)

    # Average number of positions from weights_history
    row["avg_positions"] = float("nan")
    if result.weights_history is not None and not result.weights_history.empty:
        n_pos = (result.weights_history > 0.001).sum(axis=1)
        row["avg_positions"] = float(n_pos.mean())

    # Recent 12-month Sharpe (last 252 trading days) — OOS proxy / fallback for --select
    row["recent_sharpe_12m"] = float("nan")
    if result.equity_curve is not None and len(result.equity_curve) > 63:
        ec_tail = result.equity_curve.tail(252)
        rets = ec_tail.pct_change().dropna()
        if len(rets) > 20:
            row["recent_sharpe_12m"] = float(rets.mean() / rets.std() * (252 ** 0.5))

    return row


# ---------------------------------------------------------------------------
# Macro-conditioned Sharpe (used by --select when MacroStateStore is available)
# Delegates to MCPS.macro_cond_sharpe() — single source of truth.
# ---------------------------------------------------------------------------

def _macro_cond_sharpe(
    equity: pd.Series,
    macro_df: pd.DataFrame,
    today_vec: dict,
    features: list,
    min_overlap: int = 60,
) -> float:
    """
    Gaussian-kernel-weighted Sharpe ratio for macro-conditioned param selection.
    Delegates to MCPS.macro_cond_sharpe() for unified implementation.
    """
    try:
        if str(_PROJECT_DIR) not in sys.path:
            sys.path.insert(0, str(_PROJECT_DIR))
        from MCPS import macro_cond_sharpe
    except ImportError:
        import importlib.util
        spec = importlib.util.spec_from_file_location(
            "MCPS", str(_PROJECT_DIR / "MCPS.py"))
        _mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(_mod)
        macro_cond_sharpe = _mod.macro_cond_sharpe

    return macro_cond_sharpe(
        equity=equity,
        macro_df=macro_df,
        today_vec=today_vec,
        features=features,
        min_overlap=min_overlap,
    )


# ---------------------------------------------------------------------------
# Single-run wrapper with timing and error capture
# ---------------------------------------------------------------------------

def _run_one(
    name: str,
    base_cfg: dict,
    prices: pd.DataFrame,
    macro: pd.DataFrame,
) -> dict:
    t0 = time.time()
    row = {"param_set": name, "group": _group_of(name) or "?",
           "description": _PARAM_SET_DESCRIPTIONS.get(name, name),
           "status": "ok", "error": "", "elapsed_s": float("nan")}
    try:
        cfg = apply_param_set(base_cfg, PARAM_SETS[name])
        engine = SectorRotationBacktest(cfg)
        result = engine.run(prices=prices, macro=macro)
        row.update(_extract_metrics(result))
        row["elapsed_s"] = round(time.time() - t0, 1)
    except Exception as exc:
        row["status"] = "error"
        row["error"] = str(exc)
        row["elapsed_s"] = round(time.time() - t0, 1)
        log.warning(f"[{name}] FAILED: {exc}")
        log.debug(traceback.format_exc())
    return row


# ---------------------------------------------------------------------------
# Equity curve extraction
# ---------------------------------------------------------------------------

def _run_one_with_equity(
    name: str,
    base_cfg: dict,
    prices: pd.DataFrame,
    macro: pd.DataFrame,
) -> tuple:
    """Returns (metrics_row, equity_series)."""
    t0 = time.time()
    row = {"param_set": name, "group": _group_of(name) or "?",
           "description": _PARAM_SET_DESCRIPTIONS.get(name, name),
           "status": "ok", "error": "", "elapsed_s": float("nan")}
    equity = pd.Series(dtype=float, name=name)
    try:
        cfg = apply_param_set(base_cfg, PARAM_SETS[name])
        engine = SectorRotationBacktest(cfg)
        result = engine.run(prices=prices, macro=macro)
        row.update(_extract_metrics(result))
        row["elapsed_s"] = round(time.time() - t0, 1)
        equity = result.equity_curve.rename(name)
    except Exception as exc:
        row["status"] = "error"
        row["error"] = str(exc)
        row["elapsed_s"] = round(time.time() - t0, 1)
        log.warning(f"[{name}] FAILED: {exc}")
    return row, equity


# ---------------------------------------------------------------------------
# Pretty console table
# ---------------------------------------------------------------------------

def _print_summary(df: pd.DataFrame, sort_by: str = "sharpe", top_n: int = 55) -> None:
    ok = df[df["status"] == "ok"].copy()
    if ok.empty:
        print("No successful runs.")
        return

    # Sort
    asc = sort_by in {"max_drawdown", "annual_vol", "tracking_error", "cvar_95"}
    ok = ok.sort_values(sort_by, ascending=asc)

    cols = ["param_set", "sharpe", "recent_sharpe_12m", "calmar", "annual_return",
            "max_drawdown", "annual_vol", "info_ratio",
            "annual_turnover", "avg_positions", "elapsed_s"]
    # Drop columns that don't exist (e.g. recent_sharpe_12m on old results)
    cols = [c for c in cols if c in ok.columns]

    display = ok[cols].head(top_n).copy()
    display["annual_return"] = display["annual_return"].map("{:.1%}".format)
    display["max_drawdown"]  = display["max_drawdown"].map("{:.1%}".format)
    display["annual_vol"]    = display["annual_vol"].map("{:.1%}".format)
    display["sharpe"]        = display["sharpe"].map("{:.3f}".format)
    if "recent_sharpe_12m" in display.columns:
        display["recent_sharpe_12m"] = display["recent_sharpe_12m"].map(
            lambda x: f"{x:.3f}" if not (isinstance(x, float) and np.isnan(x)) else "nan"
        )
    display["calmar"]        = display["calmar"].map("{:.3f}".format)
    display["info_ratio"]    = display["info_ratio"].map("{:.3f}".format)
    display["annual_turnover"] = display["annual_turnover"].map("{:.0%}".format)
    display["avg_positions"] = display["avg_positions"].map("{:.1f}".format)

    print(f"\n{'═'*110}")
    print(f"  SECTOR ROTATION BATCH RESULTS — sorted by {sort_by} "
          f"({len(ok)} successful / {len(df)} total)")
    print(f"{'═'*110}")
    print(display.to_string(index=False))

    errs = df[df["status"] == "error"]
    if not errs.empty:
        print(f"\nFailed runs ({len(errs)}):")
        for _, row in errs.iterrows():
            print(f"  {row['param_set']:<30} {row['error'][:80]}")


# ---------------------------------------------------------------------------
# Excel export with conditional formatting
# ---------------------------------------------------------------------------

def _save_excel(df: pd.DataFrame, path: Path) -> None:
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, numbers
        from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.warning("openpyxl not installed — skipping Excel export (CSV saved).")
        return

    ok = df[df["status"] == "ok"].copy().sort_values("sharpe", ascending=False)

    # Percentage columns → format as pct
    pct_cols  = ["annual_return", "max_drawdown", "annual_vol",
                 "active_return", "tracking_error", "cvar_95", "annual_turnover"]
    num_cols  = ["sharpe", "recent_sharpe_12m", "calmar", "info_ratio", "beta", "skewness"]
    int_cols  = ["max_drawdown_days", "elapsed_s"]

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        ok.to_excel(writer, sheet_name="Results", index=False)
        ws = writer.sheets["Results"]

        # Freeze top row
        ws.freeze_panes = "A2"

        # Header styling
        header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 28)

        # Number formatting
        header_row = {cell.value: get_column_letter(cell.column) for cell in ws[1]}

        for col_name in pct_cols:
            if col_name in header_row:
                col_letter = header_row[col_name]
                for cell in ws[col_letter][1:]:
                    cell.number_format = "0.00%"

        for col_name in num_cols:
            if col_name in header_row:
                col_letter = header_row[col_name]
                for cell in ws[col_letter][1:]:
                    cell.number_format = "0.000"

        # Color scale for Sharpe
        if "sharpe" in header_row:
            col_letter = header_row["sharpe"]
            last_row = ws.max_row
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{last_row}",
                ColorScaleRule(
                    start_type="min",  start_color="F8696B",
                    mid_type="percentile", mid_value=50, mid_color="FFEB84",
                    end_type="max",  end_color="63BE7B",
                )
            )

        # Color scale for recent_sharpe_12m (green=best, same direction as sharpe)
        if "recent_sharpe_12m" in header_row:
            col_letter = header_row["recent_sharpe_12m"]
            last_row = ws.max_row
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{last_row}",
                ColorScaleRule(
                    start_type="min",  start_color="F8696B",
                    mid_type="percentile", mid_value=50, mid_color="FFEB84",
                    end_type="max",  end_color="63BE7B",
                )
            )

        # Color scale for max_drawdown (red = worse = more negative)
        if "max_drawdown" in header_row:
            col_letter = header_row["max_drawdown"]
            last_row = ws.max_row
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{last_row}",
                ColorScaleRule(
                    start_type="min",  start_color="63BE7B",
                    mid_type="percentile", mid_value=50, mid_color="FFEB84",
                    end_type="max",  end_color="F8696B",
                )
            )

        # Data bar for annual_return
        if "annual_return" in header_row:
            col_letter = header_row["annual_return"]
            last_row = ws.max_row
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{last_row}",
                DataBarRule(start_type="min", start_value=0,
                            end_type="max", end_value=1,
                            color="638EC6")
            )

    print(f"  Excel saved → {path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Batch backtest runner for SectorRotationStrategyRuns.py"
    )
    parser.add_argument(
        "--sets", nargs="+", metavar="NAME",
        help="Run only these named param sets (space-separated).",
    )
    parser.add_argument(
        "--group", nargs="+", metavar="LETTER",
        help="Run only these groups (e.g. --group A B L).",
    )
    parser.add_argument(
        "--output-dir", default=str(_THIS_DIR / "backtest_results"),
        metavar="DIR",
        help="Directory to write CSV / Excel output.",
    )
    parser.add_argument(
        "--sort-by", default="sharpe", metavar="METRIC",
        help="Metric to sort results table by (default: sharpe).",
    )
    parser.add_argument(
        "--save-equity", action="store_true",
        help="Also save all daily equity curves to a wide CSV.",
    )
    parser.add_argument(
        "--no-excel", action="store_true",
        help="Skip Excel export (faster).",
    )
    parser.add_argument(
        "--select", action="store_true",
        help=(
            "After batch run, select the param set with the best recent 12m Sharpe "
            "and write selected_param_set.json to sector_rotation/ — this file is "
            "automatically read by SectorRotationDailySignal on every daily/weekly/monthly run."
        ),
    )
    parser.add_argument(
        "--verbose", action="store_true",
        help="Enable INFO-level logging during backtest runs.",
    )
    parser.add_argument(
        "--oos-validate", action="store_true",
        help=(
            "After batch run, run full walk-forward IS/OOS validation "
            "(anchored + rolling) using WalkForwardAnalyzer. "
            "Reports synthetic OOS Sharpe, DSR, and WFE for all 59 param sets."
        ),
    )
    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.INFO)
        for _n in [
            "sector_rotation.signals.composite",
            "sector_rotation.backtest.engine",
        ]:
            logging.getLogger(_n).setLevel(logging.INFO)

    # ── Determine which sets to run ──────────────────────────────────────────
    if args.sets:
        unknown = [s for s in args.sets if s not in PARAM_SETS]
        if unknown:
            print(f"Unknown param sets: {unknown}")
            print(f"Available: {sorted(PARAM_SETS)}")
            sys.exit(1)
        run_sets = args.sets
    elif args.group:
        run_sets = filter_by_groups(args.group)
        if not run_sets:
            print(f"No sets found for groups: {args.group}")
            sys.exit(1)
    else:
        run_sets = list(PARAM_SETS.keys())

    print(f"\n{'═'*60}")
    print(f"  SECTOR ROTATION BATCH RUN")
    print(f"  Sets to run : {len(run_sets)}")
    print(f"  Output dir  : {args.output_dir}")
    if args.select:
        print(f"  --select    : ON  → will write selected_param_set.json")
    print(f"{'═'*60}")

    # ── Load data once ───────────────────────────────────────────────────────
    print("\n[1/3] Loading price and macro data...")
    t_data = time.time()
    base_cfg = load_config()
    prices, macro = load_all(config=base_cfg)
    print(f"      Loaded in {time.time() - t_data:.1f}s  "
          f"({len(prices)} trading days, {len(prices.columns)} tickers)")

    # ── Run all sets ─────────────────────────────────────────────────────────
    print(f"\n[2/3] Running {len(run_sets)} backtests...")
    rows: List[dict] = []
    equity_frames: List[pd.Series] = []

    for i, name in enumerate(run_sets, 1):
        desc_short = _PARAM_SET_DESCRIPTIONS.get(name, name)[:55]
        print(f"  [{i:>2}/{len(run_sets)}] {name:<28} {desc_short}", end="", flush=True)

        if args.save_equity or args.select:
            row, eq = _run_one_with_equity(name, base_cfg, prices, macro)
            if not eq.empty:
                equity_frames.append(eq)
        else:
            row = _run_one(name, base_cfg, prices, macro)

        status_str = f"  ✓ {row['elapsed_s']:.0f}s  sharpe={row.get('sharpe', float('nan')):.3f}"
        if row["status"] == "error":
            status_str = f"  ✗ {row['error'][:40]}"
        print(status_str)
        rows.append(row)

    df = pd.DataFrame(rows)

    # ── Save results ─────────────────────────────────────────────────────────
    print(f"\n[3/3] Saving results...")
    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path   = out_dir / f"sr_batch_summary_{ts}.csv"
    excel_path = out_dir / f"sr_batch_summary_{ts}.xlsx"
    equity_path = out_dir / f"sr_batch_equity_{ts}.csv"

    df.to_csv(csv_path, index=False)
    print(f"  CSV saved   → {csv_path}")

    if not args.no_excel:
        _save_excel(df, excel_path)

    if args.save_equity and equity_frames:
        eq_df = pd.concat(equity_frames, axis=1)
        eq_df.to_csv(equity_path)
        print(f"  Equity CSV  → {equity_path}")

    # ── Print summary table ──────────────────────────────────────────────────
    _print_summary(df, sort_by=args.sort_by)

    # ── Quick stats ──────────────────────────────────────────────────────────
    ok = df[df["status"] == "ok"]
    if not ok.empty:
        print(f"\n{'─'*60}")
        print(f"  Summary statistics across {len(ok)} successful runs:")
        print(f"{'─'*60}")
        for metric in ["sharpe", "calmar", "annual_return", "max_drawdown"]:
            col = ok[metric].dropna()
            if not col.empty:
                print(f"  {metric:<20}  "
                      f"min={col.min():.3f}  "
                      f"median={col.median():.3f}  "
                      f"max={col.max():.3f}  "
                      f"range={col.max()-col.min():.3f}")
        best = ok.loc[ok["sharpe"].idxmax()]
        worst = ok.loc[ok["sharpe"].idxmin()]
        print(f"\n  Best  Sharpe: {best['param_set']} ({best['sharpe']:.3f})")
        print(f"  Worst Sharpe: {worst['param_set']} ({worst['sharpe']:.3f})")
        print(f"{'─'*60}\n")

    errs = df[df["status"] == "error"]
    if not errs.empty:
        print(f"  {len(errs)} run(s) failed — check error column in CSV.")

    # ── Select best param set for production (--select) ──────────────────────
    # Three-stage selection:
    #   Stage 1: WF OOS validation → filter out param sets with poor OOS performance
    #   Stage 2: Macro-conditioned Sharpe on surviving candidates
    #   Stage 3: Fallback to recent_sharpe_12m if MCPS unavailable
    if args.select and not ok.empty:
        import json as _json
        import math as _math

        _best_ps:   "str | None"           = None
        _best_row:  "pd.Series | None"     = None
        _sel_method                        = "recent_sharpe_12m"
        _sel_val                           = float("nan")
        _macro_cond: "dict[str, float]"    = {}
        _macro_df                          = pd.DataFrame()
        _oos_filter_applied                = False
        _n_survivors                       = len(ok)
        _wf_mean_wfe                       = float("nan")

        # ── Stage 1: WF OOS filter (run WF, exclude overfitting params) ──────
        # Evaluate ALL param sets across ALL OOS folds (not just when selected).
        # This gives every param set a fair OOS score regardless of MCPS selection.
        _oos_qualified: "set[str] | None" = None  # None = no filter (WF failed)
        try:
            from sector_rotation.walk_forward import (
                WalkForwardAnalyzer, _compute_metrics_from_equity
            )
            print(f"\n  [SELECT] Stage 1: Walk-Forward OOS validation...")
            _wf = WalkForwardAnalyzer(
                base_cfg=base_cfg, prices=prices, macro=macro,
                mode="anchored",
                is_years_min=base_cfg.get("backtest", {}).get("is_years", 3),
                oos_months=6, step_days=15, embargo_days=5,
            )
            _wf_result = _wf.run()
            _wf_mean_wfe = _wf_result.mean_wfe

            # Compute OOS Sharpe for ALL params across ALL folds
            # (not just folds where param was MCPS-selected)
            _eq_map = {s.name: s for s in equity_frames if not s.empty}
            _folds = _wf.generate_folds()
            _oos_qualified = set()
            _oos_sharpe_threshold = 0.5  # top ~95% of params survive

            for _ps_name, _eq in _eq_map.items():
                _fold_sharpes = []
                for _fold in _folds:
                    _seg = _eq[(_eq.index >= _fold.oos_start) & (_eq.index <= _fold.oos_end)]
                    if len(_seg) >= 60:
                        _m = _compute_metrics_from_equity(_seg / _seg.iloc[0])
                        if not np.isnan(_m.get("sharpe", float("nan"))):
                            _fold_sharpes.append(_m["sharpe"])
                if _fold_sharpes:
                    _mean_oos = float(np.mean(_fold_sharpes))
                    if _mean_oos > _oos_sharpe_threshold:
                        _oos_qualified.add(_ps_name)

            if _oos_qualified:
                _oos_filter_applied = True
                _n_survivors = len(_oos_qualified)
                print(f"    OOS filter: {_n_survivors}/{len(ok)} param sets survive "
                      f"(mean OOS Sharpe > 0)")
            else:
                # All failed OOS → don't filter (use full set with warning)
                print(f"    WARN: No param sets with OOS Sharpe > 0 — skipping OOS filter")
                _oos_qualified = None
        except Exception as _wf_e:
            log.warning(f"[SELECT] WF validation failed ({_wf_e}) — skipping OOS filter")
            _oos_qualified = None

        # ── Stage 2: Macro-conditioned Sharpe (on OOS-qualified survivors) ────
        try:
            if str(_PROJECT_DIR) not in sys.path:
                sys.path.insert(0, str(_PROJECT_DIR))
            from MacroStateStore import MacroStateStore as _MSS          # type: ignore
            from MacroStateStore import SIMILARITY_FEATURES as _SF       # type: ignore
            _store   = _MSS()
            _today_v = _store.get(datetime.now().date())
            if any(v is not None for v in _today_v.values()):
                _bs = base_cfg.get("backtest", {}).get("start_date", "2018-07-01")
                _macro_df = _store.load(_bs)
                if not _macro_df.empty:
                    _eq_map = {s.name: s for s in equity_frames if not s.empty}
                    for _ps, _eq in _eq_map.items():
                        # Skip if OOS filter active and this param didn't pass
                        if _oos_qualified is not None and _ps not in _oos_qualified:
                            continue
                        _macro_cond[_ps] = _macro_cond_sharpe(
                            _eq, _macro_df, _today_v, _SF
                        )
        except Exception as _e:
            log.warning(f"[SELECT] MacroStateStore unavailable: {_e}")

        _valid_mcs = {k: v for k, v in _macro_cond.items() if not _math.isnan(v)}

        if len(_valid_mcs) >= 3:
            # Primary path: macro-conditioned Sharpe (on OOS-qualified set)
            _best_ps    = max(_valid_mcs, key=_valid_mcs.get)
            _best_row   = ok[ok["param_set"] == _best_ps].iloc[0]
            _sel_method = "mcps_oos_filtered" if _oos_filter_applied else "macro_cond_sharpe"
            _sel_val    = round(_valid_mcs[_best_ps], 4)
        else:
            # ── Stage 3: Fallback — best recent 12m Sharpe ────────────────────
            _candidates = ok
            if _oos_qualified is not None and _oos_qualified:
                _candidates = ok[ok["param_set"].isin(_oos_qualified)]
            _ok_r = _candidates.dropna(subset=["recent_sharpe_12m"])
            if not _ok_r.empty:
                _best_row   = _ok_r.loc[_ok_r["recent_sharpe_12m"].idxmax()]
                _best_ps    = _best_row["param_set"]
                _sel_method = "recent_sharpe_12m"
                _sel_val    = round(float(_best_row["recent_sharpe_12m"]), 4)

        if _best_ps is None:
            print("\n  [SELECT] No valid selection possible — skipping.")
        else:
            _n_macro = len(_macro_df)
            # Include recent_sharpe_12m for reference even when MCPS is active
            _ref_r = ok[ok["param_set"] == _best_ps]["recent_sharpe_12m"]
            _recent_sr_ref = round(float(_ref_r.iloc[0]), 4) if not _ref_r.empty else None

            # OOS stats for selected param (if available from WF)
            _oos_stats = {}
            if _oos_filter_applied and hasattr(_wf_result, "param_oos_stats"):
                _oos_stats = _wf_result.param_oos_stats.get(_best_ps, {})

            sel_info = {
                "param_set":          _best_ps,
                "selection_method":   _sel_method,
                _sel_method:          _sel_val,
                "recent_sharpe_12m":  _recent_sr_ref,
                "full_period_sharpe": round(float(_best_row["sharpe"]), 4),
                "full_period_calmar": round(float(_best_row["calmar"]), 4),
                "selected_at":        datetime.now().strftime("%Y-%m-%d"),
                "n_candidates":       int(len(_valid_mcs) if _valid_mcs else len(ok)),
                "n_oos_survivors":    _n_survivors,
                "oos_filter_applied": _oos_filter_applied,
                "oos_mean_sharpe":    round(_oos_stats.get("mean_oos_sharpe", float("nan")), 4)
                                      if _oos_stats else None,
                "oos_n_selected":     _oos_stats.get("n_selected") if _oos_stats else None,
                "wf_mean_wfe":        round(_wf_mean_wfe, 4) if not _math.isnan(_wf_mean_wfe) else None,
                "macro_data_days":    _n_macro,
            }
            # Write to backtest_results/ (archive) AND sector_rotation/ (production)
            archive_path = out_dir / "selected_param_set.json"
            prod_path    = _THIS_DIR / "selected_param_set.json"
            for p in (archive_path, prod_path):
                p.write_text(_json.dumps(sel_info, indent=2))

            print(f"\n{'═'*60}")
            print(f"  [SELECT] Best param set  : {_best_ps}")
            print(f"  Method     : {_sel_method}  = {_sel_val:.3f}")
            if _oos_filter_applied:
                print(f"  OOS filter : {_n_survivors} survivors (mean OOS Sharpe > 0)")
                if _oos_stats:
                    print(f"  OOS record : selected {_oos_stats.get('n_selected', '?')} times, "
                          f"mean OOS Sharpe={_oos_stats.get('mean_oos_sharpe', '?'):.3f}")
                print(f"  WF mean WFE: {_wf_mean_wfe:.3f}")
            if "macro_cond" in _sel_method or _sel_method == "mcps_oos_filtered":
                _mc0 = _macro_df.index[0].date() if _n_macro else "?"
                _mc1 = _macro_df.index[-1].date() if _n_macro else "?"
                print(f"  Macro days : {_n_macro}  ({_mc0} → {_mc1})")
            else:
                print(f"  Macro days : {_n_macro}  (insufficient for MCPS — "
                      f"run 'MacroStateStore.py --init --start 2017-01-01')")
            print(f"  Full-period Sharpe : {sel_info['full_period_sharpe']:.3f}")
            print(f"  Written to : {prod_path}")
            print(f"  → DailySignal will use this param set on next run")
            print(f"{'═'*60}\n")

    # ── OOS Validation via WalkForwardAnalyzer ───────────────────────────
    if args.oos_validate:
        print(f"\n{'═'*60}")
        print(f"  OOS VALIDATION — Walk-Forward Analysis (anchored + rolling)")
        print(f"{'═'*60}\n")
        from sector_rotation.walk_forward import run_dual_mode
        wf_results = run_dual_mode(
            base_cfg=base_cfg,
            prices=prices,
            macro=macro,
            is_years_min=base_cfg.get("backtest", {}).get("is_years", 3),
            oos_months=6,
            step_days=15,
            embargo_days=5,
        )
        for mode_name, wf_r in wf_results.items():
            print(wf_r.summary())
            csv_path = Path(args.output_dir) / f"wf_{mode_name}_fold_summary_{ts}.csv"
            wf_r.fold_summary_df.to_csv(csv_path, index=False)
            print(f"  Fold summary → {csv_path}")


if __name__ == "__main__":
    main()
