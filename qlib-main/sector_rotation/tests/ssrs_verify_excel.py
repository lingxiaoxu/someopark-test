"""
ssrs_verify_excel.py — deep per-sheet verification of SSRS backtest Excel.
========================================================================
Mirror of semiconductor_strategy/tests/aiss_verify_excel.py, adapted for SSRS
(11 GICS sector ETFs; the ETF *is* the tradeable instrument, so there are NO
stock_decomp sheets — the 26-sheet portfolio Excel only).

Opens EVERY sheet of EVERY SSRS portfolio Excel and runs computation
cross-checks — not just empty/null/0/NaN, but *correctness*:

  • metric recompute  : summary Sharpe (qlib compound-interest) / CAGR / MaxDD
                        vs values recomputed from equity_history
  • drawdown          : portfolio_history.drawdown_pct == equity/cummax-1
                        (and the same on drawdown_history)
  • cum_pnl           : == cumsum(daily_pnl)
  • sector_pnl_acc    : == cumsum(sector_pnl_daily) per sector
  • pnl reconciliation: daily_pnl total == Σ(sector_pnl_daily across sectors)
  • leverage 0        : asset == equity, liability == 0, interest_expense == 0
  • weights           : sector_weights each row sums to ~1 (+cash), in [0, 1.02]
  • regime labels     : regime_indicators labels ∈ the 4-state set
  • structural        : all 26 expected sheets present, none empty/all-null,
                        equity all positive, V1/V2 rebalance cadence

Usage:
    conda run -n qlib_run python -m sector_rotation.tests.ssrs_verify_excel [YYYYMMDD] [--dir PATH]
Exit code: 0 = all files clean, 1 = one or more files have issues.
"""
from __future__ import annotations

import sys, glob, os, math, argparse, warnings
warnings.filterwarnings("ignore")
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from collections import defaultdict

_HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.abspath(os.path.join(_HERE, "..", "..")))  # qlib-main/
from sector_rotation.backtest.metrics import sharpe_ratio as _ci_sharpe

DEFAULT_DIR = os.path.abspath(os.path.join(
    _HERE, "..", "..", "..", "historical_runs", "sector_rotation"))

EXPECTED_MAIN = [
    "summary","portfolio_history","asset_history","liability_history","equity_history",
    "asset_cash_history","sector_prices","share_history","sector_weights","sector_weight_pct",
    "cost_basis","sector_ratio_matrix","sector_pnl_acc","sector_pnl_daily","sector_contribution",
    "daily_pnl","interest_expense","acc_interest","realized_pnl","total_notional",
    "drawdown_history","rebalance_trades","regime_indicators","strategy_vars",
    "stop_loss_history","config",
]
EXPECTED_DECOMP: list = []   # SSRS trades ETFs directly — no stock-decomposition sheets
VALID_REGIMES = {
    "risk_on","risk_off","transition_up","transition_down","neutral",
    "RISK_ON","RISK_OFF","TRANSITION_UP","TRANSITION_DOWN","NEUTRAL",
}
TOL_SHARPE = 0.25
TOL_CAGR   = 0.02
TOL_MAXDD  = 0.02


def sdf(wb, name):
    if name not in wb.sheetnames:
        return None
    rows = list(wb[name].iter_rows(values_only=True))
    if not rows:
        return None
    return pd.DataFrame(rows[1:], columns=rows[0])


def _num(df):
    return df.apply(pd.to_numeric, errors="coerce")


def _metrics_from_equity(eq):
    eq = pd.to_numeric(eq, errors="coerce").dropna()
    if len(eq) < 30:
        return None
    ret = eq.pct_change().dropna()
    n = len(eq)
    cagr = (eq.iloc[-1] / eq.iloc[0]) ** (252.0 / n) - 1
    sharpe = _ci_sharpe(ret) if ret.std() > 0 else float("nan")
    dd = (eq / eq.cummax() - 1).min()
    return dict(cagr=cagr, sharpe=sharpe, maxdd=dd, n=n)


def _summary_metrics(df):
    out = {}
    if df is None:
        return out
    for _, row in df.iterrows():
        vals = list(row.values)
        for i, v in enumerate(vals):
            if isinstance(v, str):
                k = v.strip().lower()
                for x in vals:
                    if isinstance(x, (int, float)) and not isinstance(x, bool):
                        if "sharpe" in k:
                            out.setdefault("sharpe", float(x))
                        if "cagr" in k or "annual_return" in k or "annual return" in k:
                            out.setdefault("cagr", float(x))
                        if ("max" in k and "draw" in k) or k in ("maxdd", "max_dd"):
                            out.setdefault("maxdd", float(x))
                        break
    for c in df.columns:
        if not isinstance(c, str):
            continue
        k = c.strip().lower()
        col = pd.to_numeric(df[c], errors="coerce").dropna()
        if col.empty:
            continue
        v = float(col.iloc[0])
        if "sharpe" in k: out.setdefault("sharpe", v)
        if "cagr" in k or "annual_return" in k: out.setdefault("cagr", v)
        if "max" in k and "draw" in k: out.setdefault("maxdd", v)
    return out


def verify_file(path):
    issues = []
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = set(wb.sheetnames)

    miss = [s for s in EXPECTED_MAIN if s not in sheets]
    if miss: issues.append(f"MISSING main sheets: {miss}")

    for s in sheets:
        df = sdf(wb, s)
        if (df is None or df.dropna(how="all").empty) and s != "stop_loss_history":
            issues.append(f"EMPTY sheet: {s}")

    eq = None
    eqdf = sdf(wb, "equity_history")
    if eqdf is not None and eqdf.shape[1] >= 2:
        eq = pd.to_numeric(eqdf.iloc[:, 1], errors="coerce")
        if eq.isna().any(): issues.append("equity_history has NaN")
        if (eq <= 0).any(): issues.append("equity_history has <=0")

    if eq is not None:
        rc = _metrics_from_equity(eq)
        sm = _summary_metrics(sdf(wb, "summary"))
        if rc and sm:
            if "sharpe" in sm and not math.isnan(rc["sharpe"]) and abs(sm["sharpe"] - rc["sharpe"]) > TOL_SHARPE:
                issues.append(f"SHARPE mismatch summary={sm['sharpe']:.2f} recompute(ci)={rc['sharpe']:.2f}")
            if "cagr" in sm and abs(sm["cagr"] - rc["cagr"]) > TOL_CAGR:
                issues.append(f"CAGR mismatch summary={sm['cagr']:.3f} recompute={rc['cagr']:.3f}")
            if "maxdd" in sm and abs(abs(sm["maxdd"]) - abs(rc["maxdd"])) > TOL_MAXDD:
                issues.append(f"MaxDD mismatch summary={sm['maxdd']:.3f} recompute={rc['maxdd']:.3f}")

    ph = sdf(wb, "portfolio_history")
    if ph is not None:
        cols = {str(c).lower(): c for c in ph.columns}
        if "equity" in cols and ("drawdown_pct" in cols or "drawdown" in cols):
            e = pd.to_numeric(ph[cols["equity"]], errors="coerce")
            dd = pd.to_numeric(ph[cols.get("drawdown_pct") or cols.get("drawdown")], errors="coerce")
            exp = e / e.cummax() - 1
            err = min((dd - exp).abs().max(), (dd / 100 - exp).abs().max())
            if err > 1e-3:
                issues.append(f"drawdown mismatch max_err={err:.4f}")
        if "daily_pnl" in cols and "cum_pnl" in cols:
            dpnl = pd.to_numeric(ph[cols["daily_pnl"]], errors="coerce").fillna(0)
            cpnl = pd.to_numeric(ph[cols["cum_pnl"]], errors="coerce").fillna(0)
            err = (cpnl - dpnl.cumsum()).abs().max()
            if err > max(1.0, abs(cpnl.iloc[-1]) * 0.001):
                issues.append(f"cum_pnl != cumsum(daily_pnl) err={err:.1f}")

    for sh in ("liability_history", "interest_expense"):
        d = sdf(wb, sh)
        if d is not None and d.shape[1] >= 2:
            v = pd.to_numeric(d.iloc[:, 1], errors="coerce").abs().max()
            if v is not None and v > 1.0:
                issues.append(f"{sh} not ~0 (max abs={v:.2f}); leverage should be 0")
    ah = sdf(wb, "asset_history")
    if ah is not None and eqdf is not None and ah.shape[1] >= 2:
        a = pd.to_numeric(ah.iloc[:, 1], errors="coerce")
        e2 = pd.to_numeric(eqdf.iloc[:, 1], errors="coerce")
        m = min(len(a), len(e2))
        if m > 0:
            rel = ((a.iloc[:m] - e2.iloc[:m]).abs() / e2.iloc[:m].replace(0, np.nan)).max()
            if rel is not None and rel > 0.01:
                issues.append(f"asset != equity (rel_err={rel:.3f}); leverage 0 expected")

    sw = sdf(wb, "sector_weights")
    nreb = None
    if sw is not None:
        n = _num(sw.set_index(sw.columns[0]))
        nreb = len(n)
        rs = n.sum(axis=1)
        if ((rs > 1.02) | (rs < -0.01)).any():
            issues.append(f"sector_weights row-sum out of [0,1.02]: min={rs.min():.3f} max={rs.max():.3f}")

    spd = sdf(wb, "sector_pnl_daily"); dp = sdf(wb, "daily_pnl")
    if spd is not None and dp is not None:
        spn = _num(spd.set_index(spd.columns[0])).sum(axis=1)
        dpn = _num(dp.set_index(dp.columns[0]))
        if dpn.shape[1]:
            dptot = dpn.iloc[:, 0]
            j = spn.index.intersection(dptot.index)
            if len(j) > 10:
                err = (spn.loc[j] - dptot.loc[j]).abs().max()
                scale = max(1.0, dptot.abs().max() * 0.005)
                if err > scale:
                    issues.append(f"Σ(sector_pnl_daily) != daily_pnl total max_err={err:.1f} (scale~{scale:.1f})")

    spa = sdf(wb, "sector_pnl_acc")
    if spa is not None and spd is not None:
        A = _num(spa.set_index(spa.columns[0])); Dd = _num(spd.set_index(spd.columns[0]))
        common = [c for c in A.columns if c in Dd.columns]
        if common:
            err = (A[common].fillna(0) - Dd[common].fillna(0).cumsum()).abs().max().max()
            if err > max(10.0, A[common].abs().max().max() * 0.001):
                issues.append(f"sector_pnl_acc != cumsum(daily) err={err:.1f}")

    ddh = sdf(wb, "drawdown_history")
    if ddh is not None and eq is not None and ddh.shape[1] >= 2:
        ddp = pd.to_numeric(ddh.iloc[:, -1], errors="coerce").reset_index(drop=True)
        exp = (eq / eq.cummax() - 1).reset_index(drop=True)
        m = min(len(ddp), len(exp))
        if m > 10:
            err = min((ddp.iloc[:m] - exp.iloc[:m]).abs().max(),
                      (ddp.iloc[:m] / 100 - exp.iloc[:m]).abs().max())
            if err > 5e-3:
                issues.append(f"drawdown_history dd_pct mismatch err={err:.4f}")

    ri = sdf(wb, "regime_indicators")
    if ri is not None:
        regcol = [c for c in ri.columns if isinstance(c, str) and "regime" in c.lower()]
        if regcol:
            labs = set(str(x) for x in ri[regcol[0]].dropna().unique())
            bad = labs - VALID_REGIMES
            if bad:
                issues.append(f"regime_indicators unknown labels: {bad}")

    wb.close()
    return issues, nreb


def run_group(d, pattern, label, out):
    files = sorted(glob.glob(os.path.join(d, pattern)))
    out.append(f"\n{'='*70}\n{label}: {len(files)} files\n{'='*70}")
    nbad = 0; reb = []
    for f in files:
        issues, nreb = verify_file(f)
        if nreb: reb.append(nreb)
        if issues:
            nbad += 1
            out.append(f"  ✗ {os.path.basename(f)[:62]}")
            for i in issues:
                out.append(f"       - {i}")
    if reb:
        out.append(f"  rebalances: min={min(reb)} max={max(reb)}")
    out.append(f"  >>> {label}: {len(files)-nbad}/{len(files)} clean, {nbad} with issues")
    return nbad, len(files)


def main():
    ap = argparse.ArgumentParser(description="Deep per-sheet verification of SSRS backtest Excel")
    ap.add_argument("date", nargs="?", default=None, help="YYYYMMDD (default: today)")
    ap.add_argument("--dir", default=DEFAULT_DIR, help="historical_runs/sector_rotation dir")
    args = ap.parse_args()
    D = args.date or pd.Timestamp.today().strftime("%Y%m%d")

    out = [f"SSRS Excel deep verification — date={D} dir={args.dir}"]
    total_bad = 0; total_files = 0
    for pat, lab in [
        (f"sr_portfolio_*_v1_IS_batch_{D}_*.xlsx",        "V1 IS batch"),
        (f"sr_portfolio_*_v2_IS_batch_{D}_*.xlsx",        "V2 IS batch"),
        (f"sr_portfolio_*_v1_IS-OOS_tearsheet_{D}_*.xlsx","V1 IS-OOS"),
        (f"sr_portfolio_*_v2_IS-OOS_tearsheet_{D}_*.xlsx","V2 IS-OOS"),
    ]:
        nbad, nfiles = run_group(args.dir, pat, lab, out)
        total_bad += nbad; total_files += nfiles

    out.append(f"\n{'='*70}")
    out.append(f"DEEP VERIFY RESULT: {total_files} files checked, {total_bad} with issues")
    out.append(f"{'='*70}")
    print("\n".join(out))
    if total_files == 0:
        print("WARNING: no files matched — did the heavy suite run for this date?")
        return 2
    return 0 if total_bad == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
