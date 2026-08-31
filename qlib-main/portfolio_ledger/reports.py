"""PnL / Risk 报告生成 — 深度 mirror MRPT/MTFS 的 PnLReport / RiskManager。

样式（字体/配色/表格/编号注释）逐项复刻 PnLReport.py 的样式库；章节结构对齐：

PnL：一、汇总 → 二、已平仓明细 → 三、持仓与生命周期明细 → 四、分红与费用明细
     → 五、简要分析 → 五(b)、组合风险指标 & 基准对比 → 六、系统成交价 vs 参考执行价对照
Risk：风险摘要 → 一、资产负债表(T/T-1D/T-1W/T-1M) → 二、利润表 → 三、资本与现金流
     → 四、敞口与集中度 → 五、Beta → 六、VaR → 七、压力测试 → 八、限额监控 → 九、诊断
     （pairs 的 双账户法/空头账簿 不适用于长仓+现金账户，不搬）

命名（用户规格）：文件名 = 数据截止日，无时分秒：
  {strategy_dir}/trading_signals/pnl_reports/pnl_report_YYYYMMDD.pdf
  {strategy_dir}/trading_signals/risk_management/risk_report_YYYYMMDD.{json,txt,pdf}
                                              + risk_workbook_YYYYMMDD.xlsx

无前视纪律：D 日报告只用 ≤D 的 account_history / ledger / Polygon 价格。
运行（需 someopark_run：reportlab/openpyxl/matplotlib）：
  cd qlib-main && python -m portfolio_ledger.reports all --backfill
"""
from __future__ import annotations

import glob
import io
import json
import logging
import os
import sys

import numpy as np
import pandas as pd

from .ledger import (STRATEGIES, _cfg, history_dir, load_ledger_rows,
                     load_store_prices)

log = logging.getLogger("portfolio_ledger.reports")

BENCH_STORE_STRATEGY = "aiss"          # SPY/SMH 均在 semi_strategy store
INITIAL_CASH = 1_000_000.0

# ═══════════════════════════════════════════════════════════════════════════
# 样式与配色 —— 复刻 PnLReport.py（copy-first）
# ═══════════════════════════════════════════════════════════════════════════
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (Image, Paragraph, SimpleDocTemplate, Spacer,
                                Table, TableStyle)


def _register_cjk() -> str:
    for path in [
        '/System/Library/Fonts/PingFang.ttc',
        '/System/Library/Fonts/STHeiti Light.ttc',
        '/Library/Fonts/Arial Unicode MS.ttf',
        '/System/Library/Fonts/Supplemental/Arial Unicode.ttf',
    ]:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont('CJK', path))
                return 'CJK'
            except Exception:
                continue
    return 'Helvetica'


FONT = _register_cjk()


def S(name, **kw):
    kw.setdefault('fontName', FONT)
    return ParagraphStyle(name, **kw)


C_HEADER  = colors.HexColor('#1a1a2e')
C_ROW_ALT = colors.HexColor('#f7f9fc')
C_GOLD    = colors.HexColor('#d4a843')
C_BORDER  = colors.HexColor('#cccccc')
C_GRAY    = colors.HexColor('#888888')

title_style  = S('TT', fontSize=15, leading=19, alignment=TA_CENTER, spaceAfter=4)
sub_style    = S('ST', fontSize=8.5, leading=12, alignment=TA_CENTER,
                 textColor=colors.HexColor('#555555'), spaceAfter=14)
h1_style     = S('H1', fontSize=10, leading=14, textColor=C_HEADER,
                 spaceBefore=12, spaceAfter=5)
footer_style = S('FT', fontSize=7, leading=9.5, alignment=TA_CENTER,
                 textColor=C_GRAY, spaceBefore=6)


def money(v, color=True) -> str:
    if v is None:
        return 'N/A'
    s = f'+{abs(v):,.2f}' if v >= 0 else f'-{abs(v):,.2f}'
    if not color:
        return s
    c = '#1a7a4a' if v >= 0 else '#c0392b'
    return f'<font color="{c}">{s}</font>'


def pct(v) -> str:
    if v is None:
        return 'N/A'
    s = f'+{abs(v):.2f}%' if v >= 0 else f'-{abs(v):.2f}%'
    c = '#1a7a4a' if v >= 0 else '#c0392b'
    return f'<font color="{c}"><b>{s}</b></font>'


def C(text, align='LEFT', header=False) -> Paragraph:
    s = ParagraphStyle('_cx', fontName=FONT, fontSize=7.5, leading=10.5,
                       alignment=(TA_RIGHT if align == 'RIGHT' else 0),
                       textColor=(colors.white if header else colors.black))
    content = f'<b>{text}</b>' if header else str(text)
    return Paragraph(content, s)


def H(text, align='LEFT') -> Paragraph:
    return C(text, align=align, header=True)


def make_table(data, col_widths, header_rows=1):
    t = Table(data, colWidths=col_widths, repeatRows=header_rows)
    n = len(data)
    cmds = [
        ('FONTNAME',      (0, 0), (-1, -1), FONT),
        ('FONTSIZE',      (0, 0), (-1, -1), 7.5),
        ('LEADING',       (0, 0), (-1, -1), 10.5),
        ('BACKGROUND',    (0, 0), (-1, header_rows - 1), C_HEADER),
        ('TEXTCOLOR',     (0, 0), (-1, header_rows - 1), colors.white),
        ('ALIGN',         (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN',         (-1, 0), (-1, -1), 'RIGHT'),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 0), (-1, -1), 3.5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3.5),
        ('LEFTPADDING',   (0, 0), (-1, -1), 5),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 5),
        ('GRID',          (0, 0), (-1, -1), 0.35, C_BORDER),
        ('LINEBELOW',     (0, header_rows - 1), (-1, header_rows - 1), 0.9, C_GOLD),
    ]
    for i in range(header_rows, n):
        bg = C_ROW_ALT if (i - header_rows) % 2 == 1 else colors.white
        cmds.append(('BACKGROUND', (0, i), (-1, i), bg))
    t.setStyle(TableStyle(cmds))
    return t


def make_kv_table(rows, label_w=9.5 * cm, val_w=5 * cm):
    sl = ParagraphStyle('_sl', fontName=FONT, fontSize=7.5, leading=10.5)
    sv = ParagraphStyle('_sv', fontName=FONT, fontSize=7.5, leading=10.5, alignment=TA_RIGHT)
    data = [[Paragraph(f'<b>{r[0]}</b>', sl), Paragraph(r[1], sv)] for r in rows]
    t = Table(data, colWidths=[label_w, val_w])
    t.setStyle(TableStyle([
        ('FONTNAME',       (0, 0), (-1, -1), FONT),
        ('ALIGN',          (1, 0), (1, -1), 'RIGHT'),
        ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',     (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING',  (0, 0), (-1, -1), 3),
        ('LEFTPADDING',    (0, 0), (-1, -1), 7),
        ('RIGHTPADDING',   (0, 0), (-1, -1), 7),
        ('LINEABOVE',      (0, 0), (-1, 0), 0.5, C_GOLD),
        ('LINEBELOW',      (0, -1), (-1, -1), 0.5, C_GOLD),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.HexColor('#fafbfc'), colors.white]),
    ]))
    return t


def note_item(num: str, text: str, W: float):
    sl = ParagraphStyle('_nl', fontName=FONT, fontSize=7.5, leading=10.5)
    st = ParagraphStyle('_nt', fontName=FONT, fontSize=7.5, leading=10.5)
    return Table(
        [[Paragraph(num, sl), Paragraph(text, st)]],
        colWidths=[0.6 * cm, W - 0.6 * cm],
        style=[('VALIGN', (0, 0), (-1, -1), 'TOP'),
               ('TOPPADDING', (0, 0), (-1, -1), 1),
               ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
               ('LEFTPADDING', (0, 0), (-1, -1), 0),
               ('RIGHTPADDING', (0, 0), (-1, -1), 0)])


# ═══════════════════════════════════════════════════════════════════════════
# 数据装配
# ═══════════════════════════════════════════════════════════════════════════

def load_account_series(strategy: str) -> dict:
    out = {}
    for fp in sorted(glob.glob(os.path.join(history_dir(strategy),
                                            f"account_{strategy}_*.json"))):
        try:
            with open(fp) as f:
                d = json.load(f)
            out[d["as_of"]] = d
        except Exception:
            continue
    return out


def equity_series(acct_series: dict, upto: str) -> pd.Series:
    days = sorted(d for d in acct_series if d <= upto)
    return pd.Series({pd.Timestamp(d): acct_series[d]["equity"] for d in days},
                     name="equity")


def load_store_open(strategy: str, tickers: list) -> pd.DataFrame:
    """Open 价宽表（section 六 参考执行价 = 次日开盘）。"""
    store = _cfg(strategy)["store_dir"]
    cols = {}
    for t in sorted(set(tickers)):
        fp = os.path.join(store, f"{t}_prices.parquet")
        if os.path.exists(fp):
            df = pd.read_parquet(fp)
            if "Open" in df.columns:
                cols[t] = df["Open"]
    px = pd.DataFrame(cols)
    if not px.empty:
        px.index = pd.DatetimeIndex(px.index).normalize()
    return px.sort_index()


def _bench_stats(bench_px: pd.DataFrame, start, end) -> dict:
    """基准同期收益/最大回撤/Sharpe（五(b) 表）。"""
    out = {}
    for b in bench_px.columns:
        s = bench_px[b].loc[start:end].dropna()
        if len(s) < 2:
            continue
        rets = s.pct_change().dropna()
        peak = s.cummax()
        out[b] = {
            "total_return_pct": (s.iloc[-1] / s.iloc[0] - 1) * 100,
            "max_dd_pct": ((s - peak) / peak).min() * 100,
            "sharpe": (rets.mean() / rets.std() * np.sqrt(252)) if rets.std() > 0 else 0.0,
        }
    return out


def _portfolio_stats(eq: pd.Series) -> dict:
    rets = eq.pct_change().dropna()
    peak = eq.cummax()
    dd = eq - peak
    dd_pct = dd / peak
    trough_i = dd.idxmin() if len(dd) else None
    peak_i = eq.loc[:trough_i].idxmax() if trough_i is not None else None
    sharpe = (rets.mean() / rets.std() * np.sqrt(252)) if len(rets) > 1 and rets.std() > 0 else 0.0
    # CDaR 95：回撤分布的 95 分位均值尾部
    cdar = float(-np.mean(np.sort(dd_pct.values)[:max(1, int(len(dd_pct) * 0.05))]) * 100) \
        if len(dd_pct) else 0.0
    return {
        "total_return_pct": (eq.iloc[-1] / eq.iloc[0] - 1) * 100 if len(eq) > 1 else 0.0,
        "max_dd": float(dd.min()) if len(dd) else 0.0,
        "max_dd_pct": float(dd_pct.min() * 100) if len(dd_pct) else 0.0,
        "max_dd_peak_date": str(peak_i.date()) if peak_i is not None else "",
        "max_dd_trough_date": str(trough_i.date()) if trough_i is not None else "",
        "sharpe": float(sharpe),
        "annual_vol_pct": float(rets.std() * np.sqrt(252) * 100) if len(rets) > 1 else 0.0,
        "cdar_95_pct": cdar,
        "rets": rets,
    }


def _equity_png(eq: pd.Series, title: str) -> bytes:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    fig, ax = plt.subplots(figsize=(7.2, 2.0), dpi=140)
    ax.plot(eq.index, eq.values, lw=1.2, color="#16213e")
    ax.fill_between(eq.index, eq.values, eq.min(), alpha=0.08, color="#16213e")
    ax.set_title(title, fontsize=8)
    ax.tick_params(labelsize=6)
    ax.grid(alpha=0.25, lw=0.4)
    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png")
    plt.close(fig)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════
# PnL Report（六节，mirror pairs）
# ═══════════════════════════════════════════════════════════════════════════

def generate_pnl_report(strategy: str, day: str, acct_series: dict,
                        prices: pd.DataFrame, opens: pd.DataFrame,
                        ledger_rows: list, bench_px: pd.DataFrame) -> str:
    a = acct_series[day]
    eq = equity_series(acct_series, day)
    first_eq = float(eq.iloc[0])
    rows_upto = [r for r in ledger_rows if r.get("date", "") <= day]
    sells = [r for r in rows_upto if r.get("side") == "SELL"]
    buys = [r for r in rows_upto if r.get("side") == "BUY"]
    divs = [r for r in rows_upto if r.get("side") == "DIV"]
    fees = [r for r in rows_upto if r.get("side") == "FEE"]
    px_row = prices.loc[pd.Timestamp(day)]
    port = _portfolio_stats(eq)
    label = strategy.upper()

    out_dir = os.path.join(_cfg(strategy)["dir"], "trading_signals", "pnl_reports")
    os.makedirs(out_dir, exist_ok=True)
    out = os.path.join(out_dir, f"pnl_report_{day.replace('-', '')}.pdf")
    doc = SimpleDocTemplate(out, pagesize=A4, topMargin=1.3 * cm, bottomMargin=1.3 * cm,
                            leftMargin=1.5 * cm, rightMargin=1.5 * cm)
    W = A4[0] - 3 * cm
    story = []

    story.append(Paragraph(f'Someo Park — {label} 组合 PnL 分析报告', title_style))
    story.append(Paragraph(
        f'报告区间：{eq.index[0].date()} ～ {day}｜账本：portfolio_ledger'
        f'（realized 定格于成交日收盘）｜价格源：Polygon store', sub_style))

    # ── 一、汇总 ─────────────────────────────────────────────────────────
    story.append(Paragraph('一、汇总', h1_style))
    subtotal = a['equity'] - first_eq
    gross_notional = sum(abs(r.get('gross', 0)) for r in sells + buys)
    period_ret = (a['equity'] / first_eq - 1) * 100
    story.append(make_kv_table([
        ['期初 equity（{}）'.format(str(eq.index[0].date())), f"${first_eq:,.2f}"],
        ['期末 equity', f"${a['equity']:,.2f}"],
        ['期间收益', pct(period_ret)],
        ['已实现（交易）', money(a['cumulative_realized'])],
        ['分红收入', money(a['cumulative_dividends'])],
        ['交易费用', money(-a['cumulative_fees'])],
        ['未实现', money(a['unrealized'])],
        ['期间小计（equity 变动）', money(subtotal)],
        ['现金 / 现金比例',
         f"${a['cash']:,.2f}（{a['cash'] / a['equity'] * 100:.1f}%）"],
        ['投资比例（长仓，无杠杆）',
         f"{a['position_value'] / a['equity'] * 100:.1f}%"],
        ['初始资本 / ROE(初始)',
         f"$1,000,000 / {pct((a['equity'] / INITIAL_CASH - 1) * 100)}"],
        ['期间成交总额（含买卖）', f"${gross_notional:,.0f}"],
    ]))
    ident = a['equity'] - a['initial_cash'] - (
        a['cumulative_realized'] + a['cumulative_dividends']
        - a['cumulative_fees'] + a['unrealized'])
    story.append(Paragraph(
        f'<super>1</super> 恒等式核对：equity − 初始 − (已实现+分红−费用+未实现) '
        f'= ${ident:+.2f}（应≈0）；已实现值取自 trade_ledger 定格值，任何一天重跑逐分不变。',
        footer_style))

    story.append(Spacer(1, 4))
    story.append(Image(io.BytesIO(_equity_png(eq, f'{label} equity (ledger basis, incl. dividends & fees)')),
                       width=W, height=4.6 * cm))

    # ── 二、已平仓明细 ────────────────────────────────────────────────────
    story.append(Paragraph(
        '二、已平仓明细（来源：trade_ledger SELL 实际成交价 = Polygon 交易日收盘）',
        h1_style))
    if sells:
        # EXIT / DECREASE：卖出后当日是否仍持有
        data = [[H('日期'), H('Ticker'), H('方向'), H('方式'), H('股数', 'RIGHT'),
                 H('成交价', 'RIGHT'), H('成本', 'RIGHT'), H('PnL', 'RIGHT'),
                 H('PnL%', 'RIGHT')]]
        pos_after = {}
        for r in rows_upto:
            if r.get('side') == 'BUY':
                pos_after[r['ticker']] = pos_after.get(r['ticker'], 0) + r['shares']
            elif r.get('side') == 'SELL':
                pos_after[r['ticker']] = pos_after.get(r['ticker'], 0) - r['shares']
        run_pos = {}
        for r in rows_upto:
            if r.get('side') == 'BUY':
                run_pos[r['ticker']] = run_pos.get(r['ticker'], 0) + r['shares']
        for r in sells:
            run_pos[r['ticker']] = run_pos.get(r['ticker'], 0)  # ensure key
        # 重放持仓判定 EXIT/DECREASE
        replay_pos = {}
        rows_sorted = sorted(rows_upto, key=lambda x: (x.get('date', ''), x.get('side', '')))
        kind_by_key = {}
        # 期初持仓（第一天快照已在 avg 成本里；从第一天 account 反推）
        first_acct = acct_series[sorted(acct_series)[0]]
        for t, p in first_acct['positions'].items():
            replay_pos[t] = p['shares']
        for r in rows_sorted:
            if r.get('side') == 'BUY':
                replay_pos[r['ticker']] = replay_pos.get(r['ticker'], 0) + r['shares']
            elif r.get('side') == 'SELL':
                replay_pos[r['ticker']] = replay_pos.get(r['ticker'], 0) - r['shares']
                kind_by_key[r['dedup_key']] = 'EXIT' if replay_pos[r['ticker']] <= 0 else 'DECREASE'
        total_realized = 0.0
        for r in sells:
            rp = r.get('realized_pnl', 0) or 0
            total_realized += rp
            cb = r.get('avg_cost_at_trade', 0) or 0
            pnl_pct = (r['price'] / cb - 1) * 100 if cb else None
            data.append([
                C(r['date']), C(r['ticker']), C('Long'),
                C(kind_by_key.get(r['dedup_key'], 'SELL')),
                C(f"{r['shares']:,}", 'RIGHT'),
                C(f"{r['price']:,.2f}", 'RIGHT'), C(f"{cb:,.2f}", 'RIGHT'),
                Paragraph(money(rp), ParagraphStyle('_p', fontName=FONT, fontSize=7.5,
                                                    leading=10.5, alignment=TA_RIGHT)),
                Paragraph(pct(pnl_pct) if pnl_pct is not None else 'N/A',
                          ParagraphStyle('_p2', fontName=FONT, fontSize=7.5,
                                         leading=10.5, alignment=TA_RIGHT)),
            ])
        data.append([C('合计'), C(''), C(''), C(''), C(''), C(''), C(''),
                     Paragraph(money(total_realized),
                               ParagraphStyle('_pt', fontName=FONT, fontSize=7.5,
                                              leading=10.5, alignment=TA_RIGHT)), C('')])
        story.append(make_table(data, [2.0*cm, 1.7*cm, 1.3*cm, 1.9*cm, 1.7*cm,
                                       2.0*cm, 2.0*cm, 2.4*cm, 2.0*cm]))
    else:
        story.append(Paragraph('（期间无平仓交易）', footer_style))

    # ── 三、持仓与生命周期明细 ────────────────────────────────────────────
    story.append(Paragraph(f'三、{label} 持仓与生命周期明细', h1_style))
    data = [[H('Ticker'), H('状态'), H('入场日'), H('股数', 'RIGHT'),
             H('平均成本', 'RIGHT'), H('现价', 'RIGHT'), H('市值', 'RIGHT'),
             H('未实现', 'RIGHT'), H('权重', 'RIGHT')]]
    unreal_terms = []
    for tk, p in sorted(a['positions'].items()):
        px = float(px_row.get(tk, np.nan))
        mv = p['shares'] * px
        up = (px - p['avg_cost']) * p['shares']
        unreal_terms.append(up)
        data.append([C(tk), C('持仓中'), C(p.get('entry_date', '')),
                     C(f"{p['shares']:,}", 'RIGHT'), C(f"{p['avg_cost']:,.2f}", 'RIGHT'),
                     C(f"{px:,.2f}", 'RIGHT'), C(f"{mv:,.0f}", 'RIGHT'),
                     Paragraph(money(up), ParagraphStyle('_u', fontName=FONT, fontSize=7.5,
                                                         leading=10.5, alignment=TA_RIGHT)),
                     C(f"{mv / a['equity'] * 100:.1f}%", 'RIGHT')])
    exited = sorted({r['ticker'] for r in sells} - set(a['positions']))
    for tk in exited:
        tr = [r for r in sells if r['ticker'] == tk]
        realized_t = sum(r.get('realized_pnl', 0) or 0 for r in tr)
        data.append([C(tk), C(f"已平仓 {tr[-1]['date'][5:]}"), C(''), C('0', 'RIGHT'),
                     C(''), C(''), C(''),
                     Paragraph(money(realized_t) + ' <font size=6>(已实现)</font>',
                               ParagraphStyle('_u2', fontName=FONT, fontSize=7.5,
                                              leading=10.5, alignment=TA_RIGHT)),
                     C('—', 'RIGHT')])
    data.append([C('现金'), C(''), C(''), C(''), C(''), C(''),
                 C(f"{a['cash']:,.0f}", 'RIGHT'), C(''),
                 C(f"{a['cash'] / a['equity'] * 100:.1f}%", 'RIGHT')])
    story.append(make_table(data, [1.8*cm, 2.3*cm, 2.0*cm, 1.7*cm, 2.0*cm,
                                   1.9*cm, 2.1*cm, 2.2*cm, 1.5*cm]))
    _sum_str = ' + '.join(money(x, color=False) for x in unreal_terms) if unreal_terms else '0'
    story.append(Paragraph(
        f'未实现 = {_sum_str} = <b>{money(a["unrealized"], color=False)}</b>；'
        f'已实现（交易） {money(a["cumulative_realized"], color=False)}；'
        f'{label} 小计 {money(subtotal, color=False)}。', footer_style))

    # ── 四、分红与费用明细 ────────────────────────────────────────────────
    story.append(Paragraph('四、分红与费用明细', h1_style))
    if divs or fees:
        data = [[H('日期'), H('类型'), H('Ticker'), H('股数', 'RIGHT'),
                 H('每股', 'RIGHT'), H('金额', 'RIGHT')]]
        for r in divs:
            data.append([C(r['date']), C('分红 (ex-date)'), C(r['ticker']),
                         C(f"{r['shares']:,}", 'RIGHT'), C(f"{r['price']:.4f}", 'RIGHT'),
                         Paragraph(money(r['gross']),
                                   ParagraphStyle('_d', fontName=FONT, fontSize=7.5,
                                                  leading=10.5, alignment=TA_RIGHT))])
        for r in fees:
            data.append([C(r['date']), C('交易费用'), C('—'), C('', 'RIGHT'), C('', 'RIGHT'),
                         Paragraph(money(r['gross']),
                                   ParagraphStyle('_f', fontName=FONT, fontSize=7.5,
                                                  leading=10.5, alignment=TA_RIGHT))])
        story.append(make_table(data, [2.2*cm, 3.0*cm, 2.0*cm, 2.2*cm, 2.4*cm, 3.0*cm]))
    else:
        story.append(Paragraph('（期间无分红与费用记录）', footer_style))

    # ── 五、简要分析 ─────────────────────────────────────────────────────
    story.append(Paragraph('五、简要分析', h1_style))
    notes = []
    direction = '盈利' if subtotal >= 0 else '亏损'
    ss_str = f'，PnL/成交名义 {subtotal / gross_notional * 100:+.2f}%' if gross_notional > 0 else ''
    notes.append(f'期间合计 {money(subtotal, color=False)}{ss_str}，'
                 f'ROE(初始现金) {(a["equity"] / INITIAL_CASH - 1) * 100:+.2f}%，整体{direction}。')
    if sells:
        by_ticker = {}
        for r in sells:
            by_ticker[r['ticker']] = by_ticker.get(r['ticker'], 0) + (r.get('realized_pnl', 0) or 0)
        best_t = max(by_ticker, key=by_ticker.get)
        worst_t = min(by_ticker, key=by_ticker.get)
        if by_ticker[best_t] > 0:
            notes.append(f'已实现最优：{best_t}（{money(by_ticker[best_t], color=False)}）。')
        if by_ticker[worst_t] < 0:
            notes.append(f'已实现最差：{worst_t}（{money(by_ticker[worst_t], color=False)}）。')
    if a['positions']:
        ups = {tk: (float(px_row.get(tk, np.nan)) - p['avg_cost']) * p['shares']
               for tk, p in a['positions'].items()}
        bo = max(ups, key=ups.get)
        if ups[bo] > 0:
            notes.append(f'最大浮盈：{bo}（当前 {money(ups[bo], color=False)}，仍持仓）。')
    if divs:
        notes.append(f'分红收入 {len(divs)} 笔合计 {money(a["cumulative_dividends"], color=False)}'
                     f'（ex-date 现金入账，参与本策略复利）。')
    reb_days = sorted({r['date'] for r in sells + buys})
    if reb_days:
        notes.append(f'期间调仓 {len(reb_days)} 次（{", ".join(d[5:] for d in reb_days)}），'
                     f'交易费用合计 {money(-a["cumulative_fees"], color=False)}。')
    notes.append(f'组合最大回撤 {money(port["max_dd"], color=False)}'
                 f'（{port["max_dd_pct"]:+.2f}%），'
                 f'峰值 {port["max_dd_peak_date"]} → 谷值 {port["max_dd_trough_date"]}。')
    notes.append(f'年化 Sharpe Ratio（账本日收益）：{port["sharpe"]:.2f}；'
                 f'年化波动 {port["annual_vol_pct"]:.2f}%。')
    bench = _bench_stats(bench_px, eq.index[0], pd.Timestamp(day))
    if bench:
        parts = [f'{b} 收益 {m["total_return_pct"]:+.2f}%，最大回撤 {m["max_dd_pct"]:.2f}%，'
                 f'Sharpe {m["sharpe"]:.2f}' for b, m in bench.items()]
        notes.append(f'基准对比（同期）：组合收益 {port["total_return_pct"]:+.2f}%；'
                     + '；'.join(parts) + '。')
    notes.append(f'现金比例 {a["cash"] / a["equity"] * 100:.1f}%——'
                 f'来自调仓目标现金权重 + 累计已实现盈利与分红（复利 sizing 下次调仓时复投）。')
    for i, n in enumerate(notes, 1):
        story.append(note_item(f'{i}.', n, W))

    # ── 五(b)、组合风险指标 & 基准对比 ────────────────────────────────────
    story.append(Paragraph('五(b)、组合风险指标 & 基准对比 <super>2</super>', h1_style))
    hdr = [H('指标')] + [H(label, 'RIGHT')] + [H(b, 'RIGHT') for b in bench]
    data = [hdr,
            [C('期间收益率')] + [Paragraph(pct(port['total_return_pct']),
                                          ParagraphStyle('_b1', fontName=FONT, fontSize=7.5,
                                                         leading=10.5, alignment=TA_RIGHT))]
            + [C(f"{bench[b]['total_return_pct']:+.2f}%", 'RIGHT') for b in bench],
            [C('最大回撤')] + [C(f"{port['max_dd_pct']:.2f}%", 'RIGHT')]
            + [C(f"{bench[b]['max_dd_pct']:.2f}%", 'RIGHT') for b in bench],
            [C('年化 Sharpe')] + [C(f"{port['sharpe']:.2f}", 'RIGHT')]
            + [C(f"{bench[b]['sharpe']:.2f}", 'RIGHT') for b in bench],
            [C('最大回撤（$）')] + [C(f"{port['max_dd']:,.2f}", 'RIGHT')]
            + [C('—', 'RIGHT') for _ in bench]]
    story.append(make_table(data, [4.2*cm] + [3.4*cm] * (1 + len(bench))))
    story.append(Paragraph(
        f'<super>2</super> 组合收益率 = equity 变动 / 期初 equity；Sharpe = 日收益均值/标准差×√252；'
        f'回撤区间 {port["max_dd_peak_date"]} → {port["max_dd_trough_date"]}；'
        f'统计样本 {max(len(eq) - 1, 0)} 个交易日。', footer_style))

    # ── 六、系统成交价 vs 参考执行价对照 ──────────────────────────────────
    story.append(Paragraph('六、系统成交价 vs 参考执行价对照 <super>3</super>', h1_style))
    trade_rows = []
    big_diff = []
    for r in sorted(sells + buys, key=lambda x: (x['date'], x['ticker'])):
        ts = pd.Timestamp(r['date'])
        future = opens.index[opens.index > ts] if not opens.empty else []
        ref_pnl = None
        if len(future) and r['ticker'] in opens.columns:
            op = opens.loc[future[0], r['ticker']]
            if not pd.isna(op):
                if r['side'] == 'SELL':
                    cb = r.get('avg_cost_at_trade', 0) or 0
                    ref_pnl = (float(op) - cb) * r['shares']
        sys_pnl = r.get('realized_pnl')
        if r['side'] == 'SELL' and sys_pnl is not None and ref_pnl is not None:
            diff = sys_pnl - ref_pnl
            flag = '差异>$3k' if abs(diff) > 3000 else ''
            if flag:
                big_diff.append(r['ticker'])
            trade_rows.append([C(r['date']), C(r['ticker']),
                               Paragraph(money(sys_pnl),
                                         ParagraphStyle('_s6', fontName=FONT, fontSize=7.5,
                                                        leading=10.5, alignment=TA_RIGHT)),
                               Paragraph(money(ref_pnl),
                                         ParagraphStyle('_s7', fontName=FONT, fontSize=7.5,
                                                        leading=10.5, alignment=TA_RIGHT)),
                               Paragraph(money(diff),
                                         ParagraphStyle('_s8', fontName=FONT, fontSize=7.5,
                                                        leading=10.5, alignment=TA_RIGHT)),
                               C(flag)])
    if trade_rows:
        data = [[H('结算日'), H('Ticker'), H('系统成交价 PnL', 'RIGHT'),
                 H('参考执行价 PnL', 'RIGHT'), H('差异(系统−参考)', 'RIGHT'), H('说明')]]
        data += trade_rows
        story.append(make_table(data, [2.2*cm, 1.8*cm, 3.2*cm, 3.2*cm, 3.2*cm, 2.2*cm]))
        note = ('系统成交价为交易日收盘（Polygon，账本口径）；参考执行价为次一交易日开盘'
                '（同店 Open）。差异反映隔夜价格漂移，不影响账本已实现值。')
        if big_diff:
            note += f' 差异较大（>$3,000）的标的：{", ".join(sorted(set(big_diff)))}。'
        story.append(Paragraph(f'<super>3</super> {note}', footer_style))
    else:
        story.append(Paragraph('（期间无已实现交易可对照）', footer_style))

    doc.build(story)
    return out


# ═══════════════════════════════════════════════════════════════════════════
# Risk Report（mirror RiskManager：摘要 + 九节 + json/txt/pdf/xlsx 四件套）
# ═══════════════════════════════════════════════════════════════════════════

def _limits_table(strategy: str, m: dict) -> list:
    """限额定义（amber/red）——mirror pairs limits 结构。长仓+现金账户阈值。"""
    vol_amber, vol_red = ((50.0, 80.0) if strategy == 'aiss'
                          else (30.0, 50.0) if strategy == 'aeus'   # 个股策略但公用事业波动低于半导体(AEUS_PLAN §6.7,D5 实测年化 ~24%)
                          else (18.0, 30.0))
    lim = [
        ('single_name_pct', m.get('max_single_name_pct') or 0, 25.0, 40.0, '%'),
        ('top3_pct',        m.get('top3_weight_pct') or 0,     60.0, 80.0, '%'),
        ('invested_pct',    m.get('invested_pct') or 0,        100.0, 105.0, '%'),
        ('var_95_1d_pct',   abs(m.get('var_95_1d_pct') or 0),  3.0, 5.0, '%'),
        ('annual_vol_pct',  m.get('annual_vol_pct') or 0,      vol_amber, vol_red, '%'),
        ('drawdown_pct',    abs(m.get('max_dd_pct') or 0),     15.0, 25.0, '%'),
    ]
    out = []
    for name, val, amber, red, fmt in lim:
        status = 'red' if val >= red else ('amber' if val >= amber else 'green')
        out.append({'name': name, 'scope': strategy, 'value': round(float(val), 3),
                    'amber': amber, 'red': red, 'fmt': fmt, 'status': status})
    return out


def _asof_lookup(acct_series: dict, target: pd.Timestamp):
    days = [d for d in sorted(acct_series) if pd.Timestamp(d) <= target]
    return acct_series[days[-1]] if days else None


def generate_risk_report(strategy: str, day: str, acct_series: dict,
                         prices: pd.DataFrame, bench_px: pd.DataFrame,
                         ledger_rows: list) -> list:
    import openpyxl
    a = acct_series[day]
    label = strategy.upper()
    eq = equity_series(acct_series, day)
    port = _portfolio_stats(eq)
    rets = port.pop('rets')
    nav = a['equity']
    px_row = prices.loc[pd.Timestamp(day)]

    mv = {t: p['shares'] * float(px_row.get(t, np.nan)) for t, p in a['positions'].items()}
    weights = {t: v / nav for t, v in mv.items() if not np.isnan(v)}
    top = sorted(weights.items(), key=lambda kv: -kv[1])
    invested = sum(v for v in mv.values() if not np.isnan(v))

    # VaR 以损失额正数表示（与 pairs RiskManager 口径一致）
    var95 = float(abs(np.percentile(rets.tail(60), 5)) * nav) if len(rets) >= 10 else None
    var99 = float(abs(np.percentile(rets.tail(60), 1)) * nav) if len(rets) >= 20 else None
    betas = {}
    for b in bench_px.columns:
        bs = bench_px[b].loc[:day].pct_change().dropna()
        joint = pd.concat([rets, bs], axis=1, join='inner').dropna()
        if len(joint) >= 10 and joint.iloc[:, 1].var() > 0:
            betas[b] = round(float(joint.iloc[:, 0].cov(joint.iloc[:, 1])
                                   / joint.iloc[:, 1].var()), 3)
    mu, sig = (rets.mean(), rets.std()) if len(rets) > 1 else (0.0, 0.0)
    kelly_full = float(mu / sig ** 2) if sig > 0 else 0.0

    summary = {
        'nav': round(nav, 2),
        'cash': a['cash'],
        'cash_pct': round(a['cash'] / nav * 100, 2),
        'invested_pct': round(invested / nav * 100, 2),
        'gross_leverage': round(invested / nav, 3),        # 长仓无杠杆 ≤1
        'net_leverage': round(invested / nav, 3),
        'net_beta': betas,
        'var_95_1d': round(var95, 2) if var95 is not None else None,
        'var_95_1d_pct': round(var95 / nav * 100, 2) if var95 is not None else None,
        'var_99_1d': round(var99, 2) if var99 is not None else None,
        'annual_vol_pct': round(port['annual_vol_pct'], 2),
        'sharpe_annual': round(port['sharpe'], 3),
        'cdar_95_pct': round(port['cdar_95_pct'], 2),
        'kelly_full': round(kelly_full, 2),
        'kelly_half': round(kelly_full / 2, 2),
        'kelly_actual_leverage': round(invested / nav, 3),
        'max_single_name': top[0][0] if top else None,
        'max_single_name_pct': round(top[0][1] * 100, 2) if top else None,
        'top3_weight_pct': round(sum(w for _, w in top[:3]) * 100, 2),
        'n_positions': len(a['positions']),
        'max_dd_pct': round(port['max_dd_pct'], 2),
        'period_return_pct': round(port['total_return_pct'], 2),
    }
    limits = _limits_table(strategy, summary)
    breaches = [l for l in limits if l['status'] != 'green']
    summary['n_breaches'] = len(breaches)
    summary['worst_breach'] = (f"{breaches[0]['name']} {breaches[0]['status']}"
                               if breaches else '')

    # 资产负债表 T / T-1D / T-1W / T-1M（真实账本快照）
    ts = pd.Timestamp(day)
    bs_cols = {'T': a,
               'T-1D': _asof_lookup(acct_series, ts - pd.Timedelta(days=1)),
               'T-1W': _asof_lookup(acct_series, ts - pd.Timedelta(days=7)),
               'T-1M': _asof_lookup(acct_series, ts - pd.Timedelta(days=30))}
    balance_sheet = {}
    for k, acc in bs_cols.items():
        if acc is None:
            continue
        balance_sheet[k] = {
            'as_of': acc['as_of'],
            'cash': acc['cash'],
            'long_securities': acc.get('position_value'),
            'total_assets': round(acc['cash'] + (acc.get('position_value') or 0), 2),
            'total_liabilities': 0.0,
            'nav': acc['equity'],
            'balance_check': round(acc['cash'] + (acc.get('position_value') or 0)
                                   - acc['equity'], 2),
        }

    # 利润表 1D/1W/MTD/ITD
    def _pnl_window(start_ts):
        sub = eq[eq.index >= start_ts]
        base = eq[eq.index < start_ts]
        base_v = float(base.iloc[-1]) if len(base) else float(eq.iloc[0])
        return round(float(sub.iloc[-1]) - base_v, 2) if len(sub) else 0.0
    income = {
        '1D': {'net_income': _pnl_window(ts)},
        '1W': {'net_income': _pnl_window(ts - pd.Timedelta(days=6))},
        'MTD': {'net_income': _pnl_window(ts.replace(day=1))},
        'ITD': {'net_income': round(nav - INITIAL_CASH, 2)},
    }
    capital_statement = {
        'initial_capital': INITIAL_CASH,
        'cumulative_realized': a['cumulative_realized'],
        'cumulative_dividends': a['cumulative_dividends'],
        'cumulative_fees': a['cumulative_fees'],
        'unrealized': a['unrealized'],
        'equity': nav,
    }
    rows_upto = [r for r in ledger_rows if r.get('date', '') <= day]
    cash_flow = {
        'buys_total': round(sum(r['gross'] for r in rows_upto if r['side'] == 'BUY'), 2),
        'sells_total': round(sum(r['gross'] for r in rows_upto if r['side'] == 'SELL'), 2),
        'dividends_total': a['cumulative_dividends'],
        'fees_total': -a['cumulative_fees'],
    }

    # 压力测试（beta 传导 + 单一标的冲击）
    beta_main = list(betas.values())[0] if betas else 1.0
    stress = {
        'market_-5pct': round(invested * beta_main * -0.05, 2),
        'market_-10pct': round(invested * beta_main * -0.10, 2),
        'top1_-20pct': round((top[0][1] * nav) * -0.20, 2) if top else 0.0,
    }

    m = {
        'report_type': f'{label} risk report',
        'signal_date': day,
        'generated_at': pd.Timestamp.now().isoformat(timespec='seconds'),
        'scope': f'{label} long-only book (portfolio_ledger)',
        'summary': summary,
        'balance_sheet': balance_sheet,
        'income_statement': income,
        'capital_statement': capital_statement,
        'cash_flow': cash_flow,
        'positions': [{'ticker': t, 'shares': a['positions'][t]['shares'],
                       'avg_cost': a['positions'][t]['avg_cost'],
                       'market_value': round(mv.get(t, 0), 2),
                       'weight_pct': round(weights.get(t, 0) * 100, 2)}
                      for t in sorted(a['positions'])],
        'stress': stress,
        'limits': limits,
        'diagnostics': {
            'price_source': 'Polygon store (strategy-isolated)',
            'ledger_identity_residual': round(
                nav - a['initial_cash'] - (a['cumulative_realized']
                                           + a['cumulative_dividends']
                                           - a['cumulative_fees'] + a['unrealized']), 2),
            'no_lookahead': True,
        },
    }

    out_dir = os.path.join(_cfg(strategy)["dir"], "trading_signals", "risk_management")
    os.makedirs(out_dir, exist_ok=True)
    stem = day.replace('-', '')
    paths = []

    # ── TXT（mirror pairs banner 格式）─────────────────────────────────────
    _f = lambda v, spec=',.0f': ('N/A' if v is None else format(v, spec))
    beta_str = '  '.join(f'{b}={v}' for b, v in betas.items()) or 'N/A'
    lines = [
        '=' * 78,
        f'  SOMEO PARK — {label} 组合风险管理报告 (RISK MANAGEMENT REPORT)',
        f'  Signal date: {day}   Generated: {m["generated_at"]}',
        f'  Scope: {label} long-only book   |   portfolio_ledger, Polygon-priced',
        '=' * 78,
        '',
        '── RISK SUMMARY ' + '─' * 58,
        f'  NAV:                ${nav:,.0f}',
        f'  投资比例:           {summary["invested_pct"]}%      现金: {summary["cash_pct"]}%',
        f'  Beta:               {beta_str}',
        f'  VaR 95% 1d:         ${_f(summary["var_95_1d"])} ({_f(summary["var_95_1d_pct"], ".2f")}% of NAV)',
        f'  Annual vol:         {summary["annual_vol_pct"]}% of NAV',
        f'  Sharpe (annual):    {summary["sharpe_annual"]}    CDaR 95%: {summary["cdar_95_pct"]}%',
        f'  Kelly: full={summary["kelly_full"]}x  half={summary["kelly_half"]}x  actual={summary["kelly_actual_leverage"]}x',
        f'  Max single-name:    {summary["max_single_name"]} ({summary["max_single_name_pct"]}% of NAV)',
        f'  Positions:          {summary["n_positions"]}      Limit breaches: {summary["n_breaches"]}'
        + (f'  (worst: {summary["worst_breach"]})' if summary['worst_breach'] else ''),
        '',
        '── BALANCE SHEET (资产负债表, T) ' + '─' * 42,
        f'  Cash:               ${a["cash"]:,.0f}',
        f'  Long securities:    ${a["position_value"]:,.0f}',
        f'  Total assets:       ${a["cash"] + a["position_value"]:,.0f}',
        f'  Total liabilities:  $0  (long-only, 无保证金/无空头)',
        f'  NAV (equity):       ${nav:,.0f}    (balance check: '
        f'{balance_sheet["T"]["balance_check"]})',
        '',
        '── LIMIT MONITOR ' + '─' * 57,
    ]
    for l in limits:
        lines.append(f'     {l["name"]:<20} {l["scope"]:<8} = {l["value"]}'
                     f' (amber {l["amber"]}, red {l["red"]}) [{l["status"]}]')
    lines += ['', '── STRESS ' + '─' * 64]
    for k, v in stress.items():
        lines.append(f'     {k:<18} → {money(v, color=False)}')
    tp = os.path.join(out_dir, f'risk_report_{stem}.txt')
    with open(tp, 'w') as f:
        f.write('\n'.join(lines) + '\n')
    paths.append(tp)

    # ── PDF（mirror RiskManager 章节）──────────────────────────────────────
    pp = os.path.join(out_dir, f'risk_report_{stem}.pdf')
    doc = SimpleDocTemplate(pp, pagesize=A4, topMargin=1.3 * cm, bottomMargin=1.3 * cm,
                            leftMargin=1.5 * cm, rightMargin=1.5 * cm)
    story = [Paragraph(f'Someo Park — {label} 组合风险管理报告', title_style),
             Paragraph(f'Signal date: {day}｜Scope: {label} long-only book｜'
                       f'portfolio_ledger（Polygon 定价）', sub_style)]
    story.append(Paragraph('风险摘要', h1_style))
    story.append(make_kv_table([
        ['NAV', f"${nav:,.2f}"],
        ['投资比例 / 现金比例', f"{summary['invested_pct']}% / {summary['cash_pct']}%"],
        ['Beta', beta_str],
        ['VaR 95% (1d)', f"${_f(summary['var_95_1d'])}（{_f(summary['var_95_1d_pct'], '.2f')}% NAV）"],
        ['年化波动', f"{summary['annual_vol_pct']}%"],
        ['年化 Sharpe / CDaR95', f"{summary['sharpe_annual']} / {summary['cdar_95_pct']}%"],
        ['Kelly full/half/actual',
         f"{summary['kelly_full']}x / {summary['kelly_half']}x / {summary['kelly_actual_leverage']}x"],
        ['最大单一标的', f"{summary['max_single_name']}（{summary['max_single_name_pct']}%）"],
        ['持仓数 / 限额告警', f"{summary['n_positions']} / {summary['n_breaches']}"
         + (f"（worst: {summary['worst_breach']}）" if summary['worst_breach'] else '')],
    ]))
    story.append(Paragraph('一、资产负债表（T / T-1D / T-1W / T-1M，真实账本快照）', h1_style))
    cols = [k for k in ('T', 'T-1D', 'T-1W', 'T-1M') if k in balance_sheet]
    data = [[H('科目')] + [H(k, 'RIGHT') for k in cols]]
    for field, zh in [('cash', '现金'), ('long_securities', '多头证券'),
                      ('total_assets', '总资产'), ('total_liabilities', '总负债'),
                      ('nav', 'NAV (equity)')]:
        data.append([C(zh)] + [C(f"{balance_sheet[k][field]:,.0f}", 'RIGHT') for k in cols])
    data.append([C('as_of')] + [C(balance_sheet[k]['as_of'], 'RIGHT') for k in cols])
    story.append(make_table(data, [3.6*cm] + [3.2*cm] * len(cols)))
    story.append(Paragraph('二、利润表', h1_style))
    data = [[H('窗口'), H('净收益', 'RIGHT')]]
    for k in ('1D', '1W', 'MTD', 'ITD'):
        data.append([C(k), Paragraph(money(income[k]['net_income']),
                                     ParagraphStyle('_i', fontName=FONT, fontSize=7.5,
                                                    leading=10.5, alignment=TA_RIGHT))])
    story.append(make_table(data, [4.0*cm, 5.0*cm]))
    story.append(Paragraph('三、资本与现金流', h1_style))
    story.append(make_kv_table([
        ['初始资本', f"${INITIAL_CASH:,.0f}"],
        ['累计已实现（交易）', money(a['cumulative_realized'])],
        ['累计分红', money(a['cumulative_dividends'])],
        ['累计费用', money(-a['cumulative_fees'])],
        ['未实现', money(a['unrealized'])],
        ['期间买入总额 / 卖出总额',
         f"${abs(cash_flow['buys_total']):,.0f} / ${cash_flow['sells_total']:,.0f}"],
        ['期末 equity', f"${nav:,.2f}"],
    ]))
    story.append(Paragraph('四、敞口与集中度（占 NAV %）', h1_style))
    data = [[H('Ticker'), H('股数', 'RIGHT'), H('市值', 'RIGHT'), H('权重', 'RIGHT')]]
    for t, w in top:
        data.append([C(t), C(f"{a['positions'][t]['shares']:,}", 'RIGHT'),
                     C(f"{mv[t]:,.0f}", 'RIGHT'), C(f"{w * 100:.2f}%", 'RIGHT')])
    data.append([C('现金'), C(''), C(f"{a['cash']:,.0f}", 'RIGHT'),
                 C(f"{summary['cash_pct']}%", 'RIGHT')])
    story.append(make_table(data, [3.2*cm, 3.2*cm, 3.6*cm, 3.0*cm]))
    story.append(Paragraph('五、风险价值 VaR 与压力测试', h1_style))
    story.append(make_kv_table([
        ['VaR 95% (1d, 60d 历史)', f"${_f(summary['var_95_1d'])}"],
        ['VaR 99% (1d)', f"${_f(summary['var_99_1d'])}"],
        ['市场 −5%（beta 传导）', money(stress['market_-5pct'])],
        ['市场 −10%', money(stress['market_-10pct'])],
        ['最大单一标的 −20%', money(stress['top1_-20pct'])],
    ]))
    story.append(Paragraph('六、限额监控', h1_style))
    data = [[H('限额'), H('当前值', 'RIGHT'), H('Amber', 'RIGHT'), H('Red', 'RIGHT'), H('状态')]]
    for l in limits:
        color = {'green': '#1a7a4a', 'amber': '#d4a843', 'red': '#c0392b'}[l['status']]
        data.append([C(l['name']), C(f"{l['value']}{l['fmt']}", 'RIGHT'),
                     C(f"{l['amber']}", 'RIGHT'), C(f"{l['red']}", 'RIGHT'),
                     Paragraph(f'<font color="{color}"><b>{l["status"].upper()}</b></font>',
                               ParagraphStyle('_l', fontName=FONT, fontSize=7.5, leading=10.5))])
    story.append(make_table(data, [4.0*cm, 2.6*cm, 2.2*cm, 2.2*cm, 2.4*cm]))
    story.append(Paragraph('七、组合诊断', h1_style))
    story.append(make_kv_table([
        ['价格源', 'Polygon store（策略隔离，禁 yfinance）'],
        ['账本恒等式残差', f"${m['diagnostics']['ledger_identity_residual']:+.2f}（应≈0）"],
        ['无前视', '本报告只用 ≤signal date 的账本/价格数据'],
        ['已实现口径', 'trade_ledger 定格值（成交日收盘），跨报告不漂移'],
    ]))
    doc.build(story)
    paths.append(pp)

    # ── JSON ────────────────────────────────────────────────────────────────
    m['pdf_path'] = pp
    jp = os.path.join(out_dir, f'risk_report_{stem}.json')
    with open(jp, 'w') as f:
        json.dump(m, f, indent=2, ensure_ascii=False)
    paths.append(jp)

    # ── XLSX ────────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'summary'
    for k, v in summary.items():
        ws.append([k, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v])
    ws2 = wb.create_sheet('balance_sheet')
    ws2.append(['col'] + cols)
    for field in ('as_of', 'cash', 'long_securities', 'total_assets',
                  'total_liabilities', 'nav'):
        ws2.append([field] + [balance_sheet[k][field] for k in cols])
    ws3 = wb.create_sheet('positions')
    ws3.append(['ticker', 'shares', 'avg_cost', 'market_value', 'weight_pct'])
    for p in m['positions']:
        ws3.append([p['ticker'], p['shares'], p['avg_cost'], p['market_value'],
                    p['weight_pct']])
    ws4 = wb.create_sheet('limits')
    ws4.append(['name', 'value', 'amber', 'red', 'status'])
    for l in limits:
        ws4.append([l['name'], l['value'], l['amber'], l['red'], l['status']])
    ws5 = wb.create_sheet('equity')
    ws5.append(['date', 'equity'])
    for d_, v in eq.items():
        ws5.append([str(d_.date()), v])
    xp = os.path.join(out_dir, f'risk_workbook_{stem}.xlsx')
    wb.save(xp)
    paths.append(xp)
    return paths


# ═══════════════════════════════════════════════════════════════════════════
# 入口
# ═══════════════════════════════════════════════════════════════════════════

def run(strategy: str, day: str | None = None, backfill: bool = False) -> int:
    acct_series = load_account_series(strategy)
    if not acct_series:
        log.warning(f"[{strategy}] 无 account_history —— 先 replay")
        return 0
    ledger_rows = load_ledger_rows(strategy)
    tickers = sorted({t for a in acct_series.values() for t in a["positions"]}
                     | {r.get('ticker') for r in ledger_rows if r.get('ticker')})
    tickers = [t for t in tickers if t]
    prices = load_store_prices(strategy, tickers)
    opens = load_store_open(strategy, tickers)
    # 2026-08-31 泛化:各策略基准在各自 store(AEUS 的 XLU/GRID 在 elec_strategy);
    # SSRS 沿用历史遗留(其 SPY/SMH 在 semi store),aiss 行为逐位不变。
    _bench_store = strategy if strategy in ("aiss", "aeus") else BENCH_STORE_STRATEGY
    bench = load_store_prices(_bench_store, _cfg(strategy)["benchmarks"])
    days = sorted(acct_series) if backfill else [day or max(acct_series)]
    n = 0
    for d in days:
        generate_pnl_report(strategy, d, acct_series, prices, opens, ledger_rows, bench)
        generate_risk_report(strategy, d, acct_series, prices, bench, ledger_rows)
        n += 1
    log.info(f"[{strategy}] 生成 {n} 天 × (pnl.pdf + risk json/txt/pdf/xlsx)")
    return n


def main():
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    args = [a for a in sys.argv[1:] if not a.startswith("-")]
    backfill = "--backfill" in sys.argv
    day = None
    if "--date" in sys.argv:
        day = sys.argv[sys.argv.index("--date") + 1]
    targets = list(STRATEGIES) if (not args or args[0] == "all") else [args[0]]
    for s in targets:
        run(s, day=day, backfill=backfill)


if __name__ == "__main__":
    main()
